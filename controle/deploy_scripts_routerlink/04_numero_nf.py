# Fase 2.7 do runbook (DEPLOY_PENDENTE_ROUTERLINK.md).
#
# Roda com (dentro da pasta controle, produção):
#   venv\Scripts\python.exe manage.py shell -c "exec(open(r'deploy_scripts_routerlink\04_numero_nf.py', encoding='utf-8').read())"
#
# Lê a planilha CRUA do fornecedor (não a de análise) — abas com colunas
# Part Number / Serial / NFe / Valor Mensal / PMB, cabeçalho na linha 4.
# Casa por Item.numero_serie e preenche Item.numero_nf. NUNCA sobrescreve
# um numero_nf já preenchido que seja diferente do da planilha — se achar
# um conflito, avisa e pula (não aplica), pra você decidir manualmente.
#
# AJUSTE ANTES DE RODAR: o caminho do arquivo abaixo (ARQUIVO) deve apontar
# pra planilha real recebida do fornecedor nesta leva de produção.
#
# COMO USAR: 1) rode com APLICAR=False, confira o resumo (bateram, sem NF,
# conflitos); 2) troque para APLICAR=True e rode de novo.

import openpyxl
from ProjetoEstoque.models import Item

ARQUIVO = r"C:\caminho\para\RELAÇÃO DE EQUIPAMENTOS CORRETA - FORNECEDOR.xlsx"
FORNECEDOR = "Routerlink"
LINHA_CABECALHO = 4  # primeira linha de dado = LINHA_CABECALHO + 1
COL_PART_NUMBER = 2
COL_SERIAL = 3
COL_NFE = 4
APLICAR = False

wb = openpyxl.load_workbook(ARQUIVO, data_only=True)
planilha = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(LINHA_CABECALHO + 1, ws.max_row + 1):
        serial = ws.cell(row=r, column=COL_SERIAL).value
        nfe = ws.cell(row=r, column=COL_NFE).value
        if not serial or not nfe:
            continue
        chave = str(serial).strip().upper()
        if str(serial).strip().lower() == "serial":  # linha de cabeçalho duplicada dentro dos dados
            continue
        planilha[chave] = str(nfe).strip()

print(f"Linhas válidas na planilha (todas as abas): {len(planilha)}")

itens = list(Item.objects.filter(fornecedor__nome__iexact=FORNECEDOR).exclude(numero_serie__isnull=True).exclude(numero_serie=""))
atualizar = []
conflitos = []
ja_igual = 0
for it in itens:
    chave = it.numero_serie.strip().upper()
    if chave not in planilha:
        continue
    nfe_planilha = planilha[chave]
    if it.numero_nf:
        if it.numero_nf.strip() == nfe_planilha:
            ja_igual += 1
        else:
            conflitos.append((it.pk, it.numero_serie, it.numero_nf, nfe_planilha))
    else:
        it.numero_nf = nfe_planilha
        atualizar.append(it)

print(f"Já tinham NF igual (nada a fazer): {ja_igual}")
print(f"Sem NF no banco — serão preenchidos: {len(atualizar)}")
print(f"CONFLITOS (NF diferente já cadastrada — NÃO tocados, revisar manualmente): {len(conflitos)}")
for c in conflitos[:20]:
    print(f"  ID {c[0]} | série {c[1]} | banco={c[2]!r} | planilha={c[3]!r}")

if APLICAR:
    Item.objects.bulk_update(atualizar, ["numero_nf"])
    print(f"\nAPLICADO — {len(atualizar)} itens atualizados.")
else:
    print("\nDRY-RUN — nada alterado. Troque APLICAR=True e rode de novo depois de conferir.")
