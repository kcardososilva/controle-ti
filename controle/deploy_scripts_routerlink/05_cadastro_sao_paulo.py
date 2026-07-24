# Fase 2.8 do runbook (DEPLOY_PENDENTE_ROUTERLINK.md).
#
# Roda com (dentro da pasta controle, produção):
#   venv\Scripts\python.exe manage.py shell -c "exec(open(r'deploy_scripts_routerlink\05_cadastro_sao_paulo.py', encoding='utf-8').read())"
#
# Cadastra os equipamentos da aba "São Paulo" da planilha crua do fornecedor
# (Part Number/Serial/NFe/Valor Mensal/PMB, cabeçalho linha 4) — filial que
# nunca tinha sido comparada contra o sistema. Cria Item + Locacao completos
# (Fornecedor, Localidade, Centro de Custo, status Backup, NF, tempo de
# contrato, data de entrada inferida). NÃO define Subtipo aqui — rode o
# script 03_subtipo.py DEPOIS deste pra preencher (evita duplicar a lógica
# de mapeamento em dois lugares).
#
# AJUSTE ANTES DE RODAR:
#   - ARQUIVO: caminho da planilha real desta leva de produção.
#   - LOCALIDADE_NOME: confirme que "TI - SP" é o nome exato cadastrado em
#     produção (o usuário confirmou esse valor especificamente para dev —
#     confira de novo lá, pode ter grafia diferente).
#
# COMO USAR: 1) rode com APLICAR=False, confira quantos seriam criados e se
# nenhum já existe no banco; 2) troque para APLICAR=True e rode de novo.

import unicodedata
from collections import Counter, defaultdict
from decimal import Decimal

from django.db import transaction
from ProjetoEstoque.models import Item, Locacao, Fornecedor, CentroCusto, Localidade

ARQUIVO = r"C:\caminho\para\RELAÇÃO DE EQUIPAMENTOS CORRETA - FORNECEDOR.xlsx"
ABA = "São Paulo"
LINHA_CABECALHO = 4
COL_PART_NUMBER, COL_SERIAL, COL_NFE, COL_VALOR, COL_PMB = 2, 3, 4, 5, 6

FORNECEDOR_NOME = "Routerlink"
CC_NUMERO, CC_DEPARTAMENTO = "12105", "TI"
LOCALIDADE_NOME = "TI - SP"
TEMPO_MESES = 36
STATUS_INICIAL = "backup"
APLICAR = False


def norm(s):
    if not s:
        return ""
    s = str(s).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


import openpyxl
ws = openpyxl.load_workbook(ARQUIVO, data_only=True)[ABA]
linhas = []
for r in range(LINHA_CABECALHO + 1, ws.max_row + 1):
    part = ws.cell(row=r, column=COL_PART_NUMBER).value
    serial = ws.cell(row=r, column=COL_SERIAL).value
    nfe = ws.cell(row=r, column=COL_NFE).value
    valor = ws.cell(row=r, column=COL_VALOR).value
    pmb = ws.cell(row=r, column=COL_PMB).value
    if not serial or str(serial).strip().lower() == "serial":
        continue
    linhas.append({
        "part": str(part).strip(), "serial": str(serial).strip(),
        "nfe": str(nfe).strip() if nfe else None, "valor": valor, "pmb": pmb,
    })
print(f"Linhas válidas na aba {ABA!r}: {len(linhas)}")

ja_existe = [l for l in linhas if Item.all_objects.filter(numero_serie=l["serial"]).exists()]
print(f"Já cadastrados no banco (não deveria haver — CONFIRME antes de aplicar): {len(ja_existe)}")
for l in ja_existe:
    print(f"  {l['serial']} ({l['part']})")

fornecedor = Fornecedor.objects.filter(nome__iexact=FORNECEDOR_NOME).first()
cc = CentroCusto.objects.filter(numero=CC_NUMERO, departamento=CC_DEPARTAMENTO).first()
localidade = Localidade.objects.filter(local=LOCALIDADE_NOME).first()

if not fornecedor:
    print(f"ERRO: Fornecedor {FORNECEDOR_NOME!r} não encontrado.")
if not cc:
    print(f"ERRO: CentroCusto numero={CC_NUMERO!r} departamento={CC_DEPARTAMENTO!r} não encontrado.")
if not localidade:
    print(f"ERRO: Localidade {LOCALIDADE_NOME!r} não encontrada — confira o nome exato em produção.")

# Base p/ inferir data_entrada: todas as locações Routerlink já existentes com data preenchida.
base = Locacao.objects.filter(equipamento__fornecedor__nome__iexact=FORNECEDOR_NOME, data_entrada__isnull=False).select_related("equipamento")
por_modelo_data = defaultdict(Counter)
global_data = Counter()
for loc in base:
    por_modelo_data[norm(loc.equipamento.modelo)][loc.data_entrada] += 1
    global_data[loc.data_entrada] += 1
data_fallback = global_data.most_common(1)[0][0] if global_data else None
print(f"\nData de entrada — fallback global (mais comum na base Routerlink): {data_fallback}")

pronto_para_aplicar = fornecedor and cc and localidade and not ja_existe and data_fallback
if not pronto_para_aplicar:
    print("\nATENÇÃO: faltando pré-requisito acima — corrija antes de marcar APLICAR=True.")

if APLICAR and pronto_para_aplicar:
    criados = []
    with transaction.atomic():
        for l in linhas:
            chave_modelo = norm(l["part"])
            data_entrada = por_modelo_data[chave_modelo].most_common(1)[0][0] if chave_modelo in por_modelo_data else data_fallback
            pmb = "sim" if (l["pmb"] and norm(l["pmb"]) == "pmb") or "pmb" in norm(l["part"]) else "nao"

            item = Item(
                nome=l["part"], modelo=l["part"], numero_serie=l["serial"], numero_nf=l["nfe"],
                fornecedor=fornecedor, locado="sim", status=STATUS_INICIAL, pmb=pmb,
                centro_custo=cc, localidade=localidade,
                observacoes=f"Cadastrado a partir de RELAÇÃO DE EQUIPAMENTOS CORRETA - FORNECEDOR.xlsx (aba {ABA}).",
            )
            item._locacao_data_inicio_override = data_entrada
            item.full_clean()
            item.save()

            locacao = Locacao(
                equipamento=item,
                valor_mensal=Decimal(str(l["valor"])).quantize(Decimal("0.01")) if l["valor"] is not None else None,
                tempo_locado=TEMPO_MESES, data_entrada=data_entrada, fornecedor=fornecedor,
            )
            locacao.full_clean()
            locacao.save()
            criados.append(item)

    print(f"\nAPLICADO — {len(criados)} itens criados. Rode 03_subtipo.py em seguida.")
else:
    print(f"\nDRY-RUN — {len(linhas)} itens SERIAM criados. Nada foi alterado.")
