# Categoria - CRUD

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import DateField, DateTimeField, IntegerField, BigIntegerField, CharField, Sum, Count, Case, When, Value as V, ExpressionWrapper, Window, DecimalField
from django.db.models import Q, Count, Prefetch, F, Sum, DecimalField, Value, CharField,Exists
from django.core.paginator import Paginator
from django.core.exceptions import FieldError
from django.db.models.functions import TruncMonth, Coalesce, Concat, Cast
from collections import defaultdict
from django.db.models import OuterRef, Subquery
from datetime import date, timedelta, datetime
from django.db import transaction
from django.utils import timezone
from django.conf import settings
from django.views.decorators.http import require_POST
from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db import models
import unicodedata
from django.db.models.functions import Coalesce, NullIf
from .models import Categoria, Subtipo, Usuario,PreventivaExecucao,Fornecedor, Localidade, Funcao, CentroCusto, Item, Locacao, Comentario, MovimentacaoItem, CicloManutencao, StatusItemChoices, LocalidadeChoices, SimNaoChoices, StatusUsuarioChoices, CheckListModelo, CheckListPergunta, Preventiva, PreventivaResposta, TipoRespostaChoices, SimNaoChoices, TipoMovimentacaoChoices, Licenca, MovimentacaoLicenca, TipoMovLicencaChoices, PeriodicidadeChoices, LicencaLote
from django.utils.dateparse import parse_date
from .forms import CategoriaForm, SubtipoForm, UsuarioForm, FornecedorForm, LocalidadeForm, FuncaoForm, CentroCustoForm, ItemForm, LocacaoForm, LicencaForm, ComentarioForm, MovimentacaoItemForm, CicloManutencaoForm, PreventivaStartForm, ChecklistModeloForm, ChecklistPerguntaForm, PreventivaStartForm, LicencaForm, MovimentacaoLicencaForm, LicencaLoteForm
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.utils.timezone import now
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import io
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.core.exceptions import ValidationError
from django.template.loader import get_template
from xhtml2pdf import pisa


# --- Helpers de compatibilidade com nomes de campos diferentes no LicencaLote ---
def _lote_total_get(lote):
    for n in ("qtd_total", "quantidade_total", "quantidade", "total", "qtd"):
        if hasattr(lote, n):
            return getattr(lote, n) or 0
    return 0

def _lote_total_set(lote, val):
    for n in ("qtd_total", "quantidade_total", "quantidade", "total", "qtd"):
        if hasattr(lote, n):
            setattr(lote, n, val); return

def _lote_total_fieldname(lote_or_cls):
    # usado para apontar erro no form
    for n in ("qtd_total", "quantidade_total", "quantidade", "total", "qtd"):
        if hasattr(lote_or_cls, n):
            return n
    return None

def _lote_disp_get(lote):
    for n in ("disponivel", "qtd_disponivel", "quantidade_disponivel", "disponiveis", "saldo", "em_estoque"):
        if hasattr(lote, n):
            return getattr(lote, n) or 0
    return 0

def _lote_disp_set(lote, val):
    for n in ("disponivel", "qtd_disponivel", "quantidade_disponivel", "disponiveis", "saldo", "em_estoque"):
        if hasattr(lote, n):
            setattr(lote, n, val); return
############### CATEGORIA ##############################

def categorias_list(request):
    categorias = Categoria.objects.all()
    return render(request, 'categoria/list.html', {'categorias': categorias})

def categoria_create(request):
    form = CategoriaForm(request.POST or None)
    if form.is_valid():
        categoria = form.save(commit=False)
        categoria.criado_por = request.user
        categoria.atualizado_por = request.user
        categoria.save()
        return redirect('categorias_list')
    return render(request, 'categoria/form.html', {'form': form})

def categoria_update(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    form = CategoriaForm(request.POST or None, instance=categoria)
    if form.is_valid():
        categoria = form.save(commit=False)
        categoria.atualizado_por = request.user
        categoria.save()
        return redirect('categorias_list')
    return render(request, 'categoria/form.html', {'form': form})

def categoria_delete(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        return redirect('categorias_list')
    return render(request, 'categoria/delete.html', {'obj': categoria})


############### SUBTIPO ##############################

@login_required
def subtipo_list(request):
    q = request.GET.get("q", "").strip()
    cat = request.GET.get("categoria", "").strip()
    alocado = request.GET.get("alocado", "").strip()

    qs = Subtipo.objects.select_related("categoria").order_by("categoria__nome", "nome")

    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(categoria__nome__icontains=q))
    if cat:
        qs = qs.filter(categoria__id=cat)
    if alocado:
        qs = qs.filter(alocado=alocado)

    context = {
        "subtipos": qs,
        "categorias": Categoria.objects.order_by("nome"),
        "alocado_choices": (("sim","Sim"),("nao","Não")),
        "request": request,  # para manter valores nos filtros
    }
    return render(request, "front/subtipo_list.html", context)

@login_required
def subtipo_create(request):
    if request.method == "POST":
        form = SubtipoForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Subtipo cadastrado com sucesso!")
            return redirect("subtipo_list")
        messages.error(request, "Verifique os campos destacados.")
    else:
        form = SubtipoForm()

    return render(request, "front/subtipo_form.html", {"form": form, "editar": False})

@login_required
def subtipo_update(request, pk):
    obj = get_object_or_404(Subtipo, pk=pk)
    if request.method == "POST":
        form = SubtipoForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Subtipo atualizado com sucesso!")
            return redirect("subtipo_list")
        messages.error(request, "Verifique os campos destacados.")
    else:
        form = SubtipoForm(instance=obj)

    return render(request, "front/subtipo_form.html", {"form": form, "editar": True, "obj": obj})

@login_required
def subtipo_delete(request, pk):
    obj = get_object_or_404(Subtipo, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Subtipo excluído com sucesso!")
        return redirect("subtipo_list")
    return render(request, "front/subtipo_confirm_delete.html", {"obj": obj})

@login_required
def subtipo_detail(request, pk):
    obj = get_object_or_404(Subtipo.objects.select_related("categoria"), pk=pk)
    return render(request, "front/subtipo_detail.html", {"obj": obj})


############### USUÁRIO ##############################

@login_required
def usuario_list(request):
    """
    Listagem de Usuários (Google Material Design).
    """
    # 1. Filtros
    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    pmb = request.GET.get("pmb", "").strip()
    cc = request.GET.get("cc", "").strip()
    loc = request.GET.get("loc", "").strip()
    func = request.GET.get("func", "").strip()

    qs = Usuario.objects.select_related("centro_custo", "localidade", "funcao").order_by("nome")

    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(email__icontains=q))
    if status:
        qs = qs.filter(status=status)
    if pmb:
        qs = qs.filter(pmb=pmb)
    if cc and cc.isdigit():
        qs = qs.filter(centro_custo_id=int(cc))
    if loc and loc.isdigit():
        qs = qs.filter(localidade_id=int(loc))
    if func and func.isdigit():
        qs = qs.filter(funcao_id=int(func))

    total_filtrado = qs.count()

    # 2. KPIs (Totais Globais)
    # Importante: Calculados sobre o total (sem filtros) para dar contexto geral ao gestor
    kpi_total = Usuario.objects.count()
    kpi_ativos = Usuario.objects.filter(status=StatusUsuarioChoices.ATIVO).count()
    kpi_desligados = Usuario.objects.filter(status=StatusUsuarioChoices.DESLIGADO).count()
    kpi_pmb = Usuario.objects.filter(pmb=SimNaoChoices.SIM).count()

    # 3. Paginação
    try:
        per_page = int(request.GET.get("pp", 20))
    except ValueError:
        per_page = 20

    paginator = Paginator(qs, per_page)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    get_copy = request.GET.copy()
    if "page" in get_copy: del get_copy["page"]
    qs_keep = get_copy.urlencode()

    context = {
        "usuarios": page_obj.object_list,
        "page_obj": page_obj,
        "total": total_filtrado,
        "qs_keep": qs_keep,
        
        # KPIs completos
        "kpi_total": kpi_total,
        "kpi_ativos": kpi_ativos,
        "kpi_desligados": kpi_desligados,
        "kpi_pmb": kpi_pmb,
        
        # Filtros
        "f_q": q, "f_status": status, "f_pmb": pmb, 
        "f_cc": int(cc) if cc.isdigit() else "",
        "f_loc": int(loc) if loc.isdigit() else "",
        "f_func": int(func) if func.isdigit() else "",

        # Listas
        "opt_status": StatusUsuarioChoices.choices,
        "opt_pmb": SimNaoChoices.choices,
        "opt_cc": CentroCusto.objects.values("id", "numero", "departamento").order_by("numero"),
        "opt_loc": Localidade.objects.values("id", "local").order_by("local"),
        "opt_func": Funcao.objects.values("id", "nome").order_by("nome"),
    }

    return render(request, "front/usuarios/usuario_list.html", context)
# CREATE
@login_required
def usuario_create(request):
    if request.method == "POST":
        form = UsuarioForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Usuário criado com sucesso!")
            return redirect("usuario_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = UsuarioForm()
    return render(request, "front/usuarios/usuario_form.html", {"form": form, "editar": False})

@login_required
def usuario_update(request, pk: int):
    obj = get_object_or_404(Usuario, pk=pk)
    if request.method == "POST":
        form = UsuarioForm(request.POST, instance=obj)
        if form.is_valid():
            sobj = form.save(commit=False)
            sobj.atualizado_por = request.user
            sobj.save()
            messages.success(request, "Usuário atualizado com sucesso!")
            return redirect("usuario_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = UsuarioForm(instance=obj)
    return render(request, "front/usuarios/usuario_form.html", {"form": form, "editar": True})


# DETAIL
@login_required
def usuario_detail(request, pk):
    """
    Dashboard detalhado do Usuário (Ativos e Licenças).
    """
    usuario = get_object_or_404(
        Usuario.objects.select_related("centro_custo", "localidade", "funcao"),
        pk=pk
    )

    # =========================================================
    # 1. ITENS (HARDWARE) - Lógica de posse atual
    # =========================================================
    # (Mantendo a lógica existente de itens que você já usa, apenas otimizando)
    movs_itens = (
        MovimentacaoItem.objects
        .select_related("item", "item__subtipo", "item__localidade", "item__centro_custo")
        .filter(item__isnull=False)
        .order_by("item_id", "-created_at", "-id")
    )
    
    # Dicionário para pegar apenas a última movimentação de cada item
    itens_ativos = []
    itens_processados = set()
    
    # Nota: Em produção com muitos dados, ideal filtrar movs apenas dos itens relevantes
    # Aqui simplificado para o exemplo contextua.
    for mov in movs_itens:
        if mov.item_id in itens_processados:
            continue
        itens_processados.add(mov.item_id)
        
        # Regra de Posse: A última movimentação deve ser para este usuário
        # e não pode ser uma "Baixa" ou "Devolução ao Estoque"
        if mov.usuario_id == usuario.pk and mov.tipo_movimentacao not in ['baixa', 'devolucao']:
            item = mov.item
            # Calcula custo do item
            custo_item = item.valor or Decimal(0)
            tipo_custo = "aquisicao"
            
            # Se for locado, tenta pegar valor da locação
            if getattr(item, 'locado', 'nao') == 'sim':
                try:
                    loc = item.locacao
                    if loc and loc.valor_mensal:
                        custo_item = loc.valor_mensal
                        tipo_custo = "locacao"
                except Exception: pass
            
            item.custo_calc = custo_item
            item.tipo_custo_calc = tipo_custo
            itens_ativos.append(item)

    # =========================================================
    # 2. LICENÇAS (SOFTWARE) - Lógica de Atribuição Ativa
    # =========================================================
    # Busca todas as movs envolvendo este usuário, ordenadas da mais recente
    movs_lic = (
        MovimentacaoLicenca.objects
        .filter(usuario=usuario)
        .select_related("licenca", "licenca__fornecedor", "lote")
        .order_by("licenca_id", "-created_at", "-id")
    )

    licencas_ativas = []
    licencas_processadas = set()

    total_lic_mensal = Decimal(0)
    total_lic_anual = Decimal(0)

    for mov in movs_lic:
        if mov.licenca_id in licencas_processadas:
            continue
        licencas_processadas.add(mov.licenca_id)

        # Regra: Só é ativo se o ÚLTIMO movimento for ATRIBUIÇÃO
        if mov.tipo == TipoMovLicencaChoices.ATRIBUICAO:
            lic = mov.licenca
            lote = mov.lote
            
            # Cálculos Financeiros baseados no Snapshot da Movimentação (valor_unitario)
            # O valor_unitario gravado é o CUSTO DO CICLO (ex: 50 mensal ou 600 anual)
            custo_base = mov.valor_unitario or Decimal(0)
            
            custo_mensal = Decimal(0)
            custo_anual = Decimal(0)
            periodicidade = ""

            # Tenta pegar periodicidade do lote (mais preciso) ou da licença
            if lote:
                periodicidade = str(lote.periodicidade).lower()
            else:
                periodicidade = str(lic.periodicidade).lower()

            # Projeção
            if periodicidade == 'anual':
                custo_mensal = custo_base / Decimal(12)
                custo_anual = custo_base
            elif periodicidade == 'semestral':
                custo_mensal = custo_base / Decimal(6)
                custo_anual = custo_base * 2
            elif periodicidade == 'trimestral':
                custo_mensal = custo_base / Decimal(3)
                custo_anual = custo_base * 4
            else:
                # Mensal (Padrão)
                custo_mensal = custo_base
                custo_anual = custo_base * 12

            total_lic_mensal += custo_mensal
            total_lic_anual += custo_anual

            licencas_ativas.append({
                "licenca": lic,
                "lote": lote,
                "data_atribuicao": mov.created_at,
                "custo_mensal": custo_mensal,
                "custo_anual": custo_anual,
                "custo_base": custo_base,
                "periodicidade_label": lote.get_periodicidade_display() if lote else lic.get_periodicidade_display()
            })

    # =========================================================
    # 3. TOTALIZADORES GERAIS
    # =========================================================
    # Soma Itens
    total_itens_loc = sum(i.custo_calc for i in itens_ativos if i.tipo_custo_calc == 'locacao')
    total_itens_aq = sum(i.custo_calc for i in itens_ativos if i.tipo_custo_calc == 'aquisicao')

    # Soma Geral (Custo Mensal Recorrente Estimado)
    # Considera Locação de Itens + Mensalidade de Licenças
    burn_rate_total = total_itens_loc + total_lic_mensal

    context = {
        "obj": usuario,
        "itens_ativos": itens_ativos,
        "licencas_ativas": licencas_ativas,
        "kpi": {
            "itens_qtd": len(itens_ativos),
            "licencas_qtd": len(licencas_ativas),
            "custo_mensal_lic": total_lic_mensal,
            "custo_anual_lic": total_lic_anual,
            "custo_mensal_loc": total_itens_loc,
            "total_aquisicao": total_itens_aq,
            "burn_rate_total": burn_rate_total
        }
    }
    return render(request, "front/usuarios/usuario_detail.html", context)


@login_required
@transaction.atomic
def licenca_devolver_rapido(request, usuario_id, licenca_id):
    """
    Ação rápida para devolver uma licença do usuário ao estoque.
    Restaura o saldo do lote original e ajusta o centro de custo.
    """
    if request.method != "POST":
        messages.warning(request, "Ação não permitida via GET.")
        return redirect("usuario_detail", pk=usuario_id)

    usuario = get_object_or_404(Usuario, pk=usuario_id)
    
    # 1. Encontrar a atribuição ativa
    # Buscamos a última movimentação. Deve ser uma ATRIBUIÇÃO para podermos devolver.
    last_mov = MovimentacaoLicenca.objects.filter(
        usuario_id=usuario_id,
        licenca_id=licenca_id
    ).order_by('-created_at', '-id').first()

    if not last_mov or last_mov.tipo != TipoMovLicencaChoices.ATRIBUICAO:
        messages.error(request, "Não foi encontrada uma atribuição ativa desta licença para este usuário.")
        return redirect("usuario_detail", pk=usuario_id)

    # 2. Identificar Lote de Origem
    lote_origem = last_mov.lote # O lote de onde saiu
    
    # Se o lote original não existir mais (deletado?), tentamos um fallback (LIFO)
    if not lote_origem:
        lote_origem = LicencaLote.objects.filter(licenca_id=licenca_id).order_by('-data_compra').first()

    # 3. Criar Movimentação de Devolução
    nova_mov = MovimentacaoLicenca(
        tipo=TipoMovLicencaChoices.DEVOLUCAO,
        licenca_id=licenca_id,
        usuario_id=usuario_id,
        lote=lote_origem,
        criado_por=request.user,
        atualizado_por=request.user,
        # Mantém o valor histórico para estorno contábil
        valor_unitario=last_mov.valor_unitario
    )

    # 4. Restaurar Estoque e Definir Destino do Custo
    if lote_origem:
        # Trava lote para atualização segura
        lote_atual = LicencaLote.objects.select_for_update().get(pk=lote_origem.pk)
        lote_atual.quantidade_disponivel += 1
        lote_atual.save()
        
        # O custo volta para o dono do lote (ex: TI, ou Depto que comprou)
        nova_mov.centro_custo_destino = lote_atual.centro_custo
    else:
        # Se não tem lote, tenta CC da licença
        from .models import Licenca
        lic = Licenca.objects.get(pk=licenca_id)
        nova_mov.centro_custo_destino = lic.centro_custo

    nova_mov.save()

    messages.success(request, f"Licença devolvida com sucesso! O saldo retornou ao estoque.")
    return redirect("usuario_detail", pk=usuario_id)




# DELETE (POST via modal)
@login_required
def usuario_delete(request, pk: int):
    obj = get_object_or_404(Usuario, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Usuário removido com sucesso.")
    else:
        messages.error(request, "Ação inválida.")
    return redirect("usuario_list")


############### FORNECEDOR ##############################

# Helper para reutilizar filtros na listagem e no PDF
def _get_fornecedores_filtrados(request):
    q = request.GET.get("q", "").strip()
    tem_contrato = request.GET.get("tem_contrato", "").strip()

    qs = Fornecedor.objects.all().order_by("nome")

    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(cnpj__icontains=q))
    
    if tem_contrato == "sim":
        qs = qs.exclude(contrato__isnull=True).exclude(contrato__exact="")
    elif tem_contrato == "nao":
        qs = qs.filter(Q(contrato__isnull=True) | Q(contrato__exact=""))
    
    return qs, q, tem_contrato

@login_required
def fornecedor_list(request):
    """
    Dashboard de Fornecedores (Enterprise Style).
    """
    qs, q, tem_contrato = _get_fornecedores_filtrados(request)
    
    forn_ids = list(qs.values_list("id", flat=True))
    total_filtrado = qs.count()

    # --- KPIs (Baseados em Itens) ---
    itens_qs = Item.objects.filter(fornecedor_id__in=forn_ids)
    
    # 1. Agregação Global (Itens Locados geram custo)
    globais = itens_qs.filter(locado="sim").aggregate(
        total_custo=Sum("locacao__valor_mensal"),
        qtd_locados=Count("id")
    )
    kpi_custo_total = globais["total_custo"] or Decimal(0)
    kpi_itens_locados = globais["qtd_locados"] or 0
    kpi_media = (kpi_custo_total / total_filtrado) if total_filtrado > 0 else 0

    # 2. Dados por Fornecedor (Mapas para performance)
    # Total de itens (locados ou não)
    dados_total = itens_qs.values("fornecedor").annotate(qtd=Count("id"))
    mapa_total = {d["fornecedor"]: d["qtd"] for d in dados_total}

    # Dados financeiros (apenas locados)
    dados_fin = itens_qs.filter(locado="sim").values("fornecedor").annotate(
        custo=Sum("locacao__valor_mensal"), 
        qtd_loc=Count("id")
    )
    mapa_custo = {d["fornecedor"]: d["custo"] or 0 for d in dados_fin}
    mapa_locados = {d["fornecedor"]: d["qtd_loc"] or 0 for d in dados_fin}

    # 3. Top Fornecedor (Maior Custo)
    kpi_top_forn = None
    kpi_top_val = 0
    if mapa_custo:
        top_id = max(mapa_custo, key=mapa_custo.get)
        kpi_top_val = mapa_custo[top_id]
        try:
            kpi_top_forn = Fornecedor.objects.get(id=top_id)
        except Fornecedor.DoesNotExist: pass

    # --- Paginação ---
    try:
        per_page = int(request.GET.get("pp", 12))
    except ValueError:
        per_page = 12

    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_num)

    # Injeta dados calculados nos objetos da página
    for f in page_obj.object_list:
        f.qtd_total_calc = mapa_total.get(f.id, 0)
        f.qtd_locados_calc = mapa_locados.get(f.id, 0)
        f.custo_calc = mapa_custo.get(f.id, 0)

    get_copy = request.GET.copy()
    if "page" in get_copy: del get_copy["page"]
    qs_keep = get_copy.urlencode()

    context = {
        "fornecedores": page_obj.object_list,
        "page_obj": page_obj,
        "total": total_filtrado,
        "qs_keep": qs_keep,
        
        # Filtros
        "f_q": q,
        "f_contrato": tem_contrato,

        # KPIs
        "kpi_custo_total": kpi_custo_total,
        "kpi_locados": kpi_itens_locados,
        "kpi_media": kpi_media,
        "kpi_top_forn": kpi_top_forn,
        "kpi_top_val": kpi_top_val,
    }

    return render(request, "front/fornecedores/fornecedor_list.html", context)

@login_required
def fornecedor_export_pdf(request):
    qs, q, tem_contrato = _get_fornecedores_filtrados(request)
    
    # Lista completa para o PDF
    fornecedores = list(qs)
    forn_ids = [f.id for f in fornecedores]
    
    # Recalcula dados
    itens_qs = Item.objects.filter(fornecedor_id__in=forn_ids)
    
    # Mapas
    mapa_total = {d["fornecedor"]: d["qtd"] for d in itens_qs.values("fornecedor").annotate(qtd=Count("id"))}
    
    dados_fin = itens_qs.filter(locado="sim").values("fornecedor").annotate(
        custo=Sum("locacao__valor_mensal"), 
        qtd_loc=Count("id")
    )
    mapa_custo = {d["fornecedor"]: d["custo"] or 0 for d in dados_fin}
    mapa_locados = {d["fornecedor"]: d["qtd_loc"] or 0 for d in dados_fin}

    total_geral_custo = 0
    for f in fornecedores:
        f.qtd_total_calc = mapa_total.get(f.id, 0)
        f.qtd_locados_calc = mapa_locados.get(f.id, 0)
        f.custo_calc = mapa_custo.get(f.id, 0)
        total_geral_custo += f.custo_calc

    context = {
        "fornecedores": fornecedores,
        "total_geral_custo": total_geral_custo,
        "filtros": {"Busca": q, "Contrato": tem_contrato},
        "usuario": request.user,
    }
    
    template_path = 'front/fornecedores/fornecedor_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_fornecedores.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err: return HttpResponse('Erro ao gerar PDF', status=500)
    return response

# CREATE
@login_required
def fornecedor_create(request):
    if request.method == "POST":
        form = FornecedorForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Fornecedor criado com sucesso!")
            return redirect("fornecedor_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = FornecedorForm()

    return render(request, "front/fornecedores/fornecedor_form.html", {"form": form, "editar": False})


# UPDATE
@login_required
def fornecedor_update(request, pk: int):
    obj = get_object_or_404(Fornecedor, pk=pk)
    if request.method == "POST":
        form = FornecedorForm(request.POST, instance=obj)
        if form.is_valid():
            sobj = form.save(commit=False)
            sobj.atualizado_por = request.user
            sobj.save()
            messages.success(request, "Fornecedor atualizado com sucesso!")
            return redirect("fornecedor_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = FornecedorForm(instance=obj)

    return render(request, "front/fornecedores/fornecedor_form.html", {"form": form, "editar": True})


# DETAIL
@login_required
def fornecedor_detail(request, pk: int):
    """
    Dashboard detalhado do Fornecedor.
    Inclui dados cadastrais, contrato, KPIs financeiros e lista de ativos fornecidos.
    """
    obj = get_object_or_404(Fornecedor, pk=pk)
    
    # 1. Busca Itens fornecidos por este parceiro
    itens_qs = (
        Item.objects
        .filter(fornecedor=obj)
        .select_related('subtipo', 'localidade', 'centro_custo')
        .order_by('-created_at')
    )

    # 2. KPIs em Tempo Real
    total_itens = itens_qs.count()
    
    # Filtra apenas itens que são locados (custo recorrente)
    locados_qs = itens_qs.filter(locado=SimNaoChoices.SIM)
    qtd_locados = locados_qs.count()
    
    # Soma o valor mensal dos contratos de locação
    custo_mensal = locados_qs.aggregate(
        total=Sum('locacao__valor_mensal')
    )['total'] or Decimal(0)

    # Valor patrimonial (Itens comprados/próprios)
    valor_patrimonial = itens_qs.exclude(locado=SimNaoChoices.SIM).aggregate(
        total=Sum('valor')
    )['total'] or Decimal(0)

    context = {
        "obj": obj,
        "itens_fornecidos": itens_qs,
        "kpi": {
            "total_itens": total_itens,
            "qtd_locados": qtd_locados,
            "custo_mensal": custo_mensal,
            "valor_aquisicao": valor_patrimonial,
        }
    }
    
    return render(request, "front/fornecedores/fornecedor_detail.html", context)


# DELETE (POST via modal)
@login_required
def fornecedor_delete(request, pk: int):
    obj = get_object_or_404(Fornecedor, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Fornecedor removido com sucesso.")
    else:
        messages.error(request, "Ação inválida.")
    return redirect("fornecedor_list")


############### LOCALIDADE ##############################

@login_required
def localidade_list(request):
    qs = Localidade.objects.all().order_by("local")

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(Q(local__icontains=q) | Q(codigo__icontains=q))

    context = {
        "localidades": qs,
        "LocalidadeChoices": LocalidadeChoices,
        "q": q,
    }
    return render(request, "front/localidade_list.html", context)

@login_required
def localidade_create(request):
    if request.method == "POST":
        form = LocalidadeForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Localidade criada com sucesso.")
            return redirect("localidade_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = LocalidadeForm()

    return render(request, "front/localidade_form.html", {"form": form, "editar": False})

@login_required
def localidade_update(request, pk):
    obj = get_object_or_404(Localidade, pk=pk)
    if request.method == "POST":
        form = LocalidadeForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Localidade atualizada com sucesso.")
            return redirect("localidade_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = LocalidadeForm(instance=obj)

    return render(request, "front/localidade_form.html", {"form": form, "editar": True, "obj": obj})

@login_required
def localidade_delete(request, pk):
    obj = get_object_or_404(Localidade, pk=pk)
    if request.method == "POST":
        nome = obj.local
        obj.delete()
        messages.success(request, f"Localidade '{nome}' excluída.")
        return redirect("localidade_list")
    # Não renderizamos página separada; exclusão via modal na listagem
    return redirect("localidade_list")

# DETAIL
@login_required
def localidade_detail(request, pk):
    obj = get_object_or_404(Localidade, pk=pk)
    
    # --- CORREÇÃO AQUI ---
    # Removido 'status' do select_related pois é um campo simples, não uma FK.
    itens = (
        Item.objects
        .filter(localidade=obj)
        .select_related('subtipo') # Apenas FKs aqui
        .order_by("nome")[:20]
    )
    itens_count = Item.objects.filter(localidade=obj).count()
    
    # Busca usuários vinculados
    usuarios = (
        Usuario.objects
        .filter(localidade=obj)
        .select_related('funcao')
        .order_by("nome")[:20]
    )
    usuarios_count = Usuario.objects.filter(localidade=obj).count()

    context = {
        "obj": obj,
        "itens": itens,
        "itens_count": itens_count,
        "usuarios": usuarios,
        "usuarios_count": usuarios_count,
    }
    return render(request, "front/localidade_detail.html", context)



#################### CENTRO DE CUSTO ########################

def _get_centros_filtrados(request):
    q = request.GET.get("q", "").strip()
    pmb = request.GET.get("pmb", "").strip()

    qs = CentroCusto.objects.all().order_by("numero")

    if q:
        qs = qs.filter(Q(numero__icontains=q) | Q(departamento__icontains=q))
    if pmb:
        qs = qs.filter(pmb=pmb)
    
    return qs, q, pmb

@login_required
def centrocusto_list(request):
    """
    Dashboard de Centros de Custo (Card View).
    """
    qs, q, pmb = _get_centros_filtrados(request)
    
    centros_ids = list(qs.values_list("id", flat=True))
    total_filtrado = qs.count()

    # --- KPIs ---
    locados_qs = Item.objects.filter(
        centro_custo_id__in=centros_ids,
        locado="sim"
    )

    agregado = locados_qs.aggregate(
        total_custo=Sum("locacao__valor_mensal"),
        total_itens=Count("id")
    )
    kpi_custo_total = agregado["total_custo"] or Decimal(0)
    kpi_itens_locados = agregado["total_itens"] or 0
    kpi_media = (kpi_custo_total / total_filtrado) if total_filtrado > 0 else 0

    # Dados por Centro
    dados_por_centro = (
        locados_qs
        .values("centro_custo")
        .annotate(custo=Sum("locacao__valor_mensal"), qtd=Count("id"))
    )
    mapa_custo = {d["centro_custo"]: d["custo"] or 0 for d in dados_por_centro}
    mapa_qtd = {d["centro_custo"]: d["qtd"] or 0 for d in dados_por_centro}

    # Top Offensor
    kpi_top_cc = None
    kpi_top_val = 0
    if mapa_custo:
        top_id = max(mapa_custo, key=mapa_custo.get)
        kpi_top_val = mapa_custo[top_id]
        try:
            kpi_top_cc = CentroCusto.objects.get(id=top_id)
        except CentroCusto.DoesNotExist: pass

    # --- Paginação ---
    try:
        per_page = int(request.GET.get("pp", 12)) # Cards geralmente pedem menos itens por pg
    except ValueError:
        per_page = 12

    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_num)

    for cc in page_obj.object_list:
        cc.custo_calc = mapa_custo.get(cc.id, 0)
        cc.qtd_calc = mapa_qtd.get(cc.id, 0)

    get_copy = request.GET.copy()
    if "page" in get_copy: del get_copy["page"]
    qs_keep = get_copy.urlencode()

    context = {
        "centros": page_obj.object_list,
        "page_obj": page_obj,
        "total": total_filtrado,
        "qs_keep": qs_keep,
        "f_q": q, "f_pmb": pmb, "opt_pmb": SimNaoChoices.choices,
        "kpi_custo_total": kpi_custo_total,
        "kpi_itens": kpi_itens_locados,
        "kpi_media": kpi_media,
        "kpi_top_cc": kpi_top_cc,
        "kpi_top_val": kpi_top_val,
    }

    return render(request, "front/centrocusto/centrocusto_list.html", context)

@login_required
def centrocusto_export_pdf(request):
    qs, q, pmb = _get_centros_filtrados(request)
    centros = list(qs)
    centros_ids = [c.id for c in centros]
    
    locados_qs = Item.objects.filter(centro_custo_id__in=centros_ids, locado="sim")
    dados = locados_qs.values("centro_custo").annotate(custo=Sum("locacao__valor_mensal"), qtd=Count("id"))
    
    mapa_custo = {d["centro_custo"]: d["custo"] or 0 for d in dados}
    mapa_qtd = {d["centro_custo"]: d["qtd"] or 0 for d in dados}
    
    total_geral = 0
    for c in centros:
        c.custo_calc = mapa_custo.get(c.id, 0)
        c.qtd_calc = mapa_qtd.get(c.id, 0)
        total_geral += c.custo_calc

    context = {
        "centros": centros,
        "total_geral": total_geral,
        "filtros": {"Busca": q, "PMB": pmb},
        "usuario": request.user,
    }
    
    template_path = 'front/centrocusto/centrocusto_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_centros.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err: return HttpResponse('Erro PDF', status=500)
    return response


# CREATE
@login_required
def centrocusto_create(request):
    if request.method == "POST":
        form = CentroCustoForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Centro de custo criado com sucesso!")
            return redirect("centrocusto_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = CentroCustoForm()

    return render(request, "front/centrocusto/centrocusto_form.html", {"form": form, "editar": False})


# UPDATE
@login_required
def centrocusto_update(request, pk: int):
    obj = get_object_or_404(CentroCusto, pk=pk)
    if request.method == "POST":
        form = CentroCustoForm(request.POST, instance=obj)
        if form.is_valid():
            sobj = form.save(commit=False)
            sobj.atualizado_por = request.user
            sobj.save()
            messages.success(request, "Centro de custo atualizado com sucesso!")
            return redirect("centrocusto_list")
        messages.error(request, "Corrija os erros do formulário.")
    else:
        form = CentroCustoForm(instance=obj)

    return render(request, "front/centrocusto/centrocusto_form.html", {"form": form, "editar": True})


# DETAIL
@login_required
def centrocusto_detail(request, pk):
    """
    Dashboard detalhado de um Centro de Custo.
    Inclui KPIs financeiros e lista de ativos vinculados.
    """
    obj = get_object_or_404(CentroCusto, pk=pk)
    
    # Busca itens vinculados
    itens_qs = (
        Item.objects
        .filter(centro_custo=obj)
        .select_related('subtipo', 'localidade')
        .order_by('nome')
    )

    # Cálculo de KPIs em tempo real
    # 1. Total de Itens
    total_itens = itens_qs.count()

    # 2. Custo Mensal (Apenas Itens Locados)
    # Soma o valor da locação se o item estiver marcado como locado
    custo_mensal = itens_qs.filter(locado=SimNaoChoices.SIM).aggregate(
        total=Sum('locacao__valor_mensal')
    )['total'] or Decimal(0)

    # 3. Valor Patrimonial (Apenas Itens Próprios/Aquisição)
    # Soma o valor de compra se NÃO for locado
    valor_patrimonial = itens_qs.exclude(locado=SimNaoChoices.SIM).aggregate(
        total=Sum('valor')
    )['total'] or Decimal(0)

    context = {
        'obj': obj,
        'itens_cc': itens_qs,
        'kpi': {
            'total_itens': total_itens,
            'custo_mensal': custo_mensal,
            'valor_patrimonial': valor_patrimonial,
        }
    }
    
    return render(request, 'front/centrocusto/centrocusto_detail.html', context)


# DELETE (POST via modal)
@login_required
def centrocusto_delete(request, pk: int):
    obj = get_object_or_404(CentroCusto, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Centro de custo removido com sucesso.")
    else:
        messages.error(request, "Ação inválida.")
    return redirect("centrocusto_list")

######################### FUNÇÃO ################################


################ ITEM #######################################

@login_required
@transaction.atomic  # Garante que tudo ou nada seja salvo 
def item_create(request):
    if request.method == "POST":
        form = ItemForm(request.POST)
        # Instancia o form de locação apenas se necessário, mas valida depois
        locacao_form = LocacaoForm(request.POST)

        if form.is_valid():
            item = form.save(commit=False)
            item.criado_por = request.user
            
            # Verifica flag de locação
            eh_locado = request.POST.get('locado') == 'sim'  # Ou use form.cleaned_data após validação básica

            if eh_locado:
                if locacao_form.is_valid():
                    # Salva o item primeiro para ter o ID
                    item.save()
                    
                    locacao = locacao_form.save(commit=False)
                    locacao.equipamento = item
                    locacao.save()
                    return redirect("equipamentos_list")
                # Se locação não for válida, cai para o render abaixo com erros exibidos
            else:
                # Não é locado, fluxo simples
                item.save()
                return redirect("equipamentos_list")

    else:
        form = ItemForm()
        locacao_form = LocacaoForm()

    return render(request, "front/equipamentos/cadastrar_equipamento.html", {
        "form": form, 
        "locacao_form": locacao_form
    })
# ITEM LIST

def _norm(s: str) -> str:
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()





# importe seus modelos/choices
# from .models import Item, Subtipo, MovimentacaoItem, StatusItemChoices

def _build_queryset_and_context(request):
    """
    Monta queryset + contexto base (padrão único para página e partial).
    """
    qs = (
        Item.objects
        .select_related("subtipo", "localidade", "centro_custo", "fornecedor")
        .order_by("nome", "id")
    )

    # Prefetch de movimentações (se for necessário em outras views)
    mov_qs = MovimentacaoItem.objects.select_related("usuario").order_by("-created_at")
    qs = qs.prefetch_related(Prefetch("movimentacoes", queryset=mov_qs, to_attr="pref_movs"))

    p = request.GET

    # -------- Filtros ----------
    nome         = (p.get("nome") or "").strip()
    subtipo_id   = (p.get("subtipo") or "").strip()
    status_code  = (p.get("status") or "").strip()
    numero_serie = (p.get("numero_serie") or "").strip()
    usuario      = (p.get("usuario") or "").strip()  # caso use em algum lugar
    localidade   = (p.get("localidade") or "").strip()
    centro_custo = (p.get("centro_custo") or "").strip()
    fornecedor   = (p.get("fornecedor") or "").strip()

    if nome:
        qs = qs.filter(nome__icontains=nome)
    if subtipo_id:
        qs = qs.filter(subtipo_id=subtipo_id)
    if status_code:
        qs = qs.filter(status=status_code)
    if numero_serie:
        qs = qs.filter(numero_serie__icontains=numero_serie)
    if usuario:
        qs = qs.filter(movimentacoes__usuario__nome__icontains=usuario)
    if localidade:
        qs = qs.filter(localidade__local__icontains=localidade)
    if centro_custo:
        qs = qs.filter(
            Q(centro_custo__departamento__icontains=centro_custo) |
            Q(centro_custo__numero__icontains=centro_custo)
        )
    if fornecedor:
        qs = qs.filter(fornecedor__nome__icontains=fornecedor)

    qs = qs.distinct()
    filtered_total = qs.count()

    # -------- Paginação --------
    try:
        per_page = int(p.get("pp") or 20)
        if per_page not in (10, 20, 50, 100):
            per_page = 20
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(qs, per_page)
    page_number = p.get("page") or 1
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # querystring preservada (sem page)
    params = p.copy()
    params.pop("page", None)
    qs_keep = params.urlencode()

    # -------- KPIs (status) --------
    total_ativos     = Item.objects.filter(status=StatusItemChoices.ATIVO).count()
    total_backup     = Item.objects.filter(status=StatusItemChoices.BACKUP).count()
    total_manutencao = Item.objects.filter(status=StatusItemChoices.MANUTENCAO).count()
    total_queimados  = Item.objects.filter(status=StatusItemChoices.DEFEITO).count()
    total_geral      = Item.objects.count()

    context = {
        "equipamentos": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "per_page": per_page,
        "qs_keep": qs_keep,
        "filtered_total": filtered_total,

        "row_start": page_obj.start_index(),  # índice base da página (para #)
        "subtipos": Subtipo.objects.all().order_by("nome"),
        "status_choices": StatusItemChoices.choices,

        "total_ativos": total_ativos,
        "total_backup": total_backup,
        "total_manutencao": total_manutencao,
        "total_queimados": total_queimados,
        "total_geral": total_geral,
    }
    return context


@login_required
def equipamentos_list(request):
    context = _build_queryset_and_context(request)

    # Verifica se é uma requisição AJAX / Partial
    is_partial = request.GET.get("partial") == "1" or request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if is_partial:
        # Renderiza apenas os fragmentos
        data = {
            "tbody": render_to_string("front/equipamentos/_tbody.html", context, request=request),
            "pagination": render_to_string("front/equipamentos/_pagination.html", context, request=request),
            "kpis": render_to_string("front/equipamentos/_kpis.html", context, request=request),
            "count": context["filtered_total"],
        }
        return JsonResponse(data)

    # Renderização Full Page
    return render(request, "front/equipamentos/equipamentos_list.html", context)

### ITEM / Equipamento detalhe 
@login_required
def equipamento_detalhe(request, pk: int):
    item = get_object_or_404(
        Item.objects.select_related(
            "subtipo", "localidade", "centro_custo", "fornecedor"
        ),
        pk=pk,
    )

    # 1) Histórico de movimentações
    movimentacoes = (
        MovimentacaoItem.objects.filter(item=item)
        .select_related(
            "usuario",
            "localidade_destino",
            "centro_custo_destino",
            "fornecedor_manutencao",
        )
        .order_by("-created_at")
    )

    historico_manutencao = movimentacoes.filter(
        tipo_movimentacao__in=[
            TipoMovimentacaoChoices.ENVIO_MANUTENCAO,
            TipoMovimentacaoChoices.RETORNO_MANUTENCAO,
        ]
    )

    ultimo_resp = "Em Estoque / Não Definido"
    if movimentacoes.exists():
        last = movimentacoes.first()
        if last.usuario:
            ultimo_resp = f"Usuário: {last.usuario.nome}"
        elif last.centro_custo_destino:
            ultimo_resp = f"Setor: {last.centro_custo_destino.departamento}"
        elif last.localidade_destino:
            ultimo_resp = f"Local: {last.localidade_destino.local}"
        elif last.fornecedor_manutencao:
            ultimo_resp = f"Externo: {last.fornecedor_manutencao.nome}"

    # 2) Financeiro / ciclo de vida
    today = timezone.localdate()
    locacao = getattr(item, "locacao", None)

    custo_manut = historico_manutencao.aggregate(t=Sum("custo"))["t"] or Decimal("0.00")

    financeiro = {
        "modo": "LOCAÇÃO" if getattr(item, "locado", "nao") == "sim" else "AQUISIÇÃO",
        "custo_aquisicao": item.valor or Decimal("0.00"),

        # custos
        "custo_manutencao": custo_manut,
        "custo_aluguel_acumulado": Decimal("0.00"),  # só locado
        "custo_tempo_empresa": Decimal("0.00"),       # tempo (aluguéis pagos ou aquisição)
        "custo_total_empresa": Decimal("0.00"),       # TEMPO + MANUTENÇÕES
        "tco": Decimal("0.00"),                       # alias do custo_total_empresa

        # valores de referência
        "valor_atual": Decimal("0.00"),               # aluguel mensal ou valor estimado
        "custo_mensal": Decimal("0.00"),

        # tempo
        "dias_na_empresa": 0,
        "meses_na_empresa": 0,
        "data_inicio": None,
        "data_fim": None,

        # contrato / vida útil
        "vida_util_perc": 0,
        "vida_util_texto": "Indefinido",
        "status_vida": "ok",
    }

    # ===== EQUIPAMENTO LOCADO =====
    if financeiro["modo"] == "LOCAÇÃO" and locacao:
        valor_mensal = locacao.valor_mensal or Decimal("0.00")
        dt_inicio = locacao.data_entrada          # campo cadastrado na locação
        tempo_locado_meses = locacao.tempo_locado or 0

        financeiro["custo_mensal"] = valor_mensal
        financeiro["valor_atual"] = valor_mensal

        if dt_inicio:
            financeiro["data_inicio"] = dt_inicio

            dias_corridos = (today - dt_inicio).days
            if dias_corridos < 0:
                dias_corridos = 0

            # meses que o equipamento já ficou na empresa
            meses_uso = dias_corridos // 30 if dias_corridos > 0 else 0

            # meses de aluguel considerados como já pagos
            if tempo_locado_meses > 0:
                meses_pagos = min(meses_uso, int(tempo_locado_meses))
            else:
                meses_pagos = meses_uso

            financeiro["dias_na_empresa"] = dias_corridos
            financeiro["meses_na_empresa"] = meses_uso

            # >>> ALUGUÉIS JÁ PAGOS <<<
            custo_aluguel_acumulado = valor_mensal * Decimal(meses_pagos)
            financeiro["custo_aluguel_acumulado"] = custo_aluguel_acumulado

            # custo do tempo = o que já foi pago em aluguel
            financeiro["custo_tempo_empresa"] = custo_aluguel_acumulado

            # >>> CUSTO NA EMPRESA (TEMPO + MANUTENÇÕES) <<<
            custo_total_empresa = custo_aluguel_acumulado + custo_manut
            financeiro["custo_total_empresa"] = custo_total_empresa
            financeiro["tco"] = custo_total_empresa

            # contrato / vida
            if tempo_locado_meses > 0:
                total_dias = int(tempo_locado_meses) * 30
                dt_fim = dt_inicio + timedelta(days=total_dias)
                financeiro["data_fim"] = dt_fim

                if total_dias > 0:
                    perc = (dias_corridos / total_dias) * 100
                    financeiro["vida_util_perc"] = min(100, max(0, int(perc)))

                restante = (dt_fim - today).days
                if restante < 0:
                    financeiro["vida_util_texto"] = "Contrato vencido"
                    financeiro["status_vida"] = "critical"
                else:
                    financeiro["vida_util_texto"] = f"{restante} dias restantes"
                    if restante < 30:
                        financeiro["status_vida"] = "warning"
            else:
                financeiro["vida_util_texto"] = "Contrato sem prazo definido"
        else:
            # sem data_entrada → não calcula aluguel, só manutenções
            financeiro["custo_total_empresa"] = custo_manut
            financeiro["tco"] = custo_manut

    # ===== EQUIPAMENTO PRÓPRIO (AQUISIÇÃO) =====
    else:
        financeiro["custo_tempo_empresa"] = financeiro["custo_aquisicao"]
        custo_total_empresa = financeiro["custo_aquisicao"] + custo_manut
        financeiro["custo_total_empresa"] = custo_total_empresa
        financeiro["tco"] = custo_total_empresa

        if item.data_compra:
            financeiro["data_inicio"] = item.data_compra

            vida_util_anos = 5
            vida_util_dias = vida_util_anos * 365
            dias_uso = (today - item.data_compra).days
            if dias_uso < 0:
                dias_uso = 0

            financeiro["dias_na_empresa"] = dias_uso
            financeiro["meses_na_empresa"] = dias_uso // 30 if dias_uso > 0 else 0

            if dias_uso < vida_util_dias:
                fator = Decimal(dias_uso) / Decimal(vida_util_dias)
                financeiro["valor_atual"] = financeiro["custo_aquisicao"] * (1 - fator)
                financeiro["vida_util_perc"] = int(
                    (dias_uso / vida_util_dias) * 100
                )
                anos_restantes = max(0, vida_util_anos - (dias_uso // 365))
                financeiro["vida_util_texto"] = f"~{anos_restantes} anos restantes"
                financeiro["data_fim"] = item.data_compra + timedelta(
                    days=vida_util_dias
                )
            else:
                financeiro["valor_atual"] = Decimal("0.00")
                financeiro["vida_util_perc"] = 100
                financeiro["vida_util_texto"] = "Totalmente depreciado"
                financeiro["status_vida"] = "warning"
                financeiro["data_fim"] = item.data_compra + timedelta(
                    days=vida_util_dias
                )

    # 3) Preventivas
    preventivas = (
        Preventiva.objects.filter(equipamento=item)
        .select_related("checklist_modelo")
        .order_by("data_proxima")
    )

    status_saude = "ok"
    for p in preventivas:
        if not p.data_proxima and p.checklist_modelo.intervalo_dias:
            base = p.data_ultima or today
            p.data_proxima = base + timedelta(days=p.checklist_modelo.intervalo_dias)

        p.atrasado = p.data_proxima and p.data_proxima < today
        if p.atrasado:
            status_saude = "critical"

    context = {
        "item": item,
        "ultimo_resp": ultimo_resp,
        "movimentacoes": movimentacoes,
        "historico_manutencao": historico_manutencao,
        "preventivas": preventivas,
        "financeiro": financeiro,
        "status_saude": status_saude,
        "locacao": locacao,
    }

    return render(request, "front/equipamentos/equipamento_detalhe.html", context)

@login_required
@transaction.atomic
def editar_equipamento(request, pk):
    item = get_object_or_404(Item, pk=pk)
    locacao_instance = getattr(item, "locacao", None)  # pode existir ou não

    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)

        # verifica se o usuário marcou o item como locado
        eh_locado = request.POST.get("locado") == "sim"

        # se for locado, sempre usa a MESMA instância de Locacao (se já existir)
        locacao_form = LocacaoForm(
            request.POST,
            instance=locacao_instance if eh_locado else None
        )

        if form.is_valid():
            item = form.save(commit=False)
            item.atualizado_por = request.user

            if eh_locado:
                # precisa validar também a locação
                if locacao_form.is_valid():
                    item.save()

                    locacao = locacao_form.save(commit=False)
                    locacao.equipamento = item
                    locacao.save()

                    return redirect("equipamentos_list")
                # se locacao_form não for válido, cai para o render lá embaixo
            else:
                # se marcou como NÃO locado e existia locação, remove
                item.save()
                if locacao_instance:
                    locacao_instance.delete()

                return redirect("equipamentos_list")

    else:
        # GET: pré-carrega os dados
        form = ItemForm(instance=item)
        locacao_form = LocacaoForm(instance=locacao_instance)

    return render(
        request,
        "front/equipamentos/cadastrar_equipamento.html",
        {
            "form": form,
            "locacao_form": locacao_form,
            "editar": True,  # para o template saber que é edição
        },
    )

@require_POST
@login_required
def equipamento_excluir(request, pk: int):
    item = get_object_or_404(Item, pk=pk)
    item.delete()
    messages.success(request, "Item excluído com sucesso.")
    return redirect("equipamentos_list")


#########################################################################################################################

# Cadastro de licença
@login_required
def cadastrar_licenca(request):
    if request.method == "POST":
        item_form = ItemForm(request.POST)
        licenca_form = LicencaForm(request.POST)

        if item_form.is_valid() and licenca_form.is_valid():
            item = item_form.save(commit=False)
            item.criado_por = request.user
            item.atualizado_por = request.user
            item.save()

            licenca = licenca_form.save(commit=False)
            licenca.criado_por = request.user
            licenca.atualizado_por = request.user
            licenca.save()
            licenca_form.save_m2m()

            return redirect("lista_licencas")
    else:
        item_form = ItemForm()
        licenca_form = LicencaForm()

    return render(request, "cadastro_licenca.html", {
        "item_form": item_form,
        "licenca_form": licenca_form
    })

################ EQUIPAMENTO ##################



#def equipamento_delete(request, pk):
    #equipamento = get_object_or_404(Equipamento, pk=pk)
    #f request.method == 'POST':
        #equipamento.delete()
        #return redirect('equipamentos_list')
    #return render(request, 'equipamento/delete.html', {'obj': equipamento})

########### LOCADO ##################

def locacoes_list(request):
    locacoes = Locacao.objects.all()
    return render(request, 'locacao/list.html', {'locacoes': locacoes})

def locacao_create(request):
    form = LocacaoForm(request.POST or None)
    if form.is_valid():
        locacao = form.save(commit=False)
        locacao.save()
        return redirect('locacoes_list')
    return render(request, 'locacao/form.html', {'form': form})

def locacao_update(request, pk):
    locacao = get_object_or_404(Locacao, pk=pk)
    form = LocacaoForm(request.POST or None, instance=locacao)
    if form.is_valid():
        locacao = form.save(commit=False)
        locacao.save()
        return redirect('locacoes_list')
    return render(request, 'locacao/form.html', {'form': form})

def locacao_delete(request, pk):
    locacao = get_object_or_404(Locacao, pk=pk)
    if request.method == 'POST':
        locacao.delete()
        return redirect('locacoes_list')
    return render(request, 'locacao/delete.html', {'obj': locacao})

@login_required
def funcao_list(request):
    q = (request.GET.get("q") or "").strip()
    funcoes = Funcao.objects.all().order_by("nome")
    if q:
        funcoes = funcoes.filter(nome__icontains=q)
    ctx = {
        "funcoes": funcoes,
        "q": q,
        "total": funcoes.count(),
    }
    return render(request, "front/funcoes/funcao_list.html", ctx)


@login_required
def funcao_form(request, pk=None):
    instance = get_object_or_404(Funcao, pk=pk) if pk else None

    if request.method == "POST":
        form = FuncaoForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save(commit=False)
            if instance is None:
                # se seu AuditModel usa esse campo
                obj.criado_por = request.user
            # idem para atualizado_por
            obj.atualizado_por = request.user
            obj.save()

            messages.success(request, "Função salva com sucesso.")
            # ✅ redireciona para a LISTA após salvar
            return redirect("funcoes_list")
    else:
        form = FuncaoForm(instance=instance)

    return render(
        request,
        "front/funcoes/funcao_form.html",
        {"form": form, "instance": instance},
    )


@login_required
def funcao_delete(request, pk):
    if request.method != "POST":
        messages.error(request, "Requisição inválida.")
        return redirect("funcoes_list")

    obj = get_object_or_404(Funcao, pk=pk)
    obj.delete()
    messages.success(request, "Função removida.")
    return redirect("funcoes_list")

### COMENTARIO ####

def comentarios_list(request):
    comentarios = Comentario.objects.all()
    return render(request, 'comentario/list.html', {'comentarios': comentarios})

def comentario_create(request):
    form = ComentarioForm(request.POST or None)
    if form.is_valid():
        comentario = form.save(commit=False)
        comentario.criado_por = request.user
        comentario.save()
        return redirect('comentarios_list')
    return render(request, 'comentario/form.html', {'form': form})

def comentario_update(request, pk):
    comentario = get_object_or_404(Comentario, pk=pk)
    form = ComentarioForm(request.POST or None, instance=comentario)
    if form.is_valid():
        form.save()
        return redirect('comentarios_list')
    return render(request, 'comentario/form.html', {'form': form})

def comentario_delete(request, pk):
    comentario = get_object_or_404(Comentario, pk=pk)
    if request.method == 'POST':
        comentario.delete()
        return redirect('comentarios_list')
    return render(request, 'comentario/delete.html', {'obj': comentario})

####################### MOVIMENTA ITEM ################
# views.py
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage
from django.db.models import Count, Q
from django.shortcuts import render

# Ajuste conforme seu projeto
# from .models import MovimentacaoItem, TipoMovimentacaoChoices

def _get_movimentacao_qs(request):
    """
    Helper: Aplica os filtros da requisição e retorna o QuerySet.
    Usado tanto na view de lista quanto na de exportação PDF.
    """
    q = (request.GET.get("q") or "").strip()
    tipo = (request.GET.get("tipo") or "").strip()
    usuario_q = (request.GET.get("usuario") or "").strip()
    numero_serie = (request.GET.get("numero_serie") or "").strip()
    centro_custo = (request.GET.get("centro_custo") or "").strip()
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")

    qs = (
        MovimentacaoItem.objects
        .select_related(
            "item", "usuario",
            "localidade_origem", "localidade_destino",
            "centro_custo_origem", "centro_custo_destino",
            "fornecedor_manutencao",
        )
        .order_by("-created_at")
    )

    if q:
        qs = qs.filter(item__nome__icontains=q)
    if tipo:
        qs = qs.filter(tipo_movimentacao=tipo)
    if usuario_q:
        qs = qs.filter(usuario__nome__icontains=usuario_q)
    if numero_serie:
        qs = qs.filter(item__numero_serie__icontains=numero_serie)
    if centro_custo:
        qs = qs.filter(
            Q(centro_custo_origem__numero__icontains=centro_custo) |
            Q(centro_custo_origem__departamento__icontains=centro_custo) |
            Q(centro_custo_destino__numero__icontains=centro_custo) |
            Q(centro_custo_destino__departamento__icontains=centro_custo)
        )
    if data_inicio:
        qs = qs.filter(created_at__date__gte=data_inicio)
    if data_fim:
        qs = qs.filter(created_at__date__lte=data_fim)
        
    return qs

@login_required
def movimentacao_list(request):
    # 1. Recupera QS Filtrado usando o Helper
    qs = _get_movimentacao_qs(request)
    total_filtrado = qs.count()

    # 2. KPIs (Mantendo sua lógica original)
    stats = dict(qs.values_list("tipo_movimentacao").annotate(c=Count("id")).order_by())
    def get_count(key): return stats.get(key, 0)

    kpi_entrada = get_count("entrada")
    kpi_saida = get_count("baixa")
    kpi_transf = get_count("transferencia") + get_count("transferencia_equipamento")
    kpi_manut = get_count("envio_manutencao") + get_count("retorno_manutencao")

    hoje = timezone.now().date()
    kpi_hoje = qs.filter(created_at__date=hoje).count()

    top_mover_data = qs.values('item__nome').annotate(total=Count('id')).order_by('-total').first()
    kpi_top_item_nome = top_mover_data['item__nome'] if top_mover_data else "-"
    kpi_top_item_qtd = top_mover_data['total'] if top_mover_data else 0

    # 3. Paginação
    try:
        per_page = int(request.GET.get("pp", 20))
    except ValueError:
        per_page = 20

    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_num)

    get_copy = request.GET.copy()
    if "page" in get_copy: del get_copy["page"]
    qs_keep = get_copy.urlencode()

    context = {
        "movimentacoes": page_obj.object_list,
        "page_obj": page_obj,
        "total": total_filtrado,
        "qs_keep": qs_keep,
        # Filtros de volta para o template
        "f_q": request.GET.get("q", ""), 
        "f_tipo": request.GET.get("tipo", ""), 
        "f_user": request.GET.get("usuario", ""), 
        "f_serie": request.GET.get("numero_serie", ""), 
        "f_cc": request.GET.get("centro_custo", ""),
        "f_ini": request.GET.get("data_inicio", ""), 
        "f_fim": request.GET.get("data_fim", ""),
        "tipos_choices": TipoMovimentacaoChoices.choices,
        "kpi": {
            "hoje": kpi_hoje,
            "top_item": kpi_top_item_nome,
            "top_item_qtd": kpi_top_item_qtd,
            "entrada": kpi_entrada,
            "saida": kpi_saida,
            "transferencias": kpi_transf,
            "manutencao": kpi_manut
        }
    }

    return render(request, "front/movimentacao_list.html", context)

@login_required
def movimentacao_export_pdf(request):
    """
    Gera PDF usando xhtml2pdf (Pisa).
    """
    # 1. Recupera dados filtrados
    qs = _get_movimentacao_qs(request)
    
    # 2. Contexto
    context = {
        'movimentacoes': qs,
        'usuario': request.user,
        'data_geracao': timezone.now(),
        'total': qs.count(),
        'filtros': {
            'inicio': request.GET.get("data_inicio"),
            'fim': request.GET.get("data_fim"),
            'tipo': request.GET.get("tipo")
        }
    }
    
    # 3. Renderiza Template
    template_path = 'front/movimentacao_pdf.html'
    template = get_template(template_path)
    html = template.render(context)
    
    # 4. Gera PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"relatorio_movimentacoes_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Criação do PDF
    pisa_status = pisa.CreatePDF(
       html, dest=response
    )

    if pisa_status.err:
       return HttpResponse('Erro ao gerar PDF <pre>' + html + '</pre>')
       
    return response




@login_required
def movimentacao_create(request):
    if request.method == "POST":
        form = MovimentacaoItemForm(request.POST, request.FILES)
        if form.is_valid():
            mov = form.save(commit=False)
            mov.criado_por = request.user
            # Limpa usuário se não for pertinente
            if mov.tipo_movimentacao in ("envio_manutencao", "retorno_manutencao", "entrada", "baixa"):
                mov.usuario = None
            mov.save()
            messages.success(request, "Movimentação realizada com sucesso!")
            return redirect("movimentacao_list")
        else:
            messages.error(request, "Verifique os erros no formulário.")
    else:
        form = MovimentacaoItemForm()

    return render(request, "front/movimentacao_form.html", {"form": form})






@login_required
def movimentacao_detail(request, pk):
    """
    Exibe os detalhes de uma movimentação com design Enterprise.
    Corrige o erro de 'kpi' fornecendo estatísticas contextuais do item.
    """
    mov = get_object_or_404(MovimentacaoItem, pk=pk)

    # 1. Dados de Origem (Tratamento de Nulos)
    origem_loc = mov.localidade_origem.local if mov.localidade_origem else "—"
    origem_cc = f"{mov.centro_custo_origem.numero} - {mov.centro_custo_origem.departamento}" if mov.centro_custo_origem else "—"

    # 2. Dados de Destino
    dest_loc = mov.localidade_destino.local if mov.localidade_destino else "—"
    dest_cc = f"{mov.centro_custo_destino.numero} - {mov.centro_custo_destino.departamento}" if mov.centro_custo_destino else "—"

    # 3. Status Visual
    if mov.tipo_movimentacao in ("retorno", "retorno_manutencao"):
        status_final = "Backup"
    elif mov.tipo_movimentacao == "transferencia_equipamento" and mov.status_transferencia:
        status_final = dict(StatusItemChoices.choices).get(mov.status_transferencia, mov.status_transferencia)
    elif mov.tipo_movimentacao == "envio_manutencao":
        status_final = "Em Manutenção"
    elif mov.tipo_movimentacao == "baixa":
        status_final = "Baixado"
    else:
        status_final = mov.item.get_status_display() if mov.item else "—"

    # 4. Impacto Visual (Badge)
    impacto_map = {
        "entrada": (f"+{mov.quantidade} (Entrada)", "st-green"),
        "baixa": (f"-{mov.quantidade} (Baixa)", "st-red"),
        "envio_manutencao": ("Saída Manutenção", "st-orange"),
        "retorno_manutencao": ("Retorno Manutenção", "st-blue"),
        "transferencia": ("Transferência de Posse", "st-gray"),
        "transferencia_equipamento": ("Transferência de Setor", "st-gray"),
    }
    impacto_texto, impacto_class = impacto_map.get(mov.tipo_movimentacao, ("Apenas Registro", "st-gray"))

    # 5. KPIs Específicos do Item (Substitui o 'kpi' global que causava erro)
    # Mostra o histórico deste item específico para contexto
    total_movs_item = MovimentacaoItem.objects.filter(item=mov.item).count()
    ultima_mov = MovimentacaoItem.objects.filter(item=mov.item).exclude(pk=pk).order_by('-created_at').first()

    context = {
        "mov": mov,
        "origem": {"loc": origem_loc, "cc": origem_cc},
        "destino": {"loc": dest_loc, "cc": dest_cc},
        "status_final": status_final,
        "impacto": {"texto": impacto_texto, "class": impacto_class},
        # Nova variável de stats para evitar o erro e dar contexto
        "stats": {
            "total_movs": total_movs_item,
            "ultima_data": ultima_mov.created_at if ultima_mov else None
        }
    }
    
    return render(request, "front/movimentacao_detail.html", context)


def movimentacao_update(request, pk):
    mov = get_object_or_404(MovimentacaoItem, pk=pk)
    form = MovimentacaoItemForm(request.POST or None, request.FILES or None, instance=mov)
    if form.is_valid():
        mov = form.save(commit=False)
        mov.atualizado_por = request.user
        mov.save()
        return redirect('movimentacoes_list')
    return render(request, 'movimentacaoitem/form.html', {'form': form})

def movimentacao_delete(request, pk):
    mov = get_object_or_404(MovimentacaoItem, pk=pk)
    if request.method == 'POST':
        mov.delete()
        return redirect('movimentacoes_list')
    return render(request, 'movimentacaoitem/delete.html', {'obj': mov})

###################### CICLO MANUTENÇÃO ##########################

# LISTA DE CICLOS
def ciclos_list(request):
    ciclos = CicloManutencao.objects.all().order_by('-created_at')
    return render(request, 'ciclomanutencao/list.html', {'ciclos': ciclos})

# CRIAR CICLO (INTELIGENTE)
def ciclo_create(request):
    form = CicloManutencaoForm(request.POST or None)

    if form.is_valid():
        ciclo = form.save(commit=False)
        item = ciclo.item

        # Verifica se já existe um ciclo aberto
        ciclo_aberto = CicloManutencao.objects.filter(item=item, data_fim__isnull=True).exists()
        if ciclo_aberto:
            messages.error(request, f"O item '{item.nome}' já possui um ciclo de manutenção em andamento.")
            return render(request, 'ciclomanutencao/form.html', {'form': form})

        ciclo.criado_por = request.user
        ciclo.atualizado_por = request.user
        item.status = 'manutencao'
        item.atualizado_por = request.user

        item.save()
        ciclo.save()

        messages.success(request, f"Ciclo de manutenção iniciado para '{item.nome}'.")
        return redirect('ciclos_list')

    return render(request, 'ciclomanutencao/form.html', {'form': form})

# ATUALIZAR CICLO (ENCERRAR)
def ciclo_update(request, pk):
    ciclo = get_object_or_404(CicloManutencao, pk=pk)
    form = CicloManutencaoForm(request.POST or None, instance=ciclo)

    if form.is_valid():
        ciclo = form.save(commit=False)
        ciclo.atualizado_por = request.user
        item = ciclo.item

        if ciclo.data_fim:
            item.status = 'ativo'  # ou 'backup', conforme regra de negócio
            item.atualizado_por = request.user
            item.save()

            messages.success(request, f"Ciclo encerrado. '{item.nome}' voltou para operação.")

        ciclo.save()
        return redirect('ciclos_list')

    return render(request, 'ciclomanutencao/form.html', {'form': form})

# EXCLUIR CICLO
def ciclo_delete(request, pk):
    ciclo = get_object_or_404(CicloManutencao, pk=pk)
    if request.method == 'POST':
        ciclo.delete()
        return redirect('ciclos_list')
    return render(request, 'ciclomanutencao/delete.html', {'obj': ciclo})

@login_required
def sobre_plataforma(request):
    # Versão/identidade opcionais via settings (defina se quiser)
    app_name = getattr(settings, "PROJECT_NAME", "Controle de Ativos")
    version = getattr(settings, "APP_VERSION", "1.0.0")
    build_date = getattr(settings, "APP_BUILD_DATE", None)

    # Métricas rápidas
    total_itens = Item.objects.count()
    total_ativos = Item.objects.filter(status=StatusItemChoices.ATIVO).count()
    total_backup = Item.objects.filter(status=StatusItemChoices.BACKUP).count()
    total_manut = Item.objects.filter(status=StatusItemChoices.MANUTENCAO).count()
    total_defeito = Item.objects.filter(status=StatusItemChoices.DEFEITO).count()

    ctx = {
        "app_name": app_name,
        "version": version,
        "build_date": build_date,
        "now": timezone.now(),
        "totais": {
            "itens": total_itens,
            "ativos": total_ativos,
            "backup": total_backup,
            "manutencao": total_manut,
            "defeitos": total_defeito,
            "usuarios": Usuario.objects.count(),
            "localidades": Localidade.objects.count(),
            "centros": CentroCusto.objects.count(),
            "fornecedores": Fornecedor.objects.count(),
            "subtipos": Subtipo.objects.count(),
            "categorias": Categoria.objects.count(),
            "funcoes": Funcao.objects.count(),
            "licencas": Licenca.objects.count(),
            "movimentacoes": MovimentacaoItem.objects.count(),
        },
    }
    return render(request, "front/sobre_plataforma.html", ctx)

# ----------------- CHECKLIST: LIST/CRUD -----------------
def _calc_proxima_para_item(item: Item, base: date | None = None) -> date | None:
    """
    Calcula a próxima data de preventiva usando 'data_limite_preventiva' (dias) do Item.
    Se o item não exige preventiva, retorna None.
    """
    if not item or item.precisa_preventiva != SimNaoChoices.SIM:
        return None
    dias = item.data_limite_preventiva or 0
    if dias <= 0:
        return None
    base = base or timezone.localdate()
    return base + timedelta(days=int(dias))

# =========================
#   CHECKLIST - LIST
# =========================
@login_required
def checklist_list(request):
    """
    Listagem de Checklists com busca e paginação.
    """
    q = (request.GET.get("q") or "").strip()

    # QuerySet Base Otimizada
    # select_related('subtipo') evita queries N+1 se você exibir o subtipo no template
    qs = (
        CheckListModelo.objects
        .select_related('subtipo') 
        .annotate(perguntas_count=Count("perguntas")) 
        .order_by("-created_at") # Mais recentes primeiro
    )

    # Filtro Seguro (Removido 'descricao' que causava erro)
    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(subtipo__nome__icontains=q))

    # Paginação (10 itens por página)
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj, # O template iterará sobre page_obj
        "q": q,
        "total": qs.count(),
    }
    return render(request, "front/preventivas/checklist_list.html", context)

# =========================
#   CHECKLIST - CREATE/EDIT
# =========================
@login_required
def checklist_form(request, pk=None):
    instance = get_object_or_404(CheckListModelo, pk=pk) if pk else None
    if request.method == "POST":
        form = ChecklistModeloForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save(commit=False)
            if instance is None:
                obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Checklist salvo com sucesso.")
            return redirect("checklist_form", pk=obj.pk)
    else:
        form = ChecklistModeloForm(instance=instance)

    perguntas = CheckListPergunta.objects.filter(checklist_modelo=instance).order_by("id") if instance else []
    return render(
        request,
        "front/preventivas/checklist_form.html",
        {"form": form, "perguntas": perguntas},
    )

@login_required
def checklist_delete(request, pk):
    if request.method != "POST":
        messages.error(request, "Requisição inválida.")
        return redirect("checklist_list")
    obj = get_object_or_404(CheckListModelo, pk=pk)
    obj.delete()
    messages.success(request, "Checklist removido.")
    return redirect("checklist_list")

# =========================
#   PERGUNTA - CREATE/EDIT
# =========================
@login_required
def pergunta_form(request, checklist_pk, pk=None):
    checklist = get_object_or_404(CheckListModelo, pk=checklist_pk)
    instance = get_object_or_404(CheckListPergunta, pk=pk, checklist_modelo=checklist) if pk else None

    if request.method == "POST":
        form = ChecklistPerguntaForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.checklist_modelo = checklist
            if instance is None:
                obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Pergunta salva com sucesso.")
            return redirect("checklist_form", pk=checklist.pk)
    else:
        form = ChecklistPerguntaForm(instance=instance)

    return render(
        request,
        "front/preventivas/pergunta_form.html",
        {"form": form, "checklist": checklist},
    )

@login_required
def pergunta_delete(request, checklist_pk, pk):
    if request.method != "POST":
        messages.error(request, "Requisição inválida.")
        return redirect("checklist_form", pk=checklist_pk)
    checklist = get_object_or_404(CheckListModelo, pk=checklist_pk)
    pergunta = get_object_or_404(CheckListPergunta, pk=pk, checklist_modelo=checklist)
    pergunta.delete()
    messages.success(request, "Pergunta removida.")
    return redirect("checklist_form", pk=checklist.pk)

# =========================
#   PREVENTIVAS - LIST
# =========================
@login_required
def preventiva_list(request):
    """
    Listagem Inteligente de Preventivas.
    Calcula datas em tempo real e suporta filtros avançados.
    """
    q = (request.GET.get("q") or "").strip()
    status_filter = request.GET.get("status")
    
    # 1. QuerySet Otimizado
    qs = (
        Preventiva.objects
        .select_related("equipamento__localidade", "checklist_modelo")
        .order_by("data_proxima")
    )

    if q:
        qs = qs.filter(
            Q(equipamento__nome__icontains=q) | 
            Q(equipamento__patrimonio__icontains=q) |
            Q(equipamento__numero_serie__icontains=q) # Busca também por NS
        )

    # 2. Processamento de Datas e KPIs (Antes da paginação para contagem correta)
    today = timezone.localdate()
    next_week = today + timedelta(days=7)
    
    # Listas para contagem
    list_vencidas = []
    list_proximas = []
    list_ok = []
    
    processed_list = []

    for p in qs:
        # LÓGICA DE CÁLCULO DE DATA (Mesma do Detalhe)
        intervalo = 0
        
        # Prioridade: Equipamento > Modelo
        if hasattr(p.equipamento, 'data_limite_preventiva') and p.equipamento.data_limite_preventiva:
            try: intervalo = int(p.equipamento.data_limite_preventiva)
            except: pass
        elif p.checklist_modelo and p.checklist_modelo.intervalo_dias:
            try: intervalo = int(p.checklist_modelo.intervalo_dias)
            except: pass
            
        # Calcula data projetada
        if intervalo > 0:
            base = p.data_ultima if p.data_ultima else today
            p.proxima_calc = base + timedelta(days=intervalo)
        else:
            p.proxima_calc = p.data_proxima # Fallback

        # Define Status
        if not p.proxima_calc:
            p.status_visual = 'indefinido'
        else:
            delta = (p.proxima_calc - today).days
            p.dias_restantes = delta
            
            if delta < 0:
                p.status_visual = 'vencida'
                list_vencidas.append(p)
            elif delta <= 7:
                p.status_visual = 'atencao'
                list_proximas.append(p)
            else:
                p.status_visual = 'ok'
                list_ok.append(p)
        
        processed_list.append(p)

    # 3. Filtragem em Memória (Pois usamos campos calculados)
    if status_filter == 'vencida':
        final_list = [x for x in processed_list if x.status_visual == 'vencida']
    elif status_filter == 'proxima':
        final_list = [x for x in processed_list if x.status_visual == 'atencao']
    elif status_filter == 'ok':
        final_list = [x for x in processed_list if x.status_visual == 'ok']
    else:
        final_list = processed_list

    # 4. KPIs
    kpi = {
        "total": len(processed_list),
        "vencidas": len(list_vencidas),
        "proximas": len(list_proximas),
        "em_dia": len(list_ok)
    }

    # 5. Paginação manual sobre a lista processada
    paginator = Paginator(final_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "preventivas": page_obj,
        "kpi": kpi,
        "today": today,
        "filter_q": q,
        "filter_status": status_filter
    }
    
    if request.GET.get('print') == 'true':
        return render(request, "front/preventivas/preventiva_list_print.html", context)

    return render(request, "front/preventivas/preventiva_list.html", context)
# =========================
#   PREVENTIVA - START
# =========================
@login_required
def preventiva_start(request, item_id=None):
    """
    Tela Inicial: Seleção de Ativo e Checklist para gerar a Ordem de Serviço.
    """
    item_instance = None
    if item_id:
        item_instance = get_object_or_404(Item.objects.select_related("subtipo"), pk=item_id)

    if request.method == "POST":
        form = PreventivaStartForm(request.POST, item_instance=item_instance)
        if form.is_valid():
            item = form.cleaned_data["item"]
            modelo = form.cleaned_data["checklist_modelo"]
            
            # --- Lógica de Periodicidade Inicial ---
            # Define a próxima data baseada no intervalo configurado
            today = timezone.localdate()
            intervalo = 0
            
            # 1. Tenta pegar do Equipamento
            if item.data_limite_preventiva:
                try: intervalo = int(item.data_limite_preventiva)
                except: pass
            
            # 2. Se não, pega do Modelo
            if intervalo <= 0 and modelo.intervalo_dias:
                try: intervalo = int(modelo.intervalo_dias)
                except: pass
            
            # Se for a primeira vez, a "próxima" é hoje (para aparecer na lista de execução agora)
            # A data futura será calculada APÓS a execução ser finalizada.
            data_inicial = today

            # Cria ou recupera a Preventiva Ativa
            prev, created = Preventiva.objects.get_or_create(
                equipamento=item,
                checklist_modelo=modelo,
                defaults={
                    "criado_por": request.user,
                    "atualizado_por": request.user,
                    "data_ultima": None,
                    "data_proxima": data_inicial, 
                }
            )
            
            if not created:
                prev.atualizado_por = request.user
                # Se estava sem data (inativa), reativa para hoje
                if not prev.data_proxima:
                    prev.data_proxima = today
                prev.save(update_fields=["atualizado_por", "data_proxima", "updated_at"])
            
            messages.success(request, f"Ordem de serviço gerada para {item.nome}.")
            return redirect("preventiva_exec", pk=prev.pk)
    else:
        form = PreventivaStartForm(item_instance=item_instance)

    return render(request, "front/preventivas/preventiva_start.html", {"form": form})

# alias para a URL com item_id
@login_required
def preventiva_start_item(request, item_id):
    return preventiva_start(request, item_id=item_id)

# =========================
#   PREVENTIVA - DETAIL
# =========================
@login_required
def preventiva_detail(request, pk):
    """
    Dashboard detalhado da Preventiva.
    Calcula prazos baseados no intervalo do equipamento e exibe histórico.
    """
    preventiva = get_object_or_404(
        Preventiva.objects.select_related('equipamento__localidade', 'equipamento__centro_custo', 'checklist_modelo'),
        pk=pk
    )

    # 1. Filtros de Data
    ini = request.GET.get("inicio")
    fim = request.GET.get("fim")
    
    dt_ini = None
    dt_fim = None
    if ini:
        try: dt_ini = datetime.strptime(ini, "%Y-%m-%d").date()
        except: pass
    if fim:
        try: dt_fim = datetime.strptime(fim, "%Y-%m-%d").date()
        except: pass

    # 2. Busca Execuções (Ordenadas da mais recente)
    exec_qs = (
        preventiva.execucoes.all()
        .select_related("criado_por")
        .order_by("-data_execucao", "-id")
    )
    
    if dt_ini: exec_qs = exec_qs.filter(data_execucao__gte=dt_ini)
    if dt_fim: exec_qs = exec_qs.filter(data_execucao__lte=dt_fim)

    # 3. Perguntas (Ordem do Checklist)
    perguntas = CheckListPergunta.objects.filter(
        checklist_modelo=preventiva.checklist_modelo
    ).order_by("ordem", "id")

    # 4. Montagem dos dados para o template (Otimizado)
    execucoes_data = []
    for ex in exec_qs:
        # Mapeia respostas para acesso rápido
        resp_map = {r.pergunta_id: r for r in ex.respostas.all()}
        
        linhas = []
        for p in perguntas:
            r = resp_map.get(p.id)
            linhas.append({
                "texto": p.texto_pergunta,
                "resposta": r.resposta if r else "-",
                "respondido_em": r.created_at if r else None,
            })
        
        execucoes_data.append({
            "obj": ex,
            "linhas": linhas
        })

    # 5. Cálculo de Próxima Data e Status (Lógica solicitada)
    today = timezone.localdate()
    
    # Prioridade: Intervalo do Equipamento > Intervalo do Modelo
    intervalo_dias = 0
    origem_intervalo = "Manual"
    
    if preventiva.equipamento.data_limite_preventiva: # Assumindo que este campo guarda dias (int)
        try:
            intervalo_dias = int(preventiva.equipamento.data_limite_preventiva)
            origem_intervalo = "Cadastro do Equipamento"
        except: pass
    elif preventiva.checklist_modelo.intervalo_dias:
        intervalo_dias = int(preventiva.checklist_modelo.intervalo_dias)
        origem_intervalo = "Modelo de Checklist"

    # Se temos a última execução e um intervalo, calculamos a projeção
    proxima_calc = None
    if preventiva.data_ultima and intervalo_dias > 0:
        proxima_calc = preventiva.data_ultima + timedelta(days=intervalo_dias)
    else:
        # Fallback para o que está salvo no banco se não der para calcular
        proxima_calc = preventiva.data_proxima

    # Definição de Status Visual
    status_prazo = "ok" # ok, warning, late
    dias_restantes = 0
    
    if proxima_calc:
        delta = (proxima_calc - today).days
        dias_restantes = delta
        if delta < 0: status_prazo = "late"
        elif delta <= 7: status_prazo = "warning"

    context = {
        "preventiva": preventiva,
        "equipamento": preventiva.equipamento,
        "execucoes": execucoes_data,
        "today": today,
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "proxima_calc": proxima_calc,
        "status_prazo": status_prazo,
        "dias_restantes": dias_restantes,
        "intervalo_info": f"{intervalo_dias} dias ({origem_intervalo})",
    }
    
    # Se for pedido de impressão, renderiza template limpo
    if request.GET.get('print') == 'true':
        return render(request, "front/preventivas/preventiva_print.html", context)

    return render(request, "front/preventivas/preventiva_detail.html", context)


# =========================
#   PREVENTIVA - EXEC
# =========================
@login_required
def preventiva_exec(request, pk):
    """
    Tela de Execução do Checklist.
    Processa respostas, salva evidências e reagenda a próxima preventiva.
    """
    preventiva = get_object_or_404(Preventiva, pk=pk)

    # Carrega perguntas na ordem definida
    perguntas = (
        CheckListPergunta.objects
        .filter(checklist_modelo=preventiva.checklist_modelo)
        .order_by('ordem', 'id')
    )

    # Processa lista de opções para o template
    for p in perguntas:
        p.opcoes_list = []
        if getattr(p, "opcoes", None):
            p.opcoes_list = [o.strip() for o in str(p.opcoes).split(",") if o.strip()]

    if request.method == "POST":
        erros = []
        respostas_bulk = []

        # 1. Validação e Coleta
        for p in perguntas:
            field_name = f"r_{p.id}"
            raw_val = request.POST.get(field_name, "").strip()
            
            # Validação Obrigatória
            if p.obrigatorio == 'sim' and not raw_val:
                erros.append(f"A pergunta '{p.texto_pergunta}' é obrigatória.")
                continue

            # Validação Numérica
            tipo = str(p.tipo_resposta).lower()
            if raw_val and tipo in ('numero', 'inteiro', 'decimal'):
                try:
                    Decimal(raw_val.replace(',', '.')) # Aceita vírgula como decimal
                except:
                    erros.append(f"Valor inválido na pergunta '{p.texto_pergunta}'.")
                    continue
            
            # Prepara objeto (ainda não salvo)
            if raw_val:
                respostas_bulk.append(PreventivaResposta(
                    preventiva=preventiva,
                    pergunta=p,
                    resposta=raw_val,
                    criado_por=request.user,
                    atualizado_por=request.user
                ))

        if erros:
            for e in erros: messages.error(request, e)
            return render(request, "front/preventivas/preventiva_exec.html", {
                "preventiva": preventiva, "perguntas": perguntas
            })

        # 2. Transação Atômica (Segurança de Dados)
        with transaction.atomic():
            hoje = timezone.now().date()
            
            # A. Cria o registro mestre da execução
            execucao = PreventivaExecucao.objects.create(
                preventiva=preventiva,
                data_execucao=hoje,
                observacao=request.POST.get("observacao", ""),
                foto_antes=request.FILES.get("foto_antes"),
                foto_depois=request.FILES.get("foto_depois"),
                criado_por=request.user,
                atualizado_por=request.user
            )

            # B. Vincula respostas à execução e salva
            for r in respostas_bulk:
                r.execucao = execucao
            PreventivaResposta.objects.bulk_create(respostas_bulk)

            # C. Atualiza status e datas da Preventiva Pai
            intervalo = 0
            if preventiva.checklist_modelo and preventiva.checklist_modelo.intervalo_dias:
                try:
                    intervalo = int(preventiva.checklist_modelo.intervalo_dias)
                except:
                    intervalo = 0
            
            preventiva.data_ultima = hoje
            # Se tem intervalo, projeta próxima. Senão, fica sem data (avulsa)
            preventiva.data_proxima = (hoje + timedelta(days=intervalo)) if intervalo > 0 else None
            
            # Atualiza flags e evidências recentes (retrocompatibilidade)
            preventiva.updated_at = timezone.now()
            if request.FILES.get("foto_antes"):
                preventiva.foto_antes = request.FILES.get("foto_antes")
            if request.FILES.get("foto_depois"):
                preventiva.foto_depois = request.FILES.get("foto_depois")
                
            preventiva.save()

        messages.success(request, "Checklist finalizado com sucesso!")
        return redirect("preventiva_detail", pk=preventiva.pk)

    return render(request, "front/preventivas/preventiva_exec.html", {
        "preventiva": preventiva,
        "perguntas": perguntas
    })
# =========================
# Helpers de datas/séries
# =========================
def _month_key(dt):
    """YYYY-MM para indexação."""
    return f"{dt.year:04d}-{dt.month:02d}"

def _last_n_month_stamps(n=12):
    """Lista de (ano, mês) dos últimos n meses, do mais antigo ao mais recente."""
    now = timezone.localtime()
    y, m = now.year, now.month
    out = []
    for _ in range(n):
        out.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(out))

def _labels_pt_br(stamps):
    """Gera labels 'Mes/AnoCurto' ex.: Jan/25 a partir de (ano, mês)."""
    nomes = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    return [f"{nomes[m-1]}/{str(y)[-2:]}" for (y, m) in stamps]

def _align_series(stamps, qs_month_count, field_name="c"):
    """
    Alinha uma série mensal (dict {'YYYY-MM': count}) aos stamps fornecidos.
    qs_month_count: queryset com values('m').annotate(c=Count(...))
                    onde 'm' = TruncMonth(), retornado como datetime.
    """
    m2v = {}
    for row in qs_month_count:
        mdt = row["m"]
        if not isinstance(mdt, datetime):
            # TruncMonth retorna datetime/tz-aware
            continue
        m2v[_month_key(timezone.localtime(mdt))] = row[field_name]
    out = []
    for (y, m) in stamps:
        out.append(int(m2v.get(f"{y:04d}-{m:02d}", 0)))
    return out


# =========================
# Helpers de custo de licença
# =========================
def _custo_mensal_lic(lic: Licenca) -> Decimal:
    cm = lic.custo_mensal()  # helper implementado na sua model
    return cm if cm is not None else Decimal("0.00")

# ==============================================================================
# 1. MOTOR DE DADOS (Funções Auxiliares)
# ==============================================================================

def _generate_month_keys(months=12):
    """
    Gera as chaves (ano, mes) e os labels para os últimos N meses.
    Retorna: (lista_chaves, lista_labels)
    Ex: ([(2023, 1), ...], ['Jan/23', ...])
    """
    today = timezone.localdate()
    keys = []
    labels = []
    
    # Começa do primeiro dia do mês atual
    curr = today.replace(day=1)
    
    names = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

    for _ in range(months):
        k = (curr.year, curr.month)
        lbl = f"{names[curr.month]}/{str(curr.year)[2:]}"
        
        keys.append(k)
        labels.append(lbl)
        
        # Volta 1 mês
        curr = (curr - timedelta(days=1)).replace(day=1)
    
    # Inverte para ficar cronológico (Antigo -> Novo)
    return list(reversed(keys)), list(reversed(labels))

def _process_chart_data(keys, queryset):
    """
    Recebe um QuerySet agrupado por mês e alinha com as chaves de data.
    IMPORTANTE: O QuerySet DEVE ter o campo anotado como 'valor'.
    """
    # 1. Transforma o QuerySet em um Dicionário de busca rápida: {(ano, mes): valor}
    data_map = {}
    for item in queryset:
        dt = item.get('m') # 'm' é o TruncMonth
        val = item.get('valor') # 'valor' é o dado padronizado
        
        if dt:
            # Garante que None vire 0 (caso de Somas nulas)
            final_val = val if val is not None else 0
            data_map[(dt.year, dt.month)] = final_val

    # 2. Monta a lista final alinhada com as keys (preenche buracos com 0)
    result = []
    for k in keys:
        result.append(data_map.get(k, 0))
        
    return result

# ==============================================================================
# 2. VIEW PRINCIPAL
# ==============================================================================

@login_required
def dashboard(request):
    """
    Dashboard Enterprise - Refatorado para estabilidade total.
    """
    # --- Configuração de Tempo ---
    month_keys, labels = _generate_month_keys(12)
    
    # Data de corte (início do período)
    first_key = month_keys[0] # (Ano, Mes) mais antigo
    start_date = timezone.datetime(year=first_key[0], month=first_key[1], day=1)
    start_date = timezone.make_aware(start_date) # Adiciona timezone se necessário

    # --- A. KPIs (Topo) ---
    kpi = {
        "total": Item.objects.count(),
        "ativos": Item.objects.filter(status='ativo').count(),
        "estoque": Item.objects.filter(status='backup').count(),
        "manutencao": Item.objects.filter(status='manutencao').count(),
        "problema": Item.objects.filter(status__in=['defeito', 'sucata', 'queimado']).count(),
    }

    # --- B. Gráfico de Movimentações (Linha do Tempo) ---
    # Base: últimos 12 meses
    mov_base = MovimentacaoItem.objects.filter(created_at__gte=start_date)

    def get_mov_series(tipo_mov):
        """Helper interno para buscar movimentações padronizadas"""
        qs = (
            mov_base.filter(tipo_movimentacao=tipo_mov)
            .annotate(m=TruncMonth('created_at'))
            .values('m')
            .annotate(valor=Count('id')) # <--- PADRONIZAÇÃO: Nome sempre será 'valor'
            .order_by('m')
        )
        return _process_chart_data(month_keys, qs)

    series_mov = {
        "entrada": get_mov_series('entrada'),
        "baixa": get_mov_series('baixa'),
        "transf": get_mov_series('transferencia'),
        "manut": get_mov_series('envio_manutencao'), # Envios para manutenção
    }

    # --- C. Gráfico de Custos (Manutenção) ---
    # Soma dos custos de manutenção nos últimos 12 meses
    qs_custo = (
        mov_base.filter(tipo_movimentacao__in=['envio_manutencao', 'retorno_manutencao'])
        .annotate(m=TruncMonth('created_at'))
        .values('m')
        .annotate(valor=Sum('custo')) # <--- PADRONIZAÇÃO: Nome sempre será 'valor'
        .order_by('m')
    )
    data_custo = _process_chart_data(month_keys, qs_custo)

    # --- D. Preventivas (Status) ---
    today = timezone.localdate()
    prev_atrasadas = Preventiva.objects.filter(data_proxima__lt=today).count()
    prev_em_dia = Preventiva.objects.filter(data_proxima__gte=today).count()
    
    # Próximas a vencer (Tabela)
    prev_proximas = (
        Preventiva.objects
        .filter(data_proxima__gte=today)
        .select_related('equipamento', 'checklist_modelo')
        .order_by('data_proxima')[:5]
    )

    # --- E. Categorias (Barras) ---
    # Função genérica para Top N
    def get_top_category(field_name, limit=5):
        qs = (
            Item.objects.values(field_name)
            .annotate(valor=Count('id')) # Padronizado
            .order_by('-valor')[:limit]
        )
        # Trata valores nulos no nome
        labels_cat = [item[field_name] or 'Não Definido' for item in qs]
        data_cat = [item['valor'] for item in qs]
        return labels_cat, data_cat

    sub_labels, sub_data = get_top_category('subtipo__nome', 5)
    loc_labels, loc_data = get_top_category('localidade__local', 8)

    # --- Contexto Final ---
    context = {
        "kpi": kpi,
        "labels": labels, # Eixo X (Jan/24, Fev/24...)
        "series": series_mov,
        "custo_data": data_custo,
        "prev_status": [prev_em_dia, prev_atrasadas],
        "prev_proximas": prev_proximas,
        "cat_subtipo": {"labels": sub_labels, "data": sub_data},
        "cat_local": {"labels": loc_labels, "data": loc_data},
    }

    return render(request, "front/dashboards/dashboard.html", context)


# ==== helpers que você já usa em outros dashboards ====
def _month_key(dt):
    return f"{dt.year:04d}-{dt.month:02d}"

def _last_n_month_stamps(n=12):
    now = timezone.localtime()
    y, m = now.year, now.month
    out = []
    for _ in range(n):
        out.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(out))

def _labels_pt_br(stamps):
    nomes = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    return [f"{nomes[m-1]}/{str(y)[-2:]}" for (y, m) in stamps]

def _align_series(stamps, qs_month_count, field_name="c"):
    """Alinha uma série mensal (values('m').annotate(c=...)) em relação aos stamps fornecidos."""
    m2v = {}
    for row in qs_month_count:
        mdt = row["m"]
        if not isinstance(mdt, datetime):
            continue
        m2v[_month_key(timezone.localtime(mdt))] = int(row[field_name] or 0)
    out = []
    for (y, m) in stamps:
        out.append(int(m2v.get(f"{y:04d}-{m:02d}", 0)))
    return out


@login_required
def preventiva_dashboard(request):
    """
    Dashboard de Preventivas:
      - KPIs: totais, no prazo, vencidas, sem agenda, executadas no mês
      - Séries (12 meses): executadas x programadas
      - Tabelas: por checklist, localidade e subtipo
      - Listas: vencidas e próximas 30 dias (+ histórico)
    Filtros: q (item/obs), status (ok|vencida|sem_agenda|""), checklist (id),
             local (icontains), subtipo (icontains), inicio/fim (opcional p/ séries)
    """
    today = timezone.localdate()
    now = timezone.localtime()

    # -------- filtros básicos --------
    q       = (request.GET.get("q") or "").strip()
    status  = (request.GET.get("status") or "").strip()       # ok | vencida | sem_agenda | ""
    chk_id  = (request.GET.get("checklist") or "").strip()
    loc     = (request.GET.get("local") or "").strip()
    subtipo = (request.GET.get("subtipo") or "").strip()

    base = (Preventiva.objects
            .select_related(
                "equipamento",
                "equipamento__localidade",
                "equipamento__subtipo",
                "checklist_modelo",
            ))

    if q:
        base = base.filter(Q(equipamento__nome__icontains=q) | Q(observacao__icontains=q))
    if chk_id.isdigit():
        base = base.filter(checklist_modelo_id=int(chk_id))
    if loc:
        base = base.filter(equipamento__localidade__local__icontains=loc)
    if subtipo:
        base = base.filter(equipamento__subtipo__nome__icontains=subtipo)

    # Guardamos uma cópia SEM o filtro de status para KPIs/listas.
    # (Se preferir que o status também afete KPIs/listas, troque 'base_kpi' por 'base' nos cálculos.)
    base_kpi = base

    # Filtro de status (apenas para a visualização geral; KPIs usam base_kpi para não “zerar”)
    if status == "ok":
        base = base.filter(data_proxima__isnull=False, data_proxima__gte=today)
    elif status == "vencida":
        base = base.filter(data_proxima__lt=today)
    elif status == "sem_agenda":
        base = base.filter(data_proxima__isnull=True)

    # -------- KPIs (usando base_kpi para refletir a situação real do parque) --------
    total             = base_kpi.count()
    vencidas_count    = base_kpi.filter(data_proxima__lt=today).count()
    sem_agenda_count  = base_kpi.filter(data_proxima__isnull=True).count()
    ok_count          = base_kpi.filter(data_proxima__isnull=False, data_proxima__gte=today).count()
    executadas_mes    = base_kpi.filter(data_ultima__year=now.year, data_ultima__month=now.month).count()

    # -------- Séries 12 meses: executadas x programadas (também a partir de base_kpi) --------
    stamps12 = _last_n_month_stamps(12)
    labels12 = _labels_pt_br(stamps12)
    start12  = timezone.make_aware(datetime(stamps12[0][0], stamps12[0][1], 1))

    exec_qs = (base_kpi.filter(data_ultima__isnull=False, data_ultima__gte=start12)
                        .annotate(m=TruncMonth("data_ultima"))
                        .values("m")
                        .annotate(c=Count("id"))
                        .order_by("m"))
    prog_qs = (base_kpi.filter(data_proxima__isnull=False, data_proxima__gte=start12)
                        .annotate(m=TruncMonth("data_proxima"))
                        .values("m")
                        .annotate(c=Count("id"))
                        .order_by("m"))

    serie_exec = _align_series(stamps12, exec_qs)
    serie_prog = _align_series(stamps12, prog_qs)

    # -------- AGG por Checklist / Localidade / Subtipo (usando base_kpi) --------
    agg_chk = (base_kpi.values("checklist_modelo_id", "checklist_modelo__nome")
                      .annotate(
                          total=Count("id"),
                          vencidas=Count("id", filter=Q(data_proxima__lt=today)),
                          ok=Count("id", filter=Q(data_proxima__isnull=False, data_proxima__gte=today)),
                          sem_agenda=Count("id", filter=Q(data_proxima__isnull=True)),
                          prox_30=Count("id", filter=Q(data_proxima__gte=today,
                                                       data_proxima__lte=today + timedelta(days=30))),
                      )
                      .order_by("-total", "checklist_modelo__nome"))

    chk_labels, chk_rates = [], []
    for r in agg_chk[:8]:
        den = (r["ok"] or 0) + (r["vencidas"] or 0)
        taxa = (100.0 * (r["ok"] or 0) / den) if den > 0 else 0.0
        chk_labels.append(r["checklist_modelo__nome"] or "Sem checklist")
        chk_rates.append(round(taxa, 2))

    agg_loc = (base_kpi.values("equipamento__localidade__local")
                      .annotate(
                          total=Count("id"),
                          vencidas=Count("id", filter=Q(data_proxima__lt=today)),
                          ok=Count("id", filter=Q(data_proxima__isnull=False, data_proxima__gte=today)),
                          sem_agenda=Count("id", filter=Q(data_proxima__isnull=True)),
                      )
                      .order_by("-vencidas", "equipamento__localidade__local"))

    agg_sub = (base_kpi.values("equipamento__subtipo__nome")
                      .annotate(
                          total=Count("id"),
                          vencidas=Count("id", filter=Q(data_proxima__lt=today)),
                          ok=Count("id", filter=Q(data_proxima__isnull=False, data_proxima__gte=today)),
                          sem_agenda=Count("id", filter=Q(data_proxima__isnull=True)),
                      )
                      .order_by("-vencidas", "equipamento__subtipo__nome"))

    # -------- Listas operacionais (calculadas corretamente) --------
    vencidas = list(
        base_kpi.filter(data_proxima__lt=today)
                .order_by("data_proxima")
                .select_related("equipamento", "equipamento__localidade", "equipamento__subtipo", "checklist_modelo")[:50]
    )
    for p in vencidas:
        p.dias_atraso = (today - p.data_proxima).days if p.data_proxima else None

    proximas = list(
        base_kpi.filter(data_proxima__isnull=False,
                        data_proxima__gte=today,
                        data_proxima__lte=today + timedelta(days=30))
                .order_by("data_proxima")
                .select_related("equipamento", "equipamento__localidade", "equipamento__subtipo", "checklist_modelo")
    )
    for p in proximas:
        p.dias_faltam = (p.data_proxima - today).days if p.data_proxima else None

    historico = (base_kpi.filter(data_ultima__isnull=False)
                        .order_by("-data_ultima", "-updated_at")
                        .select_related("equipamento", "checklist_modelo")[:20])

    checklist_opts = CheckListModelo.objects.all().order_by("nome").values("id", "nome")

    ctx = dict(
        # filtros
        q=q, status=status, checklist=chk_id, local=loc, subtipo=subtipo,
        checklist_opts=checklist_opts,

        # KPIs
        total=total,
        ok_count=ok_count,
        vencidas_count=vencidas_count,
        sem_agenda_count=sem_agenda_count,
        executadas_mes=executadas_mes,

        # Séries
        serie_labels=labels12,
        serie_exec=serie_exec,
        serie_prog=serie_prog,

        # Tabelas
        agg_chk=list(agg_chk),
        agg_loc=list(agg_loc),
        agg_sub=list(agg_sub),

        # Listas
        vencidas=vencidas,
        proximas=proximas,
        historico=historico,

        # Gráfico de adesão por checklist
        chk_labels=chk_labels,
        chk_rates=chk_rates,

        today=today,
    )
    return render(request, "front/dashboards/preventiva_dashboard.html", ctx)

# ---- helpers de data ----
def _parse_date(date_str, default):
    if not date_str: return default
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return default

def _get_meses_ciclo(periodicidade_str):
    """Auxiliar para converter periodicidade em meses"""
    if not periodicidade_str: return 1
    p = str(periodicidade_str).upper()
    if 'MEN' in p: return 1
    if 'BI' in p: return 2
    if 'TRI' in p: return 3
    if 'SEM' in p: return 6
    if 'ANU' in p: return 12
    return 1

def _get_cc_custos_data(request):
    """
    Função Helper: Processa todos os cálculos de custo por Centro de Custo.
    Retorna o dicionário de contexto para ser usado na View Web ou no PDF.
    """
    hoje = timezone.localdate()
    dt_ini = _parse_date(request.GET.get("inicio"), hoje.replace(day=1))
    dt_fim = _parse_date(request.GET.get("fim"), hoje)

    totals = {}

    def get_acc(cc_id):
        if not cc_id: return None
        if cc_id not in totals:
            totals[cc_id] = {
                "cc_obj": None,
                "qtd_usuarios": 0,
                "qtd_itens": 0,
                "qtd_licencas": 0,
                "custo_locacao": Decimal("0.00"),
                "custo_licencas": Decimal("0.00"),
                "custo_baixas": Decimal("0.00"),
            }
        return totals[cc_id]

    # 1. Locação (Hardware)
    locacoes = Locacao.objects.select_related("equipamento__centro_custo").filter(
        equipamento__status='ativo', valor_mensal__gt=0, equipamento__centro_custo__isnull=False
    )
    for loc in locacoes:
        acc = get_acc(loc.equipamento.centro_custo.id)
        if acc: acc["custo_locacao"] += (loc.valor_mensal or Decimal(0))

    # 2. Licenças (Software - Unitário)
    movs_lic = MovimentacaoLicenca.objects.select_related(
        "licenca", "usuario__centro_custo", "centro_custo_destino", "lote"
    ).filter(usuario__isnull=False).order_by("licenca_id", "usuario_id", "created_at")

    estado_atual_lic = { (m.licenca_id, m.usuario_id): m for m in movs_lic }

    for (lid, uid), mov in estado_atual_lic.items():
        if mov.tipo == TipoMovLicencaChoices.ATRIBUICAO:
            cc_id = None
            if mov.usuario and mov.usuario.centro_custo:
                cc_id = mov.usuario.centro_custo.id
            elif mov.centro_custo_destino:
                cc_id = mov.centro_custo_destino.id
            elif mov.licenca.centro_custo:
                cc_id = mov.licenca.centro_custo.id
            
            acc = get_acc(cc_id)
            if acc:
                lote = mov.lote
                custo_mensal_unit = Decimal("0.00")
                if lote:
                    c_ciclo = lote.custo_ciclo or Decimal(0)
                    meses = _get_meses_ciclo(lote.periodicidade)
                    if meses > 0: custo_mensal_unit = (c_ciclo / Decimal(meses))
                else:
                    custo_mensal_unit = mov.licenca.custo or Decimal(0) # Fallback

                acc["custo_licencas"] += custo_mensal_unit
                acc["qtd_licencas"] += 1

    # 3. Baixas (Pontual)
    baixas = MovimentacaoItem.objects.filter(
        tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
        created_at__date__gte=dt_ini, created_at__date__lte=dt_fim
    ).select_related("item__centro_custo", "centro_custo_origem")

    for b in baixas:
        cc_id = b.centro_custo_origem.id if b.centro_custo_origem else (b.item.centro_custo.id if b.item.centro_custo else None)
        acc = get_acc(cc_id)
        if acc:
            val = b.custo if b.custo is not None else (b.item.valor or Decimal(0)) * (b.quantidade or 1)
            acc["custo_baixas"] += val

    # 4. Metadados (Nomes, Contagens)
    cc_ids = list(totals.keys())
    ccs_objs = CentroCusto.objects.filter(id__in=cc_ids)
    for cc in ccs_objs:
        if cc.id in totals: totals[cc.id]["cc_obj"] = cc

    users_agg = Usuario.objects.filter(centro_custo_id__in=cc_ids, status='ativo').values('centro_custo_id').annotate(n=Count('id'))
    for u in users_agg: 
        if u['centro_custo_id'] in totals: totals[u['centro_custo_id']]['qtd_usuarios'] = u['n']

    itens_agg = Item.objects.filter(centro_custo_id__in=cc_ids, status='ativo').values('centro_custo_id').annotate(n=Count('id'))
    for i in itens_agg:
        if i['centro_custo_id'] in totals: totals[i['centro_custo_id']]['qtd_itens'] = i['n']

    # 5. Consolidação
    linhas = []
    total_geral_itens = Decimal(0)
    total_geral_lics = Decimal(0)
    total_geral_baixas = Decimal(0)

    for cc_id, dados in totals.items():
        if not dados["cc_obj"]: continue

        c_itens = dados["custo_locacao"]
        c_lics = dados["custo_licencas"]
        c_baixas = dados["custo_baixas"]
        total_mensal = c_itens + c_lics
        total_impacto = total_mensal + c_baixas

        total_geral_itens += c_itens
        total_geral_lics += c_lics
        total_geral_baixas += c_baixas

        linhas.append({
            "cc": dados["cc_obj"],
            "usuarios": dados["qtd_usuarios"],
            "itens": dados["qtd_itens"],
            "licencas": dados["qtd_licencas"],
            "custo_itens": c_itens,
            "custo_licencas": c_lics,
            "baixas": c_baixas,
            "total_mensal": total_mensal,
            "total_impacto": total_impacto
        })

    linhas.sort(key=lambda x: x["total_impacto"], reverse=True)

    # Dados de Gráfico
    chart_labels = [f"{l['cc'].numero}" for l in linhas[:10]]
    chart_itens = [float(l['custo_itens']) for l in linhas[:10]]
    chart_lics = [float(l['custo_licencas']) for l in linhas[:10]]

    return {
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "linhas": linhas,
        
        # KPIs
        "kpi_cc_count": len(linhas),
        "kpi_total_mensal": total_geral_itens + total_geral_lics,
        "kpi_total_baixas": total_geral_baixas,
        "kpi_top_cc": linhas[0]['cc'].departamento if linhas else "-",
        
        # Charts (apenas para Web)
        "js_labels": chart_labels,
        "js_itens": chart_itens,
        "js_lics": chart_lics,
        "js_mix_values": [float(total_geral_itens), float(total_geral_lics)]
    }

@login_required
def cc_custos_dashboard(request):
    """
    Dashboard de Custos por Centro de Custo (Enterprise Version)
    - Locação de Hardware (Valor Mensal)
    - Licenças de Software (Valor Unitário Mensal x Qtd Usuários)
    - Baixas/Perdas (Valor Pontual no Período)
    """
    hoje = timezone.localdate()
    dt_ini = _parse_date(request.GET.get("inicio"), hoje.replace(day=1)) # Início do mês atual
    dt_fim = _parse_date(request.GET.get("fim"), hoje)

    # Estrutura de Acumulação: { cc_id: { dados... } }
    totals = {}

    def get_acc(cc_id):
        if not cc_id: return None
        if cc_id not in totals:
            totals[cc_id] = {
                "cc_obj": None,
                "qtd_usuarios": 0,
                "qtd_itens": 0,         # Hardware alocado
                "qtd_licencas": 0,      # Assentos de software
                "custo_locacao": Decimal("0.00"),  # Mensal Recorrente
                "custo_licencas": Decimal("0.00"), # Mensal Recorrente
                "custo_baixas": Decimal("0.00"),   # Pontual (Perda/Consumo)
            }
        return totals[cc_id]

    # ==========================================================
    # 1. CUSTO DE LOCAÇÃO (Hardware Recorrente)
    # ==========================================================
    # Itens ativos que possuem contrato de locação com valor mensal
    locacoes = (
        Locacao.objects
        .select_related("equipamento__centro_custo")
        .filter(
            equipamento__status='ativo', # Apenas ativos geram custo recorrente
            valor_mensal__gt=0,
            equipamento__centro_custo__isnull=False
        )
    )
    
    for loc in locacoes:
        cc_id = loc.equipamento.centro_custo.id
        acc = get_acc(cc_id)
        if acc:
            acc["custo_locacao"] += (loc.valor_mensal or Decimal(0))

    # ==========================================================
    # 2. CUSTO DE LICENÇAS (Software Recorrente - Lógica Unitária)
    # ==========================================================
    # Busca a última movimentação de cada par (licença, usuário) para saber quem está usando o que.
    
    movs_lic = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario__centro_custo", "centro_custo_destino", "lote")
        .filter(usuario__isnull=False) # Apenas atribuições a pessoas
        .order_by("licenca_id", "usuario_id", "created_at")
    )

    # Snapshot do estado atual
    estado_atual_lic = {}
    for m in movs_lic:
        estado_atual_lic[(m.licenca_id, m.usuario_id)] = m

    for (lid, uid), mov in estado_atual_lic.items():
        # Só conta custo se estiver atribuído (não devolvido)
        if mov.tipo == TipoMovLicencaChoices.ATRIBUICAO:
            
            # Determina o CC pagante (CC do Usuário > Destino > CC da Licença)
            cc_id = None
            if mov.usuario and mov.usuario.centro_custo:
                cc_id = mov.usuario.centro_custo.id
            elif mov.centro_custo_destino:
                cc_id = mov.centro_custo_destino.id
            elif mov.licenca.centro_custo: # Fallback para TI se usuário sem setor
                cc_id = mov.licenca.centro_custo.id
            
            acc = get_acc(cc_id)
            if acc:
                # CÁLCULO FINANCEIRO UNITÁRIO
                # Usa o lote vinculado na atribuição para saber o preço exato daquela unidade
                lote = mov.lote
                custo_mensal_unitario = Decimal("0.00")
                
                if lote:
                    custo_ciclo = lote.custo_ciclo or Decimal(0)
                    meses = _get_meses_ciclo(lote.periodicidade)
                    # Valor Mensal desta unidade = Valor Compra / Meses Ciclo
                    if meses > 0:
                        custo_mensal_unitario = (custo_ciclo / Decimal(meses))
                else:
                    # Fallback: Se não tem lote (dado legado), usa custo da licença pai
                    custo_base = mov.licenca.custo or Decimal(0) # Assumindo custo mensal na licença pai
                    # Se licença pai não tem periodicidade definida, assume 1 mês ou lógica custom
                    custo_mensal_unitario = custo_base 

                acc["custo_licencas"] += custo_mensal_unitario
                acc["qtd_licencas"] += 1

    # ==========================================================
    # 3. CUSTO DE BAIXAS (Perda/Descarte - Pontual no Período)
    # ==========================================================
    baixas = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim
        )
        .select_related("item__centro_custo", "centro_custo_origem")
    )

    for b in baixas:
        # O custo da baixa vai para o setor de origem (quem perdeu o item)
        # ou o setor dono do item
        cc_id = None
        if b.centro_custo_origem:
            cc_id = b.centro_custo_origem.id
        elif b.item.centro_custo:
            cc_id = b.item.centro_custo.id
        
        acc = get_acc(cc_id)
        if acc:
            # Se a movimentação tem custo explícito, usa ele. Se não, calcula qtd * valor_item
            custo_baixa = b.custo if b.custo is not None else (b.item.valor or Decimal(0)) * (b.quantidade or 1)
            acc["custo_baixas"] += custo_baixa

    # ==========================================================
    # 4. Dados Cadastrais para Contexto (Qtd Usuários e Itens)
    # ==========================================================
    cc_ids = list(totals.keys())
    
    # Busca objetos CentroCusto
    ccs_objs = CentroCusto.objects.filter(id__in=cc_ids)
    for cc in ccs_objs:
        if cc.id in totals:
            totals[cc.id]["cc_obj"] = cc

    # Contagem de Usuários Ativos
    users_agg = Usuario.objects.filter(centro_custo_id__in=cc_ids, status='ativo').values('centro_custo_id').annotate(n=Count('id'))
    for u in users_agg:
        if u['centro_custo_id'] in totals:
            totals[u['centro_custo_id']]['qtd_usuarios'] = u['n']

    # Contagem de Itens (Hardware em posse)
    itens_agg = Item.objects.filter(centro_custo_id__in=cc_ids, status='ativo').values('centro_custo_id').annotate(n=Count('id'))
    for i in itens_agg:
        if i['centro_custo_id'] in totals:
            totals[i['centro_custo_id']]['qtd_itens'] = i['n']

    # ==========================================================
    # 5. Montagem Final da Lista
    # ==========================================================
    linhas = []
    
    total_geral_itens = Decimal(0)
    total_geral_lics = Decimal(0)
    total_geral_baixas = Decimal(0)

    for cc_id, dados in totals.items():
        if not dados["cc_obj"]: continue # Pula se CC foi deletado mas tem histórico

        c_itens = dados["custo_locacao"]
        c_lics = dados["custo_licencas"]
        c_baixas = dados["custo_baixas"]
        
        # Total Mensal Recorrente
        total_mensal = c_itens + c_lics
        # Total Geral (Impacto Financeiro no período)
        total_impacto = total_mensal + c_baixas

        total_geral_itens += c_itens
        total_geral_lics += c_lics
        total_geral_baixas += c_baixas

        linhas.append({
            "cc": dados["cc_obj"],
            "usuarios": dados["qtd_usuarios"],
            "itens": dados["qtd_itens"],
            "licencas": dados["qtd_licencas"],
            "custo_itens": c_itens,
            "custo_licencas": c_lics,
            "baixas": c_baixas,
            "total_mensal": total_mensal,
            "total_impacto": total_impacto
        })

    # Ordenar por maior impacto financeiro
    linhas.sort(key=lambda x: x["total_impacto"], reverse=True)

    # Dados para Gráficos
    chart_labels = [f"{l['cc'].numero}" for l in linhas[:10]] # Top 10 para gráfico não quebrar
    chart_itens = [float(l['custo_itens']) for l in linhas[:10]]
    chart_lics = [float(l['custo_licencas']) for l in linhas[:10]]

    context = {
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "linhas": linhas,
        
        # Totais KPI
        "kpi_cc_count": len(linhas),
        "kpi_total_mensal": total_geral_itens + total_geral_lics,
        "kpi_total_baixas": total_geral_baixas,
        "kpi_top_cc": linhas[0]['cc'].departamento if linhas else "-",

        # Charts
        "js_labels": chart_labels,
        "js_itens": chart_itens,
        "js_lics": chart_lics,
        
        # Pizza Mix (Totais Gerais)
        "js_mix_values": [float(total_geral_itens), float(total_geral_lics)]
    }

    return render(request, "front/dashboards/cc_custos_dashboard.html", context)

@login_required
def cc_custos_export_pdf(request):
    """View de Exportação PDF"""
    # 1. Recupera os mesmos dados
    data = _get_cc_custos_data(request)
    
    # 2. Adiciona dados extras para o PDF
    data['usuario'] = request.user
    data['data_geracao'] = timezone.now()
    
    # 3. Renderiza Template PDF
    template_path = 'front/dashboards/cc_custos_pdf.html'
    template = get_template(template_path)
    html = template.render(data)
    
    # 4. Gera Arquivo
    response = HttpResponse(content_type='application/pdf')
    filename = f"relatorio_custos_cc_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
       return HttpResponse('Erro ao gerar PDF')
       
    return response

    
# ===== LISTA DE LICENÇAS (com cartões) =====
# ============ LICENÇAS ============

@login_required
def licenca_list(request):
    """
    Dashboard de Licenças (Enterprise View).
    Exibe listagem com saldo consolidado em tempo real baseado nos lotes.
    """
    # --- 1. Construção do QuerySet (Eager Loading para Performance) ---
    qs = (
        Licenca.objects
        .select_related("fornecedor", "centro_custo")
        .annotate(
            # KPI por Linha: Quantos lotes existem para esta licença?
            qtd_lotes=Count("lotes", distinct=True),
            
            # KPI Crítico: Soma o saldo disponível de cada lote vinculado
            # Se não tiver lotes, retorna 0 (Coalesce)
            estoque_real=Coalesce(Sum("lotes__quantidade_disponivel"), 0)
        )
        .order_by("nome")
    )

    # --- 2. Filtros Inteligentes ---
    q = request.GET.get("q", "").strip()
    fornecedor_id = request.GET.get("fornecedor", "").strip()
    pmb_filter = request.GET.get("pmb", "").strip()
    status_estoque = request.GET.get("status", "").strip()

    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(observacao__icontains=q))
    
    if fornecedor_id and fornecedor_id.isdigit():
        qs = qs.filter(fornecedor_id=fornecedor_id)
        
    if pmb_filter:
        qs = qs.filter(pmb=pmb_filter)

    # Filtro de Status de Estoque (Baseado na anotação calculada)
    if status_estoque == "com_estoque":
        qs = qs.filter(estoque_real__gt=0)
    elif status_estoque == "sem_estoque":
        qs = qs.filter(estoque_real=0)

    # --- 3. KPIs Globais (Cards do Topo) ---
    # Calculamos totais rápidos para o gestor ter visão macro
    kpi_total_licencas = Licenca.objects.count()
    
    # Soma total de assentos disponíveis na empresa inteira
    kpi_total_assentos = LicencaLote.objects.aggregate(
        total=Coalesce(Sum('quantidade_disponivel'), 0)
    )['total']
    
    kpi_pmb = Licenca.objects.filter(pmb=SimNaoChoices.SIM).count()

    # --- 4. Paginação e Controle ---
    try:
        per_page = int(request.GET.get("pp", 15))
    except ValueError:
        per_page = 15
    
    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_num)

    # Preserva filtros na paginação (query string)
    get_copy = request.GET.copy()
    if "page" in get_copy: del get_copy["page"]
    qs_keep = get_copy.urlencode()

    # --- 5. Contexto para o Template ---
    context = {
        "page_obj": page_obj,
        "qs_keep": qs_keep,
        "total_registros": qs.count(),
        
        # KPIs
        "kpi_total": kpi_total_licencas,
        "kpi_assentos": kpi_total_assentos,
        "kpi_pmb": kpi_pmb,

        # Estado dos Filtros (para manter selecionado)
        "filter_q": q,
        "filter_fornecedor": int(fornecedor_id) if fornecedor_id.isdigit() else "",
        "filter_pmb": pmb_filter,
        "filter_status": status_estoque,
        "per_page": per_page,

        # Opções para Dropdowns
        "opt_fornecedores": Fornecedor.objects.values("id", "nome").order_by("nome"),
        "opt_pmb": SimNaoChoices.choices,
    }

    return render(request, "front/licencas/licenca_list.html", context)

@login_required
def licenca_form(request, pk=None):
    """
    View simplificada para Cadastro/Edição de Licença (4 campos).
    """
    # Se houver PK, é edição. Senão, criação.
    obj = get_object_or_404(Licenca, pk=pk) if pk else None

    if request.method == "POST":
        form = LicencaForm(request.POST, instance=obj)
        if form.is_valid():
            try:
                licenca = form.save(commit=False)
                
                # Preenche auditoria
                if not obj:
                    licenca.criado_por = request.user
                licenca.atualizado_por = request.user
                
                licenca.save()
                
                verb = "editada" if obj else "criada"
                messages.success(request, f"Licença '{licenca.nome}' {verb} com sucesso!")
                return redirect("licenca_list")
                
            except Exception as e:
                messages.error(request, f"Erro crítico ao salvar: {e}")
        else:
            messages.error(request, "Verifique os campos obrigatórios.")
    else:
        form = LicencaForm(instance=obj)

    return render(request, "front/licencas/licenca_form.html", {
        "form": form,
        "obj": obj
    })


# --- HELPER: Calcula Alocação por Centro de Custo ---
def _get_dados_cc(licenca):
    """
    Reconstitui o estado atual das licenças para agrupar por Centro de Custo.
    Lógica: Pega todas as movimentações ordenadas. 
    Se 'atribuicao' -> Adiciona usuário. Se 'devolucao' -> Remove usuário.
    """
    movs = MovimentacaoLicenca.objects.filter(licenca=licenca).select_related(
        'usuario', 'centro_custo_destino'
    ).order_by('created_at')

    # 1. Descobrir quem está ativo e qual seu custo atual
    ativos = {} # {usuario_id: MovimentacaoObj}
    
    for mov in movs:
        if mov.tipo == 'atribuicao':
            ativos[mov.usuario_id] = mov
        elif mov.tipo == 'devolucao':
            ativos.pop(mov.usuario_id, None)

    # 2. Agrupar por Centro de Custo
    cc_stats = defaultdict(lambda: {'nome': 'Não Definido', 'qtd': 0, 'total': Decimal(0)})

    for mov in ativos.values():
        cc = mov.centro_custo_destino
        cc_id = cc.id if cc else 'na'
        cc_name = cc.departamento if cc else 'Sem Centro de Custo'
        
        cc_stats[cc_id]['nome'] = cc_name
        cc_stats[cc_id]['qtd'] += 1
        cc_stats[cc_id]['total'] += (mov.valor_unitario or Decimal(0))

    # Retorna lista ordenada pelo maior valor total
    return sorted(cc_stats.values(), key=lambda x: x['total'], reverse=True)

@login_required
def licenca_detail(request, pk):
    licenca = get_object_or_404(
        Licenca.objects.select_related("fornecedor", "centro_custo"), 
        pk=pk
    )

    # ... (Lógica de Lotes e KPIs Financeiros - MANTIDA IGUAL AO ANTERIOR) ...
    # (Vou resumir a parte repetida para focar na novidade)
    
    lotes_qs = LicencaLote.objects.filter(licenca=licenca).select_related(
        "fornecedor", "centro_custo"
    ).order_by("-data_compra", "-id")
    lotes = list(lotes_qs)

    qtd_total = 0
    qtd_disp = 0
    total_investido_historico = Decimal(0)
    burn_rate_mensal = Decimal(0)
    burn_rate_anual = Decimal(0)

    for lote in lotes:
        qtd_total += lote.quantidade_total
        qtd_disp += lote.quantidade_disponivel
        unit_price = lote.custo_ciclo or Decimal(0)
        lote_investimento = unit_price * lote.quantidade_total
        total_investido_historico += lote_investimento
        lote.unitario_real = unit_price
        lote.total_investido_calc = lote_investimento
        
        em_uso = lote.quantidade_total - lote.quantidade_disponivel
        if em_uso > 0:
            periodicidade = str(lote.periodicidade).lower()
            if periodicidade == 'anual':
                burn_rate_mensal += (unit_price / 12) * em_uso
                burn_rate_anual += unit_price * em_uso
            elif periodicidade == 'semestral':
                burn_rate_mensal += (unit_price / 6) * em_uso
                burn_rate_anual += (unit_price * 2) * em_uso
            else:
                burn_rate_mensal += unit_price * em_uso
                burn_rate_anual += (unit_price * 12) * em_uso

    qtd_em_uso = max(0, qtd_total - qtd_disp)
    pct_uso = int((qtd_em_uso / qtd_total * 100)) if qtd_total > 0 else 0

    movimentacoes = MovimentacaoLicenca.objects.filter(licenca=licenca).select_related(
        "usuario", "lote", "centro_custo_destino", "criado_por"
    ).order_by("-created_at")[:50]

    # [NOVO] Agrupamento por Centro de Custo
    cc_list = _get_dados_cc(licenca)

    context = {
        "obj": licenca,
        "kpi": {
            "total": qtd_total,
            "disponivel": qtd_disp,
            "em_uso": qtd_em_uso,
            "pct_uso": pct_uso,
            "investimento_total": total_investido_historico,
            "gasto_mensal": burn_rate_mensal,
            "gasto_anual": burn_rate_anual
        },
        "lotes": lotes,
        "movimentacoes": movimentacoes,
        "cc_list": cc_list, # Enviando para o template
    }

    return render(request, "front/licencas/licenca_detail.html", context)
# --- NOVA VIEW: Exportação PDF ---
@login_required
def licenca_export_pdf(request, pk):
    licenca = get_object_or_404(Licenca, pk=pk)
    
    # Recalcula dados essenciais para o relatório (KPIs + CCs)
    # Reutilizando a lógica simplificada ou chamando o helper
    cc_list = _get_dados_cc(licenca)
    
    # Totais dos CCs
    total_alocado_valor = sum(item['total'] for item in cc_list)
    total_alocado_qtd = sum(item['qtd'] for item in cc_list)

    # Dados de Lote para Inventário
    lotes = LicencaLote.objects.filter(licenca=licenca)
    estoque_total = lotes.aggregate(t=Coalesce(Sum('quantidade_total'),0))['t']
    estoque_disp = lotes.aggregate(d=Coalesce(Sum('quantidade_disponivel'),0))['d']

    context = {
        "obj": licenca,
        "cc_list": cc_list,
        "kpi": {
            "alocado_qtd": total_alocado_qtd,
            "alocado_valor": total_alocado_valor,
            "estoque_total": estoque_total,
            "estoque_disp": estoque_disp
        },
        "user_solicitante": request.user
    }

    template_path = 'front/licencas/licenca_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_licenca_{pk}.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Erro ao gerar PDF', status=500)
    return response

# ============ MOVIMENTAÇÕES ============

@login_required
def mov_licenca_list(request):
    """
    Listagem de Movimentações de Licenças com filtros e paginação.
    """
    # Filtros
    q = (request.GET.get("q") or "").strip()
    tipo = (request.GET.get("tipo") or "").strip()

    # QuerySet Otimizado
    qs = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario", "centro_custo_destino")
        .order_by("-created_at")
    )

    if q:
        qs = qs.filter(
            Q(licenca__nome__icontains=q) | 
            Q(usuario__nome__icontains=q)
        )
    
    # Validação do Tipo (segurança)
    valid_types = [choice[0] for choice in MovimentacaoLicenca._meta.get_field("tipo").choices]
    if tipo in valid_types:
        qs = qs.filter(tipo=tipo)

    # Paginação (Padrão 20 itens)
    try:
        per_page = int(request.GET.get("pp", 20))
        per_page = max(10, min(per_page, 100)) # Limites de segurança
    except ValueError:
        per_page = 20

    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_num)

    # Preserva filtros na paginação
    get_copy = request.GET.copy()
    get_copy.pop("page", None)
    qs_keep = get_copy.urlencode()

    context = {
        "movs": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "qs_keep": qs_keep,
        "q": q,
        "tipo": tipo,
        "tipos": MovimentacaoLicenca._meta.get_field("tipo").choices,
        "total": qs.count()
    }

    return render(request, "front/licencas/mov_licenca_list.html", context)

@login_required
def mov_licenca_form(request):
    initial = {}
    if "licenca" in request.GET: initial["licenca"] = request.GET.get("licenca")
    if "usuario" in request.GET: initial["usuario"] = request.GET.get("usuario")

    if request.method == "POST":
        form = MovimentacaoLicencaForm(request.POST)
        if form.is_valid():
            try:
                mov = form.save(user=request.user)
                
                # Feedback detalhado
                lote_txt = f"Lote #{mov.lote.pk}" if mov.lote else "N/A"
                cc_txt = mov.centro_custo_destino.departamento if mov.centro_custo_destino else "N/A"
                
                messages.success(request, f"{mov.get_tipo_display()} realizada. Estoque: {lote_txt} | Custo: {cc_txt}")
                return redirect("licenca_list")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
    else:
        form = MovimentacaoLicencaForm(initial=initial)

    # JSON para Select2 (Apenas lotes com saldo)
    lotes_qs = LicencaLote.objects.filter(quantidade_disponivel__gt=0).values(
        'id', 'licenca_id', 'quantidade_disponivel', 'numero_pedido', 'data_compra'
    )
    
    lotes_dict = {}
    for l in lotes_qs:
        lid = str(l['licenca_id'])
        if lid not in lotes_dict: lotes_dict[lid] = []
        dt = l['data_compra'].strftime('%d/%m/%Y') if l['data_compra'] else "-"
        txt = f"Lote #{l['id']} - Disp: {l['quantidade_disponivel']} ({dt})"
        lotes_dict[lid].append({'id': l['id'], 'text': txt})

    context = {
        "form": form,
        "lotes_json": lotes_dict,
        "pre_selected_lote": request.POST.get("lote_id_select") or ""
    }
    return render(request, "front/licencas/mov_licenca_form.html", context)
# --- LISTA DE LOTES ---
@login_required
def licenca_lote_list(request):
    """
    Lista de Lotes com busca avançada e layout otimizado.
    """
    q = request.GET.get("q", "").strip()
    
    # QueryBase com select_related para evitar N+1 queries
    qs = (
        LicencaLote.objects
        .select_related("licenca", "fornecedor", "centro_custo")
        .order_by("-created_at")
    )

    # Filtro Textual
    if q:
        qs = qs.filter(
            Q(licenca__nome__icontains=q) | 
            Q(numero_pedido__icontains=q) | 
            Q(observacao__icontains=q)
        )

    return render(request, "front/licencas/licenca_lote_list.html", {
        "lotes": qs,
        "q": q
    })

@login_required
@transaction.atomic
def licenca_lote_form(request, pk=None):
    """
    View Inteligente para Gestão de Lotes.
    Calcula automaticamente a disponibilidade baseada na entrada.
    """
    obj = get_object_or_404(LicencaLote, pk=pk) if pk else None

    if request.method == "POST":
        form = LicencaLoteForm(request.POST, instance=obj)
        if form.is_valid():
            lote = form.save(commit=False)
            
            # --- LÓGICA DE SALDO ---
            if not obj:
                # Novo Lote: Disponível = Total
                lote.quantidade_disponivel = lote.quantidade_total
                lote.criado_por = request.user # Auditoria
            else:
                # Edição: Ajusta o disponível pela diferença do total
                # Ex: Tinha 10 (8 disp, 2 uso). Editou total para 15 (+5).
                # Novo Disp = 8 + 5 = 13. Usados continuam 2.
                diff = lote.quantidade_total - obj.quantidade_total
                lote.quantidade_disponivel = obj.quantidade_disponivel + diff
            
            lote.atualizado_por = request.user # Auditoria
            lote.save()
            
            msg = f"Lote #{lote.pk} atualizado com sucesso!" if obj else "Lote criado com sucesso!"
            messages.success(request, msg)
            return redirect("licenca_lote_list")
        else:
            messages.error(request, "Verifique os erros no formulário abaixo.")
    else:
        form = LicencaLoteForm(instance=obj)

    return render(request, "front/licencas/licenca_lote_form.html", {
        "form": form,
        "obj": obj
    })




### exportações 

def _aplicar_filtros_equipamentos(request, base_qs=None):
    """
    Aplica os mesmos filtros da tela:
      nome, subtipo, status, numero_serie, fornecedor, localidade, centro_custo
    """
    qs = base_qs or Item.objects.all()
    qs = qs.select_related("subtipo", "localidade", "fornecedor", "centro_custo")

    nome = (request.GET.get("nome") or "").strip()
    if nome:
        qs = qs.filter(nome__icontains=nome)

    subtipo = (request.GET.get("subtipo") or "").strip()
    if subtipo:
        qs = qs.filter(subtipo_id=subtipo)

    status = (request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(status=status)

    numero_serie = (request.GET.get("numero_serie") or "").strip()
    if numero_serie:
        qs = qs.filter(numero_serie__icontains=numero_serie)

    fornecedor = (request.GET.get("fornecedor") or "").strip()
    if fornecedor:
        qs = qs.filter(fornecedor__nome__icontains=fornecedor)

    localidade = (request.GET.get("localidade") or "").strip()
    if localidade:
        qs = qs.filter(localidade__local__icontains=localidade)

    centro = (request.GET.get("centro_custo") or "").strip()
    if centro:
        qs = qs.filter(
            Q(centro_custo__numero__icontains=centro) |
            Q(centro_custo__departamento__icontains=centro)
        )

    return qs.order_by("nome", "id")


@login_required
def toner_cc_export_excel(request):
    """
    Exporta para Excel o custo de TONER por Centro de Custo, no período filtrado.
    Regra: considerar MovimentacaoItem do tipo 'baixa' cujo item.subtipo contém 'toner',
    agrupando por centro_custo_destino.
    Custo = quantidade * item.valor (se None, trata como 0).
    """
    # --- período (fallback: mês atual até hoje) ---
    hoje = timezone.localdate()
    dt_ini = parse_date(request.GET.get("inicio") or "") or hoje.replace(day=1)
    dt_fim = parse_date(request.GET.get("fim") or "") or hoje

    # --- queryset base: BAIXAS de TONER com CC destino válido ---
    qs = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
            item__subtipo__nome__icontains="toner",
            centro_custo_destino__isnull=False,
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .select_related("item", "centro_custo_destino", "localidade_destino")
    )

    # --- anotações seguras (evita mixed types) ---
    qty_dec = Cast(F("quantidade"), DecimalField(max_digits=12, decimal_places=2))
    item_val = Coalesce(
        Cast(F("item__valor"), DecimalField(max_digits=12, decimal_places=2)),
        V(Decimal("0.00")),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    linha_total = qty_dec * item_val  # Decimal * Decimal
    # Para agrupar por Centro de Custo:
    grp = (
        qs.values(
            "centro_custo_destino",
            "centro_custo_destino__numero",
            "centro_custo_destino__departamento",
        )
        .annotate(
            total_qtd=Coalesce(Sum("quantidade"), V(0)),
            total_valor=Coalesce(
                Sum(Cast(linha_total, DecimalField(max_digits=18, decimal_places=2))),
                V(Decimal("0.00")),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ),
            itens_distintos=Count("item", distinct=True),
            movs=Count("id"),
        )
        .order_by("centro_custo_destino__numero", "centro_custo_destino__departamento")
    )

    # --- planilha ---
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo por CC"

    # estilos
    header_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    header_font = Font(bold=True, color="001e3a")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # cabeçalho
    ws1.append(["Período", f"{dt_ini.strftime('%d/%m/%Y')} a {dt_fim.strftime('%d/%m/%Y')}"])
    ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

    ws1.append(["Centro de Custo", "Departamento", "Movimentações", "Itens Distintos",
                "Quantidade Baixada", "Valor Total (R$)"])
    for c in range(1, 7):
        cell = ws1.cell(row=2, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # linhas
    total_qtd_geral = 0
    total_valor_geral = Decimal("0.00")

    r = 3
    for row in grp:
        cc_num = row["centro_custo_destino__numero"] or "-"
        cc_dep = row["centro_custo_destino__departamento"] or "-"
        movs = row["movs"] or 0
        itens_d = row["itens_distintos"] or 0
        qtd = int(row["total_qtd"] or 0)
        val = Decimal(row["total_valor"] or 0).quantize(Decimal("0.01"))

        ws1.cell(row=r, column=1, value=str(cc_num))
        ws1.cell(row=r, column=2, value=str(cc_dep))
        ws1.cell(row=r, column=3, value=movs)
        ws1.cell(row=r, column=4, value=itens_d)
        ws1.cell(row=r, column=5, value=qtd)
        c6 = ws1.cell(row=r, column=6, value=float(val))
        c6.number_format = 'R$ #,##0.00'

        for c in range(1, 7):
            ws1.cell(row=r, column=c).border = border

        total_qtd_geral += qtd
        total_valor_geral += val
        r += 1

    # totalizador
    ws1.append(["", "", "", "Totais:", total_qtd_geral, float(total_valor_geral)])
    ws1.cell(row=r, column=4).font = Font(bold=True)
    ws1.cell(row=r, column=5).font = Font(bold=True)
    ws1.cell(row=r, column=6).font = Font(bold=True)
    ws1.cell(row=r, column=6).number_format = 'R$ #,##0.00'
    for c in range(1, 7):
        ws1.cell(row=r, column=c).border = border

    # larguras
    widths = [20, 36, 18, 18, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # --- aba de detalhes (opcional, mas útil) ---
    ws2 = wb.create_sheet("Detalhes")
    ws2.append(["Data", "Centro de Custo", "Departamento", "Item", "Subtipo",
                "Quantidade", "Valor Unitário (R$)", "Valor Total (R$)"])

    for c in range(1, 9):
        cell = ws2.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    r = 2
    for m in qs.order_by("centro_custo_destino__numero", "created_at", "id"):
        valor_unit = Decimal(m.item.valor or 0).quantize(Decimal("0.01"))
        qtd = Decimal(m.quantidade or 0)
        val_total = (valor_unit * qtd).quantize(Decimal("0.01"))

        ws2.cell(row=r, column=1, value=m.created_at.strftime("%d/%m/%Y %H:%M"))
        ws2.cell(row=r, column=2, value=getattr(m.centro_custo_destino, "numero", "") or "-")
        ws2.cell(row=r, column=3, value=getattr(m.centro_custo_destino, "departamento", "") or "-")
        ws2.cell(row=r, column=4, value=m.item.nome if m.item_id else "-")
        ws2.cell(row=r, column=5, value=getattr(m.item.subtipo, "nome", "") if m.item_id and m.item.subtipo_id else "-")
        ws2.cell(row=r, column=6, value=int(m.quantidade or 0))
        c7 = ws2.cell(row=r, column=7, value=float(valor_unit))
        c7.number_format = 'R$ #,##0.00'
        c8 = ws2.cell(row=r, column=8, value=float(val_total))
        c8.number_format = 'R$ #,##0.00'

        for c in range(1, 9):
            ws2.cell(row=r, column=c).border = border
        r += 1

    for i, w in enumerate([18, 18, 28, 36, 20, 14, 22, 22], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- resposta HTTP (arquivo xlsx) ---
    filename = f"custo_toner_por_cc_{dt_ini.strftime('%Y%m%d')}_{dt_fim.strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@login_required
def equipamentos_exportar(request):
    # 1) Query filtrada igual à lista
    qs = _aplicar_filtros_equipamentos(request)

    # 2) Mapa de locação mensal por equipamento (sua FK é 'equipamento')
    ids = list(qs.values_list("id", flat=True))
    loc_map = dict(
        Locacao.objects
        .filter(equipamento_id__in=ids)
        .values_list("equipamento_id", "valor_mensal")
    )

    # 3) Monta planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Itens"

    header = [
        "#", "Nome", "Subtipo", "Status", "Localidade",
        "Nº Série", "Fornecedor", "Centro de Custo",
        "Locado", "Locação (R$/mês)", "Valor aquisição (R$)"
    ]
    ws.append(header)

    # Estilo do header
    hfill = PatternFill("solid", fgColor="1D4ED8")
    hfont = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDE3EE")
    hborder = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = align_center
        cell.border = hborder

    # Linhas
    def _fmt_cc(obj):
        if not obj: return "-"
        num = getattr(obj, "numero", None) or ""
        dep = getattr(obj, "departamento", None) or ""
        return f"{num} - {dep}".strip(" -")

    numero_format = 'R$ #,##0.00'

    for i, it in enumerate(qs, start=1):
        subtipo = getattr(it.subtipo, "nome", "-") if getattr(it, "subtipo", None) else "-"
        status_disp = getattr(it, "get_status_display", None)
        status_txt = status_disp() if callable(status_disp) else (it.status or "-")
        local = getattr(it.localidade, "local", "-") if getattr(it, "localidade", None) else "-"
        fornecedor = getattr(it.fornecedor, "nome", "-") if getattr(it, "fornecedor", None) else "-"
        cc_txt = _fmt_cc(getattr(it, "centro_custo", None))

        locado_flag = getattr(it, "locado", None)
        # seu choices parecem "sim"/"nao" → ajuste se for booleano
        locado_txt = "Sim" if str(locado_flag).lower() in ("sim", "true", "1") else "Não"
        loc_mensal = loc_map.get(it.id) or Decimal("0.00")

        valor_aquis = getattr(it, "valor", None) or Decimal("0.00")

        row = [
            i,
            it.nome or "",
            subtipo,
            status_txt,
            local,
            it.numero_serie or "",
            fornecedor,
            cc_txt,
            locado_txt,
            float(loc_mensal),
            float(valor_aquis),
        ]
        ws.append(row)

    # Formatação de números nos 2 últimos campos
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=10).number_format = numero_format
        ws.cell(row=r, column=11).number_format = numero_format

    # Largura das colunas (auto-ajuste simples)
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            txt = str(val) if val is not None else ""
            widths[idx] = max(widths.get(idx, 0), len(txt))
    for idx, w in widths.items():
        # um pouco de folga
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 10), 48)

    # 4) Resposta HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    now = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    filename = f"itens_filtrados_{now}.xlsx"
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _parse_dt(s, default_dt):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return default_dt

@login_required
def toner_cc_dashboard(request):
    hoje = date.today()
    dt_ini = _parse_dt(request.GET.get("inicio") or "", date(hoje.year, 1, 1))
    dt_fim = _parse_dt(request.GET.get("fim") or "", hoje)

    # --- Query Base: Baixas de Toner ---
    base = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao="baixa",
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .filter(
            Q(item__subtipo__nome__icontains="toner") |
            Q(item__categoria__nome__icontains="toner") |
            Q(item__nome__icontains="toner")
        )
        .select_related("item", "centro_custo_origem", "centro_custo_destino")
    )

    # --- Definições de Cálculo Financeiro ---
    dec14_2 = DecimalField(max_digits=14, decimal_places=2)
    qtd_dec = Cast(F("quantidade"), output_field=dec14_2)
    preco_item = Coalesce(F("item__valor"), V(Decimal("0.00"), output_field=dec14_2), output_field=dec14_2)

    custo_total_expr = Case(
        When(custo__gt=Decimal("0.00"), then=F("custo")),
        default=ExpressionWrapper(qtd_dec * preco_item, output_field=dec14_2),
        output_field=dec14_2,
    )

    # --- Agregação por Centro de Custo ---
    base_cc = base.annotate(
        cc_id=Coalesce(F("centro_custo_origem_id"), F("centro_custo_destino_id"), F("item__centro_custo_id")),
        cc_numero=Case(
            When(centro_custo_origem__isnull=False, then=F("centro_custo_origem__numero")),
            When(centro_custo_destino__isnull=False, then=F("centro_custo_destino__numero")),
            default=F("item__centro_custo__numero"),
            output_field=CharField(),
        ),
        cc_departamento=Case(
            When(centro_custo_origem__isnull=False, then=F("centro_custo_origem__departamento")),
            When(centro_custo_destino__isnull=False, then=F("centro_custo_destino__departamento")),
            default=F("item__centro_custo__departamento"),
            output_field=CharField(),
        ),
    ).filter(cc_id__isnull=False)

    por_cc_qs = (
        base_cc.values("cc_id", "cc_numero", "cc_departamento")
        .annotate(
            qtd=Coalesce(Sum("quantidade"), V(0)),
            gasto=Coalesce(Sum(custo_total_expr, output_field=dec14_2), V(Decimal("0.00"), output_field=dec14_2)),
        )
        .order_by("-gasto") # Ordenar por maior gasto para pegar o Top 1 fácil
    )

    linhas, cc_labels, cc_gasto = [], [], []
    total_geral = Decimal("0.00")
    total_qtd = 0
    
    # KPIs
    top_cc_nome = "—"
    top_cc_valor = Decimal("0.00")

    for i, r in enumerate(por_cc_qs):
        cc_nome = f'{r["cc_numero"]} - {r["cc_departamento"]}'
        gasto = Decimal(r["gasto"] or 0)
        qtd = int(r["qtd"] or 0)
        
        # Pega o Top 1
        if i == 0:
            top_cc_nome = r["cc_departamento"] or r["cc_numero"]
            top_cc_valor = gasto

        linhas.append({"cc": cc_nome, "qtd": qtd, "gasto": gasto})
        cc_labels.append(cc_nome)
        cc_gasto.append(float(gasto))
        
        total_geral += gasto
        total_qtd += qtd

    # KPI: Ticket Médio (Custo por Item Baixado)
    ticket_medio = (total_geral / total_qtd) if total_qtd > 0 else Decimal("0.00")

    # --- Top Consumidores (Usuários) ---
    por_user_qs = (
        base.values("usuario__id", "usuario__nome")
        .annotate(gasto=Coalesce(Sum(custo_total_expr, output_field=dec14_2), V(Decimal("0.00"), output_field=dec14_2)))
        .order_by("-gasto")[:10]
    )
    user_labels = [r["usuario__nome"] or "—" for r in por_user_qs]
    user_gasto  = [float(r["gasto"] or 0) for r in por_user_qs]

    # --- Exportação CSV ---
    if request.GET.get("export") == "1":
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        fname = f"gasto_toner_{dt_ini:%Y%m%d}_{dt_fim:%Y%m%d}.csv"
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        resp.write("Centro de Custo;Quantidade;Gasto (R$)\n")
        for l in linhas:
            gasto_str = f"{l['gasto']:.2f}".replace(".", ",")
            resp.write(f"{l['cc']};{l['qtd']};{gasto_str}\n")
        total_str = f"{total_geral:.2f}".replace(".", ",")
        resp.write(f"TOTAL;;{total_str}\n")
        return resp

    ctx = {
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "linhas": linhas,
        "cc_labels": cc_labels,
        "cc_gasto": cc_gasto,
        "user_labels": user_labels,
        "user_gasto": user_gasto,
        # KPIs para o template
        "kpi_total_gasto": total_geral,
        "kpi_total_qtd": total_qtd,
        "kpi_top_cc_nome": top_cc_nome,
        "kpi_ticket_medio": ticket_medio,
    }
    return render(request, "front/dashboards/dashboard_toner.html", ctx)


    ##### EXPORTAR EXCEL #################

@login_required
def custo_cc_export_excel(request):
    """Exporta para Excel a mesma tabela do cc_custos_dashboard, respeitando os filtros."""
    # ── filtros (iguais ao dashboard) ─────────────────────────────────────────────
    hoje = datetime.today().date()
    dt_ini = _parse_date(request.GET.get("inicio"), hoje - timedelta(days=30))
    dt_fim = _parse_date(request.GET.get("fim"), hoje)

    # ── imports locais (mesmos modelos usados no dashboard) ──────────────────────
    from .models import (
        Locacao, MovimentacaoLicenca, MovimentacaoItem, CentroCusto,
        Usuario, Item, TipoMovLicencaChoices, TipoMovimentacaoChoices
    )

    # ── acumulador por CC (mesma estrutura) ──────────────────────────────────────
    totals = {}  # cc_id -> dict

    def acc(cc_id):
        if not cc_id:
            return None
        if cc_id not in totals:
            totals[cc_id] = {
                "cc": None,
                "usuarios": 0,
                "itens": 0,
                "licencas_set": set(),
                "assentos": 0,
                "custo_itens": Decimal("0.00"),
                "custo_licencas": Decimal("0.00"),
                "baixas": Decimal("0.00"),
            }
        return totals[cc_id]

    # ── custo mensal de ITENS (locações) por CC ──────────────────────────────────
    loc_qs = (
        Locacao.objects
        .select_related("equipamento", "equipamento__centro_custo")
        .exclude(valor_mensal__isnull=True)
    )
    for loc in loc_qs:
        item = loc.equipamento
        cc_id = getattr(item.centro_custo, "id", None)
        if not cc_id:
            continue
        valor = loc.valor_mensal or Decimal("0.00")
        if valor > 0:
            a = acc(cc_id)
            a["custo_itens"] += valor

    # ── assentos/licenças (último evento por par licença/usuário) ───────────────
    mov_l_qs = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario__centro_custo", "centro_custo_destino", "lote")
        .order_by("licenca_id", "usuario_id", "created_at", "id")
    )

    last_by_pair = {}
    for m in mov_l_qs:
        if m.usuario_id is None:
            continue
        last_by_pair[(m.licenca_id, m.usuario_id)] = m

    for (lic_id, user_id), m in last_by_pair.items():
        if m.tipo != TipoMovLicencaChoices.ATRIBUICAO:
            continue
        cc_id = (
            getattr(getattr(m.usuario, "centro_custo", None), "id", None)
            or getattr(m.centro_custo_destino, "id", None)
            or getattr(m.licenca.centro_custo, "id", None)
        )
        if not cc_id:
            continue

        cm = m.custo_mensal_usado or Decimal("0.00")
        a = acc(cc_id)
        a["assentos"] += 1
        a["custo_licencas"] += cm
        a["licencas_set"].add(lic_id)

    # ── baixas no período ────────────────────────────────────────────────────────
    baixas_qs = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .select_related("item__centro_custo", "centro_custo_origem")
    )
    for mv in baixas_qs:
        cc_id = (
            getattr(mv.centro_custo_origem, "id", None)
            or getattr(getattr(mv.item, "centro_custo", None), "id", None)
        )
        if not cc_id:
            continue
        valor_baixa = mv.custo if mv.custo is not None else (mv.item.valor or Decimal("0.00")) * (mv.quantidade or 1)
        a = acc(cc_id)
        a["baixas"] += (valor_baixa or Decimal("0.00"))

    # ── metadados: usuários e itens por CC ───────────────────────────────────────
    cc_ids = list(totals.keys())
    ccs = {cc.id: cc for cc in CentroCusto.objects.filter(id__in=cc_ids)}

    users_count = (
        Usuario.objects
        .filter(centro_custo_id__in=cc_ids, status="ativo")
        .values("centro_custo_id")
        .annotate(n=Count("id"))
    )
    itens_count = (
        Item.objects
        .filter(centro_custo_id__in=cc_ids)
        .values("centro_custo_id")
        .annotate(n=Count("id"))
    )
    map_users = {r["centro_custo_id"]: r["n"] for r in users_count}
    map_itens = {r["centro_custo_id"]: r["n"] for r in itens_count}

    # ── LINHAS (igual ao dashboard) ──────────────────────────────────────────────
    linhas = []
    for cc_id, d in totals.items():
        cc = ccs.get(cc_id)
        if not cc:
            continue
        d["cc"] = cc
        d["usuarios"] = map_users.get(cc_id, 0)
        d["itens"] = map_itens.get(cc_id, 0)
        lic_tipos = len(d["licencas_set"])
        d["licencas"] = lic_tipos
        d["total_mensal"] = (d["custo_itens"] + d["custo_licencas"])
        d["total_geral"] = (d["total_mensal"] + d["baixas"])

        linhas.append({
            "cc": cc,
            "usuarios": d["usuarios"],
            "itens": d["itens"],
            "licencas": d["licencas"],
            "assentos": d["assentos"],
            "custo_itens": d["custo_itens"],
            "custo_licencas": d["custo_licencas"],
            "baixas": d["baixas"],
            "total_mensal": d["total_mensal"],
            "total_geral": d["total_geral"],
        })

    # mesma ordenação da tela
    linhas.sort(key=lambda x: x["total_geral"], reverse=True)

    # ── EXCEL (apenas tabela detalhamento + resumo) ──────────────────────────────
    wb = Workbook()

    # Aba 1: Detalhamento (títulos iguais aos da tabela do template)
    ws = wb.active
    ws.title = "Detalhamento"

    headers = [
        "Centro de Custo",
        "Usuários",
        "Itens",
        "Licenças",
        "Assentos ativos",
        "Custo Itens (R$/mês)",
        "Custo Licenças (R$/mês)",
        "Baixas no período (R$)",
        "Total Mensal (R$)",
        "Total Geral (R$)",
    ]
    ws.append(headers)

    # estilos
    header_fill = PatternFill("solid", fgColor="FF1D4ED8")
    header_font = Font(color="FFFFFFFF", bold=True)
    thin = Side(style="thin", color="FFCBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    int_fmt = "#,##0"
    money_fmt = "[$R$-pt-BR] #,##0.00"

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    def _cc_nome(ccobj):
        try:
            return f"{ccobj.numero} - {ccobj.departamento}"
        except Exception:
            return str(ccobj) if ccobj else "—"

    for l in linhas:
        ws.append([
            _cc_nome(l["cc"]),
            int(l["usuarios"] or 0),
            int(l["itens"] or 0),
            int(l["licencas"] or 0),
            int(l["assentos"] or 0),
            Decimal(l["custo_itens"] or 0),
            Decimal(l["custo_licencas"] or 0),
            Decimal(l["baixas"] or 0),
            Decimal(l["total_mensal"] or 0),
            Decimal(l["total_geral"] or 0),
        ])

    last_row = ws.max_row
    for r in range(2, last_row + 1):
        for col in (2, 3, 4, 5):
            ws.cell(row=r, column=col).number_format = int_fmt
            ws.cell(row=r, column=col).border = border_all
        for col in (6, 7, 8, 9, 10):
            ws.cell(row=r, column=col).number_format = money_fmt
            ws.cell(row=r, column=col).border = border_all
        ws.cell(row=r, column=1).border = border_all

    # transforma em "tabela" do Excel (zebra)
    ref = f"A1:J{last_row}"
    table = Table(displayName="tb_detalhamento", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    # larguras e freeze
    widths = [28, 12, 10, 12, 16, 22, 22, 20, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "A2"

    # Aba 2: Resumo rápido
    ws2 = wb.create_sheet("Resumo")
    ws2["A1"] = "Período"
    ws2["B1"] = f"{dt_ini:%d/%m/%Y} — {dt_fim:%d/%m/%Y}"
    ws2["A1"].font = Font(bold=True)
    ws2["A3"] = "Centros de Custo (linhas)"
    ws2["B3"] = len(linhas)
    ws2["A4"] = "Total Geral (soma)"
    if last_row >= 2:
        ws2["B4"] = f"=SUM(Detalhamento!J2:J{last_row})"
        ws2["B4"].number_format = money_fmt
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 36

    # resposta HTTP
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="custo_cc_{dt_ini:%Y%m%d}_{dt_fim:%Y%m%d}.xlsx"'
    wb.save(resp)
    return resp

try:
    from .models import LicencaLote
except Exception:
    LicencaLote = None

def _parse_date_opt(date_str):
    if not date_str: return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return None

def _get_meses_ciclo(periodicidade_str):
    """
    Converte periodicidade em meses para cálculo do custo mensal.
    """
    if not periodicidade_str: return 1
    p = str(periodicidade_str).upper()
    if 'MEN' in p: return 1
    if 'BI' in p: return 2
    if 'TRI' in p: return 3
    if 'SEM' in p: return 6
    if 'ANU' in p: return 12
    return 1 

@login_required
def licencas_dashboard(request):
    hoje = timezone.localdate()

    # --- 1. Filtros ---
    q = (request.GET.get("q") or "").strip()
    fornecedor_id = (request.GET.get("fornecedor") or "").strip()
    cc_id = (request.GET.get("centro_custo") or "").strip()
    periodicidade = (request.GET.get("periodicidade") or "").strip()
    pmb = (request.GET.get("pmb") or "").strip().lower()
    
    dt_ini = _parse_date_opt(request.GET.get("inicio"))
    dt_fim = _parse_date_opt(request.GET.get("fim"))

    # Base: Licenças com seus Lotes
    qs_licencas = Licenca.objects.select_related("fornecedor").prefetch_related('lotes', 'lotes__fornecedor')

    if q:
        qs_licencas = qs_licencas.filter(Q(nome__icontains=q) | Q(fornecedor__nome__icontains=q))
    if fornecedor_id:
        qs_licencas = qs_licencas.filter(fornecedor_id=fornecedor_id)
    if pmb in ['sim', 'nao']:
        qs_licencas = qs_licencas.filter(pmb=pmb)
    
    # Filtra Licenças baseadas nas propriedades dos seus Lotes
    if periodicidade:
        qs_licencas = qs_licencas.filter(lotes__periodicidade=periodicidade)
    if cc_id:
        qs_licencas = qs_licencas.filter(lotes__centro_custo_id=cc_id)
    if dt_ini:
        qs_licencas = qs_licencas.filter(lotes__data_compra__gte=dt_ini)
    if dt_fim:
        qs_licencas = qs_licencas.filter(lotes__data_compra__lte=dt_fim)

    qs_licencas = qs_licencas.distinct()
    licencas_list = list(qs_licencas)
    licenca_ids = [l.id for l in licencas_list]

    # --- 2. Mapeamento de Usuários Ativos (Snapshot Atual) ---
    movs_ativas = (
        MovimentacaoLicenca.objects
        .filter(licenca_id__in=licenca_ids, usuario__isnull=False)
        .select_related('usuario__centro_custo', 'centro_custo_destino')
        .order_by('licenca_id', 'usuario_id', 'created_at')
    )

    # Identifica o último status de cada usuário para cada licença
    estado_usuario = {} 
    for m in movs_ativas:
        estado_usuario[(m.licenca_id, m.usuario_id)] = m
    
    # Agrupa: { licenca_id: [lista_nomes_cc_ativos, ...] }
    uso_map_cc = {}
    
    for (lid, uid), mov in estado_usuario.items():
        if mov.tipo == 'atribuicao': # Usuário está com a licença
            if lid not in uso_map_cc: uso_map_cc[lid] = []
            
            # Define qual CC paga a conta: Destino da Mov > CC do Usuário > Indefinido
            cc_nome = "Indefinido"
            if mov.centro_custo_destino:
                cc_nome = f"{mov.centro_custo_destino.numero} - {mov.centro_custo_destino.departamento}"
            elif mov.usuario and mov.usuario.centro_custo:
                cc_nome = f"{mov.usuario.centro_custo.numero} - {mov.usuario.centro_custo.departamento}"
            
            uso_map_cc[lid].append(cc_nome)

    # --- 3. Processamento de Custos (CORRIGIDO: UNITÁRIO) ---
    
    kpi_total_licencas = len(licencas_list)
    kpi_assentos_em_uso = 0
    kpi_assentos_totais = 0
    kpi_custo_mensal = Decimal(0)
    
    cc_costs = {}   # Rateio por CC
    forn_costs = {} # Share por Fornecedor
    per_counts = {} # Contagem Periodicidade

    linhas_tabela = []
    lotes_detalhes = []

    for lic in licencas_list:
        # Filtra lotes relevantes
        lotes_da_lic = lic.lotes.all()
        if periodicidade:
            lotes_da_lic = [l for l in lotes_da_lic if l.periodicidade == periodicidade]
        if cc_id:
            lotes_da_lic = [l for l in lotes_da_lic if l.centro_custo_id == int(cc_id)]
        
        if not lotes_da_lic:
            continue

        # Variáveis acumuladoras da Licença (Soma dos Lotes)
        l_qtd_total = 0
        l_qtd_disp = 0
        l_custo_mensal_total_licenca = Decimal(0)
        l_periodicidades = set()
        
        # Para calcular custo médio ponderado unitário desta licença (caso tenha lotes com preços diferentes)
        custo_unitario_acumulado = Decimal(0)

        # --- LOOP NOS LOTES ---
        for lote in lotes_da_lic:
            qtd = lote.quantidade_total
            disp = lote.quantidade_disponivel
            
            # Conversão Temporal
            meses = _get_meses_ciclo(lote.periodicidade)
            
            # [CORREÇÃO] Custo Unitário: O valor cadastrado é por unidade
            custo_ciclo_unitario = lote.custo_ciclo or Decimal(0)
            
            # Custo Mensal de UMA unidade
            if meses > 0:
                custo_mensal_unitario = (custo_ciclo_unitario / Decimal(meses)).quantize(Decimal("0.01"))
            else:
                custo_mensal_unitario = Decimal(0)
            
            # Custo Mensal TOTAL deste Lote (para KPIs) = Unitário * Quantidade
            custo_mensal_lote_total = custo_mensal_unitario * Decimal(qtd)
            
            # Acumuladores da Licença
            l_qtd_total += qtd
            l_qtd_disp += disp
            l_custo_mensal_total_licenca += custo_mensal_lote_total
            
            # Acumula valor total para depois tirar a média unitária
            custo_unitario_acumulado += custo_mensal_lote_total 
            
            l_periodicidades.add(lote.get_periodicidade_display())

            # KPIs Globais
            kpi_assentos_totais += qtd
            
            # Gráficos Auxiliares
            p_label = lote.get_periodicidade_display()
            per_counts[p_label] = per_counts.get(p_label, 0) + 1

            # Tabela de Detalhes (Drill-down)
            lotes_detalhes.append({
                'licenca': lic.nome,
                'lote_id': lote.id,
                'fornecedor': lote.fornecedor.nome if lote.fornecedor else "-",
                'qtd': qtd,
                'disp': disp,
                'custo_unit_ciclo': custo_ciclo_unitario, # Valor de compra (unidade)
                'custo_mensal_unit': custo_mensal_unitario, # Valor mensal (unidade)
                'custo_mensal_total': custo_mensal_lote_total, # Valor total lote
                'periodicidade': p_label
            })

        # --- CÁLCULO INTELIGENTE DE RATEIO ---
        # 1. Definir o Custo Unitário Médio Mensal desta licença (Mix de Lotes)
        if l_qtd_total > 0:
            custo_medio_unitario = custo_unitario_acumulado / Decimal(l_qtd_total)
        else:
            custo_medio_unitario = Decimal(0)

        # 2. Usuários Ativos (Consumo)
        ccs_ativos = uso_map_cc.get(lic.id, [])
        qtd_ativos = len(ccs_ativos)
        kpi_assentos_em_uso += qtd_ativos

        # 3. Distribuição para Centros de Custo
        # A) Consumo: Cada usuário gera 1 * Custo Unitário para seu CC
        for cc_nome in ccs_ativos:
            cc_costs[cc_nome] = cc_costs.get(cc_nome, Decimal(0)) + custo_medio_unitario

        # B) Estoque (Ociosidade): O que sobra (Total - Ativos) gera custo para o CC Dono da Licença
        qtd_estoque = max(0, l_qtd_total - qtd_ativos)
        if qtd_estoque > 0:
            cc_estoque = "Estoque (Sem CC Definido)"
            
            # Prioridade 1: CC definido na Licença Pai
            if lic.centro_custo:
                cc_estoque = f"{lic.centro_custo.numero} - {lic.centro_custo.departamento}"
            # Prioridade 2: CC do primeiro lote encontrado (fallback)
            elif lotes_da_lic and lotes_da_lic[0].centro_custo:
                cc_estoque = f"{lotes_da_lic[0].centro_custo.numero} - {lotes_da_lic[0].centro_custo.departamento}"
            
            # Soma ao custo desse CC
            cc_costs[cc_estoque] = cc_costs.get(cc_estoque, Decimal(0)) + (custo_medio_unitario * Decimal(qtd_estoque))

        # --- Totais e Gráficos ---
        kpi_custo_mensal += l_custo_mensal_total_licenca

        f_nome = lic.fornecedor.nome if lic.fornecedor else "Indefinido"
        forn_costs[f_nome] = forn_costs.get(f_nome, Decimal(0)) + l_custo_mensal_total_licenca

        per_display = ", ".join(l_periodicidades) if l_periodicidades else "-"
        
        # Linha para Tabela Principal
        linhas_tabela.append({
            'obj': lic,
            'periodicidade_display': per_display,
            'custo_mensal_total': l_custo_mensal_total_licenca,
            'ativos': qtd_ativos,
            'total': l_qtd_total,
            'estoque': qtd_estoque
        })

    # Ordenação
    sorted_cc = sorted(cc_costs.items(), key=lambda x: x[1], reverse=True)
    
    # Opções para o select do filtro
    periodicidade_choices = LicencaLote._meta.get_field('periodicidade').choices

    context = {
        'f_q': q, 'f_forn': fornecedor_id, 'f_cc': cc_id, 
        'f_per': periodicidade, 'f_pmb': pmb, 
        'dt_ini': dt_ini, 'dt_fim': dt_fim,
        
        'fornecedores': Fornecedor.objects.all().order_by('nome'),
        'centros_custo': CentroCusto.objects.all().order_by('numero'),
        'periodicidade_choices': periodicidade_choices,

        'kpi_total': kpi_total_licencas,
        'kpi_assentos': kpi_assentos_em_uso,
        'kpi_disp': kpi_assentos_totais - kpi_assentos_em_uso,
        'kpi_custo_mensal': kpi_custo_mensal,
        'kpi_custo_anual': kpi_custo_mensal * 12,

        'linhas': linhas_tabela,
        'lotes_rows': lotes_detalhes,
        
        'cc_list': [{'label': k, 'val': v} for k, v in sorted_cc],
        'chart_forn_labels': list(forn_costs.keys()),
        'chart_forn_data': [float(v) for v in forn_costs.values()],
        'chart_per_labels': list(per_counts.keys()),
        'chart_per_data': list(per_counts.values()),
    }

    return render(request, "front/dashboards/licencas_dashboard.html", context)