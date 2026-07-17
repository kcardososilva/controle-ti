"""
Portal do Fornecedor — área isolada (sandbox) para fornecedores externos.

Camadas de isolamento (defesa em profundidade — ver CLAUDE.md):
  1. FornecedorAccessMiddleware  — restringe o grupo "Fornecedor" a /portal/
  2. @fornecedor_required        — resolve o Fornecedor do request ou 403
  3. itens_do_fornecedor(...)    — TODA query parte daqui (nunca Item.objects.all())

v1: somente visão de equipamentos + status.
Fluxo de manutenção e licenças entram em fases seguintes.
"""
from datetime import date as _date
from decimal import Decimal
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone

from ..models import (
    Item,
    Categoria,
    Localidade,
    Licenca,
    LoteEnvioFornecedor,
    LoteSeparacao,
    OrdemManutencao,
    OrdemManutencaoAnexo,
    SeparacaoItem,
    StatusItemChoices,
    StatusLoteEnvioFornecedorChoices,
    StatusOrdemManutencaoChoices,
    StatusSeparacaoChoices,
    TipoMovimentacaoChoices,
    TipoSeparacaoChoices,
)


# ─── Helpers de segurança ─────────────────────────────────────────────────────

def fornecedor_do_request(request):
    """Retorna o Fornecedor vinculado ao usuário logado (perfil ativo), ou None."""
    perfil = getattr(request.user, "perfil_fornecedor", None)
    if perfil is not None and perfil.ativo:
        return perfil.fornecedor
    return None


def fornecedor_required(view_func):
    """
    Garante que o request tem um Fornecedor ativo vinculado e injeta
    `request.fornecedor`. Deve decorar TODA view do portal.

    Quando o usuário está no grupo "Fornecedor" mas ainda não tem um
    PerfilFornecedor ativo, mostra uma página orientativa (em vez de um 403
    cru) pedindo para o TI concluir a configuração do acesso.
    """
    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        fornecedor = fornecedor_do_request(request)
        if fornecedor is None:
            return render(request, "front/portal/portal_sem_acesso.html", status=403)
        request.fornecedor = fornecedor
        return view_func(request, *args, **kwargs)
    return _wrapped


def itens_do_fornecedor(fornecedor):
    """
    Núcleo de segurança: o conjunto de itens visíveis a um fornecedor.
      • itens fornecidos por ele (Item.fornecedor)
      • itens enviados para manutenção sob sua responsabilidade
        (MovimentacaoItem.fornecedor_manutencao em um envio_manutencao)
      • itens em remessa (Envio/Devolução) endereçados a ele
        (SeparacaoItem.fornecedor)
    Toda view do portal DEVE partir desta função — nunca de Item.objects.all().
    """
    return (
        Item.objects
        .filter(
            Q(fornecedor=fornecedor)
            | Q(
                movimentacoes__tipo_movimentacao=TipoMovimentacaoChoices.ENVIO_MANUTENCAO,
                movimentacoes__fornecedor_manutencao=fornecedor,
            )
            | Q(separacoes__fornecedor=fornecedor)
        )
        .distinct()
    )


# Ordem de exibição dos status nos cards do painel.
_STATUS_ORDEM = [
    StatusItemChoices.ATIVO,
    StatusItemChoices.MANUTENCAO,
    StatusItemChoices.DEFEITO,
    StatusItemChoices.BACKUP,
    StatusItemChoices.PAUSADO,
]


# ─── Views ────────────────────────────────────────────────────────────────────

@fornecedor_required
def portal_home(request):
    """Painel inicial do fornecedor: KPIs de equipamentos por status."""
    qs = itens_do_fornecedor(request.fornecedor)

    counts = {
        row["status"]: row["n"]
        for row in qs.values("status").annotate(n=Count("id"))
    }
    status_cards = [
        {
            "slug": s.value,
            "label": s.label,
            "count": counts.get(s.value, 0),
        }
        for s in _STATUS_ORDEM
    ]

    total = qs.count()

    recentes = (
        qs.select_related("subtipo", "localidade").order_by("-updated_at")[:6]
    )

    # Resumo de manutenção do fornecedor
    SOM = StatusOrdemManutencaoChoices
    os_qs = OrdemManutencao.objects.filter(fornecedor=request.fornecedor)
    qtd_os_abertas = os_qs.exclude(status__in=[SOM.CONCLUIDO, SOM.CANCELADO, SOM.DESCARTADO]).count()
    qtd_os_acao = os_qs.filter(status__in=[
        SOM.AGUARDANDO_RECEBIMENTO, SOM.RECEBIDO, SOM.EM_AVALIACAO,
        SOM.EM_REPARO, SOM.REPARADO, SOM.SEM_REPARO,
        SOM.SEM_CONDICOES, SOM.DESCARTE_LOCAL_APROVADO,
    ]).count()

    context = {
        "fornecedor": request.fornecedor,
        "total_itens": total,
        "status_cards": status_cards,
        "recentes": recentes,
        "qtd_os_abertas": qtd_os_abertas,
        "qtd_os_acao": qtd_os_acao,
        "active_nav": "home",
    }
    return render(request, "front/portal/portal_home.html", context)


def _portal_itens_filtrados(request):
    """Aplica busca + filtros (status, localidade, categoria) ao escopo do fornecedor.
    Usado pela listagem e pela exportação para garantir o mesmo resultado."""
    qs = (
        itens_do_fornecedor(request.fornecedor)
        .select_related("subtipo", "subtipo__categoria", "localidade")
        .order_by("nome")
    )

    q = request.GET.get("q", "").strip()
    marca = request.GET.get("marca", "").strip()
    modelo = request.GET.get("modelo", "").strip()
    serie = request.GET.get("serie", "").strip()
    status = request.GET.get("status", "").strip()
    localidade = request.GET.get("localidade", "").strip()
    categoria = request.GET.get("categoria", "").strip()

    if q:
        qs = qs.filter(nome__icontains=q)
    if marca:
        qs = qs.filter(marca__icontains=marca)
    if modelo:
        qs = qs.filter(modelo__icontains=modelo)
    if serie:
        qs = qs.filter(numero_serie__icontains=serie)
    if status:
        qs = qs.filter(status=status)
    if localidade.isdigit():
        qs = qs.filter(localidade_id=int(localidade))
    if categoria.isdigit():
        qs = qs.filter(subtipo__categoria_id=int(categoria))

    filtros = {
        "q": q, "marca": marca, "modelo": modelo, "serie": serie,
        "status": status, "localidade": localidade, "categoria": categoria,
    }
    return qs, filtros


@fornecedor_required
def portal_equipamentos_list(request):
    """Lista de equipamentos do fornecedor — busca + filtros separados + export."""
    qs, filtros = _portal_itens_filtrados(request)

    base = itens_do_fornecedor(request.fornecedor)
    localidades = Localidade.objects.filter(id__in=base.values("localidade")).order_by("local")
    categorias = Categoria.objects.filter(id__in=base.values("subtipo__categoria")).order_by("nome")

    paginator = Paginator(qs, 15)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    get_copy = request.GET.copy()
    get_copy.pop("page", None)
    qs_keep = get_copy.urlencode()

    from services.lote_envio_fornecedor_service import LoteEnvioFornecedorService
    ids_devolver = LoteEnvioFornecedorService.itens_aguardando_devolucao_ids(list(page_obj))

    context = {
        "fornecedor": request.fornecedor,
        "page_obj": page_obj,
        "total": paginator.count,
        "f_q": filtros["q"],
        "f_marca": filtros["marca"],
        "f_modelo": filtros["modelo"],
        "f_serie": filtros["serie"],
        "f_status": filtros["status"],
        "f_localidade": filtros["localidade"],
        "f_categoria": filtros["categoria"],
        "tem_filtro": any(filtros.values()),
        "status_choices": StatusItemChoices.choices,
        "localidades": localidades,
        "categorias": categorias,
        "qs_keep": qs_keep,
        "ids_devolver": ids_devolver,
        "active_nav": "equipamentos",
    }
    return render(request, "front/portal/portal_equipamentos_list.html", context)


@fornecedor_required
def portal_equipamentos_export(request):
    """Exporta para Excel os equipamentos do fornecedor conforme os filtros aplicados."""
    qs, _ = _portal_itens_filtrados(request)
    return _portal_itens_xlsx(list(qs), request.fornecedor)


def _portal_itens_xlsx(itens, fornecedor):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND_DARK, BRAND, SOFT, ZEBRA, INK = "0B3D6E", "0071E3", "E5F0FB", "F4F9FE", "1F2733"
    hair = Side(style="thin", color="CFE0F2")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    f_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color="5B6B7F")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Calibri", size=10, color=INK)
    fill_title = PatternFill("solid", fgColor=BRAND_DARK)
    fill_sub = PatternFill("solid", fgColor=SOFT)
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left", vertical="center", indent=1)

    STATUS_FILL = {
        "ativo": ("E6F4EA", "1E8E3E"), "manutencao": ("FEF1E0", "B35A00"),
        "defeito": ("FCE8E6", "D93025"), "backup": ("E5F0FB", "0B5BB5"),
        "pausado": ("F0F0F2", "5B6B7F"), "descarte": ("E8EAEE", "475569"),
    }

    header = ["#", "Equipamento", "Nº Série", "Marca", "Modelo", "Categoria", "Tipo", "Localidade", "Status"]
    ncols = len(header)
    center_cols = {1, 9}

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    ws.sheet_view.showGridLines = False

    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]; c.value = "MEUS EQUIPAMENTOS"; c.font = f_title; c.fill = fill_title; c.alignment = a_left
    ws.row_dimensions[1].height = 34
    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
    ws.merge_cells(f"A2:{last}2")
    c2 = ws["A2"]
    c2.value = f"{fornecedor.nome}  ·  {len(itens)} equipamento(s)  ·  gerado em {gerado}"
    c2.font = f_sub; c2.fill = fill_sub; c2.alignment = a_left
    ws.row_dimensions[2].height = 18

    HEADER_ROW = 3
    for ci, h in enumerate(header, 1):
        cc = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cc.fill = fill_header; cc.font = f_header; cc.border = border
        cc.alignment = a_center if ci in center_cols else a_left
    ws.row_dimensions[HEADER_ROW].height = 26

    row = HEADER_ROW + 1
    for i, it in enumerate(itens, start=1):
        categoria = it.subtipo.categoria.nome if (it.subtipo and it.subtipo.categoria) else ""
        tipo = it.subtipo.nome if it.subtipo else ""
        local = it.localidade.local if it.localidade else ""
        valores = [i, it.nome, it.numero_serie or "", it.marca or "", it.modelo or "",
                   categoria, tipo, local, it.get_status_display()]
        zebra = (i % 2 == 0)
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = border; cell.font = f_cell
            cell.alignment = a_center if ci in center_cols else a_left
            if ci == 9:
                bg, fg = STATUS_FILL.get(it.status, ("F0F0F2", "5B6B7F"))
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=fg)
                cell.alignment = a_center
            elif zebra:
                cell.fill = fill_zebra
        row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last}{max(row - 1, HEADER_ROW)}"

    widths = {}
    for r_ in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        for idx, val in enumerate(r_, start=1):
            widths[idx] = max(widths.get(idx, 0), len(str(val)) if val is not None else 0)
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 11), 46)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    now = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="meus_equipamentos_{now}.xlsx"'
    return resp


@fornecedor_required
def portal_equipamento_detail(request, pk: int):
    """Ficha read-only de um equipamento — 404 se não pertencer ao fornecedor."""
    item = get_object_or_404(
        itens_do_fornecedor(request.fornecedor).select_related(
            "subtipo", "subtipo__categoria", "localidade"
        ),
        pk=pk,
    )

    manutencoes = (
        item.movimentacoes
        .filter(
            tipo_movimentacao__in=[
                TipoMovimentacaoChoices.ENVIO_MANUTENCAO,
                TipoMovimentacaoChoices.RETORNO_MANUTENCAO,
            ]
        )
        .order_by("-created_at")[:20]
    )

    # OS de manutenção aberta para este item sob responsabilidade do fornecedor
    os_aberta = (
        OrdemManutencao.objects
        .filter(item=item, fornecedor=request.fornecedor)
        .exclude(status__in=[
            StatusOrdemManutencaoChoices.CONCLUIDO,
            StatusOrdemManutencaoChoices.CANCELADO,
            StatusOrdemManutencaoChoices.DESCARTADO,
        ])
        .order_by("-created_at")
        .first()
    )

    # Histórico completo de OS deste item sob responsabilidade do fornecedor —
    # inclui orçamentos (reparo/troca/descarte) e o substituto, quando houver.
    ordens_manutencao = (
        OrdemManutencao.objects
        .filter(item=item, fornecedor=request.fornecedor)
        .select_related("item_substituto")
        .prefetch_related("orcamentos")
        .order_by("-created_at")
    )

    # Histórico de locação (congelável por status) — registro do fornecedor.
    loc_periodos = list(item.locacao_periodos.all()) if item.eh_locado else []
    loc_total = sum((p.valor_acumulado for p in loc_periodos), Decimal("0.00"))
    loc_atual = next((p for p in loc_periodos if p.em_andamento), None)

    # Troca antecipada já em andamento (OS aberta ou rascunho no lote de envio) —
    # evita o fornecedor mandar um substituto duplicado por engano.
    troca_pendente = False
    if item.status == StatusItemChoices.DEFEITO:
        from services.lote_envio_fornecedor_service import LoteEnvioFornecedorService
        troca_pendente = LoteEnvioFornecedorService.item_tem_troca_pendente(item)

    context = {
        "fornecedor": request.fornecedor,
        "item": item,
        "manutencoes": manutencoes,
        "os_aberta": os_aberta,
        "ordens_manutencao": ordens_manutencao,
        "troca_pendente": troca_pendente,
        "loc_periodos": loc_periodos,
        "loc_total": loc_total,
        "loc_atual": loc_atual,
        "loc_congelado": item.eh_locado and loc_atual is None,
        "active_nav": "equipamentos",
    }
    return render(request, "front/portal/portal_equipamento_detail.html", context)


def _portal_manutencao_nav(ordem):
    """Aba do menu que fica ativa no detalhe da OS (troca antecipada tem aba própria)."""
    return "troca_antecipada" if ordem.troca_antecipada else "manutencao"


# ─── Manutenção (Ordens de Serviço conduzidas pelo fornecedor) ────────────────

_OS_ABERTAS_EXCLUI = [
    StatusOrdemManutencaoChoices.CONCLUIDO,
    StatusOrdemManutencaoChoices.CANCELADO,
    StatusOrdemManutencaoChoices.DESCARTADO,
]


# Status em que a próxima ação é do FORNECEDOR (para destacar "pendentes de você").
_FORNECEDOR_ACAO_STATUS = {
    StatusOrdemManutencaoChoices.AGUARDANDO_RECEBIMENTO,
    StatusOrdemManutencaoChoices.RECEBIDO,
    StatusOrdemManutencaoChoices.EM_AVALIACAO,
    StatusOrdemManutencaoChoices.APROVADO,
    StatusOrdemManutencaoChoices.REPROVADO,
    StatusOrdemManutencaoChoices.EM_REPARO,
    StatusOrdemManutencaoChoices.REPARADO,
    StatusOrdemManutencaoChoices.SEM_REPARO,
    StatusOrdemManutencaoChoices.TROCA_APROVADA,
    StatusOrdemManutencaoChoices.TROCA_REPROVADA,
    StatusOrdemManutencaoChoices.TROCA_DANO_REPROVADA,
    StatusOrdemManutencaoChoices.DESCARTE_AVALIACAO_APROVADA,
    StatusOrdemManutencaoChoices.DESCARTE_AVALIACAO_REPROVADA,
    StatusOrdemManutencaoChoices.DESCARTE_LOCAL_APROVADO,
    StatusOrdemManutencaoChoices.TROCA_ANT_DEFEITUOSO_ENVIADO,
    StatusOrdemManutencaoChoices.TROCA_ANT_DEFEITUOSO_RECEBIDO,
}


def _parse_data_filtro(valor):
    """Lê uma data de filtro (input type=date, formato ISO yyyy-mm-dd). Silenciosa
    em valor ausente/ inválido — filtro por data nunca deve derrubar a tela."""
    valor = (valor or "").strip()
    if not valor:
        return None
    try:
        return _date.fromisoformat(valor)
    except ValueError:
        return None


@fornecedor_required
def portal_manutencao_list(request):
    """Fila de ordens de manutenção do fornecedor (abertas + histórico).

    O histórico aceita filtros (nome, modelo, nº de série, período) — pensado
    para o fornecedor localizar rapidamente OS's antigas na hora de montar um
    Lote de Manutenção. As "em aberto" não são filtradas: são poucas e o
    fornecedor precisa vê-las todas para agir."""
    base = (
        OrdemManutencao.objects
        .filter(fornecedor=request.fornecedor)
        .select_related("item", "item_substituto")
    )
    abertas = list(base.exclude(status__in=_OS_ABERTAS_EXCLUI))

    # ── Filtros do histórico ────────────────────────────────────────────────
    f_nome = (request.GET.get("nome") or "").strip()
    f_modelo = (request.GET.get("modelo") or "").strip()
    f_ns = (request.GET.get("ns") or "").strip()
    f_data_de = _parse_data_filtro(request.GET.get("data_de"))
    f_data_ate = _parse_data_filtro(request.GET.get("data_ate"))

    historico_qs = base.filter(status__in=_OS_ABERTAS_EXCLUI)
    if f_nome:
        historico_qs = historico_qs.filter(item__nome__icontains=f_nome)
    if f_modelo:
        historico_qs = historico_qs.filter(item__modelo__icontains=f_modelo)
    if f_ns:
        historico_qs = historico_qs.filter(item__numero_serie__icontains=f_ns)
    if f_data_de:
        historico_qs = historico_qs.filter(finalizada_em__date__gte=f_data_de)
    if f_data_ate:
        historico_qs = historico_qs.filter(finalizada_em__date__lte=f_data_ate)
    historico_qs = historico_qs.order_by("-finalizada_em", "-created_at")

    paginator = Paginator(historico_qs, 15)
    historico_page = paginator.get_page(request.GET.get("page", 1))
    for o in historico_page:
        o.valor_calc = o.valor_manutencao

    get_copy = request.GET.copy()
    get_copy.pop("page", None)
    qs_keep = get_copy.urlencode()

    # Marca as OS que aguardam a ação do fornecedor (a coluna prioritária da tela).
    qtd_acao = 0
    for o in abertas:
        o.precisa_acao = o.status in _FORNECEDOR_ACAO_STATUS
        if o.precisa_acao:
            qtd_acao += 1

    # Resumo por status (somente os que têm ordens) — melhora a visão geral.
    counts = {row["status"]: row["n"] for row in base.values("status").annotate(n=Count("id"))}
    resumo_status = [
        {"slug": s.value, "label": s.label, "count": counts.get(s.value, 0)}
        for s in StatusOrdemManutencaoChoices
        if counts.get(s.value, 0)
    ]

    context = {
        "fornecedor": request.fornecedor,
        "abertas": abertas,
        "historico_page": historico_page,
        "qtd_abertas": len(abertas),
        "qtd_acao": qtd_acao,
        "resumo_status": resumo_status,
        "total_os": base.count(),
        "active_nav": "manutencao",
        "f_nome": f_nome,
        "f_modelo": f_modelo,
        "f_ns": f_ns,
        "f_data_de": request.GET.get("data_de", ""),
        "f_data_ate": request.GET.get("data_ate", ""),
        "tem_filtro": any([f_nome, f_modelo, f_ns, f_data_de, f_data_ate]),
        "qs_keep": qs_keep,
    }
    return render(request, "front/portal/portal_manutencao_list.html", context)


_PORTAL_SEP_TITULOS = {
    TipoSeparacaoChoices.ENVIO: "Remessa para Envio",
    TipoSeparacaoChoices.DEVOLUCAO: "Remessa para Devolução",
}


def _portal_separacao_list(request, tipo):
    """Visão somente-leitura: quem cria/organiza/despacha a remessa é sempre o
    time interno (ver ProjetoEstoque/views/separacoes.py) — o fornecedor só
    acompanha o que já foi endereçado a ele."""
    from services.separacao_service import SeparacaoService

    soltos = (
        SeparacaoItem.objects
        .filter(fornecedor=request.fornecedor, tipo=tipo, lote__isnull=True)
        .exclude(status=StatusSeparacaoChoices.CANCELADO)
        .select_related("item", "item__locacao")
        .order_by("-created_at")
    )
    lotes = (
        LoteSeparacao.objects
        .filter(fornecedor=request.fornecedor, tipo=tipo)
        .prefetch_related("itens")
        .order_by("-created_at")
    )
    soltos = list(soltos)
    for s in soltos:
        s.badge_contrato = SeparacaoService.badge_contrato(s.item)

    return {
        "fornecedor": request.fornecedor,
        "tipo": tipo,
        "titulo": _PORTAL_SEP_TITULOS[tipo],
        "soltos": soltos,
        "lotes": lotes,
        "active_nav": "separacao_envio" if tipo == TipoSeparacaoChoices.ENVIO else "separacao_devolucao",
    }


@fornecedor_required
def portal_separacao_envio_list(request):
    context = _portal_separacao_list(request, TipoSeparacaoChoices.ENVIO)
    return render(request, "front/portal/portal_separacao_list.html", context)


@fornecedor_required
def portal_separacao_devolucao_list(request):
    context = _portal_separacao_list(request, TipoSeparacaoChoices.DEVOLUCAO)
    return render(request, "front/portal/portal_separacao_list.html", context)


@fornecedor_required
def portal_separacao_lote_detail(request, pk: int):
    """Documento completo do lote: dados do equipamento, contrato de Locação
    (na Devolução) e, quando já despachado para manutenção, a Ordem de
    Manutenção correspondente — tudo em um único lugar para o fornecedor."""
    from services.separacao_service import SeparacaoService

    lote = get_object_or_404(
        LoteSeparacao.objects.select_related("fornecedor"),
        pk=pk, fornecedor=request.fornecedor,
    )

    if request.method == "POST" and request.POST.get("acao") == "confirmar_recebimento_lote":
        from services.ordem_manutencao_service import OrdemManutencaoService

        if lote.tipo != TipoSeparacaoChoices.ENVIO:
            messages.error(request, "Esta ação só se aplica a remessas de envio.")
        else:
            pendentes = list(
                SeparacaoItem.objects.filter(lote=lote)
                .select_related("movimentacao_despacho")
            )
            confirmados = 0
            for sep in pendentes:
                if not sep.movimentacao_despacho_id:
                    continue
                ordem = sep.movimentacao_despacho.ordens_manutencao.first()
                if ordem and ordem.status == StatusOrdemManutencaoChoices.AGUARDANDO_RECEBIMENTO:
                    try:
                        OrdemManutencaoService.transicionar(
                            ordem=ordem, novo_status=StatusOrdemManutencaoChoices.RECEBIDO,
                            user=request.user, ator="fornecedor",
                        )
                        confirmados += 1
                    except ValidationError:
                        pass
            if confirmados:
                messages.success(request, f"{confirmados} equipamento(s) confirmado(s) como recebido(s).")
            else:
                messages.error(request, "Não há itens desta remessa aguardando confirmação de recebimento.")
        return redirect("portal_separacao_lote_detail", pk=lote.pk)

    itens = list(
        lote.itens
        .select_related(
            "item", "item__categoria", "item__subtipo", "item__localidade",
            "item__centro_custo", "item__locacao", "movimentacao_despacho",
        )
        .order_by("-created_at")
    )
    qtd_aguardando_recebimento = 0
    for i in itens:
        i.info = SeparacaoService.info_equipamento(i)
        i.badge_contrato = SeparacaoService.badge_contrato(i.item)
        i.ordem = None
        if i.movimentacao_despacho_id:
            i.ordem = i.movimentacao_despacho.ordens_manutencao.first()
        if i.ordem and i.ordem.status == StatusOrdemManutencaoChoices.AGUARDANDO_RECEBIMENTO:
            qtd_aguardando_recebimento += 1

    context = {
        "fornecedor": request.fornecedor,
        "lote": lote,
        "itens": itens,
        "qtd_aguardando_recebimento": qtd_aguardando_recebimento,
        "titulo": _PORTAL_SEP_TITULOS.get(lote.tipo, "Remessa"),
        "active_nav": "separacao_envio" if lote.tipo == TipoSeparacaoChoices.ENVIO else "separacao_devolucao",
    }
    return render(request, "front/portal/portal_separacao_lote_detail.html", context)


@fornecedor_required
def portal_manutencao_detail(request, pk: int):
    """Detalhe da OS + ações de transição do fornecedor."""
    ordem = get_object_or_404(
        OrdemManutencao.objects.select_related(
            "item", "item__subtipo", "item_substituto", "fornecedor", "movimentacao_origem",
            "devolucao_localidade",
        ),
        pk=pk,
        fornecedor=request.fornecedor,
    )

    if request.method == "POST":
        acao = request.POST.get("acao", "")

        # ── Upload de Nota Fiscal (o fornecedor pode anexar quantas quiser) ──
        if acao == "anexar_nf":
            arquivos = request.FILES.getlist("nf")
            if not arquivos:
                messages.error(request, "Selecione ao menos um arquivo de NF.")
            else:
                descricao = request.POST.get("descricao", "").strip()
                for arq in arquivos:
                    OrdemManutencaoAnexo.objects.create(
                        ordem=ordem,
                        arquivo=arq,
                        origem=OrdemManutencaoAnexo.OrigemAnexo.FORNECEDOR,
                        descricao=descricao,
                        criado_por=request.user,
                        atualizado_por=request.user,
                    )
                messages.success(request, f"{len(arquivos)} nota(s) fiscal(is) anexada(s).")
            return redirect("portal_manutencao_detail", pk=pk)

        # ── Exclusão de Nota Fiscal (correção de erro) ───────────────────────
        # O fornecedor só pode excluir NF que ele mesmo anexou (origem=fornecedor).
        if acao == "excluir_nf":
            anexo = ordem.anexos.filter(
                pk=request.POST.get("anexo_id"),
                origem=OrdemManutencaoAnexo.OrigemAnexo.FORNECEDOR,
            ).first()
            if not anexo:
                messages.error(request, "Nota fiscal não encontrada ou não pode ser excluída por você.")
            else:
                if anexo.arquivo:
                    anexo.arquivo.delete(save=False)  # remove o arquivo físico também
                anexo.delete()
                messages.success(request, "Nota fiscal excluída.")
            return redirect("portal_manutencao_detail", pk=pk)

        # ── Desfazer o último envio (voltar ao formulário anterior) ─────────
        # Não é navegação: reverte de verdade o status da OS para o estágio
        # anterior e apaga a proposta que ainda não foi decidida pelo TI, para
        # o fornecedor corrigir e reenviar (ex.: digitou o valor errado).
        if acao == "desfazer":
            from services.ordem_manutencao_service import OrdemManutencaoService

            try:
                OrdemManutencaoService.desfazer_ultima_proposta(ordem=ordem, user=request.user)
                messages.success(request, "Envio desfeito — o formulário anterior foi reaberto para correção.")
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
            return redirect("portal_manutencao_detail", pk=pk)

        # ── Separar o retorno do reparo concluído em um Lote de Envio ────────
        # (mesmo mecanismo já usado para troca antecipada/equipamento novo —
        # NF, múltiplos itens, envio em lote. Não é uma transição de status:
        # a OS só vira DEVOLVIDO quando o lote for de fato enviado.)
        if acao == "separar_lote":
            from services.lote_envio_fornecedor_service import LoteEnvioFornecedorService

            lote_id = (request.POST.get("lote_id") or "").strip()
            try:
                item_lote = LoteEnvioFornecedorService.adicionar_item_reparo_concluido(
                    fornecedor=request.fornecedor,
                    user=request.user,
                    ordem=ordem,
                    localidade_devolucao_id=request.POST.get("localidade_devolucao"),
                    valor_avaliacao_tecnica=request.POST.get("valor_avaliacao_tecnica", ""),
                    lote_id=(lote_id if lote_id.isdigit() else None),
                    lote_nome_novo=request.POST.get("lote_nome_novo", ""),
                )
                messages.success(
                    request,
                    f'Equipamento separado no lote "{item_lote.lote.nome}" — '
                    f"anexe a NF e envie ao TI quando estiver pronto.",
                )
                return redirect("portal_lote_envio_detail", pk=item_lote.lote_id)
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
                return redirect("portal_manutencao_detail", pk=pk)

        # ── Transição de status conduzida pelo fornecedor ────────────────────
        from services.ordem_manutencao_service import OrdemManutencaoService
        extra = {
            "diagnostico": request.POST.get("diagnostico", ""),
            "nome": request.POST.get("nome", ""),
            "numero_serie": request.POST.get("numero_serie", ""),
            "marca": request.POST.get("marca", ""),
            "modelo": request.POST.get("modelo", ""),
            "contrato": request.POST.get("contrato", ""),
            "valor": request.POST.get("valor", ""),
            "data": request.POST.get("data", ""),
            "locado": request.POST.get("locado", ""),
            "tempo_contrato_meses": request.POST.get("tempo_contrato_meses", ""),
            "localidade_devolucao": request.POST.get("localidade_devolucao", ""),
            "localidade_substituto": request.POST.get("localidade_substituto", ""),
            "reparo_valor": request.POST.get("reparo_valor", ""),
            "valor_orcamento": request.POST.get("valor_orcamento", ""),
            "valor_conserto": request.POST.get("valor_conserto", ""),
            "valor_total": request.POST.get("valor_total", ""),
            "valor_avaliacao_tecnica": request.POST.get("valor_avaliacao_tecnica", ""),
            "valor_equipamento_danificado": request.POST.get("valor_equipamento_danificado", ""),
            "tem_garantia": request.POST.get("tem_garantia", ""),
            "garantia_dias": request.POST.get("garantia_dias", ""),
        }
        try:
            OrdemManutencaoService.transicionar(
                ordem=ordem,
                novo_status=acao,
                user=request.user,
                observacao=request.POST.get("observacao", ""),
                ator="fornecedor",
                **extra,
            )
            messages.success(request, "Status atualizado com sucesso.")
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        return redirect("portal_manutencao_detail", pk=pk)

    from services.lote_envio_fornecedor_service import LoteEnvioFornecedorService
    from services.ordem_manutencao_service import OrdemManutencaoService

    eventos = ordem.eventos.select_related("criado_por").all()
    anexos = ordem.anexos.select_related("criado_por").all()

    context = {
        "fornecedor": request.fornecedor,
        "ordem": ordem,
        "eventos": eventos,
        "anexos": anexos,
        "localidades": Localidade.objects.order_by("local"),
        "active_nav": _portal_manutencao_nav(ordem),
        "pode_separar_lote": LoteEnvioFornecedorService.ordem_elegivel_para_retorno(ordem),
        "rascunho_lote_atual": LoteEnvioFornecedorService.rascunho_retorno_ativo(ordem),
        "lotes_abertos": LoteEnvioFornecedorService.lotes_abertos(request.fornecedor),
        # Desfazer o envio (voltar ao formulário anterior) — só nas propostas
        # ainda aguardando decisão do TI (ver DESFAZAVEIS no service).
        "pode_desfazer": OrdemManutencaoService.pode_desfazer(ordem),
    }
    return render(request, "front/portal/portal_manutencao_detail.html", context)


# ─── Troca antecipada de equipamento ──────────────────────────────────────────

# Estágios abertos da troca antecipada (para métricas/telas).
_TROCA_ANT_ABERTAS = [
    StatusOrdemManutencaoChoices.TROCA_ANT_SUBSTITUTO_ENVIADO,
    StatusOrdemManutencaoChoices.TROCA_ANT_SUBSTITUTO_RECEBIDO,
    StatusOrdemManutencaoChoices.TROCA_ANT_DEFEITUOSO_ENVIADO,
    StatusOrdemManutencaoChoices.TROCA_ANT_DEFEITUOSO_RECEBIDO,
]


def _itens_elegiveis_troca(fornecedor):
    """Equipamentos elegíveis para troca antecipada: do fornecedor, com status
    DEFEITO, sem outra ordem de manutenção aberta e sem rascunho pendente no
    lote de envio (evita duplicar a troca por engano)."""
    from ..models import LoteEnvioFornecedorItem, StatusItemLoteEnvioChoices

    com_os_aberta = (
        OrdemManutencao.objects
        .filter(fornecedor=fornecedor)
        .exclude(status__in=_OS_ABERTAS_EXCLUI)
        .values_list("item_id", flat=True)
    )
    com_rascunho = (
        LoteEnvioFornecedorItem.objects
        .filter(lote__fornecedor=fornecedor, status=StatusItemLoteEnvioChoices.RASCUNHO)
        .values_list("item_defeituoso_id", flat=True)
    )
    return (
        itens_do_fornecedor(fornecedor)
        .filter(status=StatusItemChoices.DEFEITO)
        .exclude(pk__in=list(com_os_aberta))
        .exclude(pk__in=list(com_rascunho))
        .select_related("subtipo", "localidade")
        .order_by("nome")
    )


@fornecedor_required
def portal_troca_antecipada_list(request):
    """Tela dedicada da Troca Antecipada — acompanha em detalhe os processos em
    andamento e o histórico, com o estágio de cada troca."""
    base = (
        OrdemManutencao.objects
        .filter(fornecedor=request.fornecedor, troca_antecipada=True)
        .select_related("item", "item_substituto")
        .order_by("-created_at")
    )
    andamento = base.exclude(status__in=_OS_ABERTAS_EXCLUI)
    historico = base.filter(status__in=_OS_ABERTAS_EXCLUI)[:50]

    counts = {row["status"]: row["n"] for row in base.values("status").annotate(n=Count("id"))}
    context = {
        "fornecedor": request.fornecedor,
        "andamento": andamento,
        "historico": historico,
        "qtd_andamento": andamento.count(),
        "qtd_total": base.count(),
        "qtd_elegiveis": _itens_elegiveis_troca(request.fornecedor).count(),
        "counts": counts,
        "active_nav": "troca_antecipada",
    }
    return render(request, "front/portal/portal_troca_antecipada_list.html", context)


@fornecedor_required
def portal_troca_antecipada_nova(request):
    """Monta o RASCUNHO de uma troca antecipada dentro do Lote de Envio do
    fornecedor (a caixinha ao lado do form) — nenhuma OrdemManutencao/Item é
    criada ainda. O fornecedor só abre a troca de fato (`abrir_troca_antecipada`)
    quando clicar em "Enviar" no lote (ver `portal_lote_envio_fornecedor.py`).
    Só equipamentos com status DEFEITO e sem troca já pendente."""
    from services.lote_envio_fornecedor_service import LoteEnvioFornecedorService

    if request.method == "POST":
        item_id = (request.POST.get("item_defeituoso") or "").strip()
        # Só itens elegíveis (DEFEITO, sem OS aberta e sem rascunho) podem ser escolhidos.
        item = (
            _itens_elegiveis_troca(request.fornecedor).filter(pk=item_id).first()
            if item_id.isdigit() else None
        )
        if item is None:
            messages.error(request, "Selecione um equipamento válido (com status Defeito) para substituir.")
        else:
            lote_id = (request.POST.get("lote_id") or "").strip()
            try:
                item_lote = LoteEnvioFornecedorService.adicionar_item_troca_antecipada(
                    fornecedor=request.fornecedor,
                    user=request.user,
                    item_defeituoso=item,
                    sub_modelo=request.POST.get("sub_modelo", ""),
                    sub_serie=request.POST.get("sub_serie", ""),
                    sub_marca=request.POST.get("sub_marca", ""),
                    sub_data_contrato=request.POST.get("sub_data_contrato", ""),
                    lote_id=(lote_id if lote_id.isdigit() else None),
                    lote_nome_novo=request.POST.get("lote_nome_novo", ""),
                )
                messages.success(
                    request,
                    f"\"{item.nome}\" adicionado ao lote de envio. Quando quiser, envie o item ou o lote inteiro ao TI.",
                )
                lote_origem = (request.GET.get("lote") or "").strip()
                if lote_origem.isdigit():
                    return redirect("portal_lote_envio_detail", pk=item_lote.lote_id)
                return redirect("portal_troca_antecipada_nova")
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))

    lote_preselect_id = (request.GET.get("lote") or "").strip()
    lote_preselect = None
    if lote_preselect_id.isdigit():
        lote_preselect = (
            LoteEnvioFornecedor.objects
            .filter(pk=lote_preselect_id, fornecedor=request.fornecedor, status=StatusLoteEnvioFornecedorChoices.ABERTO)
            .first()
        )
    lotes_abertos = list(
        LoteEnvioFornecedorService.lotes_abertos(request.fornecedor).prefetch_related("itens", "anexos")
    )
    context = {
        "fornecedor": request.fornecedor,
        "itens": _itens_elegiveis_troca(request.fornecedor),
        "lotes_abertos": lotes_abertos,
        "lote_preselect": lote_preselect,
        "active_nav": "troca_antecipada",
    }
    return render(request, "front/portal/portal_troca_antecipada_nova.html", context)


# ─── Central de Ajuda ─────────────────────────────────────────────────────────

@fornecedor_required
def portal_ajuda(request):
    """Central de ajuda do Portal — guia de uso para o fornecedor."""
    context = {
        "fornecedor": request.fornecedor,
        "active_nav": "ajuda",
    }
    return render(request, "front/portal/portal_ajuda.html", context)


@fornecedor_required
def portal_ajuda_diagrama(request):
    """Diagrama de casos de uso do módulo de Manutenção — todos os processos,
    atores (Fornecedor/TI) e decisões, espelhando OrdemManutencaoService."""
    context = {
        "fornecedor": request.fornecedor,
        "active_nav": "ajuda",
    }
    return render(request, "front/portal/portal_ajuda_diagrama.html", context)


# ─── Notificações do fornecedor (sino do Portal) ──────────────────────────────

@fornecedor_required
def portal_notificacoes_marcar_lidas(request):
    """Marca como lidas as notificações do fornecedor (ao abrir o sino do Portal)."""
    from django.http import JsonResponse
    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)
    from ..models import Notificacao
    Notificacao.objects.filter(fornecedor=request.fornecedor, lida_fornecedor=False).update(lida_fornecedor=True)
    return JsonResponse({"ok": True})


# ─── Licenças (somente leitura) ───────────────────────────────────────────────

@fornecedor_required
def portal_licencas_list(request):
    """Licenças vinculadas ao fornecedor — somente leitura, com agregados de lotes."""
    licencas = (
        Licenca.objects
        .filter(fornecedor=request.fornecedor)
        .annotate(
            total_assentos=Coalesce(Sum("lotes__quantidade_total"), 0),
            saldo_disponivel=Coalesce(Sum("lotes__quantidade_disponivel"), 0),
            qtd_lotes=Count("lotes", distinct=True),
        )
        .order_by("nome")
    )

    q = request.GET.get("q", "").strip()
    if q:
        licencas = licencas.filter(nome__icontains=q)

    paginator = Paginator(licencas, 15)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    get_copy = request.GET.copy()
    get_copy.pop("page", None)
    qs_keep = get_copy.urlencode()

    context = {
        "fornecedor": request.fornecedor,
        "page_obj": page_obj,
        "total": paginator.count,
        "f_q": q,
        "qs_keep": qs_keep,
        "active_nav": "licencas",
    }
    return render(request, "front/portal/portal_licencas_list.html", context)
