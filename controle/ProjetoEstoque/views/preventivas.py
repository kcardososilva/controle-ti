from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, OuterRef, Q, Subquery
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from ..forms import (
    ChecklistModeloForm,
    ChecklistPerguntaForm,
    PreventivaStartForm,
)
from ..models import (
    GRUPO_FORNECEDOR,
    CheckListModelo,
    CheckListPergunta,
    Item,
    Localidade,
    Preventiva,
    PreventivaExecucao,
    PreventivaResposta,
    SimNaoChoices,
    StatusItemChoices,
)


# =========================================================
# HELPERS - PROGRAMAÇÃO DE PREVENTIVA
# =========================================================

def _to_int(value, default=0) -> int:
    """Converte valores de intervalo com segurança."""
    try:
        if value in (None, ""):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _intervalo_preventiva(item=None, checklist=None) -> tuple[int, str]:
    """
    Define o intervalo oficial da programação.

    Regra adotada para não alterar radicalmente a estrutura atual:
    1. Primeiro usa Item.data_limite_preventiva, pois é específico do ativo.
    2. Se o item não tiver intervalo, usa CheckListModelo.intervalo_dias.
    3. Se nenhum existir, retorna 0 e deixa a preventiva sem programação calculada.
    """
    item_intervalo = _to_int(getattr(item, "data_limite_preventiva", None), 0)
    if item_intervalo > 0:
        return item_intervalo, "Cadastro do equipamento"

    checklist_intervalo = _to_int(getattr(checklist, "intervalo_dias", None), 0)
    if checklist_intervalo > 0:
        return checklist_intervalo, "Modelo de checklist"

    return 0, "Sem intervalo configurado"


def _calc_proxima(data_ultima=None, data_proxima=None, intervalo_dias: int = 0) -> date | None:
    """
    Calcula a próxima execução.

    - Se já houve execução e existe intervalo, calcula data_ultima + intervalo.
    - Se nunca houve execução, usa data_proxima já gravada.
    - Se nada estiver configurado, retorna None.
    """
    if data_ultima and intervalo_dias > 0:
        return data_ultima + timedelta(days=intervalo_dias)
    return data_proxima


def _classificar_programacao(proxima: date | None, hoje: date | None = None, janela_alerta: int = 7) -> dict:
    """Retorna status visual, label, classe CSS e dias restantes."""
    hoje = hoje or timezone.localdate()

    if not proxima:
        return {
            "status": "indefinido",
            "label": "Sem data",
            "css": "st-indef",
            "dias": None,
            "prioridade": 99,
        }

    dias = (proxima - hoje).days

    if dias < 0:
        return {
            "status": "vencida",
            "label": "Vencida",
            "css": "st-vencida",
            "dias": dias,
            "prioridade": 1,
        }

    if dias <= janela_alerta:
        return {
            "status": "atencao",
            "label": "Atenção",
            "css": "st-atencao",
            "dias": dias,
            "prioridade": 2,
        }

    return {
        "status": "ok",
        "label": "Em dia",
        "css": "st-ok",
        "dias": dias,
        "prioridade": 3,
    }


def _normalizar_decimal(valor: str) -> str:
    """Valida números aceitando vírgula decimal."""
    valor = (valor or "").strip().replace(".", "").replace(",", ".")
    try:
        return str(Decimal(valor))
    except (InvalidOperation, ValueError):
        raise ValueError("Número inválido")


def _preparar_opcoes_perguntas(perguntas):
    """Anexa lista de opções calculada para perguntas de escolha."""
    for pergunta in perguntas:
        pergunta.opcoes_list = []
        if getattr(pergunta, "opcoes", None):
            pergunta.opcoes_list = [
                opcao.strip()
                for opcao in str(pergunta.opcoes).split(",")
                if opcao.strip()
            ]
    return perguntas


def _aplicar_status_preventiva(preventiva: Preventiva, hoje: date | None = None) -> Preventiva:
    """Anexa propriedades calculadas em memória para uso nos templates."""
    hoje = hoje or timezone.localdate()
    intervalo, origem = _intervalo_preventiva(preventiva.equipamento, preventiva.checklist_modelo)
    proxima_auto = _calc_proxima(preventiva.data_ultima, preventiva.data_proxima, intervalo)
    # data_agendamento sobrepõe o cálculo automático; marca a diferença para o template
    agendamento = getattr(preventiva, "data_agendamento", None)
    proxima = agendamento if agendamento else proxima_auto

    preventiva.intervalo_dias_calc = intervalo
    preventiva.intervalo_origem_calc = origem
    preventiva.proxima_calc = proxima
    preventiva.tem_agendamento = bool(agendamento)

    if getattr(preventiva, "pausada", False):
        preventiva.status_visual = "pausada"
        preventiva.status_label = "Pausada"
        preventiva.status_css = "st-pausada"
        preventiva.dias_restantes = getattr(preventiva, "dias_restantes_pausa", None)
        preventiva.prioridade_visual = 4
    else:
        status = _classificar_programacao(proxima, hoje)
        preventiva.status_visual = status["status"]
        preventiva.status_label = status["label"]
        preventiva.status_css = status["css"]
        preventiva.dias_restantes = status["dias"]
        preventiva.prioridade_visual = status["prioridade"]

    return preventiva


def _pode_editar_execucao(execucao: "PreventivaExecucao", user) -> bool:
    """
    Só o técnico que executou pode editar/excluir a execução. Segue a mesma
    regra de identidade usada no bloqueio de conflito de horário: técnico
    explícito se houver, senão quem criou o registro (execuções antigas sem
    `tecnico` preenchido).
    """
    if not user or not user.is_authenticated:
        return False
    if execucao.tecnico_id:
        return execucao.tecnico_id == user.id
    return execucao.criado_por_id == user.id


def _tecnicos_disponiveis(excluir_id=None):
    """Usuários ativos da equipe de TI (fora do grupo Fornecedor), para
    seleção de técnico responsável/auxiliares. Compartilhado pelas telas de
    execução, edição e planos de agendamento."""
    from django.contrib.auth import get_user_model
    AuthUser = get_user_model()
    qs = (
        AuthUser.objects
        .filter(is_active=True)
        .exclude(groups__name=GRUPO_FORNECEDOR)
        .order_by("first_name", "last_name", "username")
    )
    if excluir_id:
        qs = qs.exclude(pk=excluir_id)
    return qs


def _resolver_tecnicos_auxiliares(request, tecnico_principal):
    """
    Valida a lista de técnicos auxiliares enviada no POST (preventivas
    realizadas em dupla): precisam ser usuários ativos da equipe de TI e
    diferentes do técnico principal da execução. Retorna (auxiliares, erros).
    """
    from django.contrib.auth import get_user_model
    AuthUser = get_user_model()

    ids_raw = request.POST.getlist("tecnicos_auxiliares")
    ids = {int(i) for i in ids_raw if str(i).isdigit()}
    if tecnico_principal:
        ids.discard(tecnico_principal.id)

    if not ids:
        return [], []

    auxiliares = list(
        AuthUser.objects
        .filter(pk__in=ids, is_active=True)
        .exclude(groups__name=GRUPO_FORNECEDOR)
    )
    encontrados = {u.id for u in auxiliares}
    erros = ["Um ou mais técnicos auxiliares selecionados são inválidos."] if (ids - encontrados) else []
    return auxiliares, erros


def _conflito_horario(tecnico, data_exec, hora_ini, hora_fim, excluir_pk=None):
    """
    Localiza uma PreventivaExecucao conflitante (mesmo técnico, mesmo dia,
    horário sobreposto). Considera o técnico como responsável principal
    (ou execução legada sem `tecnico`, via `criado_por`) e também como
    auxiliar de outra execução — em ambos os papéis ele está fisicamente
    ocupado naquele horário.
    """
    qs = (
        PreventivaExecucao.objects
        .filter(data_execucao=data_exec)
        .filter(
            Q(tecnico=tecnico)
            | Q(tecnico__isnull=True, criado_por=tecnico)
            | Q(tecnicos_auxiliares=tecnico)
        )
        .exclude(hora_inicio__isnull=True)
        .exclude(hora_fim__isnull=True)
        .filter(hora_inicio__lt=hora_fim, hora_fim__gt=hora_ini)
        .select_related("preventiva__equipamento")
        .order_by("hora_inicio")
        .distinct()
    )
    if excluir_pk:
        qs = qs.exclude(pk=excluir_pk)
    return qs.first()


def _recalcular_agregados_preventiva(preventiva: Preventiva) -> None:
    """
    Recalcula `data_ultima`/`data_proxima`/`dentro_do_prazo` da Preventiva a
    partir do histórico real de execuções restantes.

    Necessário após editar ou excluir uma `PreventivaExecucao`, pois ela pode
    não ser a mais recente — `preventiva.data_ultima` não pode continuar
    refletindo um registro que foi alterado/apagado.
    """
    ultima = preventiva.execucoes.order_by("-data_execucao", "-id").first()
    preventiva.data_ultima = ultima.data_execucao if ultima else None
    preventiva.recomputar_prazo()  # usa preventiva.data_ultima (já atualizado acima)
    preventiva.save(update_fields=["data_ultima", "data_proxima", "dentro_do_prazo", "updated_at"])


# =========================================================
# CHECKLIST - LISTAGEM
# =========================================================
@login_required
def checklist_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        CheckListModelo.objects
        .select_related("subtipo", "subtipo__categoria")
        .annotate(perguntas_count=Count("perguntas"))
        .order_by("-created_at", "nome")
    )

    if q:
        qs = qs.filter(
            Q(nome__icontains=q)
            | Q(subtipo__nome__icontains=q)
            | Q(subtipo__categoria__nome__icontains=q)
        )

    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "q": q,
        "total": qs.count(),
    }
    return render(request, "front/preventivas/checklist_list.html", context)


# =========================================================
# CHECKLIST - CRIAR / EDITAR
# =========================================================
@login_required
def checklist_form(request, pk=None):
    instance = get_object_or_404(CheckListModelo, pk=pk) if pk else None

    if request.method == "POST":
        form = ChecklistModeloForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save(commit=False)
            if instance is None:
                obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Checklist salvo com sucesso.")
            return redirect("checklist_form", pk=obj.pk)
    else:
        form = ChecklistModeloForm(instance=instance)

    perguntas = (
        CheckListPergunta.objects
        .filter(checklist_modelo=instance)
        .order_by("ordem", "id")
        if instance else []
    )

    return render(
        request,
        "front/preventivas/checklist_form.html",
        {"form": form, "perguntas": perguntas},
    )


@login_required
@require_POST
def checklist_delete(request, pk):
    obj = get_object_or_404(CheckListModelo, pk=pk)
    obj.delete()
    messages.success(request, "Checklist removido com sucesso.")
    return redirect("checklist_list")


# =========================================================
# PERGUNTA - CRIAR / EDITAR
# =========================================================
@login_required
def pergunta_form(request, checklist_pk, pk=None):
    checklist = get_object_or_404(CheckListModelo, pk=checklist_pk)
    instance = (
        get_object_or_404(CheckListPergunta, pk=pk, checklist_modelo=checklist)
        if pk else None
    )

    if request.method == "POST":
        form = ChecklistPerguntaForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.checklist_modelo = checklist
            if instance is None:
                obj.criado_por = request.user
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, "Pergunta salva com sucesso.")
            return redirect("checklist_form", pk=checklist.pk)
    else:
        form = ChecklistPerguntaForm(instance=instance)

    return render(
        request,
        "front/preventivas/pergunta_form.html",
        {"form": form, "checklist": checklist},
    )


@login_required
@require_POST
def pergunta_delete(request, checklist_pk, pk):
    checklist = get_object_or_404(CheckListModelo, pk=checklist_pk)
    pergunta = get_object_or_404(CheckListPergunta, pk=pk, checklist_modelo=checklist)
    pergunta.delete()
    messages.success(request, "Pergunta removida com sucesso.")
    return redirect("checklist_form", pk=checklist.pk)


# =========================================================
# PROGRAMAÇÃO AUTOMÁTICA DE PREVENTIVAS
# =========================================================
@login_required
@require_POST
def preventiva_sincronizar_programacao(request):
    """
    Cria programações pendentes para itens marcados com precisa_preventiva='sim'.

    Sem alteração radical de estrutura:
    - Não cria tabela nova.
    - Não muda models.
    - Usa Preventiva existente como programação ativa.
    - Tenta usar checklist do mesmo subtipo; se não houver, usa checklist genérico sem subtipo.
    """
    hoje = timezone.localdate()

    itens = (
        Item.objects
        .filter(precisa_preventiva=SimNaoChoices.SIM)
        .select_related("subtipo", "localidade", "centro_custo")
        .order_by("nome")
    )

    checklists_ativos = (
        CheckListModelo.objects
        .filter(ativo=SimNaoChoices.SIM)
        .select_related("subtipo")
        .order_by("subtipo__nome", "nome")
    )

    _STATUS_PAUSANTES = {
        StatusItemChoices.PAUSADO,
        StatusItemChoices.BACKUP,
        StatusItemChoices.ESTOQUE,
        StatusItemChoices.MANUTENCAO,
        StatusItemChoices.DEFEITO,
        StatusItemChoices.DESCARTE,
        StatusItemChoices.DEVOLVIDO,
    }

    criadas = 0
    existentes = 0
    sem_checklist = 0
    sem_intervalo = 0
    pausadas_sync = 0
    retomadas_sync = 0
    resync = 0

    with transaction.atomic():
        for item in itens:
            checklist = None

            if item.subtipo_id:
                checklist = checklists_ativos.filter(subtipo=item.subtipo).first()

            if checklist is None:
                checklist = checklists_ativos.filter(subtipo__isnull=True).first()

            if checklist is None:
                sem_checklist += 1
                continue

            intervalo, _origem = _intervalo_preventiva(item, checklist)
            if intervalo <= 0:
                sem_intervalo += 1
                continue

            prev, created = Preventiva.objects.get_or_create(
                equipamento=item,
                checklist_modelo=checklist,
                defaults={
                    "data_ultima": None,
                    "data_proxima": hoje,
                    "dentro_do_prazo": True,
                    "criado_por": request.user,
                    "atualizado_por": request.user,
                },
            )

            if created:
                criadas += 1
            else:
                existentes += 1
                # Recalcula e PERSISTE data_proxima = data efetiva (agendamento OU
                # data_ultima + intervalo). Corrige o campo defasado/nulo, que fazia
                # dashboards/alertas mostrarem vencidas erradas.
                antes = prev.data_proxima
                prev.sincronizar_data_proxima(hoje, salvar=False)
                if prev.data_proxima != antes:
                    prev.atualizado_por = request.user
                    prev.save(update_fields=["data_proxima", "dentro_do_prazo", "atualizado_por", "updated_at"])
                    resync += 1

            # Sincroniza estado de pausa conforme status atual do equipamento.
            item_pausante = item.status in _STATUS_PAUSANTES
            if item_pausante and not prev.pausada:
                prev.pausar()
                pausadas_sync += 1
            elif not item_pausante and prev.pausada:
                prev.retomar()
                retomadas_sync += 1

    messages.success(
        request,
        (
            f"Programação sincronizada. Criadas: {criadas}. "
            f"Já existentes: {existentes}. Datas recalculadas: {resync}. "
            f"Sem checklist: {sem_checklist}. Sem intervalo: {sem_intervalo}. "
            f"Pausadas: {pausadas_sync}. Retomadas: {retomadas_sync}."
        ),
    )
    return redirect("preventiva_list")


# =========================================================
# PREVENTIVAS - LISTAGEM / PROGRAMAÇÃO
# =========================================================
@login_required
def preventiva_list(request):
    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    localidade_filter = (request.GET.get("localidade") or "").strip()
    checklist_filter = (request.GET.get("checklist") or "").strip()
    janela_alerta = _to_int(request.GET.get("janela"), 7)
    janela_alerta = janela_alerta if janela_alerta in (7, 15, 30, 60) else 7

    ultima_execucao_qs = (
        PreventivaExecucao.objects
        .filter(preventiva=OuterRef("pk"))
        .order_by("-data_execucao", "-id")
    )

    qs = (
        Preventiva.objects
        .select_related(
            "equipamento",
            "equipamento__localidade",
            "equipamento__centro_custo",
            "equipamento__subtipo",
            "checklist_modelo",
            "atualizado_por",
        )
        .annotate(
            ultimo_executor_username=Subquery(ultima_execucao_qs.values("criado_por__username")[:1]),
            ultimo_executor_nome=Subquery(ultima_execucao_qs.values("criado_por__first_name")[:1]),
            ultima_execucao_data=Subquery(ultima_execucao_qs.values("data_execucao")[:1]),
        )
        .order_by("data_proxima", "equipamento__nome", "id")
    )

    if q:
        qs = qs.filter(
            Q(equipamento__nome__icontains=q)
            | Q(equipamento__numero_serie__icontains=q)
            | Q(equipamento__marca__icontains=q)
            | Q(equipamento__modelo__icontains=q)
            | Q(equipamento__localidade__local__icontains=q)
            | Q(equipamento__centro_custo__departamento__icontains=q)
            | Q(checklist_modelo__nome__icontains=q)
        )

    if localidade_filter:
        qs = qs.filter(equipamento__localidade_id=localidade_filter)

    if checklist_filter:
        qs = qs.filter(checklist_modelo_id=checklist_filter)

    hoje = timezone.localdate()
    processadas = []

    kpi = {
        "total": 0,
        "vencidas": 0,
        "proximas": 0,
        "proximas_30": 0,
        "em_dia": 0,
        "sem_data": 0,
    }

    for preventiva in qs:
        _aplicar_status_preventiva(preventiva, hoje)

        # Reclassifica atenção conforme a janela escolhida pelo usuário (ignora pausadas).
        if not getattr(preventiva, "pausada", False) and preventiva.proxima_calc:
            status = _classificar_programacao(preventiva.proxima_calc, hoje, janela_alerta)
            preventiva.status_visual = status["status"]
            preventiva.status_label = status["label"]
            preventiva.status_css = status["css"]
            preventiva.dias_restantes = status["dias"]
            preventiva.prioridade_visual = status["prioridade"]

        kpi["total"] += 1

        if preventiva.status_visual == "pausada":
            pass  # não conta em nenhuma categoria de prazo
        elif preventiva.status_visual == "vencida":
            kpi["vencidas"] += 1
        elif preventiva.status_visual == "atencao":
            kpi["proximas"] += 1
        elif preventiva.status_visual == "ok":
            kpi["em_dia"] += 1
        else:
            kpi["sem_data"] += 1

        if not getattr(preventiva, "pausada", False) and preventiva.proxima_calc and 0 <= (preventiva.proxima_calc - hoje).days <= 30:
            kpi["proximas_30"] += 1

        processadas.append(preventiva)

    if status_filter:
        if status_filter == "proxima":
            processadas = [p for p in processadas if p.status_visual == "atencao"]
        elif status_filter in ("vencida", "ok", "indefinido", "pausada"):
            processadas = [p for p in processadas if p.status_visual == status_filter]

    processadas.sort(key=lambda p: (p.prioridade_visual, p.proxima_calc or date.max, p.equipamento.nome))

    paginator = Paginator(processadas, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "preventivas": page_obj,
        "kpi": kpi,
        "today": hoje,
        "filter_q": q,
        "filter_status": status_filter,
        "filter_localidade": localidade_filter,
        "filter_checklist": checklist_filter,
        "janela_alerta": janela_alerta,
        "localidades": Localidade.objects.order_by("local"),
        "checklists": CheckListModelo.objects.order_by("nome"),
    }

    if request.GET.get("print") == "true":
        return render(request, "front/preventivas/preventiva_list_print.html", context)

    return render(request, "front/preventivas/preventiva_list.html", context)


# =========================================================
# PREVENTIVA - START
# =========================================================
@login_required
def preventiva_start(request, item_id=None):
    item_instance = None
    if item_id:
        item_instance = get_object_or_404(Item.objects.select_related("subtipo"), pk=item_id)

    if request.method == "POST":
        form = PreventivaStartForm(request.POST, item_instance=item_instance)
        if form.is_valid():
            item = form.cleaned_data["item"]
            checklist = form.cleaned_data["checklist_modelo"]
            hoje = timezone.localdate()

            intervalo, origem = _intervalo_preventiva(item, checklist)
            data_inicial = hoje

            if intervalo <= 0:
                messages.warning(
                    request,
                    "Preventiva criada sem intervalo definido. Configure a periodicidade no equipamento ou no checklist.",
                )

            with transaction.atomic():
                preventiva, created = Preventiva.objects.get_or_create(
                    equipamento=item,
                    checklist_modelo=checklist,
                    defaults={
                        "data_ultima": None,
                        "data_proxima": data_inicial,
                        "dentro_do_prazo": True,
                        "criado_por": request.user,
                        "atualizado_por": request.user,
                    },
                )

                if not created:
                    preventiva.atualizado_por = request.user
                    if not preventiva.data_proxima and not preventiva.data_ultima:
                        preventiva.data_proxima = data_inicial
                    preventiva.save(update_fields=["data_proxima", "atualizado_por", "updated_at"])

            messages.success(
                request,
                f"Preventiva programada para {item.nome}. Intervalo: {intervalo or 'não definido'} dias ({origem}).",
            )
            return redirect("preventiva_exec", pk=preventiva.pk)
    else:
        form = PreventivaStartForm(item_instance=item_instance)

    return render(request, "front/preventivas/preventiva_start.html", {"form": form})


@login_required
def preventiva_start_item(request, item_id):
    return preventiva_start(request, item_id=item_id)


# =========================================================
# PREVENTIVA - DETALHE
# =========================================================
@login_required
def preventiva_detail(request, pk):
    preventiva = get_object_or_404(
        Preventiva.objects.select_related(
            "equipamento",
            "equipamento__localidade",
            "equipamento__centro_custo",
            "equipamento__subtipo",
            "checklist_modelo",
        ),
        pk=pk,
    )

    dt_ini = None
    dt_fim = None

    ini = (request.GET.get("inicio") or "").strip()
    fim = (request.GET.get("fim") or "").strip()

    if ini:
        try:
            dt_ini = datetime.strptime(ini, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "Data inicial inválida. Filtro ignorado.")

    if fim:
        try:
            dt_fim = datetime.strptime(fim, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "Data final inválida. Filtro ignorado.")

    exec_qs = (
        preventiva.execucoes
        .select_related("criado_por", "tecnico")
        .prefetch_related("respostas", "respostas__pergunta", "tecnicos_auxiliares")
        .order_by("-data_execucao", "-id")
    )

    if dt_ini:
        exec_qs = exec_qs.filter(data_execucao__gte=dt_ini)
    if dt_fim:
        exec_qs = exec_qs.filter(data_execucao__lte=dt_fim)

    perguntas = list(
        CheckListPergunta.objects
        .filter(checklist_modelo=preventiva.checklist_modelo)
        .order_by("ordem", "id")
    )

    execucoes_data = []
    total_nao_conforme = 0

    for execucao in exec_qs:
        resp_map = {resposta.pergunta_id: resposta for resposta in execucao.respostas.all()}
        linhas = []

        for pergunta in perguntas:
            resposta_obj = resp_map.get(pergunta.id)
            resposta = resposta_obj.resposta if resposta_obj else "-"

            if str(resposta).strip().lower() in {"nao", "não", "n"}:
                total_nao_conforme += 1

            linhas.append({
                "texto": pergunta.texto_pergunta,
                "tipo": pergunta.tipo_resposta,
                "resposta": resposta,
                "respondido_em": resposta_obj.created_at if resposta_obj else None,
            })

        execucoes_data.append({
            "obj": execucao,
            "linhas": linhas,
            "pode_editar": _pode_editar_execucao(execucao, request.user),
        })

    hoje = timezone.localdate()
    _aplicar_status_preventiva(preventiva, hoje)

    context = {
        "preventiva": preventiva,
        "equipamento": preventiva.equipamento,
        "execucoes": execucoes_data,
        "today": hoje,
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "proxima_calc": preventiva.proxima_calc,
        "status_prazo": preventiva.status_visual,
        "status_css": preventiva.status_css,
        "dias_restantes": preventiva.dias_restantes,
        "intervalo_info": f"{preventiva.intervalo_dias_calc} dias ({preventiva.intervalo_origem_calc})",
        "total_execucoes": len(execucoes_data),
        "total_nao_conforme": total_nao_conforme,
    }

    if request.GET.get("print") == "true":
        return render(request, "front/preventivas/preventiva_print.html", context)

    return render(request, "front/preventivas/preventiva_detail.html", context)


# =========================================================
# PREVENTIVA - EXECUÇÃO
# =========================================================
@login_required
def preventiva_exec(request, pk):
    preventiva = get_object_or_404(
        Preventiva.objects.select_related(
            "equipamento",
            "equipamento__localidade",
            "equipamento__centro_custo",
            "checklist_modelo",
        ),
        pk=pk,
    )

    perguntas = list(
        CheckListPergunta.objects
        .filter(checklist_modelo=preventiva.checklist_modelo)
        .order_by("ordem", "id")
    )
    _preparar_opcoes_perguntas(perguntas)

    if request.method == "POST":
        erros = []
        respostas_bulk = []
        hoje = timezone.localdate()
        tecnico_alvo = preventiva.tecnico or request.user

        # ── Preventiva executada em dupla: técnico(s) auxiliar(es) que
        # estiveram junto e cujas horas também devem ser contabilizadas ──────
        auxiliares, erros_aux = _resolver_tecnicos_auxiliares(request, tecnico_alvo)
        erros.extend(erros_aux)

        # ── Apontamento de horas trabalhadas pelo técnico ─────────────────────
        data_exec_str = (request.POST.get("data_execucao") or "").strip()
        hora_ini_str = (request.POST.get("hora_inicio") or "").strip()
        hora_fim_str = (request.POST.get("hora_fim") or "").strip()

        data_exec = hoje
        if data_exec_str:
            try:
                data_exec = datetime.strptime(data_exec_str, "%Y-%m-%d").date()
            except ValueError:
                erros.append("Data da execução inválida.")
        if data_exec and data_exec > hoje:
            erros.append("A data da execução não pode ser futura.")

        hora_ini = hora_fim = None
        if not hora_ini_str or not hora_fim_str:
            erros.append("Informe a hora de início e a hora de término do serviço.")
        else:
            try:
                hora_ini = datetime.strptime(hora_ini_str, "%H:%M").time()
            except ValueError:
                erros.append("Hora de início inválida.")
            try:
                hora_fim = datetime.strptime(hora_fim_str, "%H:%M").time()
            except ValueError:
                erros.append("Hora de término inválida.")
            if hora_ini and hora_fim and PreventivaExecucao.calcular_duracao_minutos(hora_ini, hora_fim) == 0:
                erros.append("A hora de término deve ser diferente da hora de início.")

        # ── Conflito de horário: nem o técnico principal nem os auxiliares
        # podem ter 2 execuções ao mesmo tempo, no mesmo dia ────────────────
        if hora_ini and hora_fim and hora_fim > hora_ini:
            for tec in [tecnico_alvo, *auxiliares]:
                conflito = _conflito_horario(tec, data_exec, hora_ini, hora_fim)
                if conflito:
                    nome_tec = tec.get_full_name() or tec.username
                    erros.append(
                        f"Conflito de horário: {nome_tec} já tem uma execução registrada em "
                        f"{data_exec:%d/%m/%Y} das {conflito.hora_inicio:%H:%M} às "
                        f"{conflito.hora_fim:%H:%M} ({conflito.preventiva.equipamento.nome}). "
                        f"Escolha um horário livre, por exemplo a partir das {conflito.hora_fim:%H:%M}."
                    )
        elif hora_ini and hora_fim and hora_fim < hora_ini:
            messages.warning(
                request,
                "Horário cruzando a meia-noite: verifique manualmente se não há "
                "conflito com outra execução do mesmo técnico.",
            )

        if not perguntas:
            erros.append("Este checklist não possui perguntas cadastradas.")

        for pergunta in perguntas:
            field_name = f"r_{pergunta.id}"
            raw_val = (request.POST.get(field_name) or "").strip()
            tipo = str(pergunta.tipo_resposta or "").lower()

            if pergunta.obrigatorio == SimNaoChoices.SIM and not raw_val:
                erros.append(f"A pergunta '{pergunta.texto_pergunta}' é obrigatória.")
                continue

            if not raw_val:
                continue

            if tipo in {"numero", "inteiro", "decimal"}:
                try:
                    raw_val = _normalizar_decimal(raw_val)
                except ValueError:
                    erros.append(f"Valor numérico inválido na pergunta '{pergunta.texto_pergunta}'.")
                    continue

            if tipo in {"booleano", "sim_nao", "sn"} and raw_val not in {"sim", "nao"}:
                erros.append(f"Resposta inválida na pergunta '{pergunta.texto_pergunta}'.")
                continue

            if tipo in {"escolha", "opcao", "choice"}:
                opcoes_validas = getattr(pergunta, "opcoes_list", [])
                if opcoes_validas and raw_val not in opcoes_validas:
                    erros.append(f"Opção inválida na pergunta '{pergunta.texto_pergunta}'.")
                    continue

            respostas_bulk.append(
                PreventivaResposta(
                    preventiva=preventiva,
                    pergunta=pergunta,
                    resposta=raw_val,
                    criado_por=request.user,
                    atualizado_por=request.user,
                )
            )

        if erros:
            for erro in erros:
                messages.error(request, erro)
            _aplicar_status_preventiva(preventiva, hoje)
            return render(
                request,
                "front/preventivas/preventiva_exec.html",
                {
                    "preventiva": preventiva,
                    "perguntas": perguntas,
                    "today": hoje,
                    "apontamento": {
                        "data_execucao": data_exec_str,
                        "hora_inicio": hora_ini_str,
                        "hora_fim": hora_fim_str,
                    },
                    "tecnicos_disponiveis": _tecnicos_disponiveis(excluir_id=tecnico_alvo.id),
                    "tecnicos_auxiliares_ids": [t.id for t in auxiliares],
                },
            )

        with transaction.atomic():
            intervalo, _origem = _intervalo_preventiva(preventiva.equipamento, preventiva.checklist_modelo)
            proxima = data_exec + timedelta(days=intervalo) if intervalo > 0 else None

            # Snapshot de desempenho antes de limpar o agendamento
            data_agendada_snap = preventiva.data_agendamento
            no_prazo_snap = (data_agendada_snap is None) or (data_exec <= data_agendada_snap)

            execucao = PreventivaExecucao.objects.create(
                preventiva=preventiva,
                data_execucao=data_exec,
                observacao=(request.POST.get("observacao") or "").strip(),
                foto_antes=request.FILES.get("foto_antes"),
                foto_depois=request.FILES.get("foto_depois"),
                foto_antes_2=request.FILES.get("foto_antes_2"),
                foto_depois_2=request.FILES.get("foto_depois_2"),
                tecnico=tecnico_alvo,
                data_agendada=data_agendada_snap,
                no_prazo=no_prazo_snap,
                hora_inicio=hora_ini,
                hora_fim=hora_fim,
                criado_por=request.user,
                atualizado_por=request.user,
            )

            if auxiliares:
                execucao.tecnicos_auxiliares.set(auxiliares)

            for resposta in respostas_bulk:
                resposta.execucao = execucao

            if respostas_bulk:
                PreventivaResposta.objects.bulk_create(respostas_bulk)

            preventiva.data_ultima = data_exec
            preventiva.data_proxima = proxima
            preventiva.data_agendamento = None  # agendamento consumido pela execução
            preventiva.dentro_do_prazo = True if proxima is None else hoje <= proxima
            preventiva.atualizado_por = request.user

            foto_antes = request.FILES.get("foto_antes")
            foto_depois = request.FILES.get("foto_depois")
            foto_antes_2 = request.FILES.get("foto_antes_2")
            foto_depois_2 = request.FILES.get("foto_depois_2")

            update_fields = [
                "data_ultima",
                "data_proxima",
                "data_agendamento",
                "dentro_do_prazo",
                "atualizado_por",
                "updated_at",
            ]

            if foto_antes:
                preventiva.foto_antes = foto_antes
                update_fields.append("foto_antes")

            if foto_depois:
                preventiva.foto_depois = foto_depois
                update_fields.append("foto_depois")

            if foto_antes_2:
                preventiva.foto_antes_2 = foto_antes_2
                update_fields.append("foto_antes_2")

            if foto_depois_2:
                preventiva.foto_depois_2 = foto_depois_2
                update_fields.append("foto_depois_2")

            preventiva.save(update_fields=update_fields)

        duracao_txt = execucao.duracao_formatada
        sufixo_aux = ""
        if auxiliares:
            nomes_aux = ", ".join(a.get_full_name() or a.username for a in auxiliares)
            sufixo_aux = f" Horas também contabilizadas para: {nomes_aux}."
        if duracao_txt:
            messages.success(
                request,
                f"Execução registrada ({duracao_txt} de serviço) e próxima preventiva reagendada com sucesso.{sufixo_aux}",
            )
        else:
            messages.success(request, f"Execução registrada e próxima preventiva reagendada com sucesso.{sufixo_aux}")
        return redirect("preventiva_detail", pk=preventiva.pk)

    hoje = timezone.localdate()
    _aplicar_status_preventiva(preventiva, hoje)

    return render(
        request,
        "front/preventivas/preventiva_exec.html",
        {
            "preventiva": preventiva,
            "perguntas": perguntas,
            "today": hoje,
            "tecnicos_disponiveis": _tecnicos_disponiveis(excluir_id=(preventiva.tecnico_id or request.user.id)),
            "tecnicos_auxiliares_ids": [],
        },
    )


# =========================================================
# PREVENTIVA - EXECUÇÃO: EDIÇÃO / EXCLUSÃO
# (restrito ao técnico que executou — ver _pode_editar_execucao)
# =========================================================
@login_required
def preventiva_execucao_editar(request, execucao_pk):
    execucao = get_object_or_404(
        PreventivaExecucao.objects.select_related(
            "preventiva",
            "preventiva__equipamento",
            "preventiva__equipamento__localidade",
            "preventiva__equipamento__centro_custo",
            "preventiva__checklist_modelo",
            "tecnico",
            "criado_por",
        ),
        pk=execucao_pk,
    )
    preventiva = execucao.preventiva

    if not _pode_editar_execucao(execucao, request.user):
        messages.error(request, "Apenas o técnico que executou esta preventiva pode editá-la.")
        return redirect("preventiva_detail", pk=preventiva.pk)

    perguntas = list(
        CheckListPergunta.objects
        .filter(checklist_modelo=preventiva.checklist_modelo)
        .order_by("ordem", "id")
    )
    _preparar_opcoes_perguntas(perguntas)

    respostas_map = {r.pergunta_id: r.resposta for r in execucao.respostas.all()}
    for pergunta in perguntas:
        pergunta.valor_atual = respostas_map.get(pergunta.id, "")

    if request.method == "POST":
        erros = []
        hoje = timezone.localdate()
        tecnico_alvo = execucao.tecnico or execucao.criado_por or request.user

        auxiliares, erros_aux = _resolver_tecnicos_auxiliares(request, tecnico_alvo)
        erros.extend(erros_aux)

        data_exec_str = (request.POST.get("data_execucao") or "").strip()
        hora_ini_str = (request.POST.get("hora_inicio") or "").strip()
        hora_fim_str = (request.POST.get("hora_fim") or "").strip()

        data_exec = execucao.data_execucao
        if data_exec_str:
            try:
                data_exec = datetime.strptime(data_exec_str, "%Y-%m-%d").date()
            except ValueError:
                erros.append("Data da execução inválida.")
        if data_exec and data_exec > hoje:
            erros.append("A data da execução não pode ser futura.")

        hora_ini = hora_fim = None
        if not hora_ini_str or not hora_fim_str:
            erros.append("Informe a hora de início e a hora de término do serviço.")
        else:
            try:
                hora_ini = datetime.strptime(hora_ini_str, "%H:%M").time()
            except ValueError:
                erros.append("Hora de início inválida.")
            try:
                hora_fim = datetime.strptime(hora_fim_str, "%H:%M").time()
            except ValueError:
                erros.append("Hora de término inválida.")
            if hora_ini and hora_fim and PreventivaExecucao.calcular_duracao_minutos(hora_ini, hora_fim) == 0:
                erros.append("A hora de término deve ser diferente da hora de início.")

        # ── Conflito de horário (mesma regra do registro, excluindo a própria
        # execução sendo editada; nem o técnico principal nem os auxiliares
        # podem ter outra execução no mesmo horário) ─────────────────────────
        if hora_ini and hora_fim and hora_fim > hora_ini:
            for tec in [tecnico_alvo, *auxiliares]:
                conflito = _conflito_horario(tec, data_exec, hora_ini, hora_fim, excluir_pk=execucao.pk)
                if conflito:
                    nome_tec = tec.get_full_name() or tec.username
                    erros.append(
                        f"Conflito de horário: {nome_tec} já tem outra execução registrada em "
                        f"{data_exec:%d/%m/%Y} das {conflito.hora_inicio:%H:%M} às "
                        f"{conflito.hora_fim:%H:%M} ({conflito.preventiva.equipamento.nome}). "
                        f"Escolha um horário livre, por exemplo a partir das {conflito.hora_fim:%H:%M}."
                    )
        elif hora_ini and hora_fim and hora_fim < hora_ini:
            messages.warning(
                request,
                "Horário cruzando a meia-noite: verifique manualmente se não há "
                "conflito com outra execução do mesmo técnico.",
            )

        if not perguntas:
            erros.append("Este checklist não possui perguntas cadastradas.")

        # Preenche já a observação no objeto em memória (não persistida ainda),
        # para que uma eventual re-exibição por erro mostre o que foi digitado.
        execucao.observacao = (request.POST.get("observacao") or "").strip()

        respostas_valores = {}
        for pergunta in perguntas:
            field_name = f"r_{pergunta.id}"
            raw_val = (request.POST.get(field_name) or "").strip()
            tipo = str(pergunta.tipo_resposta or "").lower()

            if pergunta.obrigatorio == SimNaoChoices.SIM and not raw_val:
                erros.append(f"A pergunta '{pergunta.texto_pergunta}' é obrigatória.")
                continue

            if not raw_val:
                continue

            if tipo in {"numero", "inteiro", "decimal"}:
                try:
                    raw_val = _normalizar_decimal(raw_val)
                except ValueError:
                    erros.append(f"Valor numérico inválido na pergunta '{pergunta.texto_pergunta}'.")
                    continue

            if tipo in {"booleano", "sim_nao", "sn"} and raw_val not in {"sim", "nao"}:
                erros.append(f"Resposta inválida na pergunta '{pergunta.texto_pergunta}'.")
                continue

            if tipo in {"escolha", "opcao", "choice"}:
                opcoes_validas = getattr(pergunta, "opcoes_list", [])
                if opcoes_validas and raw_val not in opcoes_validas:
                    erros.append(f"Opção inválida na pergunta '{pergunta.texto_pergunta}'.")
                    continue

            respostas_valores[pergunta.id] = raw_val

        if erros:
            for erro in erros:
                messages.error(request, erro)
            for pergunta in perguntas:
                pergunta.valor_atual = respostas_valores.get(pergunta.id, "")
            _aplicar_status_preventiva(preventiva, hoje)
            return render(
                request,
                "front/preventivas/preventiva_exec.html",
                {
                    "preventiva": preventiva,
                    "perguntas": perguntas,
                    "today": hoje,
                    "execucao": execucao,
                    "apontamento": {
                        "data_execucao": data_exec_str,
                        "hora_inicio": hora_ini_str,
                        "hora_fim": hora_fim_str,
                    },
                    "tecnicos_disponiveis": _tecnicos_disponiveis(excluir_id=tecnico_alvo.id),
                    "tecnicos_auxiliares_ids": [t.id for t in auxiliares],
                },
            )

        with transaction.atomic():
            execucao.data_execucao = data_exec
            execucao.hora_inicio = hora_ini
            execucao.hora_fim = hora_fim
            # observacao já foi setada acima (antes da checagem de erros)
            if execucao.data_agendada:
                execucao.no_prazo = data_exec <= execucao.data_agendada

            for campo in ("foto_antes", "foto_depois", "foto_antes_2", "foto_depois_2"):
                arquivo = request.FILES.get(campo)
                if arquivo:
                    setattr(execucao, campo, arquivo)

            execucao.atualizado_por = request.user
            execucao.save()
            execucao.tecnicos_auxiliares.set(auxiliares)

            existentes = {r.pergunta_id: r for r in execucao.respostas.all()}
            for pergunta in perguntas:
                valor = respostas_valores.get(pergunta.id)
                resposta_obj = existentes.get(pergunta.id)
                if resposta_obj is not None:
                    if valor is None:
                        resposta_obj.delete()
                    elif resposta_obj.resposta != valor:
                        resposta_obj.resposta = valor
                        resposta_obj.atualizado_por = request.user
                        resposta_obj.save(update_fields=["resposta", "atualizado_por", "updated_at"])
                elif valor is not None:
                    PreventivaResposta.objects.create(
                        preventiva=preventiva,
                        execucao=execucao,
                        pergunta=pergunta,
                        resposta=valor,
                        criado_por=request.user,
                        atualizado_por=request.user,
                    )

            _recalcular_agregados_preventiva(preventiva)

        messages.success(request, "Execução atualizada com sucesso.")
        return redirect("preventiva_detail", pk=preventiva.pk)

    hoje = timezone.localdate()
    _aplicar_status_preventiva(preventiva, hoje)
    return render(
        request,
        "front/preventivas/preventiva_exec.html",
        {
            "preventiva": preventiva,
            "perguntas": perguntas,
            "today": hoje,
            "execucao": execucao,
            "apontamento": {
                "data_execucao": execucao.data_execucao.isoformat(),
                "hora_inicio": execucao.hora_inicio.strftime("%H:%M") if execucao.hora_inicio else "",
                "hora_fim": execucao.hora_fim.strftime("%H:%M") if execucao.hora_fim else "",
            },
            "tecnicos_disponiveis": _tecnicos_disponiveis(
                excluir_id=(execucao.tecnico_id or execucao.criado_por_id or request.user.id)
            ),
            "tecnicos_auxiliares_ids": list(execucao.tecnicos_auxiliares.values_list("id", flat=True)),
        },
    )


@login_required
def preventiva_execucao_excluir(request, execucao_pk):
    execucao = get_object_or_404(
        PreventivaExecucao.objects.select_related(
            "preventiva", "preventiva__equipamento", "tecnico", "criado_por",
        ),
        pk=execucao_pk,
    )
    preventiva = execucao.preventiva

    if not _pode_editar_execucao(execucao, request.user):
        messages.error(request, "Apenas o técnico que executou esta preventiva pode excluí-la.")
        return redirect("preventiva_detail", pk=preventiva.pk)

    if request.method == "POST":
        senha = request.POST.get("senha") or ""
        if not request.user.check_password(senha):
            messages.error(request, "Senha incorreta. A execução não foi excluída.")
            return redirect("preventiva_execucao_excluir", execucao_pk=execucao.pk)

        with transaction.atomic():
            execucao.delete()
            _recalcular_agregados_preventiva(preventiva)

        messages.success(request, "Execução excluída com sucesso.")
        return redirect("preventiva_detail", pk=preventiva.pk)

    return render(
        request,
        "front/preventivas/preventiva_execucao_excluir.html",
        {"execucao": execucao, "preventiva": preventiva},
    )


@login_required
def preventiva_agendadas(request):
    """
    Lista todas as preventivas com data_agendamento definida,
    agrupadas e ordenadas por proximidade da data agendada.
    """
    hoje = timezone.localdate()

    base_qs = Preventiva.objects.filter(data_agendamento__isnull=False)

    kpi_total    = base_qs.count()
    kpi_vencidas = base_qs.filter(data_agendamento__lt=hoje).count()
    kpi_hoje     = base_qs.filter(data_agendamento=hoje).count()
    kpi_semana   = base_qs.filter(
        data_agendamento__gte=hoje,
        data_agendamento__lte=hoje + timedelta(days=7),
    ).count()

    q                = (request.GET.get("q") or "").strip()
    localidade_filter = (request.GET.get("localidade") or "").strip()
    periodo_filter   = (request.GET.get("periodo") or "").strip()

    ultima_exec_qs = (
        PreventivaExecucao.objects
        .filter(preventiva=OuterRef("pk"))
        .order_by("-data_execucao", "-id")
    )

    qs = base_qs.select_related(
        "equipamento",
        "equipamento__localidade",
        "equipamento__centro_custo",
        "equipamento__subtipo",
        "checklist_modelo",
        "atualizado_por",
    ).annotate(
        ultimo_executor_username=Subquery(ultima_exec_qs.values("criado_por__username")[:1]),
        ultimo_executor_nome=Subquery(ultima_exec_qs.values("criado_por__first_name")[:1]),
        ultima_execucao_data=Subquery(ultima_exec_qs.values("data_execucao")[:1]),
    )

    if q:
        qs = qs.filter(
            Q(equipamento__nome__icontains=q)
            | Q(equipamento__numero_serie__icontains=q)
            | Q(equipamento__localidade__local__icontains=q)
            | Q(checklist_modelo__nome__icontains=q)
        )
    if localidade_filter:
        qs = qs.filter(equipamento__localidade_id=localidade_filter)
    if periodo_filter == "vencidas":
        qs = qs.filter(data_agendamento__lt=hoje)
    elif periodo_filter == "hoje":
        qs = qs.filter(data_agendamento=hoje)
    elif periodo_filter == "semana":
        qs = qs.filter(data_agendamento__gte=hoje, data_agendamento__lte=hoje + timedelta(days=7))

    qs = qs.order_by("data_agendamento", "equipamento__nome")

    preventivas = list(qs)
    for p in preventivas:
        _aplicar_status_preventiva(p, hoje)
        p.dias_agendamento = (p.data_agendamento - hoje).days

    paginator = Paginator(preventivas, 25)
    page_obj  = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "kpi": {
            "total":    kpi_total,
            "vencidas": kpi_vencidas,
            "hoje":     kpi_hoje,
            "semana":   kpi_semana,
        },
        "today":            hoje,
        "filter_q":         q,
        "filter_localidade": localidade_filter,
        "filter_periodo":   periodo_filter,
        "localidades":      Localidade.objects.order_by("local"),
        "total":            len(preventivas),
    }
    return render(request, "front/preventivas/preventiva_agendadas.html", context)


@login_required
def preventiva_plano(request):
    """
    Plano de Preventivas — agendamento em lote por equipamento.

    GET : lista todos os equipamentos com precisa_preventiva='sim',
          exibindo o estado da preventiva associada (quando existir).
    POST acao='agendar' : cria Preventiva via get_or_create e define data_agendamento.
    POST acao='limpar'  : limpa data_agendamento das preventivas dos itens selecionados.
    """
    from django.utils.dateparse import parse_date
    from types import SimpleNamespace

    from django.contrib.auth import get_user_model
    AuthUser = get_user_model()

    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    localidade_filter = (request.GET.get("localidade") or "").strip()
    checklist_filter = (request.GET.get("checklist") or "").strip()
    tecnico_filter = (request.GET.get("tecnico") or "").strip()

    if request.method == "POST":
        item_ids_raw = request.POST.getlist("item_ids")
        data_str = (request.POST.get("data_agendamento") or "").strip()
        acao = (request.POST.get("acao") or "").strip()

        if acao not in ("agendar", "limpar"):
            messages.error(request, "Ação inválida.")
            return redirect("preventiva_plano")

        item_ids = [int(i) for i in item_ids_raw if str(i).isdigit()]
        if not item_ids:
            messages.warning(request, "Nenhum equipamento selecionado.")
            return redirect("preventiva_plano")

        if acao == "limpar":
            hoje = timezone.localdate()
            prevs = (
                Preventiva.objects
                .filter(equipamento_id__in=item_ids)
                .select_related("equipamento", "checklist_modelo")
            )
            count = 0
            with transaction.atomic():
                for pv in prevs:
                    pv.data_agendamento = None
                    pv.tecnico = None
                    # data_proxima volta a refletir a data automática (data_ultima + intervalo)
                    pv.sincronizar_data_proxima(hoje, salvar=False)
                    pv.atualizado_por = request.user
                    pv.save(update_fields=[
                        "data_agendamento", "tecnico", "data_proxima",
                        "dentro_do_prazo", "atualizado_por", "updated_at",
                    ])
                    count += 1
            messages.success(request, f"Agendamento removido de {count} preventiva(s).")
            return redirect("preventiva_plano")

        # acao == "agendar"
        if not data_str:
            messages.error(request, "Informe uma data para agendar os equipamentos selecionados.")
            return redirect("preventiva_plano")
        data = parse_date(data_str)
        if not data:
            messages.error(request, "Data inválida.")
            return redirect("preventiva_plano")

        # Técnico responsável (opcional) — usuário Django ativo (equipe de TI)
        tecnico_obj = None
        tecnico_id = (request.POST.get("tecnico_id") or "").strip()
        if tecnico_id.isdigit():
            from django.contrib.auth import get_user_model
            AuthUser = get_user_model()
            tecnico_obj = AuthUser.objects.filter(pk=int(tecnico_id), is_active=True).first()

        itens = Item.objects.filter(pk__in=item_ids).select_related("subtipo")
        checklists_ativos = (
            CheckListModelo.objects
            .filter(ativo=SimNaoChoices.SIM)
            .select_related("subtipo")
            .order_by("subtipo__nome", "nome")
        )
        hoje = timezone.localdate()
        count = 0
        sem_checklist = 0

        with transaction.atomic():
            for item in itens:
                checklist = None
                if item.subtipo_id:
                    checklist = checklists_ativos.filter(subtipo=item.subtipo).first()
                if checklist is None:
                    checklist = checklists_ativos.filter(subtipo__isnull=True).first()
                if checklist is None:
                    sem_checklist += 1
                    continue

                preventiva, created = Preventiva.objects.get_or_create(
                    equipamento=item,
                    checklist_modelo=checklist,
                    defaults={
                        "data_ultima": None,
                        "dentro_do_prazo": True,
                        "criado_por": request.user,
                        "atualizado_por": request.user,
                    },
                )
                preventiva.data_agendamento = data
                preventiva.atualizado_por = request.user
                campos = ["data_agendamento", "data_proxima", "dentro_do_prazo", "atualizado_por", "updated_at"]
                if tecnico_obj is not None:
                    preventiva.tecnico = tecnico_obj
                    campos.append("tecnico")
                # data_proxima passa a refletir a data efetiva (o agendamento)
                preventiva.sincronizar_data_proxima(hoje, salvar=False)
                preventiva.save(update_fields=campos)
                count += 1

        if sem_checklist:
            messages.warning(
                request,
                f"{sem_checklist} equipamento(s) sem checklist configurado foram ignorados.",
            )
        if count:
            messages.success(
                request,
                f"{count} equipamento(s) agendado(s) para {data.strftime('%d/%m/%Y')}.",
            )
        return redirect("preventiva_plano")

    # ── GET — listar equipamentos com precisa_preventiva='sim' ────────────────
    items_qs = (
        Item.objects
        .filter(precisa_preventiva=SimNaoChoices.SIM)
        .select_related("localidade", "subtipo", "centro_custo")
        .order_by("nome")
    )

    if q:
        items_qs = items_qs.filter(
            Q(nome__icontains=q)
            | Q(numero_serie__icontains=q)
            | Q(marca__icontains=q)
            | Q(modelo__icontains=q)
            | Q(localidade__local__icontains=q)
            | Q(centro_custo__departamento__icontains=q)
        )

    if localidade_filter:
        items_qs = items_qs.filter(localidade_id=localidade_filter)

    if checklist_filter:
        ids_com_checklist = set(
            Preventiva.objects.filter(
                equipamento__in=items_qs,
                checklist_modelo_id=checklist_filter,
            ).values_list("equipamento_id", flat=True)
        )
        items_qs = items_qs.filter(pk__in=ids_com_checklist)

    # Pré-carrega a preventiva mais urgente por item (evita N+1)
    prev_map = {}
    for p in (
        Preventiva.objects
        .filter(equipamento__in=items_qs)
        .select_related("checklist_modelo", "equipamento", "tecnico")
        .order_by("equipamento_id", "data_proxima")
    ):
        if p.equipamento_id not in prev_map:
            prev_map[p.equipamento_id] = p

    hoje = timezone.localdate()
    rows = []

    for item in items_qs:
        p = prev_map.get(item.pk)

        row = SimpleNamespace(
            pk=item.pk,
            equipamento=item,
            checklist_modelo=p.checklist_modelo if p else None,
            data_ultima=p.data_ultima if p else None,
            data_proxima=p.data_proxima if p else None,
            data_agendamento=p.data_agendamento if p else None,
            pausada=getattr(p, "pausada", False) if p else False,
            status_visual="indefinido",
            status_label="Sem preventiva",
            status_css="st-indef",
            dias_restantes=None,
            prioridade_visual=98,
            intervalo_dias_calc=0,
            intervalo_origem_calc="Sem intervalo configurado",
            proxima_calc=None,
            tem_agendamento=False,
            tecnico=(p.tecnico if p else None),
        )

        if p:
            _aplicar_status_preventiva(p, hoje)
            row.status_visual = p.status_visual
            row.status_label = p.status_label
            row.status_css = p.status_css
            row.dias_restantes = p.dias_restantes
            row.prioridade_visual = p.prioridade_visual
            row.intervalo_dias_calc = p.intervalo_dias_calc
            row.intervalo_origem_calc = p.intervalo_origem_calc
            row.proxima_calc = p.proxima_calc
            row.tem_agendamento = p.tem_agendamento
        else:
            intervalo, origem = _intervalo_preventiva(item, None)
            row.intervalo_dias_calc = intervalo
            row.intervalo_origem_calc = origem

        rows.append(row)

    if status_filter:
        if status_filter == "proxima":
            rows = [r for r in rows if r.status_visual == "atencao"]
        elif status_filter == "agendada":
            rows = [r for r in rows if r.tem_agendamento]
        elif status_filter in ("vencida", "ok", "indefinido", "pausada"):
            rows = [r for r in rows if r.status_visual == status_filter]

    if tecnico_filter.isdigit():
        _tid = int(tecnico_filter)
        rows = [r for r in rows if r.tecnico and r.tecnico.pk == _tid]
    elif tecnico_filter == "sem":
        rows = [r for r in rows if not r.tecnico]

    rows.sort(
        key=lambda r: (r.prioridade_visual, r.proxima_calc or date.max, r.equipamento.nome)
    )

    paginator = Paginator(rows, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "today": hoje,
        "filter_q": q,
        "filter_status": status_filter,
        "filter_localidade": localidade_filter,
        "filter_checklist": checklist_filter,
        "filter_tecnico": tecnico_filter,
        "localidades": Localidade.objects.order_by("local"),
        "checklists": CheckListModelo.objects.order_by("nome"),
        "tecnicos": AuthUser.objects.filter(is_active=True).exclude(groups__name=GRUPO_FORNECEDOR).order_by("first_name", "last_name", "username"),
        "total": len(rows),
    }
    return render(request, "front/preventivas/preventiva_plano.html", context)


@login_required
def preventiva_agendar(request, pk):
    """AJAX — define ou limpa data_agendamento de uma preventiva."""
    from django.http import JsonResponse
    from django.utils.dateparse import parse_date

    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método não permitido."}, status=405)

    preventiva = get_object_or_404(Preventiva, pk=pk)
    data_str = (request.POST.get("data_agendamento") or "").strip()

    if data_str:
        data = parse_date(data_str)
        if not data:
            return JsonResponse({"ok": False, "erro": "Data inválida."}, status=400)
        preventiva.data_agendamento = data
    else:
        preventiva.data_agendamento = None

    preventiva.atualizado_por = request.user
    # Mantém data_proxima = data efetiva (agendamento, ou data automática ao limpar)
    preventiva.sincronizar_data_proxima(timezone.localdate(), salvar=False)
    preventiva.save(update_fields=[
        "data_agendamento", "data_proxima", "dentro_do_prazo", "atualizado_por", "updated_at",
    ])

    return JsonResponse({
        "ok": True,
        "data_agendamento": preventiva.data_agendamento.strftime("%Y-%m-%d") if preventiva.data_agendamento else None,
        "data_agendamento_fmt": preventiva.data_agendamento.strftime("%d/%m/%Y") if preventiva.data_agendamento else None,
    })


# =========================================================
# DESEMPENHO DO TÉCNICO + MINHAS ATIVIDADES
# =========================================================

def _parse_data_opt(valor, default):
    from django.utils.dateparse import parse_date
    d = parse_date(valor) if valor else None
    return d or default


_MESES_PT_PREV = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                  "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def _fmt_minutos(minutos) -> str:
    """Formata uma quantidade de minutos como '1h 30min', '2h', '45min' ou '—'."""
    if not minutos:
        return "0min" if minutos == 0 else "—"
    minutos = int(minutos)
    horas, mins = divmod(minutos, 60)
    if horas and mins:
        return f"{horas}h {mins}min"
    if horas:
        return f"{horas}h"
    return f"{mins}min"


@login_required
def tecnico_desempenho(request):
    """Medidores de desempenho dos técnicos quanto às atividades agendadas/executadas."""
    from django.contrib.auth import get_user_model
    from django.db.models import Sum
    from django.db.models.functions import Coalesce, TruncMonth

    AuthUser = get_user_model()
    hoje = timezone.localdate()

    inicio = _parse_data_opt(request.GET.get("inicio"), date(hoje.year, 1, 1))
    fim = _parse_data_opt(request.GET.get("fim"), hoje)
    if inicio > fim:
        inicio, fim = fim, inicio

    # Execuções no período, agrupadas pelo técnico efetivo (snapshot ou executor)
    execs = (
        PreventivaExecucao.objects
        .filter(data_execucao__gte=inicio, data_execucao__lte=fim)
        .annotate(tec=Coalesce("tecnico_id", "criado_por_id"))
    )

    por_exec = {
        row["tec"]: row
        for row in execs.values("tec").annotate(
            total=Count("id"),
            com_agenda=Count("id", filter=Q(data_agendada__isnull=False)),
            no_prazo=Count("id", filter=Q(data_agendada__isnull=False, no_prazo=True)),
            minutos=Coalesce(Sum("duracao_minutos"), 0),
            com_horas=Count("id", filter=Q(duracao_minutos__isnull=False)),
        )
        if row["tec"] is not None
    }

    # Atribuições atuais (pendentes agendadas) por técnico
    por_atrib = {
        row["tecnico_id"]: row
        for row in (
            Preventiva.objects
            .filter(tecnico__isnull=False, data_agendamento__isnull=False, pausada=False)
            .values("tecnico_id")
            .annotate(
                atribuidas=Count("id"),
                atrasadas=Count("id", filter=Q(data_agendamento__lt=hoje)),
            )
        )
    }

    # Participações como técnico AUXILIAR (preventivas em dupla): contam para
    # as execuções e as horas do colega, sem afetar pontualidade/atribuições
    # (que seguem ligadas apenas ao técnico oficialmente responsável).
    por_aux = {
        row["tecnicos_auxiliares"]: row
        for row in (
            PreventivaExecucao.objects
            .filter(data_execucao__gte=inicio, data_execucao__lte=fim, tecnicos_auxiliares__isnull=False)
            .values("tecnicos_auxiliares")
            .annotate(
                total=Count("id", distinct=True),
                minutos=Coalesce(Sum("duracao_minutos"), 0),
                com_horas=Count("id", filter=Q(duracao_minutos__isnull=False), distinct=True),
            )
        )
    }

    fornecedor_ids = set(
        AuthUser.objects.filter(groups__name=GRUPO_FORNECEDOR).values_list("id", flat=True)
    )
    ids = (set(por_exec) | set(por_atrib) | set(
        AuthUser.objects.filter(is_active=True).values_list("id", flat=True)
    )) - fornecedor_ids
    users = {u.id: u for u in AuthUser.objects.filter(id__in=ids)}

    tecnicos = []
    for uid in ids:
        u = users.get(uid)
        if not u:
            continue
        ex = por_exec.get(uid, {})
        at = por_atrib.get(uid, {})
        aux = por_aux.get(uid, {})
        execucoes_proprias = int(ex.get("total", 0))
        com_agenda = int(ex.get("com_agenda", 0))
        no_prazo = int(ex.get("no_prazo", 0))
        minutos_proprios = int(ex.get("minutos", 0) or 0)
        com_horas = int(ex.get("com_horas", 0)) + int(aux.get("com_horas", 0))
        atribuidas = int(at.get("atribuidas", 0))
        atrasadas = int(at.get("atrasadas", 0))
        execucoes_auxiliar = int(aux.get("total", 0))
        minutos_auxiliar = int(aux.get("minutos", 0) or 0)
        execucoes = execucoes_proprias + execucoes_auxiliar
        minutos = minutos_proprios + minutos_auxiliar
        if execucoes == 0 and atribuidas == 0:
            continue
        pontualidade = round(no_prazo / com_agenda * 100) if com_agenda else None
        media_min = round(minutos / com_horas) if com_horas else None
        tecnicos.append({
            "id": uid,
            "nome": u.get_full_name() or u.username,
            "username": u.username,
            "execucoes": execucoes,
            "execucoes_auxiliar": execucoes_auxiliar,
            "com_agenda": com_agenda,
            "no_prazo": no_prazo,
            "atribuidas": atribuidas,
            "atrasadas": atrasadas,
            "pendentes": max(0, atribuidas - atrasadas),
            "pontualidade": pontualidade,
            "minutos": minutos,
            "minutos_auxiliar": minutos_auxiliar,
            "com_horas": com_horas,
            "horas_dec": round(minutos / 60, 1),
            "tempo_fmt": _fmt_minutos(minutos),
            "media_fmt": _fmt_minutos(media_min) if media_min is not None else None,
        })

    tecnicos.sort(key=lambda t: (-t["execucoes"], -t["atribuidas"], t["nome"]))

    tot_exec = sum(t["execucoes"] for t in tecnicos)
    tot_com_agenda = sum(t["com_agenda"] for t in tecnicos)
    tot_no_prazo = sum(t["no_prazo"] for t in tecnicos)
    tot_atribuidas = sum(t["atribuidas"] for t in tecnicos)
    tot_atrasadas = sum(t["atrasadas"] for t in tecnicos)
    tot_minutos = sum(t["minutos"] for t in tecnicos)
    tot_com_horas = sum(t["com_horas"] for t in tecnicos)
    pont_global = round(tot_no_prazo / tot_com_agenda * 100) if tot_com_agenda else None

    mensal = {
        (r["m"].year, r["m"].month): int(r["c"] or 0)
        for r in execs.annotate(m=TruncMonth("data_execucao")).values("m").annotate(c=Count("id"))
        if r["m"]
    }
    mes_labels, mes_exec = [], []
    cursor = date(inicio.year, inicio.month, 1)
    limite = date(fim.year, fim.month, 1)
    guard = 0
    while cursor <= limite and guard < 60:
        mes_labels.append(f"{_MESES_PT_PREV[cursor.month]}/{str(cursor.year)[2:]}")
        mes_exec.append(mensal.get((cursor.year, cursor.month), 0))
        cursor = date(cursor.year + 1, 1, 1) if cursor.month == 12 else date(cursor.year, cursor.month + 1, 1)
        guard += 1

    context = {
        "inicio": inicio,
        "fim": fim,
        "kpi": {
            "execucoes": tot_exec,
            "com_agenda": tot_com_agenda,
            "no_prazo": tot_no_prazo,
            "pontualidade": pont_global,
            "atribuidas": tot_atribuidas,
            "atrasadas": tot_atrasadas,
            "tecnicos": len(tecnicos),
            "minutos": tot_minutos,
            "horas_dec": round(tot_minutos / 60, 1),
            "tempo_fmt": _fmt_minutos(tot_minutos),
            "com_horas": tot_com_horas,
            "media_fmt": _fmt_minutos(round(tot_minutos / tot_com_horas)) if tot_com_horas else None,
        },
        "tecnicos": tecnicos,
        "chart_tec_labels": [t["nome"] for t in tecnicos],
        "chart_tec_exec": [t["execucoes"] for t in tecnicos],
        "chart_tec_pont": [t["pontualidade"] if t["pontualidade"] is not None else 0 for t in tecnicos],
        "chart_tec_atrasadas": [t["atrasadas"] for t in tecnicos],
        "chart_tec_horas": [t["horas_dec"] for t in tecnicos],
        "mes_labels": mes_labels,
        "mes_exec": mes_exec,
    }
    return render(request, "front/preventivas/tecnico_desempenho.html", context)


@login_required
def minhas_atividades(request):
    """
    Atividades (preventivas agendadas) atribuídas a um técnico.
    Padrão: o usuário logado. Staff/superuser pode escolher outro técnico (?tecnico=<id>).
    """
    from django.contrib.auth import get_user_model
    AuthUser = get_user_model()
    hoje = timezone.localdate()

    pode_escolher = request.user.is_staff or request.user.is_superuser
    alvo = request.user
    if pode_escolher:
        tid = (request.GET.get("tecnico") or "").strip()
        if tid.isdigit():
            escolhido = AuthUser.objects.filter(pk=int(tid)).first()
            if escolhido:
                alvo = escolhido

    base = (
        Preventiva.objects
        .filter(tecnico=alvo, data_agendamento__isnull=False, pausada=False)
        .select_related("equipamento", "equipamento__localidade", "equipamento__centro_custo", "checklist_modelo")
        .order_by("data_agendamento", "equipamento__nome")
    )

    atividades = list(base)
    for p in atividades:
        dias = (p.data_agendamento - hoje).days
        p.dias_agendamento = dias
        if dias < 0:
            p.classe_prazo, p.label_prazo = "vencida", "Vencida"
        elif dias == 0:
            p.classe_prazo, p.label_prazo = "hoje", "Hoje"
        elif dias <= 7:
            p.classe_prazo, p.label_prazo = "semana", "Esta semana"
        else:
            p.classe_prazo, p.label_prazo = "futura", "Programada"

    kpi = {
        "total": len(atividades),
        "vencidas": sum(1 for p in atividades if p.dias_agendamento < 0),
        "hoje": sum(1 for p in atividades if p.dias_agendamento == 0),
        "semana": sum(1 for p in atividades if 0 <= p.dias_agendamento <= 7),
    }

    recentes = list(
        PreventivaExecucao.objects
        .filter(Q(tecnico=alvo) | Q(tecnico__isnull=True, criado_por=alvo) | Q(tecnicos_auxiliares=alvo))
        .select_related("preventiva", "preventiva__equipamento")
        .order_by("-data_execucao", "-id")
        .distinct()[:8]
    )
    for e in recentes:
        e.papel_auxiliar = (e.tecnico_id != alvo.id) and not (e.tecnico_id is None and e.criado_por_id == alvo.id)

    context = {
        "today": hoje,
        "alvo": alvo,
        "alvo_nome": alvo.get_full_name() or alvo.username,
        "is_proprio": (alvo == request.user),
        "pode_escolher": pode_escolher,
        "tecnicos": AuthUser.objects.filter(is_active=True).exclude(groups__name=GRUPO_FORNECEDOR).order_by("first_name", "last_name", "username") if pode_escolher else [],
        "atividades": atividades,
        "kpi": kpi,
        "recentes": recentes,
    }
    return render(request, "front/preventivas/minhas_atividades.html", context)


_DIAS_SEMANA_PT = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
                   "Sexta-feira", "Sábado", "Domingo"]


@login_required
def preventiva_minha_agenda(request):
    """
    Agenda pessoal de execuções do usuário logado, agrupada por dia/hora.
    Estritamente pessoal: sempre request.user, sem seleção de outro técnico.
    """
    hoje = timezone.localdate()
    fim = _parse_data_opt(request.GET.get("fim"), hoje)
    inicio = _parse_data_opt(request.GET.get("inicio"), fim - timedelta(days=30))
    if inicio > fim:
        inicio, fim = fim, inicio

    execs = list(
        PreventivaExecucao.objects
        .filter(
            Q(tecnico=request.user)
            | Q(tecnico__isnull=True, criado_por=request.user)
            | Q(tecnicos_auxiliares=request.user)
        )
        .filter(data_execucao__gte=inicio, data_execucao__lte=fim)
        .select_related(
            "preventiva", "preventiva__equipamento",
            "preventiva__equipamento__localidade", "preventiva__checklist_modelo",
        )
        .order_by("-data_execucao", "hora_inicio", "id")
        .distinct()
    )
    for execucao in execs:
        execucao.papel_auxiliar = (execucao.tecnico_id != request.user.id) and not (
            execucao.tecnico_id is None and execucao.criado_por_id == request.user.id
        )

    por_dia = defaultdict(list)
    for execucao in execs:
        por_dia[execucao.data_execucao].append(execucao)

    dias = []
    for data_dia in sorted(por_dia.keys(), reverse=True):
        itens = por_dia[data_dia]
        minutos = sum(e.duracao_minutos or 0 for e in itens)
        dias.append({
            "data": data_dia,
            "dia_semana": _DIAS_SEMANA_PT[data_dia.weekday()],
            "execucoes": itens,
            "total_execucoes": len(itens),
            "total_fmt": _fmt_minutos(minutos),
        })

    total_minutos = sum(e.duracao_minutos or 0 for e in execs)
    com_agenda = sum(1 for e in execs if e.data_agendada)
    no_prazo = sum(1 for e in execs if e.data_agendada and e.no_prazo)

    context = {
        "inicio": inicio,
        "fim": fim,
        "hoje": hoje,
        "dias": dias,
        "kpi": {
            "total_execucoes": len(execs),
            "dias_com_atividade": len(dias),
            "tempo_total_fmt": _fmt_minutos(total_minutos),
            "pontualidade": round(no_prazo / com_agenda * 100) if com_agenda else None,
        },
    }
    return render(request, "front/preventivas/preventiva_minha_agenda.html", context)


@login_required
def apontamentos_horas_export(request):
    """
    Exporta (Excel) os apontamentos de horas das execuções de preventiva,
    consolidados por técnico e detalhados por execução, no período filtrado.
    Aceita ?inicio=YYYY-MM-DD&fim=YYYY-MM-DD&tecnico=<id>.
    """
    from io import BytesIO

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    hoje = timezone.localdate()
    inicio = _parse_data_opt(request.GET.get("inicio"), date(hoje.year, 1, 1))
    fim = _parse_data_opt(request.GET.get("fim"), hoje)
    if inicio > fim:
        inicio, fim = fim, inicio

    tecnico_id = (request.GET.get("tecnico") or "").strip()

    execs = (
        PreventivaExecucao.objects
        .filter(data_execucao__gte=inicio, data_execucao__lte=fim)
        .select_related(
            "preventiva",
            "preventiva__equipamento",
            "preventiva__equipamento__localidade",
            "preventiva__checklist_modelo",
            "tecnico",
            "criado_por",
        )
        .prefetch_related("tecnicos_auxiliares")
        .order_by("data_execucao", "id")
    )

    if tecnico_id.isdigit():
        _tid = int(tecnico_id)
        execs = execs.filter(
            Q(tecnico_id=_tid) | Q(tecnico__isnull=True, criado_por_id=_tid) | Q(tecnicos_auxiliares=_tid)
        ).distinct()

    def _nome_tecnico(ex):
        u = ex.tecnico or ex.criado_por
        if not u:
            return "—"
        return u.get_full_name() or u.username

    def _nomes_auxiliares(ex):
        return ", ".join(a.get_full_name() or a.username for a in ex.tecnicos_auxiliares.all())

    # ── Estilos ──────────────────────────────────────────────────────────────
    BRAND = "0071E3"
    DARK = "0A2540"
    GREEN = "047857"
    GRAY = "8A8A8E"
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=9, color="FFFFFF")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Calibri", size=10, color="1D1D1F")
    f_bold = Font(name="Calibri", size=10, bold=True, color="1D1D1F")
    fill_title = PatternFill("solid", fgColor=DARK)
    fill_sub = PatternFill("solid", fgColor="EEF2F7")
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor="F7F9FC")
    a_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    a_center = Alignment(horizontal="center", vertical="center")
    a_right = Alignment(horizontal="right", vertical="center")

    def faixa_titulo(ws, ncols, titulo, subtitulo):
        last = get_column_letter(ncols)
        ws.merge_cells(f"A1:{last}1")
        c1 = ws["A1"]
        c1.value = titulo
        c1.font = f_title
        c1.fill = fill_title
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A2:{last}2")
        c2 = ws["A2"]
        c2.value = subtitulo
        c2.font = Font(name="Calibri", size=9, color="334155")
        c2.fill = fill_sub
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18
        ws.sheet_view.showGridLines = False

    def cabecalho(ws, row, headers, center_cols=()):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = fill_header
            c.font = f_header
            c.border = border
            c.alignment = a_center if ci in center_cols else a_left
        ws.row_dimensions[row].height = 22

    periodo_txt = f"{inicio:%d/%m/%Y} a {fim:%d/%m/%Y}"
    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")

    # ── Agrega por técnico em Python (inclui participações como auxiliar,
    # já que as horas trabalhadas em dupla contam para os dois técnicos) ─────
    resumo = {}
    detalhe = list(execs)
    for ex in detalhe:
        nome = _nome_tecnico(ex)
        r = resumo.setdefault(nome, {"execucoes": 0, "com_horas": 0, "minutos": 0})
        r["execucoes"] += 1
        if ex.duracao_minutos:
            r["com_horas"] += 1
            r["minutos"] += int(ex.duracao_minutos)

        for auxiliar in ex.tecnicos_auxiliares.all():
            nome_aux = auxiliar.get_full_name() or auxiliar.username
            r_aux = resumo.setdefault(nome_aux, {"execucoes": 0, "com_horas": 0, "minutos": 0})
            r_aux["execucoes"] += 1
            if ex.duracao_minutos:
                r_aux["com_horas"] += 1
                r_aux["minutos"] += int(ex.duracao_minutos)

    wb = Workbook()

    # =========================================================================
    # ABA 1 — RESUMO POR TÉCNICO
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Resumo por Tecnico"
    headers1 = ["Técnico", "Execuções", "Com apontamento", "Tempo total", "Horas (decimal)", "Média / execução"]
    faixa_titulo(ws1, len(headers1),
                 "APONTAMENTO DE HORAS — PREVENTIVAS",
                 f"Santa Colomba Agropecuária  ·  Período {periodo_txt}  ·  Gerado em {gerado}")
    hr = 4
    cabecalho(ws1, hr, headers1, center_cols=(2, 3, 4, 5, 6))
    row = hr + 1
    tot_exec = tot_com = tot_min = 0
    for i, (nome, r) in enumerate(sorted(resumo.items(), key=lambda kv: -kv[1]["minutos"])):
        media = round(r["minutos"] / r["com_horas"]) if r["com_horas"] else 0
        valores = [
            nome,
            r["execucoes"],
            r["com_horas"],
            _fmt_minutos(r["minutos"]),
            round(r["minutos"] / 60, 2),
            _fmt_minutos(media) if r["com_horas"] else "—",
        ]
        zebra = (i % 2 == 1)
        for ci, val in enumerate(valores, 1):
            c = ws1.cell(row=row, column=ci, value=val)
            c.border = border
            c.font = f_cell
            c.alignment = a_left if ci == 1 else a_center
            if ci == 5:
                c.number_format = "0.00"
            if zebra:
                c.fill = fill_zebra
        tot_exec += r["execucoes"]
        tot_com += r["com_horas"]
        tot_min += r["minutos"]
        row += 1

    # rodapé total
    if resumo:
        valores = ["TOTAL", tot_exec, tot_com, _fmt_minutos(tot_min), round(tot_min / 60, 2),
                   _fmt_minutos(round(tot_min / tot_com)) if tot_com else "—"]
        for ci, val in enumerate(valores, 1):
            c = ws1.cell(row=row, column=ci, value=val)
            c.font = f_bold
            c.border = border
            c.fill = fill_sub
            c.alignment = a_left if ci == 1 else a_center
            if ci == 5:
                c.number_format = "0.00"
    else:
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers1))
        c = ws1.cell(row=row, column=1, value="Nenhum apontamento no período selecionado.")
        c.font = f_cell
        c.alignment = a_center

    for i, w in enumerate([34, 12, 16, 16, 16, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = f"A{hr + 1}"

    # =========================================================================
    # ABA 2 — APONTAMENTOS DETALHADOS
    # =========================================================================
    ws2 = wb.create_sheet(title="Apontamentos")
    headers2 = [
        "Data", "Técnico", "Equipamento", "Nº Série", "Localidade", "Checklist",
        "Início", "Término", "Duração", "Horas (dec.)", "No prazo", "Observação",
        "Auxiliar(es)",
    ]
    faixa_titulo(ws2, len(headers2),
                 "APONTAMENTOS DETALHADOS",
                 f"{len(detalhe)} execução(ões) no período {periodo_txt}")
    hr = 4
    cabecalho(ws2, hr, headers2, center_cols=(1, 7, 8, 9, 10, 11))
    row = hr + 1
    for i, ex in enumerate(detalhe):
        eq = ex.preventiva.equipamento if ex.preventiva else None
        horas_dec = ex.duracao_horas
        valores = [
            ex.data_execucao.strftime("%d/%m/%Y") if ex.data_execucao else "—",
            _nome_tecnico(ex),
            eq.nome if eq else "—",
            (eq.numero_serie or "—") if eq else "—",
            (eq.localidade.local if eq and eq.localidade else "—"),
            ex.preventiva.checklist_modelo.nome if ex.preventiva and ex.preventiva.checklist_modelo else "—",
            ex.hora_inicio.strftime("%H:%M") if ex.hora_inicio else "—",
            ex.hora_fim.strftime("%H:%M") if ex.hora_fim else "—",
            ex.duracao_formatada or "—",
            horas_dec if horas_dec is not None else "—",
            "Sim" if ex.no_prazo else "Não",
            (ex.observacao or "").strip(),
            _nomes_auxiliares(ex) or "—",
        ]
        zebra = (i % 2 == 1)
        for ci, val in enumerate(valores, 1):
            c = ws2.cell(row=row, column=ci, value=val)
            c.border = border
            c.font = f_cell
            if ci in (1, 7, 8, 9, 11):
                c.alignment = a_center
            elif ci == 10:
                c.alignment = a_center
                if isinstance(val, float):
                    c.number_format = "0.00"
            else:
                c.alignment = a_left
            if ci == 11:
                c.font = Font(name="Calibri", size=10, bold=True,
                              color=GREEN if ex.no_prazo else "B91C1C")
            if zebra:
                c.fill = fill_zebra
        row += 1

    for i, w in enumerate([12, 26, 30, 16, 22, 24, 9, 9, 13, 12, 10, 40, 26], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = f"A{hr + 1}"

    # ── Resposta ─────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arquivo = f"apontamentos_horas_{inicio:%Y%m%d}_{fim:%Y%m%d}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response
