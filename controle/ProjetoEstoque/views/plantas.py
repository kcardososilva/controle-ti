import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

_GRUPO_TV = 'Visualizador TV'


def _is_tv_user(user) -> bool:
    return (
        user.is_authenticated
        and not user.is_staff
        and not user.is_superuser
        and user.groups.filter(name=_GRUPO_TV).exists()
    )

from django.core.cache import cache

from ..forms import PlantaProjetoForm
from ..models import Item, Localidade, PlantaProjeto, PlantaLayoutHistorico
from services import prtg_service

_KPIS_CACHE_KEY = 'planta_kpis_globais'


# ── Lista de Plantas ──────────────────────────────────────────────────────────

@login_required
def planta_list(request):
    localidade_id = request.GET.get("localidade")
    qs = PlantaProjeto.objects.select_related("localidade", "criado_por").order_by(
        "localidade__local", "nome"
    )
    if localidade_id:
        qs = qs.filter(localidade_id=localidade_id)

    localidades = Localidade.objects.order_by("local")

    # KPIs globais — cacheados por 60s, invalidados ao salvar layout
    kpis = cache.get(_KPIS_CACHE_KEY)
    if kpis is None:
        todas = list(PlantaProjeto.objects.all())
        kpis = {
            "total_plantas":   len(todas),
            "total_elementos": sum(p.total_elementos for p in todas),
            "total_com_prtg":  sum(p.elementos_com_prtg for p in todas),
        }
        cache.set(_KPIS_CACHE_KEY, kpis, 60)
    total_plantas   = kpis["total_plantas"]
    total_elementos = kpis["total_elementos"]
    total_com_prtg  = kpis["total_com_prtg"]
    total_sem_prtg  = total_elementos - total_com_prtg

    # Evaluate queryset once; build PRTG IDs map for client-side status display
    plantas_list = list(qs)
    prtg_ids_map = {}
    previews_map = {}
    _tipos_ocultos = {"texto", "forma"}
    for p in plantas_list:
        elements = p.layout.get("elements", []) or []
        connections = p.layout.get("connections", []) or []
        prtg_ids_map[str(p.pk)] = [
            str(e["prtg_objid"]) for e in elements if e.get("prtg_objid")
        ]
        # Mini-mapa: só nós de dispositivos posicionados (exclui textos/formas)
        nodes = [
            {
                "id": str(e.get("id", "")),
                "x":  e.get("x", 0),
                "y":  e.get("y", 0),
                "o":  str(e["prtg_objid"]) if e.get("prtg_objid") else None,
            }
            for e in elements
            if e.get("type") not in _tipos_ocultos
            and isinstance(e.get("x"), (int, float))
            and isinstance(e.get("y"), (int, float))
        ]
        if nodes:
            _ids = {n["id"] for n in nodes}
            links = [
                [str(c.get("from", "")), str(c.get("to", ""))]
                for c in connections
                if str(c.get("from", "")) in _ids and str(c.get("to", "")) in _ids
            ]
            previews_map[str(p.pk)] = {"n": nodes, "l": links}

    return render(request, "front/plantas/planta_list.html", {
        "plantas":         plantas_list,
        "localidades":     localidades,
        "localidade_sel":  localidade_id,
        "total_plantas":   total_plantas,
        "total_elementos": total_elementos,
        "total_com_prtg":  total_com_prtg,
        "total_sem_prtg":  total_sem_prtg,
        "prtg_ok":         prtg_service.is_configured(),
        "prtg_ids_json":   prtg_ids_map,
        "previews_json":   previews_map,
    })


# ── Criar Planta ──────────────────────────────────────────────────────────────

@login_required
def planta_create(request):
    if request.method == "POST":
        form = PlantaProjetoForm(request.POST, request.FILES)
        if form.is_valid():
            planta = form.save(commit=False)
            planta.criado_por = request.user
            planta.atualizado_por = request.user
            planta.layout = {"elements": [], "connections": [], "canvas": {"width": 1400, "height": 900}}
            planta.save()
            messages.success(request, f'Planta "{planta.nome}" criada com sucesso!')
            return redirect("planta_editor", pk=planta.pk)
    else:
        form = PlantaProjetoForm()
    return render(request, "front/plantas/planta_form.html", {
        "form":  form,
        "titulo": "Nova Planta",
        "modo":  "criar",
    })


# ── Editar Metadados da Planta ────────────────────────────────────────────────

@login_required
def planta_update(request, pk):
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    if request.method == "POST":
        form = PlantaProjetoForm(request.POST, request.FILES, instance=planta)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.atualizado_por = request.user
            obj.save()
            messages.success(request, f'Planta "{planta.nome}" atualizada com sucesso!')
            return redirect("planta_viewer", pk=planta.pk)
    else:
        form = PlantaProjetoForm(instance=planta)
    return render(request, "front/plantas/planta_form.html", {
        "form":   form,
        "planta": planta,
        "titulo": f"Editar — {planta.nome}",
        "modo":   "editar",
    })


# ── Excluir Planta ────────────────────────────────────────────────────────────

@login_required
def planta_delete(request, pk):
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    if request.method == "POST":
        nome = planta.nome
        planta.delete()
        messages.success(request, f'Planta "{nome}" excluída.')
        return redirect("planta_list")
    return render(request, "front/plantas/planta_confirm_delete.html", {"planta": planta})


# ── Editor Canvas ─────────────────────────────────────────────────────────────

_TIPOS_ELEMENTOS = [
    ("camera",       "Câmera",       "#ef4444", "camera"),
    ("access_point", "Access Point", "#0ea5e9", "wifi"),
    ("switch",       "Switch",       "#2563eb", "network-wired"),
    ("meraki",       "Meraki",       "#67b346", "cloud"),
    ("starlink",     "Starlink",     "#4361ee", "satellite-dish"),
    ("caixa_emenda", "Caixa de Emenda", "#0891b2", "circle-nodes"),
    ("rack",         "Rack",         "#475569", "server"),
    ("desktop",      "Desktop",      "#10b981", "desktop"),
    ("impressora",   "Impressora",   "#f97316", "print"),
    ("nobreak",      "Nobreak",      "#f59e0b", "bolt"),
    ("fonte",        "Fonte",        "#f59e0b", "plug"),
    ("servidor",     "Servidor",     "#8b5cf6", "cloud"),
    ("ponto_rede",   "Ponto de Rede","#6366f1", "circle-dot"),
    ("texto",        "Texto",        "#1d1d1f", "font"),
    ("quadro",       "Quadro/Área",  "#2563eb", "square"),
    ("circulo",      "Círculo/Zona", "#8b5cf6", "circle"),
]

@login_required
def planta_editor(request, pk):
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    return render(request, "front/plantas/planta_editor.html", {
        "planta":          planta,
        "layout_json":     planta.layout,
        "prtg_ok":         prtg_service.is_configured(),
        "tipos_elementos": _TIPOS_ELEMENTOS,
    })


# ── Visualizador com Status PRTG ──────────────────────────────────────────────

@login_required
def planta_viewer(request, pk):
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    return render(request, "front/plantas/planta_viewer.html", {
        "planta":      planta,
        "layout_json": planta.layout,
        "prtg_ok":     prtg_service.is_configured(),
    })


# ── API: Salvar Layout (AJAX POST) ────────────────────────────────────────────

_ALLOWED_EL_KEYS = {
    "id", "type", "label", "x", "y", "x2", "y2", "color",
    "width", "height", "strokeWidth", "dash", "arrowEnd", "arrowStart",
    "item_id", "prtg_objid", "ip", "observacoes", "fontSize",
    "fontFamily", "fontBold", "fontItalic",
    "fillOpacity", "borderWidth", "borderColor", "borderStyle", "cornerRadius", "fillType",
    "shapeKind", "zIndex", "locked", "rotation", "groupId",
}
_ALLOWED_CN_KEYS = {
    "id", "from", "fromEdge", "to", "toEdge", "type", "label",
    "cpx", "cpy", "cp1x", "cp1y", "cp2x", "cp2y", "waypoints",
    "color", "strokeWidth", "dash", "arrow",
}
_MAX_HISTORICO = 20


@login_required
def planta_salvar_layout(request, pk):
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método não permitido"}, status=405)
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    try:
        payload = json.loads(request.body)
        if not isinstance(payload.get("elements"), list):
            raise ValueError("Campo 'elements' ausente ou inválido.")
        if not isinstance(payload.get("connections"), list):
            raise ValueError("Campo 'connections' ausente ou inválido.")

        # Detecção de conflito: cliente envia a versão que conhecia
        client_version = payload.get("client_version")
        if client_version is not None:
            try:
                client_version = int(client_version)
            except (TypeError, ValueError):
                client_version = None

        if client_version is not None and client_version != planta.layout_version:
            editado_por = ""
            if planta.atualizado_por:
                editado_por = (
                    planta.atualizado_por.get_full_name()
                    or planta.atualizado_por.username
                )
            return JsonResponse({
                "ok": False,
                "conflito": True,
                "versao_atual": planta.layout_version,
                "editado_por": editado_por or "outro usuário",
            })

        # Sanitização
        clean_elements = []
        for el in payload["elements"]:
            clean_el = {k: v for k, v in el.items() if k in _ALLOWED_EL_KEYS}
            clean_el["label"] = str(clean_el.get("label", ""))[:120]
            clean_el["observacoes"] = str(clean_el.get("observacoes", ""))[:500]
            clean_elements.append(clean_el)
        clean_connections = [
            {k: v for k, v in cn.items() if k in _ALLOWED_CN_KEYS}
            for cn in payload["connections"]
        ]
        novo_layout = {
            "elements":    clean_elements,
            "connections": clean_connections,
            "canvas":      payload.get("canvas", {"width": 1400, "height": 900}),
        }

        # Criar entrada no histórico se solicitado explicitamente
        if payload.get("nova_versao"):
            descricao = str(payload.get("descricao", ""))[:200]
            PlantaLayoutHistorico.objects.create(
                planta=planta,
                versao=planta.layout_version,
                layout=planta.layout,
                salvo_por=request.user,
                descricao=descricao,
            )
            # Manter apenas os últimos N
            ids_antigos = (
                PlantaLayoutHistorico.objects
                .filter(planta=planta)
                .order_by("-versao")
                .values_list("id", flat=True)[_MAX_HISTORICO:]
            )
            if ids_antigos:
                PlantaLayoutHistorico.objects.filter(id__in=list(ids_antigos)).delete()

        nova_versao = planta.layout_version + 1
        planta.layout = novo_layout
        planta.layout_version = nova_versao
        planta.atualizado_por = request.user
        planta.save(update_fields=["layout", "layout_version", "atualizado_por", "updated_at"])
        cache.delete(_KPIS_CACHE_KEY)
        return JsonResponse({"ok": True, "versao": nova_versao})
    except (json.JSONDecodeError, ValueError, KeyError) as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=400)


# ── API: Verificar versão atual (polling de conflito) ─────────────────────────

@login_required
def planta_check_version(request, pk):
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    editado_por = ""
    if planta.atualizado_por:
        editado_por = (
            planta.atualizado_por.get_full_name()
            or planta.atualizado_por.username
        )
    return JsonResponse({"versao": planta.layout_version, "editado_por": editado_por})


# ── API: Listar histórico de versões ─────────────────────────────────────────

@login_required
def planta_historico_api(request, pk):
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    entries = (
        PlantaLayoutHistorico.objects
        .filter(planta=planta)
        .select_related("salvo_por")
        .order_by("-versao")[:_MAX_HISTORICO]
    )
    data = []
    for e in entries:
        salvo_por_nome = "—"
        if e.salvo_por:
            salvo_por_nome = e.salvo_por.get_full_name() or e.salvo_por.username
        data.append({
            "id":          e.id,
            "versao":      e.versao,
            "salvo_por":   salvo_por_nome,
            "salvo_em":    e.salvo_em.strftime("%d/%m/%Y %H:%M"),
            "descricao":   e.descricao,
            "n_elementos": len(e.layout.get("elements", [])),
            "n_conexoes":  len(e.layout.get("connections", [])),
        })
    return JsonResponse({"ok": True, "historico": data})


# ── API: Restaurar versão do histórico ────────────────────────────────────────

@login_required
def planta_restaurar_versao(request, pk, hist_pk):
    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    entrada = get_object_or_404(PlantaLayoutHistorico, pk=hist_pk, planta=planta)

    # Salva o estado atual no histórico antes de restaurar
    PlantaLayoutHistorico.objects.get_or_create(
        planta=planta,
        versao=planta.layout_version,
        defaults={
            "layout":    planta.layout,
            "salvo_por": request.user,
            "descricao": f"Backup antes de restaurar v{entrada.versao}",
        },
    )
    nova_versao = planta.layout_version + 1
    planta.layout = entrada.layout
    planta.layout_version = nova_versao
    planta.atualizado_por = request.user
    planta.save(update_fields=["layout", "layout_version", "atualizado_por", "updated_at"])
    cache.delete(_KPIS_CACHE_KEY)
    return JsonResponse({"ok": True, "versao": nova_versao})


# ── Gerenciamento de Acesso TV (somente staff/superusuários) ─────────────────

@login_required
def planta_tv_gerenciar(request):
    """Página de gerenciamento de acesso ao modo TV — restrito a staff/superusuários."""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "Acesso restrito a administradores.")
        return redirect("planta_list")

    from django.contrib.auth.models import Group
    from django.contrib.auth import get_user_model
    AuthUser = get_user_model()

    grupo_tv, _ = Group.objects.get_or_create(name=_GRUPO_TV)
    usuarios_tv = (
        grupo_tv.user_set.all()
        .prefetch_related("plantas_tv_autorizadas__localidade")
        .order_by("first_name", "last_name", "username")
    )
    usuarios_disponiveis = (
        AuthUser.objects.filter(is_active=True)
        .exclude(pk__in=grupo_tv.user_set.values("pk"))
        .order_by("first_name", "last_name", "username")
    )
    todas_plantas = (
        PlantaProjeto.objects.select_related("localidade")
        .order_by("localidade__local", "nome")
    )
    return render(request, "front/plantas/planta_tv_gerenciar.html", {
        "usuarios_tv":          list(usuarios_tv),
        "usuarios_disponiveis": list(usuarios_disponiveis),
        "todas_plantas":        list(todas_plantas),
    })


@login_required
def planta_tv_gerenciar_acao(request):
    """Endpoint JSON para adicionar/remover usuários do grupo TV e salvar plantas autorizadas."""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "erro": "Acesso negado."}, status=403)
    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)

    from django.contrib.auth.models import Group
    from django.contrib.auth import get_user_model
    AuthUser = get_user_model()

    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"ok": False, "erro": "JSON inválido."}, status=400)

    action  = payload.get("action")
    user_id = payload.get("user_id")
    grupo_tv, _ = Group.objects.get_or_create(name=_GRUPO_TV)

    try:
        target = AuthUser.objects.get(pk=user_id)
    except AuthUser.DoesNotExist:
        return JsonResponse({"ok": False, "erro": "Usuário não encontrado."}, status=404)

    if action == "adicionar_grupo":
        target.groups.add(grupo_tv)
        return JsonResponse({
            "ok": True,
            "user_id": target.pk,
            "nome": target.get_full_name() or target.username,
            "username": target.username,
            "email": target.email,
        })

    if action == "remover_grupo":
        target.groups.remove(grupo_tv)
        target.plantas_tv_autorizadas.clear()
        return JsonResponse({"ok": True})

    if action == "salvar_plantas":
        raw_ids = payload.get("plant_ids", [])
        plant_ids = [int(x) for x in raw_ids if str(x).isdigit()]
        plantas = PlantaProjeto.objects.filter(pk__in=plant_ids)
        target.plantas_tv_autorizadas.set(plantas)
        return JsonResponse({"ok": True, "total": plantas.count()})

    return JsonResponse({"ok": False, "erro": "Ação desconhecida."}, status=400)


# ── Seletor TV (acessível ao grupo Visualizador TV) ──────────────────────────

@login_required
def planta_tv_lista(request):
    """Página standalone com a lista de plantas disponíveis no modo TV.
    Usuários do grupo 'Visualizador TV' veem apenas as plantas autorizadas para eles.
    Demais usuários autenticados veem todas."""
    user = request.user
    qs = PlantaProjeto.objects.select_related("localidade").order_by("localidade__local", "nome")
    if _is_tv_user(user):
        qs = qs.filter(visualizadores_tv=user)

    plantas_list = list(qs)
    prtg_ids_map = {
        str(p.pk): [str(e["prtg_objid"]) for e in p.layout.get("elements", []) if e.get("prtg_objid")]
        for p in plantas_list
    }

    return render(request, "front/plantas/planta_tv_lista.html", {
        "plantas":       plantas_list,
        "prtg_ok":       prtg_service.is_configured(),
        "prtg_ids_json": prtg_ids_map,
        "status_url":    reverse("prtg_status_api"),
    })


# ── Modo TV ─────────────────────────────────────────────────────────────────

@login_required
def planta_tv(request, pk):
    planta = get_object_or_404(PlantaProjeto, pk=pk)
    # Usuários TV só podem ver plantas autorizadas para eles
    if _is_tv_user(request.user) and not planta.visualizadores_tv.filter(pk=request.user.pk).exists():
        return redirect("planta_tv_lista")
    interval = max(10, min(300, int(request.GET.get("interval", 30))))
    return render(request, "front/plantas/planta_tv.html", {
        "planta":      planta,
        "layout_json": planta.layout,
        "interval":    interval,
        "prtg_ok":     prtg_service.is_configured(),
    })


# ── API: Status PRTG (proxy seguro) ──────────────────────────────────────────

@login_required
def prtg_status_api(request):
    """
    Proxy server-side para PRTG.
    Credenciais NUNCA chegam ao browser — o cliente recebe apenas status filtrado.
    Resultado cacheado 30s no servidor via FileBasedCache.
    """
    if not prtg_service.is_configured():
        return JsonResponse({"ok": False, "erro": "PRTG não configurado.", "devices_map": {}})
    devices_map = prtg_service.get_devices_map()
    return JsonResponse({"ok": True, "devices_map": devices_map})


# ── API: Buscar Devices PRTG (autocomplete editor) ───────────────────────────

@login_required
def prtg_search_api(request):
    q = request.GET.get("q", "").strip()
    if len(q) < 2:
        return JsonResponse({"ok": True, "results": []})
    results = prtg_service.search_devices(q)
    return JsonResponse({"ok": True, "results": results})


# ── Monitor PRTG — cards com status em tempo real ────────────────────────────

@login_required
def prtg_monitor(request):
    """Página de monitoramento: todos os dispositivos PRTG em cards com auto-refresh."""
    import re as _re
    from collections import defaultdict

    if not prtg_service.is_configured():
        return render(request, 'front/plantas/prtg_monitor.html', {'prtg_ok': False})

    devices_map = prtg_service.get_devices_map()
    devices = list(devices_map.values())

    # ── Correlacionar dispositivos PRTG com itens do estoque por nome ─────────
    def _norm(s):
        return _re.sub(r'[^a-z0-9]', '', (s or '').lower())

    all_items = list(
        Item.objects.select_related('localidade', 'subtipo')
        .filter(status='ativo')
        .values('id', 'nome', 'numero_serie', 'marca', 'modelo',
                'localidade__local', 'subtipo__nome')
    )
    items_by_norm = {}
    for it in all_items:
        key = _norm(it['nome'])
        if key and key not in items_by_norm:
            items_by_norm[key] = it

    def _find_item(dev_name):
        dn = _norm(dev_name)
        if not dn:
            return None
        if dn in items_by_norm:
            return items_by_norm[dn]
        for k, it in items_by_norm.items():
            if len(k) >= 4 and (k in dn or dn in k):
                return it
        return None

    for dev in devices:
        dev['item_match'] = _find_item(dev['name'])

    # ── KPIs ──────────────────────────────────────────────────────────────────
    total   = len(devices)
    online  = sum(1 for d in devices if d['status_slug'] == 'up')
    offline = sum(1 for d in devices if d['status_slug'] == 'down')
    warning = sum(1 for d in devices if d['status_slug'] in ('warning', 'unusual'))
    paused  = sum(1 for d in devices if 'paused' in d['status_slug'])
    unknown = total - online - offline - warning - paused

    # ── Agrupar por grupo PRTG ────────────────────────────────────────────────
    groups: dict = defaultdict(list)
    for dev in sorted(devices, key=lambda d: d['name'].lower()):
        groups[dev.get('group') or 'Sem Grupo'].append(dev)
    groups_list = sorted(groups.items(), key=lambda g: g[0].lower())

    return render(request, 'front/plantas/prtg_monitor.html', {
        'devices':     devices,
        'groups_list': groups_list,
        'total':   total,
        'online':  online,
        'offline': offline,
        'warning': warning,
        'paused':  paused,
        'unknown': unknown,
        'prtg_ok': True,
        'status_url': reverse('prtg_status_api'),
    })


# ── Classificação de site a partir do nome do dispositivo PRTG ────────────────
# O grupo PRTG identifica o TIPO (Switch, AP, Servidor…); o SITE fica no nome,
# seguindo a convenção <TIPO>-...-SC-<SITE>-... (ex.: SW-A-SC-KTL-ADM).
def _prtg_site(name: str) -> str:
    import re as _re
    u = (name or "").upper()
    toks = set(_re.split(r"[-_ .]+", u))
    if "RDM" in toks or "RIO DO MEIO" in u or "RIODOMEIO" in u:
        return "RDM"
    if "MAMBAI" in u or "MAMBAÍ" in u:
        return "Mambai"
    if toks & {"PIN", "PNR", "PINHEIROS"} or "PINHEIRO" in u:
        return "PIN"
    if toks & {"KTL", "KLT"} or "KARITEL" in u or "SCKR" in u:
        return "KTL"
    return "Outros"


def _prtg_status_label(slug: str):
    """Retorna (rótulo, cor_hex) para o status efetivo do dispositivo."""
    s = (slug or "").lower()
    mapa = {
        "up":      ("Online",   "34C759"),
        "down":    ("Offline",  "FF3B30"),
        "warning": ("Instável", "FF9500"),
        "unusual": ("Instável", "FF9500"),
    }
    if s in mapa:
        return mapa[s]
    if s.startswith("paused"):
        return ("Pausado", "8E8E93")
    return ("Desconhecido", "9AA0A6")


# ── Exportação Excel: todos os dispositivos PRTG por site ─────────────────────

@login_required
def prtg_monitor_export(request):
    """
    Exporta TODOS os dispositivos PRTG e seus status em uma planilha Excel
    profissional, seccionada na mesma aba por site: KTL (Karitel), RDM (Rio do
    Meio), PIN (Pinheiros), Mambaí e Outros.
    """
    import re as _re
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if not prtg_service.is_configured():
        return redirect("prtg_monitor")

    devices = list(prtg_service.get_devices_map().values())

    # ── Correlação com o estoque por nome (mesma heurística do monitor) ───────
    def _norm(s):
        return _re.sub(r"[^a-z0-9]", "", (s or "").lower())

    itens = list(
        Item.objects.select_related("localidade")
        .filter(status="ativo")
        .values("nome", "localidade__local")
    )
    itens_por_norm = {}
    for it in itens:
        k = _norm(it["nome"])
        if k and k not in itens_por_norm:
            itens_por_norm[k] = it

    def _find_item(dev_name):
        n = _norm(dev_name)
        if not n:
            return None
        if n in itens_por_norm:
            return itens_por_norm[n]
        for k, it in itens_por_norm.items():
            if len(k) >= 4 and (k in n or n in k):
                return it
        return None

    # ── Classificar dispositivos por site ────────────────────────────────────
    SITES = [
        ("KTL",    "Karitel",     "1D4ED8"),
        ("RDM",    "Rio do Meio", "0D9488"),
        ("PIN",    "Pinheiros",   "7C3AED"),
        ("Mambai", "Mambaí",      "EA580C"),
        ("Outros", "Outros",      "475569"),
    ]
    por_site = {key: [] for key, _, _ in SITES}
    for dev in devices:
        por_site[_prtg_site(dev.get("name"))].append(dev)
    for key in por_site:
        por_site[key].sort(key=lambda d: (d.get("name") or "").lower())

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitor PRTG"
    ws.sheet_view.showGridLines = False

    COLS = ["Dispositivo", "IP / Host", "Tipo (Grupo PRTG)", "Status",
            "Detalhe", "Item Vinculado", "Localidade"]
    WIDTHS = [34, 18, 27, 13, 24, 30, 24]
    NCOL = len(COLS)
    last_col = get_column_letter(NCOL)
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(style="thin", color="D9DEE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    r = 1
    # Título
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    c = ws.cell(row=r, column=1, value="MONITOR PRTG — EQUIPAMENTOS E STATUS")
    c.font = Font(bold=True, color="FFFFFF", size=18, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor="0A2540")
    ws.row_dimensions[r].height = 34
    r += 1

    # Subtítulo
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
    c = ws.cell(row=r, column=1,
                value=f"Santa Colomba Agropecuária  ·  Gerado em {gerado}  ·  {len(devices)} dispositivos monitorados")
    c.font = Font(color="5B6B7F", size=10, italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor="EEF2F7")
    ws.row_dimensions[r].height = 18
    r += 2

    # ── Faixa de KPIs ────────────────────────────────────────────────────────
    total = len(devices)
    online = sum(1 for d in devices if d.get("status_slug") == "up")
    offline = sum(1 for d in devices if d.get("status_slug") == "down")
    instavel = sum(1 for d in devices if d.get("status_slug") in ("warning", "unusual"))
    pausado = sum(1 for d in devices if str(d.get("status_slug") or "").startswith("paused"))
    kpis = [
        ("TOTAL", total, "334155"),
        ("ONLINE", online, "1E8E3E"),
        ("OFFLINE", offline, "D93025"),
        ("INSTÁVEL", instavel, "E8830C"),
        ("PAUSADO", pausado, "5F6B7A"),
    ]
    # distribui 5 KPIs em pares de colunas (cada KPI ocupa ao menos 1 coluna)
    for idx, (lbl, val, color) in enumerate(kpis):
        col = idx + 1
        cl = ws.cell(row=r, column=col, value=lbl)
        cl.font = Font(bold=True, color="FFFFFF", size=9)
        cl.alignment = center
        cl.fill = PatternFill("solid", fgColor=color)
        cv = ws.cell(row=r + 1, column=col, value=val)
        cv.font = Font(bold=True, color=color, size=20)
        cv.alignment = center
        cv.fill = PatternFill("solid", fgColor="F4F6F9")
        cv.border = border
        cl.border = border
    ws.row_dimensions[r].height = 16
    ws.row_dimensions[r + 1].height = 30
    r += 3

    freeze_row = r  # congela título + KPIs

    # ── Seções por site ──────────────────────────────────────────────────────
    for key, nome, color in SITES:
        lista = por_site[key]
        if key == "Outros" and not lista:
            continue

        on = sum(1 for d in lista if d.get("status_slug") == "up")
        off = sum(1 for d in lista if d.get("status_slug") == "down")

        # Cabeçalho da seção
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        cs = ws.cell(row=r, column=1,
                     value=f"  {nome.upper()}  —  {len(lista)} dispositivo(s)   •   {on} online   •   {off} offline")
        cs.font = Font(bold=True, color="FFFFFF", size=12)
        cs.alignment = Alignment(horizontal="left", vertical="center")
        cs.fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[r].height = 26
        r += 1

        # Cabeçalho da tabela
        for ci, titulo in enumerate(COLS, start=1):
            hc = ws.cell(row=r, column=ci, value=titulo)
            hc.font = Font(bold=True, color="FFFFFF", size=10)
            hc.alignment = center if titulo == "Status" else Alignment(horizontal="left", vertical="center")
            hc.fill = PatternFill("solid", fgColor="334155")
            hc.border = border
        ws.row_dimensions[r].height = 20
        r += 1

        if not lista:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
            ec = ws.cell(row=r, column=1, value="Nenhum dispositivo classificado para este site.")
            ec.font = Font(color="9AA0A6", italic=True)
            ec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ec.border = border
            r += 1
            r += 1  # espaçamento entre seções
            continue

        for i, dev in enumerate(lista):
            it = _find_item(dev.get("name"))
            label, scolor = _prtg_status_label(dev.get("status_slug"))
            zebra = "FFFFFF" if i % 2 == 0 else "F6F8FB"
            valores = [
                dev.get("name") or "—",
                dev.get("host") or "—",
                dev.get("group") or "—",
                label,
                dev.get("statustext") or "—",
                (it["nome"] if it else "—"),
                (it["localidade__local"] if it and it.get("localidade__local") else "—"),
            ]
            for ci, val in enumerate(valores, start=1):
                cc = ws.cell(row=r, column=ci, value=val)
                cc.border = border
                if ci == 4:  # Status — célula colorida
                    cc.fill = PatternFill("solid", fgColor=scolor)
                    cc.font = Font(bold=True, color="FFFFFF", size=10)
                    cc.alignment = center
                else:
                    cc.fill = PatternFill("solid", fgColor=zebra)
                    cc.font = Font(color="1F2733", size=10)
                    cc.alignment = left
                    if ci == 2:  # IP/Host monoespaçado
                        cc.font = Font(color="334155", size=10, name="Consolas")
            ws.row_dimensions[r].height = 18
            r += 1

        r += 1  # linha em branco entre seções

    ws.freeze_panes = f"A{freeze_row}"

    # ── Resposta ─────────────────────────────────────────────────────────────
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome_arq = f"monitor_prtg_{timezone.localtime():%Y%m%d_%H%M}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{nome_arq}"'
    return resp


# ── API: Buscar Items do sistema (autocomplete editor) ────────────────────────

@login_required
def item_search_api(request):
    q = request.GET.get("q", "").strip()
    if len(q) < 2:
        return JsonResponse({"ok": True, "results": []})
    qs = Item.objects.select_related("localidade", "subtipo")
    if q.startswith("id:"):
        try:
            item_id = int(q[3:])
            qs = qs.filter(pk=item_id)
        except ValueError:
            return JsonResponse({"ok": True, "results": []})
    else:
        qs = qs.filter(nome__icontains=q, status="ativo")
    itens = qs.values("id", "nome", "numero_serie", "marca", "modelo",
                      "status", "localidade__local", "subtipo__nome")[:20]
    results = [
        {
            "id":         i["id"],
            "nome":       i["nome"],
            "serie":      i["numero_serie"] or "",
            "marca":      i["marca"] or "",
            "modelo":     i["modelo"] or "",
            "localidade": i["localidade__local"] or "",
            "subtipo":    i["subtipo__nome"] or "",
        }
        for i in itens
    ]
    return JsonResponse({"ok": True, "results": results})
