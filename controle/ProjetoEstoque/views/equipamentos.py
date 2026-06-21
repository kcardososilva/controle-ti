from decimal import Decimal
from datetime import date, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_POST
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count, Sum, F, Prefetch, Exists
from django.db.models import OuterRef, Subquery
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.core.exceptions import FieldDoesNotExist, ValidationError
from django.db import transaction, IntegrityError
from django.template.loader import render_to_string
from django.utils import timezone

from ..models import (
    Item, Subtipo, Categoria, Localidade, CentroCusto, Fornecedor,
    Locacao, SimNaoChoices, StatusItemChoices, ItemLote,
    MovimentacaoItem, TipoMovimentacaoChoices, Preventiva, PlantaProjeto,
    ItemStatusHistorico, ItemPRTGHistorico,
)
from ..forms import ItemForm, LocacaoForm, LoteEstoqueCreateForm
from services.importador_planilha import ImportadorPlanilhaService
from services.item_create_service import ItemCreateService

def adicionar_erros_validacao_no_form(form, erro):
    if hasattr(erro, "message_dict"):
        for campo, mensagens in erro.message_dict.items():
            campo_form = campo if campo in form.fields else None

            for mensagem in mensagens:
                form.add_error(campo_form, mensagem)
    else:
        form.add_error(None, erro)


@login_required
def item_create(request):
    if request.method == "POST":
        form = ItemForm(request.POST)
        locacao_form = LocacaoForm(request.POST)
        lote_form = LoteEstoqueCreateForm(request.POST)

        form_valido = form.is_valid()

        eh_locado = False
        eh_consumo = False

        if form_valido:
            eh_locado = form.cleaned_data.get("locado") == SimNaoChoices.SIM
            eh_consumo = form.cleaned_data.get("item_consumo") == SimNaoChoices.SIM

        locacao_valida = locacao_form.is_valid() if eh_locado else True
        lote_valido = lote_form.is_valid() if eh_consumo else True

        if form_valido and locacao_valida and lote_valido:
            try:
                ItemCreateService.criar_item(
                    item_form=form,
                    locacao_form=locacao_form,
                    lote_form=lote_form,
                    user=request.user,
                )

                messages.success(request, "Item cadastrado com sucesso.")
                return redirect("equipamentos_list")

            except ValidationError as e:
                adicionar_erros_validacao_no_form(form, e)
                messages.error(request, "Erro de validação. Verifique os dados informados.")

            except IntegrityError as e:
                messages.error(
                    request,
                    f"Erro de integridade ao cadastrar o item. Nenhum dado foi gravado. Detalhe: {str(e)}"
                )

            except Exception as e:
                messages.error(
                    request,
                    f"Erro inesperado ao cadastrar o item: {str(e)}"
                )
        else:
            messages.error(request, "Verifique os campos obrigatórios antes de salvar.")

    else:
        form = ItemForm()
        locacao_form = LocacaoForm()
        lote_form = LoteEstoqueCreateForm()

    return render(request, "front/equipamentos/cadastrar_equipamento.html", {
        "form": form,
        "locacao_form": locacao_form,
        "lote_form": lote_form,
        "editar": False,
    })

# ITEM LIST
from django.core.exceptions import FieldDoesNotExist
def _model_has_field(model, field_name):
    try:
        model._meta.get_field(field_name)
        return True
    except FieldDoesNotExist:
        return False


def _safe_int(value, default=20):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_decimal(value):
    if value is None:
        return Decimal("0.00")

    try:
        return Decimal(value)
    except Exception:
        return Decimal("0.00")


def _tem_valor(value):
    if value is None:
        return False

    if isinstance(value, str) and not value.strip():
        return False

    return True


def _fk_label(obj, *attrs):
    if not obj:
        return None

    for attr in attrs:
        value = getattr(obj, attr, None)

        if _tem_valor(value):
            return value

    return str(obj)


def _build_related_q(base_model, fk_field, related_fields, search_value):
    """
    Monta Q seguro para ForeignKey, evitando FieldError quando o model relacionado
    não possui algum campo sugerido.
    """
    if not search_value:
        return Q()

    try:
        related_model = base_model._meta.get_field(fk_field).remote_field.model
    except Exception:
        return Q()

    query = Q()

    for field in related_fields:
        if _model_has_field(related_model, field):
            query |= Q(**{f"{fk_field}__{field}__icontains": search_value})

    return query


def _build_nested_q(related_model, prefix, related_fields, search_value):
    """
    Monta Q seguro para relações aninhadas.
    Exemplo: vinculos_lote__lote__fornecedor__nome
    """
    if not search_value or not related_model:
        return Q()

    query = Q()

    for field in related_fields:
        if _model_has_field(related_model, field):
            query |= Q(**{f"{prefix}__{field}__icontains": search_value})

    return query


def _status_lote(percentual, disponivel):
    if disponivel <= 0:
        return "zerado"

    if percentual <= 20:
        return "critico"

    if percentual <= 50:
        return "atencao"

    return "ok"


def _enriquecer_itens_com_lotes(itens):
    """
    Adiciona atributos calculados para uso no template:
    - lotes_count
    - lote_total_entrada
    - lote_total_disponivel
    - lote_valor_disponivel
    - lote_percentual
    - lote_status
    - tipo_operacional
    """

    for item in itens:
        vinculos = list(getattr(item, "vinculos_lote").all())

        total_entrada = 0
        total_disponivel = 0
        valor_disponivel = Decimal("0.00")

        for vinculo in vinculos:
            qtd_entrada = vinculo.quantidade_entrada or 0
            qtd_disponivel = vinculo.quantidade_disponivel or 0
            custo_unitario = _safe_decimal(vinculo.custo_unitario)

            total_entrada += qtd_entrada
            total_disponivel += qtd_disponivel
            valor_disponivel += Decimal(qtd_disponivel) * custo_unitario

        percentual = 0

        if total_entrada > 0:
            percentual = int((total_disponivel / total_entrada) * 100)

        item.lotes_count = len(vinculos)
        item.lote_total_entrada = total_entrada
        item.lote_total_disponivel = total_disponivel
        item.lote_valor_disponivel = valor_disponivel
        item.lote_percentual = percentual
        item.lote_status = _status_lote(percentual, total_disponivel)

        if getattr(item, "item_consumo", "nao") == "sim":
            item.tipo_operacional = "Consumo"
            item.tipo_operacional_icon = "fa-boxes-stacked"
            item.tipo_operacional_class = "tp-consumo"

        elif getattr(item, "locado", "nao") == "sim":
            item.tipo_operacional = "Locado"
            item.tipo_operacional_icon = "fa-file-contract"
            item.tipo_operacional_class = "tp-locado"

        elif getattr(item, "tem_lote", False) or item.lotes_count > 0:
            item.tipo_operacional = "Com lote"
            item.tipo_operacional_icon = "fa-box"
            item.tipo_operacional_class = "tp-lote"

        else:
            item.tipo_operacional = "Patrimônio"
            item.tipo_operacional_icon = "fa-cube"
            item.tipo_operacional_class = "tp-patrimonio"

        item.localidade_label = _fk_label(item.localidade, "local", "nome", "descricao")
        item.centro_custo_label = _fk_label(item.centro_custo, "departamento", "nome", "descricao")
        item.centro_custo_numero = _fk_label(item.centro_custo, "numero", "codigo")
        item.fornecedor_label = _fk_label(item.fornecedor, "nome", "razao_social")
        item.subtipo_label = _fk_label(item.subtipo, "nome", "descricao")

    return itens


def _subtipos_queryset():
    qs = Subtipo.objects.all()

    if _model_has_field(Subtipo, "nome"):
        return qs.order_by("nome")

    return qs.order_by("id")


def _aplicar_filtros_itens(request, qs):
    """Aplica todos os filtros GET ao queryset de Item. Fonte única de verdade — usada pela listagem e pela exportação."""
    nome = request.GET.get("nome", "").strip()
    numero_serie = request.GET.get("numero_serie", "").strip()
    subtipo = request.GET.get("subtipo", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor = request.GET.get("fornecedor", "").strip()
    localidade = request.GET.get("localidade", "").strip()
    centro_custo = request.GET.get("centro_custo", "").strip()
    tipo_item = request.GET.get("tipo_item", "").strip()
    estoque = request.GET.get("estoque", "").strip()

    if nome:
        qs = qs.filter(nome__icontains=nome)
    if numero_serie:
        qs = qs.filter(numero_serie__icontains=numero_serie)
    if subtipo:
        qs = qs.filter(subtipo_id=subtipo)
    if status:
        qs = qs.filter(status=status)

    if fornecedor:
        fornecedor_q = _build_related_q(Item, "fornecedor", ["nome", "razao_social", "fantasia"], fornecedor)
        try:
            fornecedor_model = Item._meta.get_field("fornecedor").remote_field.model
        except Exception:
            fornecedor_model = None
        fornecedor_lote_q = _build_nested_q(
            fornecedor_model, "vinculos_lote__lote__fornecedor",
            ["nome", "razao_social", "fantasia"], fornecedor,
        )
        query = fornecedor_q | fornecedor_lote_q
        if query:
            qs = qs.filter(query).distinct()

    if localidade:
        query = _build_related_q(Item, "localidade", ["local", "nome", "descricao"], localidade)
        if query:
            qs = qs.filter(query).distinct()

    if centro_custo:
        query = _build_related_q(
            Item, "centro_custo",
            ["departamento", "nome", "numero", "codigo", "descricao"], centro_custo,
        )
        if query:
            qs = qs.filter(query).distinct()

    if tipo_item == "consumo":
        qs = qs.filter(item_consumo="sim")
    elif tipo_item == "locado":
        qs = qs.filter(locado="sim")
    elif tipo_item == "com_lote":
        qs = qs.filter(Q(tem_lote=True) | Q(vinculos_lote__isnull=False)).distinct()
    elif tipo_item == "patrimonio":
        qs = qs.exclude(item_consumo="sim").exclude(locado="sim")

    if estoque == "com_saldo":
        qs = qs.filter(quantidade__gt=0)
    elif estoque == "zerado":
        qs = qs.filter(Q(quantidade__lte=0) | Q(quantidade__isnull=True))

    return qs


def _build_queryset_and_context(request):
    qs = (
        Item.objects
        .select_related(
            "subtipo",
            "localidade",
            "centro_custo",
            "fornecedor",
        )
        .prefetch_related(
            Prefetch(
                "vinculos_lote",
                queryset=(
                    ItemLote.objects
                    .select_related("lote", "lote__fornecedor")
                    .order_by("-lote__data_entrada", "-created_at")
                )
            )
        )
        .order_by("-created_at")
    )

    qs = _aplicar_filtros_itens(request, qs)

    filtered_total = qs.count()

    per_page = _safe_int(request.GET.get("pp"), 20)

    if per_page not in [10, 20, 50, 100]:
        per_page = 20

    paginator = Paginator(qs, per_page)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)

    itens = list(page_obj.object_list)
    _enriquecer_itens_com_lotes(itens)

    filtered_ids = list(qs.values_list("id", flat=True))

    vinculos_filtrados = (
        ItemLote.objects
        .filter(item_id__in=filtered_ids)
        .select_related("item", "lote")
    )

    total_lote_entrada = 0
    total_lote_disponivel = 0
    valor_lote_disponivel = Decimal("0.00")

    for vinculo in vinculos_filtrados:
        qtd_entrada = vinculo.quantidade_entrada or 0
        qtd_disponivel = vinculo.quantidade_disponivel or 0
        custo_unitario = _safe_decimal(vinculo.custo_unitario)

        total_lote_entrada += qtd_entrada
        total_lote_disponivel += qtd_disponivel
        valor_lote_disponivel += Decimal(qtd_disponivel) * custo_unitario

    estoque_total = qs.aggregate(total=Sum("quantidade"))["total"] or 0

    kpis = {
        "total": filtered_total,
        "ativos": qs.filter(status="ativo").count(),
        "backup": qs.filter(status="backup").count(),
        "manutencao": qs.filter(status__in=["manutencao", "correcao"]).count(),
        "defeito": qs.filter(status__in=["defeito", "queimado"]).count(),
        "consumo": qs.filter(item_consumo="sim").count(),
        "locados": qs.filter(locado="sim").count(),
        "com_lote": qs.filter(Q(tem_lote=True) | Q(vinculos_lote__isnull=False)).distinct().count(),
        "estoque_total": estoque_total,
        "lote_entrada": total_lote_entrada,
        "lote_disponivel": total_lote_disponivel,
        "valor_lote_disponivel": valor_lote_disponivel,
    }

    status_choices = Item._meta.get_field("status").choices

    context = {
        "itens": itens,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_paginated": page_obj.has_other_pages(),
        "filtered_total": filtered_total,
        "per_page": per_page,
        "subtipos": _subtipos_queryset(),
        "status_choices": status_choices,
        "kpis": kpis,
        "tipo_item": request.GET.get("tipo_item", ""),
        "estoque": request.GET.get("estoque", ""),
    }

    return context


@login_required
def equipamentos_list(request):
    context = _build_queryset_and_context(request)

    is_partial = (
        request.GET.get("partial") == "1"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if is_partial:
        view_mode = request.GET.get("view", "list")
        data = {
            "pagination": render_to_string(
                "front/equipamentos/_pagination.html",
                context,
                request=request,
            ),
            "kpis": render_to_string(
                "front/equipamentos/_kpis.html",
                context,
                request=request,
            ),
            "count": context["filtered_total"],
        }
        if view_mode == "gallery":
            data["gallery"] = render_to_string(
                "front/equipamentos/_gallery.html",
                context,
                request=request,
            )
        else:
            data["tbody"] = render_to_string(
                "front/equipamentos/_tbody.html",
                context,
                request=request,
            )

        return JsonResponse(data)

    return render(request, "front/equipamentos/equipamentos_list.html", context)

## IMPORTAR PLANILHA ITEM
@login_required
def importar_planilha(request):
    if request.method != "POST":
        return JsonResponse({
            "ok": False,
            "mensagem": "Método não permitido."
        }, status=405)

    arquivo = request.FILES.get("arquivo")

    if not arquivo:
        return JsonResponse({
            "ok": False,
            "mensagem": "Selecione um arquivo."
        }, status=400)

    try:
        service = ImportadorPlanilhaService(arquivo, atualizar_sem_serie=True)
        resultado = service.executar()

        return JsonResponse({
            "ok": True,
            "mensagem": "Importação concluída com sucesso.",
            "resultado": resultado
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "mensagem": f"Erro ao importar: {str(e)}"
        }, status=500)

### ITEM / Equipamento detalhe 
def _tem_valor(valor):
    """
    Retorna True apenas para valores úteis.
    Evita exibir campos nulos, vazios ou sem relação com o item.
    """
    if valor is None:
        return False

    if isinstance(valor, str) and not valor.strip():
        return False

    return True


def _str_fk(obj, *attrs):
    """
    Busca o primeiro atributo existente/preenchido em uma FK.
    Exemplo: _str_fk(item.localidade, "local", "nome")
    """
    if not obj:
        return None

    for attr in attrs:
        value = getattr(obj, attr, None)

        if _tem_valor(value):
            return value

    return str(obj)


def _add_info(lista, label, valor, icon="fa-circle-info", mono=False, destaque=False):
    """
    Adiciona campo somente se o valor existir.
    """
    if _tem_valor(valor):
        lista.append({
            "label": label,
            "valor": valor,
            "icon": icon,
            "mono": mono,
            "destaque": destaque,
        })


def _safe_decimal(valor):
    if valor is None:
        return Decimal("0.00")

    try:
        return Decimal(valor)
    except Exception:
        return Decimal("0.00")


def _get_locacao(item):
    """
    Evita erro caso a relação reversa de locação não exista.
    """
    try:
        return getattr(item, "locacao", None)
    except Exception:
        return None


@login_required
def item_monitoracao(request, pk: int):
    """AJAX — conectividade PRTG do equipamento (histórico e status em tempo real)."""
    from collections import defaultdict

    item = get_object_or_404(Item, pk=pk)

    try:
        dias = int(request.GET.get('periodo', '30'))
    except (ValueError, TypeError):
        dias = 30
    if dias not in (7, 14, 30, 90, 0):
        dias = 30

    agora  = timezone.now()
    inicio = agora - timedelta(days=dias) if dias > 0 else None

    # ── Descobrir prtg_objid via layout das plantas (ou histórico) ─────────
    from services.prtg_monitor_service import (
        prtg_objid_do_item, registrar_evento, periodos_e_totais, _quando_transicao,
    )
    prtg_objid = prtg_objid_do_item(item.pk)

    if not prtg_objid:
        return JsonResponse({
            'ok':            True,
            'encontrado':    False,
            'periodo_dias':  dias,
            'periodo_label': 'Histórico completo' if dias == 0 else f'Últimos {dias} dias',
            'info':          None,
            'status_atual':  None,
            'periodos':      [],
            'totais':        {},
            'historico':     [],
        })

    # ── Consultar PRTG ao vivo e auto-registrar mudança ────────────────────
    prtg_info        = None
    prtg_status_atual = None
    try:
        from services.prtg_service import get_devices_map
        dev = get_devices_map().get(prtg_objid)
        if dev:
            prtg_status_atual = dev['status_slug']
            # Grava o evento se o status mudou (fallback ao coletor agendado),
            # datando pelo momento real da transição reportado pelo PRTG.
            registrar_evento(
                prtg_objid, prtg_status_atual,
                quando=_quando_transicao(dev, agora),
                device_nome=dev.get('name', ''),
                device_host=dev.get('host', ''),
                device_grupo=dev.get('group', ''),
                item_id=item.pk,
            )
            prtg_info = {
                'objid':       prtg_objid,
                'nome':        dev['name'],
                'host':        dev['host'],
                'status_slug': prtg_status_atual,
                'statustext':  dev['statustext'],
                'css_color':   dev['css_color'],
                'ping_status': dev.get('ping_status'),
                'uptime_pct':  dev.get('uptime_pct'),
            }
    except Exception:
        pass  # PRTG indisponível — mantém histórico já gravado

    # ── Construir períodos contínuos (lógica compartilhada com o relatório) ──
    periodos, totais_dd = periodos_e_totais(prtg_objid, inicio, agora)
    for p in periodos:
        p['inicio'] = p['inicio'].strftime('%d/%m/%Y %H:%M')
        p['fim']    = p['fim'].strftime('%d/%m/%Y %H:%M')
        p['dias']   = round(p['dias'], 1)

    historico = []
    for ev in ItemPRTGHistorico.objects.filter(prtg_objid=prtg_objid).order_by('-registrado_em')[:100]:
        historico.append({
            'status_anterior': ev.status_anterior,
            'status_novo':     ev.status_novo,
            'registrado_em':   ev.registrado_em.strftime('%d/%m/%Y %H:%M'),
        })

    return JsonResponse({
        'ok':            True,
        'encontrado':    True,
        'periodo_dias':  dias,
        'periodo_label': 'Histórico completo' if dias == 0 else f'Últimos {dias} dias',
        'info':          prtg_info,
        'status_atual':  prtg_status_atual,
        'periodos':      periodos,
        'totais':        dict(totais_dd),
        'historico':     historico,
    })


# ─────────────────────────────────────────────────────────────
# Relatório de monitoração (disponibilidade PRTG)
# ─────────────────────────────────────────────────────────────

_MON_STATUS_LABEL = {
    "up": "Online", "down": "Offline", "warning": "Instável",
    "unusual": "Instável", "collecting": "Coletando", "unknown": "Desconhecido",
    "no_probe": "Sem probe", "paused_by_user": "Pausado", "paused_by_dep": "Pausado (dep.)",
    "paused_by_sched": "Pausado", "paused_until": "Pausado", "not_licensed": "Sem licença",
}


def _monitoracao_xlsx(linhas, periodo_label):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND_DARK, BRAND, SOFT, ZEBRA = "0B3D6E", "0071E3", "E5F0FB", "F4F9FE"
    INK = "1F2733"
    hair = Side(style="thin", color="CFE0F2")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    f_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color="5B6B7F")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Calibri", size=10, color=INK)
    fill_title = PatternFill("solid", fgColor=BRAND_DARK)
    fill_sub = PatternFill("solid", fgColor=SOFT)
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left", vertical="center")
    a_left_ind = Alignment(horizontal="left", vertical="center", indent=1)
    dt_fmt = "DD/MM/YYYY HH:MM"

    STATUS_FILL = {
        "up": ("E6F4EA", "1E8E3E"), "down": ("FCE8E6", "D93025"),
        "warning": ("FEF1E0", "B35A00"), "unusual": ("FEF1E0", "B35A00"),
    }

    header = ["#", "Equipamento / Device", "Nº Série / Host", "Grupo / Localidade", "Status Atual",
              "Disponibilidade %", "Dias Observados", "Tempo Online (dias)",
              "Tempo Offline (h)", "Instável (dias)", "Quedas",
              "Monitorado Desde", "Último Evento"]
    ncols = len(header)
    center_cols = {1, 5, 6, 7, 8, 9, 10, 11, 12, 13}

    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoração PRTG"
    ws.sheet_view.showGridLines = False

    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]; c.value = "RELATÓRIO DE MONITORAÇÃO — PRTG"; c.font = f_title; c.fill = fill_title; c.alignment = a_left_ind
    ws.row_dimensions[1].height = 34
    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
    ws.merge_cells(f"A2:{last}2")
    c2 = ws["A2"]
    c2.value = f"Santa Colomba Agropecuária  ·  {periodo_label}  ·  {len(linhas)} equipamento(s)  ·  gerado em {gerado}"
    c2.font = f_sub; c2.fill = fill_sub; c2.alignment = a_left_ind
    ws.row_dimensions[2].height = 18

    HEADER_ROW = 3
    for ci, h in enumerate(header, 1):
        cc = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cc.fill = fill_header; cc.font = f_header; cc.border = border
        cc.alignment = a_center if ci in center_cols else a_left
    ws.row_dimensions[HEADER_ROW].height = 26

    row = HEADER_ROW + 1
    for i, r in enumerate(linhas, start=1):
        ult = r["ultimo_evento"]
        if ult is not None:
            ult = timezone.localtime(ult).replace(tzinfo=None)
        desde = r["monitorado_desde"]
        if desde is not None:
            desde = timezone.localtime(desde).replace(tzinfo=None)
        status_label = _MON_STATUS_LABEL.get(r["status_atual"], r["status_atual"] or "—")
        valores = [i, r["titulo"], r["subtitulo"] or "", r["localidade"] or "", status_label,
                   r["pct_up"] if r["pct_up"] is not None else "—",
                   r["dias_observados"], r["dias_up"], r["horas_down"],
                   r["dias_warning"], r["quedas"], desde, ult]
        zebra = (i % 2 == 0)
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = border
            cell.font = f_cell
            cell.alignment = a_center if ci in center_cols else a_left
            if ci in (12, 13) and val:
                cell.number_format = dt_fmt
                cell.alignment = a_center
            if ci == 6 and isinstance(val, (int, float)):
                cell.number_format = '0.0"%"'
            if ci == 5:
                bg, fg = STATUS_FILL.get(r["status_atual"], ("F0F0F2", "5B6B7F"))
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=fg)
                cell.alignment = a_center
            elif zebra:
                cell.fill = fill_zebra
        row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last}{max(row - 1, HEADER_ROW)}"

    widths = {}
    for r_ in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        for idx, val in enumerate(r_, start=1):
            widths[idx] = max(widths.get(idx, 0), len(str(val)) if val is not None else 0)
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 11), 46)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    now = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="monitoracao_prtg_{now}.xlsx"'
    return resp


@login_required
def monitoracao_relatorio(request):
    """Relatório de disponibilidade (uptime/downtime) dos equipamentos monitorados no PRTG."""
    from services.prtg_monitor_service import relatorio_monitoracao

    try:
        dias = int(request.GET.get("periodo", "30"))
    except (ValueError, TypeError):
        dias = 30
    if dias not in (7, 14, 30, 90, 0):
        dias = 30

    linhas, resumo = relatorio_monitoracao(dias)

    f_status = (request.GET.get("status") or "").strip()  # online | offline | quedas
    q = (request.GET.get("q") or "").strip()
    if f_status == "online":
        linhas = [r for r in linhas if r["online"]]
    elif f_status == "offline":
        linhas = [r for r in linhas if r["offline"]]
    elif f_status == "quedas":
        linhas = [r for r in linhas if r["quedas"] > 0]
    if q:
        ql = q.lower()
        linhas = [
            r for r in linhas
            if ql in (r["titulo"] or "").lower()
            or ql in (r["subtitulo"] or "").lower()
            or ql in (r["device_nome"] or "").lower()
            or ql in (r["device_host"] or "").lower()
        ]

    for r in linhas:
        r["status_label"] = _MON_STATUS_LABEL.get(r["status_atual"], r["status_atual"] or "—")

    periodo_label = "Histórico completo" if dias == 0 else f"Últimos {dias} dias"

    if request.GET.get("export") == "xlsx":
        return _monitoracao_xlsx(linhas, periodo_label)

    context = {
        "linhas": linhas,
        "resumo": resumo,
        "periodo_dias": dias,
        "periodo_label": periodo_label,
        "f_status": f_status,
        "f_q": q,
        "total_filtrado": len(linhas),
    }
    return render(request, "front/equipamentos/monitoracao_relatorio.html", context)


@login_required
def equipamento_qr(request, pk: int):
    item = get_object_or_404(Item, pk=pk)
    detalhe_url = request.build_absolute_uri(f"/equipamentos/{pk}/")
    return render(request, "front/equipamentos/equipamento_qr.html", {
        "item": item,
        "detalhe_url": detalhe_url,
    })


@login_required
def equipamento_detalhe(request, pk: int):
    item = get_object_or_404(
        Item.objects.select_related(
            "subtipo",
            "localidade",
            "centro_custo",
            "fornecedor",
            "criado_por",
            "atualizado_por",
        ),
        pk=pk,
    )

    today = timezone.localdate()

    movimentacoes = (
        MovimentacaoItem.objects
        .filter(item=item)
        .select_related(
            "usuario",
            "lote",
            "lote__fornecedor",
            "localidade_origem",
            "localidade_destino",
            "centro_custo_origem",
            "centro_custo_destino",
            "fornecedor_manutencao",
            "criado_por",
        )
        .order_by("-created_at")
    )

    historico_manutencao = movimentacoes.filter(
        tipo_movimentacao__in=[
            TipoMovimentacaoChoices.ENVIO_MANUTENCAO,
            TipoMovimentacaoChoices.RETORNO_MANUTENCAO,
        ]
    )

    custo_manut = historico_manutencao.aggregate(t=Sum("custo"))["t"] or Decimal("0.00")

    # =========================================================
    # Lotes vinculados ao item
    # =========================================================
    vinculos_lote = (
        ItemLote.objects
        .filter(item=item)
        .select_related(
            "lote",
            "lote__fornecedor",
        )
        .order_by("-lote__data_entrada", "-created_at")
    )

    lotes_cards = []
    total_entrada_lotes = 0
    total_disponivel_lotes = 0
    valor_total_lotes = Decimal("0.00")
    valor_disponivel_lotes = Decimal("0.00")

    for vinculo in vinculos_lote:
        lote = vinculo.lote

        quantidade_entrada = vinculo.quantidade_entrada or 0
        quantidade_disponivel = vinculo.quantidade_disponivel or 0
        custo_unitario = _safe_decimal(vinculo.custo_unitario or getattr(lote, "custo_unitario", None))

        valor_total = Decimal(quantidade_entrada) * custo_unitario
        valor_disponivel = Decimal(quantidade_disponivel) * custo_unitario

        total_entrada_lotes += quantidade_entrada
        total_disponivel_lotes += quantidade_disponivel
        valor_total_lotes += valor_total
        valor_disponivel_lotes += valor_disponivel

        percentual_saldo = 0
        if quantidade_entrada > 0:
            percentual_saldo = int((quantidade_disponivel / quantidade_entrada) * 100)

        movimentacoes_lote = movimentacoes.filter(lote=lote)
        ultima_mov_lote = movimentacoes_lote.first()

        if quantidade_disponivel <= 0:
            status_lote = "zerado"
        elif percentual_saldo <= 20:
            status_lote = "critico"
        elif percentual_saldo <= 50:
            status_lote = "atencao"
        else:
            status_lote = "ok"

        lotes_cards.append({
            "vinculo": vinculo,
            "lote": lote,
            "quantidade_entrada": quantidade_entrada,
            "quantidade_disponivel": quantidade_disponivel,
            "custo_unitario": custo_unitario,
            "valor_total": valor_total,
            "valor_disponivel": valor_disponivel,
            "percentual_saldo": percentual_saldo,
            "status_lote": status_lote,
            "movimentacoes_count": movimentacoes_lote.count(),
            "ultima_movimentacao": ultima_mov_lote,
        })

    possui_lotes = vinculos_lote.exists()

    # =========================================================
    # Detentor atual
    # =========================================================
    ultimo_resp = "Em estoque / Não definido"

    ultima_mov = movimentacoes.first()
    if ultima_mov:
        if ultima_mov.usuario:
            ultimo_resp = f"Usuário: {ultima_mov.usuario.nome}"
        elif ultima_mov.centro_custo_destino:
            cc_nome = _str_fk(ultima_mov.centro_custo_destino, "departamento", "nome")
            ultimo_resp = f"Setor: {cc_nome}"
        elif ultima_mov.localidade_destino:
            local_nome = _str_fk(ultima_mov.localidade_destino, "local", "nome")
            ultimo_resp = f"Local: {local_nome}"
        elif ultima_mov.fornecedor_manutencao:
            fornecedor_nome = _str_fk(ultima_mov.fornecedor_manutencao, "nome", "razao_social")
            ultimo_resp = f"Externo: {fornecedor_nome}"

    # =========================================================
    # Locação / Financeiro
    # =========================================================
    locacao = _get_locacao(item)

    financeiro = {
        "modo": "LOCAÇÃO" if getattr(item, "locado", "nao") == "sim" else "AQUISIÇÃO",
        "custo_aquisicao": _safe_decimal(item.valor),
        "custo_manutencao": custo_manut,
        "custo_aluguel_acumulado": Decimal("0.00"),
        "custo_tempo_empresa": Decimal("0.00"),
        "custo_total_empresa": Decimal("0.00"),
        "tco": Decimal("0.00"),
        "valor_atual": Decimal("0.00"),
        "custo_mensal": Decimal("0.00"),
        "dias_na_empresa": 0,
        "meses_na_empresa": 0,
        "data_inicio": None,
        "data_fim": None,
        "vida_util_perc": 0,
        "vida_util_texto": None,
        "status_vida": "ok",
    }

    if financeiro["modo"] == "LOCAÇÃO" and locacao:
        valor_mensal = _safe_decimal(getattr(locacao, "valor_mensal", None))
        dt_inicio = getattr(locacao, "data_entrada", None)
        tempo_locado_meses = getattr(locacao, "tempo_locado", 0) or 0

        financeiro["custo_mensal"] = valor_mensal
        financeiro["valor_atual"] = valor_mensal

        if dt_inicio:
            financeiro["data_inicio"] = dt_inicio

            dias_corridos = max(0, (today - dt_inicio).days)
            meses_uso = dias_corridos // 30 if dias_corridos > 0 else 0

            if tempo_locado_meses > 0:
                meses_pagos = min(meses_uso, int(tempo_locado_meses))
            else:
                meses_pagos = meses_uso

            custo_aluguel_acumulado = valor_mensal * Decimal(meses_pagos)

            financeiro["dias_na_empresa"] = dias_corridos
            financeiro["meses_na_empresa"] = meses_uso
            financeiro["custo_aluguel_acumulado"] = custo_aluguel_acumulado
            financeiro["custo_tempo_empresa"] = custo_aluguel_acumulado
            financeiro["custo_total_empresa"] = custo_aluguel_acumulado + custo_manut
            financeiro["tco"] = financeiro["custo_total_empresa"]

            if tempo_locado_meses > 0:
                total_dias = int(tempo_locado_meses) * 30
                dt_fim = dt_inicio + timedelta(days=total_dias)
                financeiro["data_fim"] = dt_fim

                financeiro["vida_util_perc"] = min(100, max(0, int((dias_corridos / total_dias) * 100)))

                restante = (dt_fim - today).days
                if restante < 0:
                    financeiro["vida_util_texto"] = "Contrato vencido"
                    financeiro["status_vida"] = "critical"
                else:
                    financeiro["vida_util_texto"] = f"{restante} dias restantes"
                    if restante < 30:
                        financeiro["status_vida"] = "warning"
            else:
                financeiro["vida_util_texto"] = "Contrato sem prazo definido"
        else:
            financeiro["custo_total_empresa"] = custo_manut
            financeiro["tco"] = custo_manut

    else:
        financeiro["custo_tempo_empresa"] = financeiro["custo_aquisicao"]
        financeiro["custo_total_empresa"] = financeiro["custo_aquisicao"] + custo_manut
        financeiro["tco"] = financeiro["custo_total_empresa"]

        if item.data_compra:
            financeiro["data_inicio"] = item.data_compra

            vida_util_anos = 5
            vida_util_dias = vida_util_anos * 365
            dias_uso = max(0, (today - item.data_compra).days)

            financeiro["dias_na_empresa"] = dias_uso
            financeiro["meses_na_empresa"] = dias_uso // 30 if dias_uso > 0 else 0

            if dias_uso < vida_util_dias:
                fator = Decimal(dias_uso) / Decimal(vida_util_dias)
                financeiro["valor_atual"] = financeiro["custo_aquisicao"] * (Decimal("1.00") - fator)
                financeiro["vida_util_perc"] = int((dias_uso / vida_util_dias) * 100)
                anos_restantes = max(0, vida_util_anos - (dias_uso // 365))
                financeiro["vida_util_texto"] = f"~{anos_restantes} anos restantes"
                financeiro["data_fim"] = item.data_compra + timedelta(days=vida_util_dias)
            else:
                financeiro["valor_atual"] = Decimal("0.00")
                financeiro["vida_util_perc"] = 100
                financeiro["vida_util_texto"] = "Totalmente depreciado"
                financeiro["status_vida"] = "warning"
                financeiro["data_fim"] = item.data_compra + timedelta(days=vida_util_dias)

    # =========================================================
    # Preventivas
    # =========================================================
    preventivas = (
        Preventiva.objects
        .filter(equipamento=item)
        .select_related("checklist_modelo")
        .order_by("data_proxima")
    )

    _JANELA_ATENCAO = 7
    status_saude = "ok" if not preventivas else "sem_data"
    for p in preventivas:
        # Prioridade: data_limite_preventiva do item → intervalo_dias do checklist
        intervalo = 0
        try:
            intervalo = int(item.data_limite_preventiva or 0)
        except (TypeError, ValueError):
            pass
        if intervalo <= 0 and p.checklist_modelo:
            try:
                intervalo = int(p.checklist_modelo.intervalo_dias or 0)
            except (TypeError, ValueError):
                pass

        # Calcula a data efetiva da próxima preventiva
        if intervalo > 0 and p.data_ultima:
            p.proxima_calc = p.data_ultima + timedelta(days=intervalo)
        else:
            p.proxima_calc = p.data_proxima

        if p.proxima_calc:
            dias = (p.proxima_calc - today).days
            p.atrasado = dias < 0
            p.atencao  = 0 <= dias <= _JANELA_ATENCAO
            if p.atrasado:
                status_saude = "critical"
            elif p.atencao and status_saude not in ("critical",):
                status_saude = "atencao"
            elif not p.atrasado and not p.atencao and status_saude == "sem_data":
                status_saude = "ok"
        else:
            p.atrasado = False
            p.atencao  = False

    # =========================================================
    # Ficha técnica dinâmica
    # Só adiciona campos preenchidos.
    # =========================================================
    dados_item = []

    _add_info(dados_item, "Nome", item.nome, "fa-cube", destaque=True)
    _add_info(dados_item, "Número de Série", item.numero_serie, "fa-barcode", mono=True)
    _add_info(dados_item, "Marca", item.marca, "fa-industry")
    _add_info(dados_item, "Modelo", item.modelo, "fa-microchip")
    _add_info(dados_item, "Subtipo", _str_fk(item.subtipo, "nome", "descricao"), "fa-tags")
    _add_info(dados_item, "Localidade Atual", _str_fk(item.localidade, "local", "nome"), "fa-location-dot")
    _add_info(
        dados_item,
        "Centro de Custo Atual",
        (
            f"{_str_fk(item.centro_custo, 'departamento', 'nome')} "
            f"({_str_fk(item.centro_custo, 'numero', 'codigo')})"
            if item.centro_custo else None
        ),
        "fa-building"
    )
    _add_info(dados_item, "Fornecedor Principal", _str_fk(item.fornecedor, "nome", "razao_social"), "fa-truck-field")
    _add_info(dados_item, "Número Pedido / NF", item.numero_pedido, "fa-file-invoice", mono=True)
    _add_info(dados_item, "Data de Compra", item.data_compra, "fa-calendar-days")
    _add_info(dados_item, "Valor Unitário", item.valor, "fa-coins")
    _add_info(dados_item, "Item de Consumo", item.get_item_consumo_display() if hasattr(item, "get_item_consumo_display") else item.item_consumo, "fa-box")
    _add_info(dados_item, "PMB", item.get_pmb_display() if hasattr(item, "get_pmb_display") else item.pmb, "fa-shield-halved")
    _add_info(dados_item, "Locado", item.get_locado_display() if hasattr(item, "get_locado_display") else item.locado, "fa-file-contract")
    _add_info(dados_item, "Preventiva", item.get_precisa_preventiva_display() if hasattr(item, "get_precisa_preventiva_display") else item.precisa_preventiva, "fa-screwdriver-wrench")
    _add_info(dados_item, "Periodicidade Preventiva", item.data_limite_preventiva, "fa-clock")

    dados_locacao = []
    pos_contrato = None  # dados de pós-contrato para exibição destacada
    if financeiro["modo"] == "LOCAÇÃO" and locacao:
        _add_info(dados_locacao, "Fornecedor Locador", _str_fk(getattr(locacao, "fornecedor", None), "nome", "razao_social"), "fa-truck")
        _add_info(dados_locacao, "Data de Entrada", getattr(locacao, "data_entrada", None), "fa-calendar-days")
        _add_info(dados_locacao, "Vencimento do Contrato", locacao.data_vencimento, "fa-calendar-xmark")
        _add_info(dados_locacao, "Tempo Contratado", f"{locacao.tempo_locado} meses" if locacao.tempo_locado else None, "fa-hourglass-half")
        _add_info(dados_locacao, "Valor Mensal", getattr(locacao, "valor_mensal", None), "fa-money-bill-wave")
        _add_info(dados_locacao, "Contrato", getattr(locacao, "contrato", None), "fa-file-signature")
        _add_info(dados_locacao, "Observações da Locação", getattr(locacao, "observacoes", None), "fa-note-sticky")

        if locacao.dias_pos_contrato is not None:
            meses, dias = locacao.meses_e_dias_pos_contrato or (0, 0)
            pos_contrato = {
                "dias_total": locacao.dias_pos_contrato,
                "meses": meses,
                "dias_restantes": dias,
                "data_vencimento": locacao.data_vencimento,
                "status_item": item.get_status_display(),
            }
        elif item.status == "pausado" and locacao.contrato_vencido:
            # contrato vencido mas item pausado → contagem encerrada
            pos_contrato = {
                "pausado": True,
                "data_vencimento": locacao.data_vencimento,
            }

    auditoria = []
    _add_info(auditoria, "Criado em", getattr(item, "created_at", None), "fa-calendar-plus")
    _add_info(auditoria, "Criado por", getattr(item, "criado_por", None), "fa-user-plus")
    _add_info(auditoria, "Última edição", getattr(item, "updated_at", None), "fa-pen-to-square")
    _add_info(auditoria, "Atualizado por", getattr(item, "atualizado_por", None), "fa-user-pen")

    # =========================================================
    # Plantas onde este item está mapeado
    # =========================================================
    item_id_str = str(item.pk)
    plantas_mapeadas = [
        p for p in PlantaProjeto.objects.select_related("localidade").only(
            "pk", "nome", "localidade__local", "updated_at"
        )
        if any(str(e.get("item_id", "")) == item_id_str for e in p.layout.get("elements", []))
    ]

    # =========================================================
    # NinjaOne RMM — somente se o item tiver NinjaDevice vinculado
    # =========================================================
    ninja_device = None
    ninja_snapshots_hoje = []
    try:
        from ProjetoEstoque.models import NinjaDevice, NinjaDeviceSnapshot
        ninja_device = NinjaDevice.objects.filter(item=item).select_related("item").first()
        if ninja_device:
            ninja_snapshots_hoje = list(
                NinjaDeviceSnapshot.objects
                .filter(device=ninja_device, timestamp__date=today)
                .order_by("timestamp")
                .values("timestamp", "is_online", "current_user", "ip_address")
            )
    except Exception:
        pass

    context = {
        "item": item,
        "ultimo_resp": ultimo_resp,
        "movimentacoes": movimentacoes,
        "historico_manutencao": historico_manutencao,
        "preventivas": preventivas,
        "financeiro": financeiro,
        "status_saude": status_saude,
        "locacao": locacao,

        "dados_item": dados_item,
        "dados_locacao": dados_locacao,
        "pos_contrato": pos_contrato,
        "auditoria": auditoria,

        "possui_lotes": possui_lotes,
        "lotes_cards": lotes_cards,
        "total_entrada_lotes": total_entrada_lotes,
        "total_disponivel_lotes": total_disponivel_lotes,
        "valor_total_lotes": valor_total_lotes,
        "valor_disponivel_lotes": valor_disponivel_lotes,
        "plantas_mapeadas": plantas_mapeadas,
        "monitoracao_url": f"/equipamentos/{item.pk}/monitoracao/",
        "ninja_device": ninja_device,
        "ninja_snapshots_hoje": ninja_snapshots_hoje,
    }

    return render(request, "front/equipamentos/equipamento_detalhe.html", context)

def adicionar_erros_validacao_no_form(form, erro):
    if hasattr(erro, "message_dict"):
        for campo, mensagens in erro.message_dict.items():
            campo_form = campo if campo in form.fields else None

            for mensagem in mensagens:
                form.add_error(campo_form, mensagem)
    else:
        form.add_error(None, erro)


def preencher_auditoria(obj, user, criando=False):
    if criando and hasattr(obj, "criado_por") and not getattr(obj, "criado_por_id", None):
        obj.criado_por = user

    if hasattr(obj, "atualizado_por"):
        obj.atualizado_por = user


@login_required
def item_update(request, pk):
    item = get_object_or_404(Item, pk=pk)

    item_lote = (
        ItemLote.objects
        .select_related("lote")
        .filter(item=item)
        .order_by("-created_at")
        .first()
    )

    lote = item_lote.lote if item_lote else None

    try:
        locacao = item.locacao
    except Exception:
        locacao = None

    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)

        form_valido = form.is_valid()

        eh_locado = False
        eh_consumo = False

        if form_valido:
            eh_locado = form.cleaned_data.get("locado") == SimNaoChoices.SIM
            eh_consumo = form.cleaned_data.get("item_consumo") == SimNaoChoices.SIM

        locacao_form = LocacaoForm(
            request.POST,
            instance=locacao if locacao else None
        )

        lote_form = LoteEstoqueCreateForm(
            request.POST,
            instance=lote if lote else None
        )

        locacao_valida = locacao_form.is_valid() if eh_locado else True
        lote_valido = lote_form.is_valid() if eh_consumo else True

        _STATUS_PAUSANTES = {
            StatusItemChoices.PAUSADO,
            StatusItemChoices.BACKUP,
            StatusItemChoices.MANUTENCAO,
            StatusItemChoices.DEFEITO,
        }

        if form_valido and locacao_valida and lote_valido:
            try:
                with transaction.atomic():
                    status_anterior = item.status
                    item_editado = form.save(commit=False)

                    preencher_auditoria(item_editado, request.user, criando=False)

                    if eh_consumo:
                        lote_editado = lote_form.save(commit=False)
                        preencher_auditoria(
                            lote_editado,
                            request.user,
                            criando=lote_editado.pk is None
                        )

                        lote_editado.full_clean()
                        lote_editado.save()

                        item_editado.tem_lote = True
                        item_editado.quantidade = lote_editado.quantidade
                        item_editado.valor = lote_editado.custo_unitario
                        item_editado.fornecedor = lote_editado.fornecedor
                        item_editado.numero_pedido = lote_editado.numero_nf
                        item_editado.data_compra = lote_editado.data_entrada

                    else:
                        if item_lote:
                            raise ValidationError(
                                "Este item possui lote vinculado. Não é permitido alterar de item de consumo para item normal sem tratar o estoque/lote."
                            )

                        item_editado.tem_lote = False

                    if eh_locado:
                        item_editado.data_compra = None
                        item_editado.numero_pedido = None

                    if item_editado.precisa_preventiva == SimNaoChoices.NAO:
                        item_editado.data_limite_preventiva = None

                    item_editado.full_clean()
                    item_editado.save()

                    if status_anterior != item_editado.status:
                        preventivas = item_editado.preventivas.all()
                        if item_editado.status in _STATUS_PAUSANTES and status_anterior not in _STATUS_PAUSANTES:
                            for prev in preventivas:
                                prev.pausar()
                        elif item_editado.status == StatusItemChoices.ATIVO and status_anterior in _STATUS_PAUSANTES:
                            for prev in preventivas:
                                prev.retomar()

                    if eh_consumo:
                        if item_lote:
                            quantidade_antiga = item_lote.quantidade_entrada or 0
                            quantidade_nova = lote_editado.quantidade or 0
                            diferenca = quantidade_nova - quantidade_antiga

                            nova_qtd_disponivel = (item_lote.quantidade_disponivel or 0) + diferenca

                            if nova_qtd_disponivel < 0:
                                raise ValidationError(
                                    "A quantidade do lote não pode ser menor que a quantidade já movimentada/consumida."
                                )

                            item_lote.lote = lote_editado
                            item_lote.quantidade_entrada = quantidade_nova
                            item_lote.quantidade_disponivel = nova_qtd_disponivel
                            item_lote.custo_unitario = lote_editado.custo_unitario

                            preencher_auditoria(item_lote, request.user, criando=False)
                            item_lote.full_clean()
                            item_lote.save()

                        else:
                            novo_item_lote = ItemLote(
                                item=item_editado,
                                lote=lote_editado,
                                quantidade_entrada=lote_editado.quantidade,
                                quantidade_disponivel=lote_editado.quantidade,
                                custo_unitario=lote_editado.custo_unitario,
                            )

                            preencher_auditoria(novo_item_lote, request.user, criando=True)
                            novo_item_lote.full_clean()
                            novo_item_lote.save()

                    if eh_locado:
                        locacao_editada = locacao_form.save(commit=False)
                        locacao_editada.equipamento = item_editado
                        locacao_editada.fornecedor = item_editado.fornecedor

                        preencher_auditoria(
                            locacao_editada,
                            request.user,
                            criando=locacao_editada.pk is None
                        )

                        locacao_editada.full_clean()
                        locacao_editada.save()

                    messages.success(request, "Item atualizado com sucesso.")
                    return redirect("equipamentos_list")

            except ValidationError as e:
                adicionar_erros_validacao_no_form(form, e)
                messages.error(request, "Erro de validação. Verifique os dados informados.")

            except IntegrityError as e:
                messages.error(
                    request,
                    f"Erro de integridade ao atualizar o item. Nenhum dado foi gravado. Detalhe: {str(e)}"
                )

            except Exception as e:
                messages.error(
                    request,
                    f"Erro inesperado ao atualizar o item: {str(e)}"
                )

        else:
            messages.error(request, "Verifique os campos obrigatórios antes de salvar.")

    else:
        form = ItemForm(instance=item)

        locacao_form = LocacaoForm(
            instance=locacao if item.locado == SimNaoChoices.SIM else None
        )

        lote_form = LoteEstoqueCreateForm(
            instance=lote if item.item_consumo == SimNaoChoices.SIM else None
        )

    return render(request, "front/equipamentos/cadastrar_equipamento.html", {
        "form": form,
        "locacao_form": locacao_form,
        "lote_form": lote_form,
        "editar": True,
        "item": item,
    })

@require_POST
@login_required
@permission_required("ProjetoEstoque.delete_item", raise_exception=True)
def equipamento_excluir(request, pk: int):
    item = get_object_or_404(Item, pk=pk)
    item.delete()
    messages.success(request, "Item excluído com sucesso.")
    return redirect("equipamentos_list")

