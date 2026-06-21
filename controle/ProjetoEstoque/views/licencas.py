from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.db.models import Q, Count, Sum, F
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..models import (
    SimNaoChoices,
    Licenca, LicencaLote, MovimentacaoLicenca,
    TipoMovLicencaChoices, PeriodicidadeChoices,
    Usuario, CentroCusto, Fornecedor,
)
from ..forms import LicencaForm, MovimentacaoLicencaForm, LicencaLoteForm


def _normalizar_custos_lote(lote):
    """
    Calcula custos unitários mensais e anuais de um LicencaLote.
    Centralizado aqui para evitar duplicação entre licenca_detail e licenca_export_excel.
    """
    periodicidade = str(lote.periodicidade or "").lower()
    qtd_lote = int(lote.quantidade_total or 0)
    custo_ciclo_lote = Decimal(lote.custo_ciclo or 0)

    if qtd_lote <= 0:
        return {
            "unitario_ciclo": Decimal("0.00"),
            "mensal_unit": Decimal("0.00"),
            "anual_unit": Decimal("0.00"),
            "mensal_total": Decimal("0.00"),
            "anual_total": Decimal("0.00"),
        }

    if periodicidade == "mensal":
        custo_mensal_lote_base = custo_ciclo_lote
    elif periodicidade == "trimestral":
        custo_mensal_lote_base = (custo_ciclo_lote / Decimal("3")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    elif periodicidade == "semestral":
        custo_mensal_lote_base = (custo_ciclo_lote / Decimal("6")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    elif periodicidade == "anual":
        custo_mensal_lote_base = (custo_ciclo_lote / Decimal("12")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    else:
        custo_mensal_lote_base = custo_ciclo_lote.quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

    unitario_ciclo = (custo_ciclo_lote / Decimal(qtd_lote)).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    mensal_unit = (custo_mensal_lote_base / Decimal(qtd_lote)).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    anual_unit = (mensal_unit * Decimal("12")).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    mensal_total = (mensal_unit * Decimal(qtd_lote)).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    anual_total = (mensal_total * Decimal("12")).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )

    return {
        "unitario_ciclo": unitario_ciclo,
        "mensal_unit": mensal_unit,
        "anual_unit": anual_unit,
        "mensal_total": mensal_total,
        "anual_total": anual_total,
    }


@login_required
def licenca_list(request):
    """
    Dashboard de Licenças (Enterprise View).
    Exibe listagem com saldo consolidado em tempo real baseado nos lotes.
    """
    # --- 1. Construção do QuerySet (Eager Loading para Performance) ---
    qs = (
        Licenca.objects
        .select_related("fornecedor", "centro_custo")
        .annotate(
            # KPI por Linha: Quantos lotes existem para esta licença?
            qtd_lotes=Count("lotes", distinct=True),
            
            # KPI Crítico: Soma o saldo disponível de cada lote vinculado
            # Se não tiver lotes, retorna 0 (Coalesce)
            estoque_real=Coalesce(Sum("lotes__quantidade_disponivel"), 0)
        )
        .order_by("nome")
    )

    # --- 2. Filtros Inteligentes ---
    q = request.GET.get("q", "").strip()
    fornecedor_id = request.GET.get("fornecedor", "").strip()
    pmb_filter = request.GET.get("pmb", "").strip()
    status_estoque = request.GET.get("status", "").strip()

    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(observacao__icontains=q))
    
    if fornecedor_id and fornecedor_id.isdigit():
        qs = qs.filter(fornecedor_id=fornecedor_id)
        
    if pmb_filter:
        qs = qs.filter(pmb=pmb_filter)

    # Filtro de Status de Estoque (Baseado na anotação calculada)
    if status_estoque == "com_estoque":
        qs = qs.filter(estoque_real__gt=0)
    elif status_estoque == "sem_estoque":
        qs = qs.filter(estoque_real=0)

    # --- 3. KPIs Globais (Cards do Topo) ---
    # Calculamos totais rápidos para o gestor ter visão macro
    kpi_total_licencas = Licenca.objects.count()
    
    # Soma total de assentos disponíveis na empresa inteira
    kpi_total_assentos = LicencaLote.objects.aggregate(
        total=Coalesce(Sum('quantidade_disponivel'), 0)
    )['total']
    
    kpi_pmb = Licenca.objects.filter(pmb=SimNaoChoices.SIM).count()

    # --- 4. Paginação e Controle ---
    try:
        per_page = int(request.GET.get("pp", 15))
    except ValueError:
        per_page = 15
    
    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_num)

    # Preserva filtros na paginação (query string)
    get_copy = request.GET.copy()
    if "page" in get_copy: del get_copy["page"]
    qs_keep = get_copy.urlencode()

    # --- 5. Contexto para o Template ---
    context = {
        "page_obj": page_obj,
        "qs_keep": qs_keep,
        "total_registros": qs.count(),
        
        # KPIs
        "kpi_total": kpi_total_licencas,
        "kpi_assentos": kpi_total_assentos,
        "kpi_pmb": kpi_pmb,

        # Estado dos Filtros (para manter selecionado)
        "filter_q": q,
        "filter_fornecedor": int(fornecedor_id) if fornecedor_id.isdigit() else "",
        "filter_pmb": pmb_filter,
        "filter_status": status_estoque,
        "per_page": per_page,

        # Opções para Dropdowns
        "opt_fornecedores": Fornecedor.objects.values("id", "nome").order_by("nome"),
        "opt_pmb": SimNaoChoices.choices,
    }

    # AJAX partial render (view toggle + pagination)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        view_mode = request.GET.get('view', 'list')
        data = {
            'count': context['total_registros'],
            'pagination': render_to_string('front/licencas/_lic_pagination.html', context, request=request),
        }
        if view_mode == 'gallery':
            data['gallery'] = render_to_string('front/licencas/_lic_gallery.html', context, request=request)
        else:
            data['tbody'] = render_to_string('front/licencas/_lic_rows.html', context, request=request)
        return JsonResponse(data)

    return render(request, "front/licencas/licenca_list.html", context)

@login_required
def licenca_form(request, pk=None):
    """
    View simplificada para Cadastro/Edição de Licença (4 campos).
    """
    # Se houver PK, é edição. Senão, criação.
    obj = get_object_or_404(Licenca, pk=pk) if pk else None

    if request.method == "POST":
        form = LicencaForm(request.POST, instance=obj)
        if form.is_valid():
            try:
                licenca = form.save(commit=False)
                
                # Preenche auditoria
                if not obj:
                    licenca.criado_por = request.user
                licenca.atualizado_por = request.user
                
                licenca.save()
                
                verb = "editada" if obj else "criada"
                messages.success(request, f"Licença '{licenca.nome}' {verb} com sucesso!")
                return redirect("licenca_list")
                
            except Exception as e:
                messages.error(request, f"Erro crítico ao salvar: {e}")
        else:
            messages.error(request, "Verifique os campos obrigatórios.")
    else:
        form = LicencaForm(instance=obj)

    return render(request, "front/licencas/licenca_form.html", {
        "form": form,
        "obj": obj
    })


# --- HELPER: Calcula Alocação por Centro de Custo ---
def _get_dados_cc(licenca):
    """
    Reconstitui o estado atual das licenças para agrupar por Centro de Custo.
    Lógica: Pega todas as movimentações ordenadas. 
    Se 'atribuicao' -> Adiciona usuário. Se 'devolucao' -> Remove usuário.
    """
    movs = MovimentacaoLicenca.objects.filter(licenca=licenca).select_related(
        'usuario', 'centro_custo_destino'
    ).order_by('created_at')

    # 1. Descobrir quem está ativo e qual seu custo atual
    ativos = {} # {usuario_id: MovimentacaoObj}
    
    for mov in movs:
        if mov.tipo == 'atribuicao':
            ativos[mov.usuario_id] = mov
        elif mov.tipo == 'devolucao':
            ativos.pop(mov.usuario_id, None)

    # 2. Agrupar por Centro de Custo
    cc_stats = defaultdict(lambda: {'nome': 'Não Definido', 'qtd': 0, 'total': Decimal(0)})

    for mov in ativos.values():
        cc = mov.centro_custo_destino
        cc_id = cc.id if cc else 'na'
        cc_name = cc.departamento if cc else 'Sem Centro de Custo'
        
        cc_stats[cc_id]['nome'] = cc_name
        cc_stats[cc_id]['qtd'] += 1
        cc_stats[cc_id]['total'] += (mov.valor_unitario or Decimal(0))

    # Retorna lista ordenada pelo maior valor total
    return sorted(cc_stats.values(), key=lambda x: x['total'], reverse=True)

@login_required
def licenca_detail(request, pk):
    licenca = get_object_or_404(
        Licenca.objects.select_related("fornecedor", "centro_custo"),
        pk=pk
    )

    lotes_qs = (
        LicencaLote.objects.filter(licenca=licenca)
        .select_related("fornecedor", "centro_custo")
        .order_by("-data_compra", "-id")
    )
    lotes = list(lotes_qs)

    qtd_total = 0
    qtd_disp = 0

    total_investido_historico = Decimal("0.00")
    burn_rate_mensal = Decimal("0.00")
    burn_rate_anual = Decimal("0.00")

    # Normalização dos lotes + KPIs
    for lote in lotes:
        qtd_lote = int(lote.quantidade_total or 0)
        qtd_disponivel = int(lote.quantidade_disponivel or 0)
        em_uso = max(0, qtd_lote - qtd_disponivel)

        qtd_total += qtd_lote
        qtd_disp += qtd_disponivel

        custos = _normalizar_custos_lote(lote)

        lote.unitario_real = custos["unitario_ciclo"]
        lote.custo_mensal_unit = custos["mensal_unit"]
        lote.custo_anual_unit = custos["anual_unit"]
        lote.total_mensal_calc = custos["mensal_total"]
        lote.total_investido_calc = custos["anual_total"]

        total_investido_historico += custos["anual_total"]

        if em_uso > 0:
            burn_rate_mensal += (custos["mensal_unit"] * Decimal(em_uso)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            burn_rate_anual += (custos["anual_unit"] * Decimal(em_uso)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

    qtd_em_uso = max(0, qtd_total - qtd_disp)
    pct_uso = int((qtd_em_uso / qtd_total) * 100) if qtd_total > 0 else 0

    # Histórico completo + paginação
    movimentacoes_qs = (
        MovimentacaoLicenca.objects.filter(licenca=licenca)
        .select_related("usuario", "lote", "centro_custo_destino", "criado_por")
        .order_by("-created_at")
    )

    movimentacoes_lista = []
    for mov in movimentacoes_qs:
        mensal = Decimal("0.00")
        anual = Decimal("0.00")

        if mov.lote:
            custos_mov = _normalizar_custos_lote(mov.lote)
            mensal = custos_mov["mensal_unit"]
            anual = custos_mov["anual_unit"]

        mov.custo_mensal_exibicao = mensal
        mov.custo_anual_exibicao = anual
        movimentacoes_lista.append(mov)

    paginator = Paginator(movimentacoes_lista, 10)
    page_number = request.GET.get("page")
    movimentacoes_page = paginator.get_page(page_number)

    cc_list = _get_dados_cc(licenca)

    context = {
        "obj": licenca,
        "kpi": {
            "total": qtd_total,
            "disponivel": qtd_disp,
            "em_uso": qtd_em_uso,
            "pct_uso": pct_uso,
            "investimento_total": total_investido_historico,
            "gasto_mensal": burn_rate_mensal,
            "gasto_anual": burn_rate_anual,
        },
        "lotes": lotes,
        "movimentacoes": movimentacoes_page,
        "cc_list": cc_list,
    }

    return render(request, "front/licencas/licenca_detail.html", context)


@login_required
@permission_required("ProjetoEstoque.view_licenca", raise_exception=True)
def licenca_export_excel(request, pk):
    licenca = get_object_or_404(
        Licenca.objects.select_related("fornecedor", "centro_custo"),
        pk=pk
    )

    lotes_qs = (
        LicencaLote.objects.filter(licenca=licenca)
        .select_related("fornecedor", "centro_custo")
        .order_by("-data_compra", "-id")
    )

    movimentacoes_qs = (
        MovimentacaoLicenca.objects.filter(licenca=licenca)
        .select_related("usuario", "lote", "centro_custo_destino", "criado_por")
        .order_by("-created_at")
    )

    cc_list = _get_dados_cc(licenca)
    lotes = list(lotes_qs)
    movs = list(movimentacoes_qs)

    # =========================================================================
    # KPIs consolidados (mesma base do detalhe da licença)
    # =========================================================================
    qtd_total = sum(int(l.quantidade_total or 0) for l in lotes)
    qtd_disp = sum(int(l.quantidade_disponivel or 0) for l in lotes)
    qtd_em_uso = max(0, qtd_total - qtd_disp)
    pct_uso = int((qtd_em_uso / qtd_total) * 100) if qtd_total > 0 else 0

    investimento_total = Decimal("0.00")
    gasto_mensal = Decimal("0.00")
    gasto_anual = Decimal("0.00")
    for l in lotes:
        custos = _normalizar_custos_lote(l)
        investimento_total += custos["anual_total"]
        em_uso = max(0, int(l.quantidade_total or 0) - int(l.quantidade_disponivel or 0))
        if em_uso > 0:
            gasto_mensal += (custos["mensal_unit"] * Decimal(em_uso)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            gasto_anual += (custos["anual_unit"] * Decimal(em_uso)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    colaboradores_ativos = len({
        m.usuario_id for m in movs if m.tipo == "atribuicao" and m.usuario_id
    } - {
        m.usuario_id for m in movs if m.tipo != "atribuicao" and m.usuario_id
    })

    # =========================================================================
    # PALETA / ESTILOS PROFISSIONAIS
    # =========================================================================
    BRAND_DARK = "0A2540"   # faixa de título
    BRAND = "1D4ED8"        # cabeçalhos de tabela
    SOFT = "EEF2F7"         # subtítulo / faixas claras
    ZEBRA = "F4F7FB"        # linhas alternadas
    INK = "1F2733"
    MUTED = "5B6B7F"
    GREEN = "1E8E3E"
    RED = "D93025"
    GRAY = "6B7280"

    hair = Side(style="thin", color="DCE3EC")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    f_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color=MUTED)
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_bold = Font(name="Calibri", size=10, bold=True, color=INK)
    f_cell = Font(name="Calibri", size=10, color=INK)
    fill_title = PatternFill("solid", fgColor=BRAND_DARK)
    fill_sub = PatternFill("solid", fgColor=SOFT)
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left", vertical="center")
    a_right = Alignment(horizontal="right", vertical="center")
    a_left_ind = Alignment(horizontal="left", vertical="center", indent=1)
    BRL = 'R$ #,##0.00'

    def faixa_titulo(ws, ncols, titulo, subtitulo):
        last = get_column_letter(ncols)
        ws.merge_cells(f"A1:{last}1")
        c = ws["A1"]
        c.value = titulo
        c.font = f_title
        c.fill = fill_title
        c.alignment = a_left_ind
        ws.row_dimensions[1].height = 34
        ws.merge_cells(f"A2:{last}2")
        c2 = ws["A2"]
        c2.value = subtitulo
        c2.font = f_sub
        c2.fill = fill_sub
        c2.alignment = a_left_ind
        ws.row_dimensions[2].height = 18
        ws.sheet_view.showGridLines = False

    def cabecalho_tabela(ws, row, headers, center_cols=()):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = fill_header
            c.font = f_header
            c.border = border
            c.alignment = a_center if ci in center_cols else a_left
        ws.row_dimensions[row].height = 22

    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")

    wb = Workbook()

    # =========================================================================
    # ABA 1 — RESUMO (capa executiva com KPIs)
    # =========================================================================
    wsr = wb.active
    wsr.title = "Resumo"
    faixa_titulo(wsr, 6, f"LICENÇA — {licenca.nome.upper()}",
                 f"Santa Colomba Agropecuária  ·  Relatório gerado em {gerado}")

    # Bloco de informações
    info = [
        ("Fornecedor", licenca.fornecedor.nome if licenca.fornecedor else "—"),
        ("Centro de Custo padrão", licenca.centro_custo.departamento if licenca.centro_custo else "—"),
        ("Nº de lotes", len(lotes)),
        ("Colaboradores com licença ativa", colaboradores_ativos),
    ]
    r = 4
    for label, val in info:
        cl = wsr.cell(row=r, column=1, value=label)
        cl.font = f_bold
        cl.alignment = a_left
        wsr.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cv = wsr.cell(row=r, column=2, value=val)
        cv.font = f_cell
        cv.alignment = a_left
        r += 1
    r += 1

    # KPIs (cartões 2 linhas: rótulo + valor)
    kpis = [
        ("ASSENTOS TOTAIS", qtd_total, "334155", None),
        ("EM USO", qtd_em_uso, GREEN, None),
        ("DISPONÍVEIS", qtd_disp, BRAND, None),
        ("OCUPAÇÃO", pct_uso / 100, "7C3AED", "0%"),
        ("INVESTIMENTO TOTAL", float(investimento_total), "0A2540", BRL),
        ("GASTO MENSAL", float(gasto_mensal), "EA580C", BRL),
    ]
    for idx, (lbl, val, color, fmt) in enumerate(kpis):
        col = idx + 1
        cl = wsr.cell(row=r, column=col, value=lbl)
        cl.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        cl.fill = PatternFill("solid", fgColor=color)
        cl.alignment = a_center
        cl.border = border
        cv = wsr.cell(row=r + 1, column=col, value=val)
        cv.font = Font(name="Calibri", size=16, bold=True, color=color)
        cv.fill = PatternFill("solid", fgColor="F4F6F9")
        cv.alignment = a_center
        cv.border = border
        if fmt:
            cv.number_format = fmt
    wsr.row_dimensions[r].height = 16
    wsr.row_dimensions[r + 1].height = 30
    for c in range(1, 7):
        wsr.column_dimensions[get_column_letter(c)].width = 20

    # =========================================================================
    # ABA 2 — USUÁRIOS E ALOCAÇÕES
    # =========================================================================
    ws1 = wb.create_sheet(title="Usuarios e Alocacoes")
    headers_ws1 = [
        "Ação", "Colaborador", "E-mail", "Matrícula", "Lote",
        "Periodicidade", "Centro de Custo Destino", "Nº CC",
        "Custo Mensal", "Custo Anual", "Responsável", "Data", "Hora",
    ]
    faixa_titulo(ws1, len(headers_ws1), f"USUÁRIOS E ALOCAÇÕES — {licenca.nome.upper()}",
                 f"{len(movs)} movimentação(ões) registrada(s)")
    hr = 4
    cabecalho_tabela(ws1, hr, headers_ws1, center_cols=(1, 5, 6, 8, 9, 10, 12, 13))
    row = hr + 1
    tot_mensal_mov = Decimal("0.00")
    tot_anual_mov = Decimal("0.00")
    for i, mov in enumerate(movs):
        mensal = anual = Decimal("0.00")
        if mov.lote:
            custos = _normalizar_custos_lote(mov.lote)
            mensal, anual = custos["mensal_unit"], custos["anual_unit"]
        is_saida = mov.tipo == "atribuicao"
        valores = [
            "Saída" if is_saida else "Devolução",
            mov.usuario.nome if mov.usuario else "—",
            (mov.usuario.email or "—") if mov.usuario else "—",
            (mov.usuario.matricula or "—") if mov.usuario else "—",
            f"#{mov.lote.pk}" if mov.lote else "—",
            mov.lote.get_periodicidade_display() if mov.lote else "—",
            mov.centro_custo_destino.departamento if mov.centro_custo_destino else "—",
            mov.centro_custo_destino.numero if mov.centro_custo_destino else "—",
            float(mensal), float(anual),
            mov.criado_por.username if mov.criado_por else "—",
            mov.created_at.strftime("%d/%m/%Y") if mov.created_at else "—",
            mov.created_at.strftime("%H:%M") if mov.created_at else "—",
        ]
        zebra = (i % 2 == 1)
        for ci, val in enumerate(valores, 1):
            c = ws1.cell(row=row, column=ci, value=val)
            c.border = border
            c.font = f_cell
            if ci == 1:  # badge de ação colorido
                c.fill = PatternFill("solid", fgColor=GREEN if is_saida else GRAY)
                c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                c.alignment = a_center
            elif ci in (9, 10):
                c.number_format = BRL
                c.alignment = a_right
                if zebra:
                    c.fill = fill_zebra
            elif ci in (5, 6, 8, 12, 13):
                c.alignment = a_center
                if zebra:
                    c.fill = fill_zebra
            else:
                c.alignment = a_left
                if zebra:
                    c.fill = fill_zebra
        if is_saida:
            tot_mensal_mov += mensal
            tot_anual_mov += anual
        row += 1
    # rodapé de totais
    if movs:
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ct = ws1.cell(row=row, column=1, value="TOTAL (saídas ativas)")
        ct.font = f_bold
        ct.alignment = a_right
        ct.fill = fill_sub
        cm = ws1.cell(row=row, column=9, value=float(tot_mensal_mov))
        ca = ws1.cell(row=row, column=10, value=float(tot_anual_mov))
        for cc_ in (cm, ca):
            cc_.font = f_bold
            cc_.number_format = BRL
            cc_.alignment = a_right
            cc_.fill = fill_sub
        for col in range(1, len(headers_ws1) + 1):
            ws1.cell(row=row, column=col).border = border
    widths_ws1 = [12, 28, 30, 14, 9, 15, 28, 10, 15, 15, 18, 12, 9]
    for i, w in enumerate(widths_ws1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = f"A{hr + 1}"

    # =========================================================================
    # ABA 3 — CENTROS DE CUSTO
    # =========================================================================
    ws2 = wb.create_sheet(title="Centros de Custo")
    headers_ws2 = ["Departamento / Centro de Custo", "Qtd. Assentos", "Gasto Consolidado"]
    faixa_titulo(ws2, len(headers_ws2), f"CENTROS DE CUSTO — {licenca.nome.upper()}",
                 "Distribuição de assentos e custo por centro de custo")
    hr = 4
    cabecalho_tabela(ws2, hr, headers_ws2, center_cols=(2, 3))
    row = hr + 1
    total_assentos = 0
    total_gasto = Decimal("0.00")
    for i, cc in enumerate(cc_list):
        qtd = cc.get("qtd", 0) or 0
        total = Decimal(cc.get("total", 0) or 0)
        zebra = (i % 2 == 1)
        c1 = ws2.cell(row=row, column=1, value=cc.get("nome", "—"))
        c2 = ws2.cell(row=row, column=2, value=qtd)
        c3 = ws2.cell(row=row, column=3, value=float(total))
        c1.alignment, c2.alignment, c3.alignment = a_left, a_center, a_right
        c3.number_format = BRL
        for c in (c1, c2, c3):
            c.border = border
            c.font = f_cell
            if zebra:
                c.fill = fill_zebra
        total_assentos += qtd
        total_gasto += total
        row += 1
    c1 = ws2.cell(row=row, column=1, value="TOTAL")
    c2 = ws2.cell(row=row, column=2, value=total_assentos)
    c3 = ws2.cell(row=row, column=3, value=float(total_gasto))
    c1.alignment, c2.alignment, c3.alignment = a_left, a_center, a_right
    c3.number_format = BRL
    for c in (c1, c2, c3):
        c.font = f_bold
        c.border = border
        c.fill = fill_sub
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 22
    ws2.freeze_panes = f"A{hr + 1}"

    # =========================================================================
    # ABA 4 — LOTES
    # =========================================================================
    ws3 = wb.create_sheet(title="Lotes")
    headers_ws3 = [
        "Lote", "Data Compra", "Pedido / NF", "Periodicidade",
        "Qtd. Total", "Qtd. Disponível", "Custo Ciclo/Lic.",
        "Mensal/Lic.", "Anual/Lic.", "Mensal do Lote", "Anual do Lote",
    ]
    faixa_titulo(ws3, len(headers_ws3), f"LOTES — {licenca.nome.upper()}",
                 f"{len(lotes)} lote(s) · investimento anual total R$ {investimento_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    hr = 4
    cabecalho_tabela(ws3, hr, headers_ws3, center_cols=(1, 4, 5, 6, 7, 8, 9, 10, 11))
    row = hr + 1
    for i, lote in enumerate(lotes):
        custos = _normalizar_custos_lote(lote)
        zebra = (i % 2 == 1)
        valores = [
            f"#{lote.pk}",
            lote.data_compra.strftime("%d/%m/%Y") if lote.data_compra else "—",
            lote.numero_pedido or "—",
            lote.get_periodicidade_display() if lote.periodicidade else "—",
            lote.quantidade_total or 0,
            lote.quantidade_disponivel or 0,
            float(custos["unitario_ciclo"]),
            float(custos["mensal_unit"]),
            float(custos["anual_unit"]),
            float(custos["mensal_total"]),
            float(custos["anual_total"]),
        ]
        for ci, val in enumerate(valores, 1):
            c = ws3.cell(row=row, column=ci, value=val)
            c.border = border
            c.font = f_cell
            if ci in (7, 8, 9, 10, 11):
                c.number_format = BRL
                c.alignment = a_right
            elif ci in (1, 4, 5, 6):
                c.alignment = a_center
            else:
                c.alignment = a_left
            if zebra and ci not in ():
                c.fill = fill_zebra
        row += 1
    widths_ws3 = [10, 14, 18, 16, 12, 14, 16, 15, 15, 16, 16]
    for i, w in enumerate(widths_ws3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = f"A{hr + 1}"

    # =========================================================================
    # RESPOSTA
    # =========================================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    slug = licenca.nome.replace(" ", "_").replace("/", "-")
    nome_arquivo = f"licenca_{licenca.pk}_{slug}_{timezone.localtime():%Y%m%d}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response

# ============ MOVIMENTAÇÕES ============

@login_required
def mov_licenca_list(request):
    """
    Listagem de Movimentações de Licenças com filtros e paginação.
    """
    # Filtros
    q = (request.GET.get("q") or "").strip()
    tipo = (request.GET.get("tipo") or "").strip()

    # QuerySet Otimizado
    qs = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario", "centro_custo_destino")
        .order_by("-created_at")
    )

    if q:
        qs = qs.filter(
            Q(licenca__nome__icontains=q) | 
            Q(usuario__nome__icontains=q)
        )
    
    # Validação do Tipo (segurança)
    valid_types = [choice[0] for choice in MovimentacaoLicenca._meta.get_field("tipo").choices]
    if tipo in valid_types:
        qs = qs.filter(tipo=tipo)

    # Paginação (Padrão 20 itens)
    try:
        per_page = int(request.GET.get("pp", 20))
        per_page = max(10, min(per_page, 100)) # Limites de segurança
    except ValueError:
        per_page = 20

    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_num)

    # Preserva filtros na paginação
    get_copy = request.GET.copy()
    get_copy.pop("page", None)
    qs_keep = get_copy.urlencode()

    context = {
        "movs": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "qs_keep": qs_keep,
        "q": q,
        "tipo": tipo,
        "tipos": MovimentacaoLicenca._meta.get_field("tipo").choices,
        "total": qs.count()
    }

    return render(request, "front/licencas/mov_licenca_list.html", context)

@login_required
def mov_licenca_form(request):
    initial = {}
    if "licenca" in request.GET: initial["licenca"] = request.GET.get("licenca")
    if "usuario" in request.GET: initial["usuario"] = request.GET.get("usuario")

    if request.method == "POST":
        form = MovimentacaoLicencaForm(request.POST)
        if form.is_valid():
            try:
                mov = form.save(user=request.user)
                
                # Feedback detalhado
                lote_txt = f"Lote #{mov.lote.pk}" if mov.lote else "N/A"
                cc_txt = mov.centro_custo_destino.departamento if mov.centro_custo_destino else "N/A"
                
                messages.success(request, f"{mov.get_tipo_display()} realizada. Estoque: {lote_txt} | Custo: {cc_txt}")
                return redirect("licenca_list")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
    else:
        form = MovimentacaoLicencaForm(initial=initial)

    # JSON para Select2 (Apenas lotes com saldo)
    lotes_qs = LicencaLote.objects.filter(quantidade_disponivel__gt=0).values(
        'id', 'licenca_id', 'quantidade_disponivel', 'numero_pedido', 'data_compra'
    )
    
    lotes_dict = {}
    for l in lotes_qs:
        lid = str(l['licenca_id'])
        if lid not in lotes_dict: lotes_dict[lid] = []
        dt = l['data_compra'].strftime('%d/%m/%Y') if l['data_compra'] else "-"
        txt = f"Lote #{l['id']} - Disp: {l['quantidade_disponivel']} ({dt})"
        lotes_dict[lid].append({'id': l['id'], 'text': txt})

    context = {
        "form": form,
        "lotes_json": lotes_dict,
        "pre_selected_lote": request.POST.get("lote_id_select") or ""
    }
    return render(request, "front/licencas/mov_licenca_form.html", context)
# --- LISTA DE LOTES ---
@login_required
def licenca_lote_list(request):
    """
    Lista de Lotes com busca avançada e layout otimizado.
    """
    q = request.GET.get("q", "").strip()
    
    # QueryBase com select_related para evitar N+1 queries
    qs = (
        LicencaLote.objects
        .select_related("licenca", "fornecedor", "centro_custo")
        .order_by("-created_at")
    )

    # Filtro Textual
    if q:
        qs = qs.filter(
            Q(licenca__nome__icontains=q) | 
            Q(numero_pedido__icontains=q) | 
            Q(observacao__icontains=q)
        )

    return render(request, "front/licencas/licenca_lote_list.html", {
        "lotes": qs,
        "q": q
    })

@login_required
@transaction.atomic
def licenca_lote_form(request, pk=None):
    """
    View Inteligente para Gestão de Lotes.
    Calcula automaticamente a disponibilidade baseada na entrada.
    """
    obj = get_object_or_404(LicencaLote, pk=pk) if pk else None

    if request.method == "POST":
        form = LicencaLoteForm(request.POST, instance=obj)
        if form.is_valid():
            lote = form.save(commit=False)
            
            # --- LÓGICA DE SALDO ---
            if not obj:
                # Novo Lote: Disponível = Total
                lote.quantidade_disponivel = lote.quantidade_total
                lote.criado_por = request.user # Auditoria
            else:
                # Edição: Ajusta o disponível pela diferença do total
                # Ex: Tinha 10 (8 disp, 2 uso). Editou total para 15 (+5).
                # Novo Disp = 8 + 5 = 13. Usados continuam 2.
                diff = lote.quantidade_total - obj.quantidade_total
                lote.quantidade_disponivel = obj.quantidade_disponivel + diff
            
            lote.atualizado_por = request.user # Auditoria
            lote.save()
            
            msg = f"Lote #{lote.pk} atualizado com sucesso!" if obj else "Lote criado com sucesso!"
            messages.success(request, msg)
            return redirect("licenca_lote_list")
        else:
            messages.error(request, "Verifique os erros no formulário abaixo.")
    else:
        form = LicencaLoteForm(instance=obj)

    return render(request, "front/licencas/licenca_lote_form.html", {
        "form": form,
        "obj": obj
    })




### exportações 

