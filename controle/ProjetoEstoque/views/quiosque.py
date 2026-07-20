"""
quiosque.py — Módulo Quiosque (app Android corporativo).

Dois grupos de views:

  • API do dispositivo (/api/quiosque/...): consumidas pelo APK Android.
    São @csrf_exempt e autenticadas por TOKEN do device (nunca @login_required).
    Contrato em docs/MODULO_QUIOSQUE_ANDROID.md (Seção 4).

  • Dashboard interno (/quiosque/...): telas web para o TI gerenciar os celulares.
    Essas sim são @login_required.

Toda a regra de negócio fica em services/quiosque_service.py.
"""
import json
from datetime import date
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from services import quiosque_service as qs


def _parse_dia_param(raw: str) -> "date | None":
    """Converte `?dia=YYYY-MM-DD` num `date`, só se estiver dentro da janela de
    retenção do histórico (RETENCAO_DIAS). Qualquer entrada inválida/fora da
    janela é ignorada silenciosamente (cai para a visão padrão) — os links da
    própria tela nunca geram um valor fora do range, então isto só protege
    contra URL editada manualmente."""
    if not raw:
        return None
    try:
        dia = date.fromisoformat(raw)
    except (TypeError, ValueError):
        return None
    hoje = timezone.localdate()
    if dia > hoje or (hoje - dia).days >= qs.RETENCAO_DIAS:
        return None
    return dia


# ──────────────────────────────────────────────────────────────────────────────
# Infra da API do dispositivo
# ──────────────────────────────────────────────────────────────────────────────

def _json_body(request) -> dict:
    """Lê o corpo JSON da requisição; cai para POST form se não for JSON."""
    if request.body:
        try:
            data = json.loads(request.body.decode("utf-8"))
            if isinstance(data, dict):
                return data
        except (ValueError, UnicodeDecodeError):
            pass
    return {k: v for k, v in request.POST.items()}


def kiosk_token_required(view):
    """Autentica o device por `Authorization: Bearer <token>` + `X-Device-UUID`."""
    @wraps(view)
    def wrapper(request, *args, **kwargs):
        auth = request.headers.get("Authorization", "")
        token = auth[7:].strip() if auth[:7].lower() == "bearer " else ""
        device_uuid = request.headers.get("X-Device-UUID", "").strip()
        device = qs.autenticar(token, device_uuid)
        if device is None:
            return JsonResponse({"ok": False, "erro": "Não autorizado."}, status=401)
        request.kiosk_device = device
        return view(request, *args, **kwargs)
    return csrf_exempt(wrapper)


# ──────────────────────────────────────────────────────────────────────────────
# API do dispositivo
# ──────────────────────────────────────────────────────────────────────────────

@csrf_exempt
def kiosk_enroll(request):
    """POST /api/quiosque/enroll/ — matrícula do device (código de uso único)."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método não permitido."}, status=405)

    dados = _json_body(request)
    try:
        res = qs.enroll(codigo_matricula=dados.get("codigo_matricula", ""), dados=dados)
    except qs.EnrollConflict as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=409)
    except ValueError as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=400)
    except Exception:  # noqa: BLE001
        return JsonResponse({"ok": False, "erro": "Falha no enrollment."}, status=500)

    device = res["device"]
    return JsonResponse({
        "ok": True,
        "device_uuid": str(device.device_uuid),
        "token": res["token"],
        "config": res["config"],
    })


@kiosk_token_required
def kiosk_checkin(request):
    """POST /api/quiosque/checkin/ — telemetria periódica (heartbeat)."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método não permitido."}, status=405)
    resp = qs.registrar_checkin(request.kiosk_device, _json_body(request), request)
    return JsonResponse(resp)


@kiosk_token_required
def kiosk_config(request):
    """GET /api/quiosque/config/ — configuração atual do device."""
    return JsonResponse({"ok": True, "config": qs.config_dict(request.kiosk_device)})


@kiosk_token_required
def kiosk_comando_ack(request, pk: int):
    """POST /api/quiosque/comando/<id>/ack/ — confirmação de execução de comando."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método não permitido."}, status=405)
    dados = _json_body(request)
    ok = qs.registrar_ack_comando(
        request.kiosk_device, pk, dados.get("status", "executado"), dados.get("detalhe", "")
    )
    return JsonResponse({"ok": ok}, status=200 if ok else 404)


@kiosk_token_required
def kiosk_atualizacao_apk(request):
    """GET /api/quiosque/atualizacao/apk/ — download do .apk atual para
    auto-atualização de um aparelho JÁ matriculado (Device Owner).

    Diferente de `kiosk_instalador_download`: aquela é a rota de provisionamento
    manual por QR Code (aparelho novo, sem token, token opaco de curta duração);
    esta é para a frota já em campo, autenticada do mesmo jeito que o /checkin/
    (Bearer + X-Device-UUID) — sem necessidade de emitir mais um token por download.
    """
    from django.http import FileResponse

    atual = qs.apk_atual()
    if atual is None:
        return JsonResponse({"ok": False, "erro": "Nenhum instalador disponível."}, status=404)
    caminho = qs.apk_dir() / atual["nome"]
    return FileResponse(
        open(caminho, "rb"),
        content_type="application/vnd.android.package-archive",
        filename=atual["nome"],
    )


def kiosk_instalador_download(request, token):
    """GET /quiosque/instalador/baixar/<token>/ — download do .apk.

    Rota PÚBLICA de propósito: quem baixa é o celular ainda sem o app instalado,
    escaneando o QR Code, sem sessão logada no sistema. A proteção é o token
    (opaco, validade curta, revogável), não @login_required — mesmo modelo do
    código de matrícula em kiosk_enroll. Resposta sempre genérica (404) para
    token inválido, expirado ou revogado, para não dar pista a quem tentar
    adivinhar tokens.
    """
    from django.http import FileResponse, Http404

    if request.method != "GET":
        raise Http404()

    link = qs.resolver_instalador(token)
    if link is None:
        raise Http404()
    caminho = qs.caminho_instalador(link.nome_arquivo)
    if caminho is None:
        raise Http404()

    qs.registrar_download_instalador(link, request.META.get("REMOTE_ADDR"))
    return FileResponse(
        open(caminho, "rb"),
        as_attachment=True,
        filename=link.nome_arquivo,
        content_type="application/vnd.android.package-archive",
    )


# ──────────────────────────────────────────────────────────────────────────────
# Dashboard interno (TI)
# ──────────────────────────────────────────────────────────────────────────────

def _resolver_itens_por_serial(devices):
    """Associa a cada device o Item do estoque cujo número de série bate com o
    serial do aparelho. Anexa `d.equip_serial` (Item ou None) em cada device.
    Uma única query para todos os seriais."""
    from ProjetoEstoque.models import Item

    seriais = {(d.serial or "").strip() for d in devices if (d.serial or "").strip()}
    mapa = {}
    if seriais:
        for it in Item.objects.filter(numero_serie__in=seriais).only("pk", "nome", "numero_serie"):
            mapa.setdefault((it.numero_serie or "").strip(), it)
    for d in devices:
        s = (d.serial or "").strip()
        d.equip_serial = mapa.get(s) if s else None
    return mapa


def _anexar_matricula(devices):
    """Anexa a cada device a matrícula ATUAL (a mais recente usada no enrollment).
    Mostra a descrição/código da matrícula para facilitar a identificação. Reflete
    automaticamente a última matrícula recebida quando o aparelho é rematriculado."""
    from ProjetoEstoque.models import KioskMatricula

    ids = [d.pk for d in devices if d.pk]
    mapa = {}
    if ids:
        # Ordenado por device + mais recente; o primeiro de cada device é o atual.
        for m in (KioskMatricula.objects
                  .filter(device_id__in=ids)
                  .order_by("device_id", "-usado_em", "-criado_em")):
            mapa.setdefault(m.device_id, m)
    for d in devices:
        d.matricula_atual = mapa.get(d.pk)
    return mapa


# ─────────────────────────────────────────────────────────────
# Exportação do histórico de check-ins (Excel)
# ─────────────────────────────────────────────────────────────

# Teto de linhas do relatório — defesa contra um device configurado com
# intervalo de check-in muito baixo (mínimo aceito: 5s) gerando dezenas de
# milhares de linhas numa janela de 5 dias e travando a geração do arquivo.
_CHECKINS_XLSX_MAX_LINHAS = 20000


def _checkins_xlsx(device, checkins: list, periodo_label: str, dia):
    """Gera o .xlsx do histórico de check-ins de UM dispositivo, no mesmo
    padrão visual do relatório de monitoração PRTG (`_monitoracao_xlsx` em
    equipamentos.py) — mantém a identidade visual dos relatórios do sistema."""
    from datetime import datetime as _datetime
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND_DARK, BRAND, SOFT, ZEBRA = "0B3D6E", "0071E3", "E5F0FB", "F4F9FE"
    INK = "1F2733"
    hair = Side(style="thin", color="CFE0F2")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    f_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color="5B6B7F")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Calibri", size=10, color=INK)
    fill_title = PatternFill("solid", fgColor=BRAND_DARK)
    fill_sub = PatternFill("solid", fgColor=SOFT)
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left", vertical="center")
    a_left_ind = Alignment(horizontal="left", vertical="center", indent=1)
    dt_fmt = "DD/MM/YYYY HH:MM:SS"

    telemetria_wifi = bool(device.telemetria_wifi)
    header = ["#", "Coletado em", "Recebido em", "Bateria %", "Carregando", "Rede", "Wi-Fi (SSID)"]
    if telemetria_wifi:
        header += ["RSSI Wi-Fi (dBm)", "Nível sinal (0-4)", "Velocidade (Mbps)", "Banda (GHz)"]
    header += ["Online", "Latitude", "Longitude", "Precisão (m)"]
    ncols = len(header)
    center_cols = {1, 4, 5, 8, 9, 10, 11} if telemetria_wifi else {1, 4, 5, 8}

    wb = Workbook()
    ws = wb.active
    ws.title = "Check-ins"
    ws.sheet_view.showGridLines = False

    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = "HISTÓRICO DE CHECK-INS — QUIOSQUE"
    c.font = f_title; c.fill = fill_title; c.alignment = a_left_ind
    ws.row_dimensions[1].height = 34

    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
    nome_device = device.apelido or device.modelo or "Dispositivo"
    ws.merge_cells(f"A2:{last}2")
    c2 = ws["A2"]
    c2.value = f"{nome_device}  ·  {periodo_label}  ·  {len(checkins)} registro(s)  ·  gerado em {gerado}"
    c2.font = f_sub; c2.fill = fill_sub; c2.alignment = a_left_ind
    ws.row_dimensions[2].height = 18

    HEADER_ROW = 3
    for ci, h in enumerate(header, 1):
        cc = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cc.fill = fill_header; cc.font = f_header; cc.border = border
        cc.alignment = a_center if ci in center_cols else a_left
    ws.row_dimensions[HEADER_ROW].height = 26

    row = HEADER_ROW + 1
    for i, c_ in enumerate(checkins, start=1):
        coletado = timezone.localtime(c_.quando).replace(tzinfo=None)
        recebido = timezone.localtime(c_.registrado_em).replace(tzinfo=None)
        valores = [i, coletado, recebido,
                   c_.bateria if c_.bateria is not None else "—",
                   "Sim" if c_.carregando else "Não",
                   c_.rede or "—", c_.ssid or "—"]
        if telemetria_wifi:
            valores += [
                c_.wifi_rssi_dbm if c_.wifi_rssi_dbm is not None else "—",
                c_.wifi_nivel if c_.wifi_nivel is not None else "—",
                c_.wifi_velocidade_mbps if c_.wifi_velocidade_mbps is not None else "—",
                c_.wifi_banda_ghz or "—",
            ]
        valores += [
            "Sim" if c_.online else "Não",
            c_.latitude if c_.latitude is not None else "—",
            c_.longitude if c_.longitude is not None else "—",
            round(c_.precisao_m) if c_.precisao_m is not None else "—",
        ]
        zebra = (i % 2 == 0)
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = border
            cell.font = f_cell
            cell.alignment = a_center if ci in center_cols else a_left
            if ci in (2, 3) and isinstance(val, _datetime):
                cell.number_format = dt_fmt
                cell.alignment = a_center
            elif zebra:
                cell.fill = fill_zebra
        row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last}{max(row - 1, HEADER_ROW)}"

    widths = {}
    for r_ in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        for idx, val in enumerate(r_, start=1):
            widths[idx] = max(widths.get(idx, 0), len(str(val)) if val is not None else 0)
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 11), 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    now = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    slug = "".join(ch if ch.isalnum() else "-" for ch in nome_device.lower()).strip("-") or "dispositivo"
    sufixo_dia = f"_{dia.isoformat()}" if dia else ""
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="checkins_{slug}{sufixo_dia}_{now}.xlsx"'
    return resp


@login_required
def quiosque_dashboard(request):
    from ProjetoEstoque.models import KioskDevice, KioskMatricula

    f_status = (request.GET.get("status") or "").strip()  # online | offline | sem_local
    q = (request.GET.get("q") or "").strip()

    devices = list(KioskDevice.objects.filter(ativo=True).select_related("item"))

    kpi = {
        "total": len(devices),
        "online": sum(1 for d in devices if d.online),
        "offline": sum(1 for d in devices if not d.online),
        "sem_local": sum(1 for d in devices if not d.tem_localizacao),
        "matriculas": KioskMatricula.objects.filter(usado=False).count(),
    }

    if f_status == "online":
        devices = [d for d in devices if d.online]
    elif f_status == "offline":
        devices = [d for d in devices if not d.online]
    elif f_status == "sem_local":
        devices = [d for d in devices if not d.tem_localizacao]
    if q:
        ql = q.lower()
        devices = [
            d for d in devices
            if ql in (d.apelido or "").lower() or ql in (d.modelo or "").lower()
            or ql in (d.serial or "").lower() or ql in (d.fabricante or "").lower()
            or ql in (d.mac or "").lower()
        ]

    devices.sort(key=lambda d: (d.online is False, (d.apelido or d.modelo or "").lower()))
    _resolver_itens_por_serial(devices)
    _anexar_matricula(devices)

    return render(request, "front/quiosque/quiosque_dashboard.html", {
        "devices": devices,
        "kpi": kpi,
        "f_status": f_status,
        "f_q": q,
        "total_filtrado": len(devices),
        "offline_apos": KioskDevice.OFFLINE_APOS,
    })


@login_required
def quiosque_detalhe(request, pk: int):
    from ProjetoEstoque.models import KioskDevice

    device = get_object_or_404(KioskDevice.objects.select_related("item"), pk=pk)

    # Equipamento do estoque resolvido pelo número de série do aparelho.
    _resolver_itens_por_serial([device])
    # Matrícula atual (descrição que identifica o aparelho).
    _anexar_matricula([device])

    dias_disponiveis = qs.dias_disponiveis_checkin(device)
    dia_selecionado = _parse_dia_param(request.GET.get("dia"))

    # Ordenado pelo instante REAL de coleta (coletado_em), não pela chegada ao
    # servidor (Meta.ordering = -registrado_em): quando o app entrega uma fila
    # offline em rajada, vários check-ins chegam quase juntos e a ordem de
    # chegada não reflete a ordem real dos eventos — mesmo problema já corrigido
    # para o traço do mapa em `montar_trilha`. Sem isso, a tabela mostraria os
    # horários (coluna "Quando" = mesma coalescência) fora de ordem cronológica.
    from django.db.models.functions import Coalesce

    checkins = device.checkins.annotate(_quando=Coalesce("coletado_em", "registrado_em"))
    if dia_selecionado:
        inicio, fim = qs.intervalo_dia_local(dia_selecionado)
        checkins = checkins.filter(_quando__gte=inicio, _quando__lt=fim)
    checkins = checkins.order_by("-_quando")

    paginator = Paginator(checkins, 30)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    comandos = device.comandos.all()[:20]

    # Traço de rota (deslocamento) para o mapa do detalhe. A montagem fica no
    # service: ordena por horário real de coleta, descarta fixes de GPS ruins e
    # saltos impossíveis — deixando o caminho fiel ao percorrido (ordem antigo→recente).
    # Com um dia selecionado, cobre o dia inteiro (não só a janela recente).
    trilha = qs.montar_trilha(device, dia=dia_selecionado)
    mapa = qs.montar_mapa_dict(device, trilha, dia=dia_selecionado)
    # "Ao vivo" só faz sentido enquanto o dia em exibição ainda pode receber
    # novos check-ins (sem filtro, ou filtrando o próprio dia de hoje) — um dia
    # passado é histórico fechado, não precisa (nem deve) ficar sondando o servidor.
    pode_atualizar_ao_vivo = dia_selecionado is None or dia_selecionado == timezone.localdate()

    resumo_dia = qs.montar_resumo_dia(device, dia_selecionado, trilha) if dia_selecionado else None

    # Geolocalização dos check-ins exibidos NESTA página da tabela. Permite focar
    # o ponto EXATO no mapa ao clicar na linha (chaveado pelo id do check-in).
    # Serializado via json_script no template → sempre com ponto decimal, sem o
    # problema de localização pt-BR (vírgula) que quebraria o parseFloat no JS.
    geo_pagina = {
        c.pk: {
            "lat": c.latitude,
            "lon": c.longitude,
            "precisao": c.precisao_m,
            "quando": timezone.localtime(c.quando).strftime("%d/%m/%Y %H:%M:%S"),
            "bateria": c.bateria,
            "rede": c.rede,
            "online": c.online,
        }
        for c in page_obj.object_list
        if c.latitude is not None and c.longitude is not None
    }

    # Percentuais/limiares para as barras de RAM e armazenamento (display-only,
    # mesmo padrão de p.atrasado/p.atencao calculados na view para preventivas).
    # ram_pouca já vem pronto do Android (MemoryInfo.lowMemory) — não recalculamos
    # limiar de RAM aqui. Para armazenamento não há sinal do SO, então usamos o
    # limiar sugerido no informe do app (crítico <1GB livre, atenção <2GB livre).
    ram_pct = None
    if device.ram_total_mb and device.ram_usada_mb is not None:
        ram_pct = min(100, round(device.ram_usada_mb / device.ram_total_mb * 100))

    armazenamento_pct = None
    armazenamento_nivel = ""
    armazenamento_livre_gb = None
    armazenamento_total_gb = None
    if device.armazenamento_total_mb and device.armazenamento_usado_mb is not None:
        armazenamento_pct = min(100, round(device.armazenamento_usado_mb / device.armazenamento_total_mb * 100))
        armazenamento_total_gb = round(device.armazenamento_total_mb / 1024, 1)
    if device.armazenamento_livre_mb is not None:
        armazenamento_livre_gb = round(device.armazenamento_livre_mb / 1024, 1)
        if device.armazenamento_livre_mb < 1024:
            armazenamento_nivel = "danger"
        elif device.armazenamento_livre_mb < 2048:
            armazenamento_nivel = "warn"

    return render(request, "front/quiosque/quiosque_detalhe.html", {
        "device": device,
        "page_obj": page_obj,
        "checkins": page_obj.object_list,
        "total_checkins": paginator.count,
        "comandos": comandos,
        "mapa": mapa,
        "geo_pagina": geo_pagina,
        "offline_apos": KioskDevice.OFFLINE_APOS,
        "ram_pct": ram_pct,
        "armazenamento_pct": armazenamento_pct,
        "armazenamento_nivel": armazenamento_nivel,
        "armazenamento_livre_gb": armazenamento_livre_gb,
        "armazenamento_total_gb": armazenamento_total_gb,
        "dias_disponiveis": dias_disponiveis,
        "dia_selecionado": dia_selecionado,
        "resumo_dia": resumo_dia,
        "pode_atualizar_ao_vivo": pode_atualizar_ao_vivo,
        "retencao_dias": qs.RETENCAO_DIAS,
    })


@login_required
def quiosque_mapa_atualizar(request, pk: int):
    """AJAX GET — snapshot atual do device para atualizar a tela de detalhe SEM
    reload: status/bateria/rede/último check-in (barra-resumo) + posição/traço
    do mapa. É o que faz o "minimapa" do detalhe deixar de ser uma foto estática
    do carregamento da página e passar a refletir novos check-ins que cheguem
    enquanto a tela estiver aberta (polling no template, ver quiosque_detalhe.html).
    Respeita o mesmo filtro de dia da página (`?dia=`); histórico de um dia
    passado não muda, então o template nem chama este endpoint nesse caso."""
    from ProjetoEstoque.models import KioskDevice

    device = get_object_or_404(KioskDevice, pk=pk)
    dia = _parse_dia_param(request.GET.get("dia"))

    trilha = qs.montar_trilha(device, dia=dia)
    mapa = qs.montar_mapa_dict(device, trilha, dia=dia)

    return JsonResponse({
        "ok": True,
        "online": device.online,
        "bateria": device.ultima_bateria,
        "rede": device.ultima_rede,
        "ultimo_checkin_label": timezone.localtime(device.ultimo_checkin).strftime("%d/%m/%Y %H:%M") if device.ultimo_checkin else None,
        "ultimo_checkin_ts_ms": int(device.ultimo_checkin.timestamp() * 1000) if device.ultimo_checkin else 0,
        "mapa": mapa,
    })


@login_required
def quiosque_checkins_exportar(request, pk: int):
    """GET — exporta o histórico de check-ins do dispositivo em Excel. Respeita
    o filtro de dia (`?dia=`) quando presente; sem filtro, exporta a janela
    completa de retenção (RETENCAO_DIAS)."""
    from django.db.models.functions import Coalesce
    from ProjetoEstoque.models import KioskDevice

    device = get_object_or_404(KioskDevice, pk=pk)
    dia = _parse_dia_param(request.GET.get("dia"))

    checkins = device.checkins.annotate(_quando=Coalesce("coletado_em", "registrado_em"))
    if dia:
        inicio, fim = qs.intervalo_dia_local(dia)
        checkins = checkins.filter(_quando__gte=inicio, _quando__lt=fim)
        periodo_label = f"Dia {dia.strftime('%d/%m/%Y')}"
    else:
        periodo_label = f"Últimos {qs.RETENCAO_DIAS} dias"
    checkins = list(checkins.order_by("-_quando")[:_CHECKINS_XLSX_MAX_LINHAS])

    return _checkins_xlsx(device, checkins, periodo_label, dia)


@login_required
def quiosque_indicadores(request):
    """Painel gerencial do módulo Quiosque — indicadores consolidados da frota
    para apresentação (RH / gestão de TI): saúde dos aparelhos, adoção do
    provisionamento e composição do parque. Retrato executivo, sem filtro ou
    paginação (o detalhe operacional de cada aparelho fica em quiosque_dashboard)."""
    return render(request, "front/quiosque/quiosque_indicadores.html", qs.montar_indicadores_gerenciais())


@login_required
def quiosque_matriculas(request):
    from django.db.models import Q
    from ProjetoEstoque.models import KioskMatricula, KioskInstaladorLink

    if request.method == "POST":
        descricao = (request.POST.get("descricao") or "").strip()
        try:
            validade = int(request.POST.get("validade_horas") or 72)
        except (TypeError, ValueError):
            validade = 72
        m = qs.criar_matricula(descricao=descricao, validade_horas=validade, user=request.user)
        messages.success(request, f"Código de matrícula gerado: {m.codigo}")
        return redirect(f"{reverse('quiosque_matriculas')}?novo={m.pk}")

    agora = timezone.now()
    # "Disponível" = não usada E ainda válida (sem expiração ou ainda dentro do prazo);
    # "Expirada" = não usada mas fora do prazo. Contagens reais (não limitadas aos 100 exibidos).
    validas = Q(expira_em__isnull=True) | Q(expira_em__gt=agora)
    matriculas = KioskMatricula.objects.select_related("device", "criado_por").all()[:100]

    # Código recém-gerado nesta navegação (?novo=<pk>, setado no redirect acima):
    # usado só para destacar o botão do QR Code na tabela — não altera nenhuma regra.
    try:
        novo_pk = int(request.GET.get("novo") or 0) or None
    except (TypeError, ValueError):
        novo_pk = None

    return render(request, "front/quiosque/quiosque_matriculas.html", {
        "matriculas": matriculas,
        "total": KioskMatricula.objects.count(),
        "disponiveis": KioskMatricula.objects.filter(usado=False).filter(validas).count(),
        "usadas": KioskMatricula.objects.filter(usado=True).count(),
        "expiradas": KioskMatricula.objects.filter(usado=False, expira_em__lte=agora).count(),
        "apk": qs.apk_atual(),
        "versao_registrada": qs.versao_apk_registrada(),
        "instaladores": KioskInstaladorLink.objects.select_related("criado_por").all()[:15],
        "versoes_anteriores": qs.versoes_anteriores(),
        "novo_pk": novo_pk,
    })


@login_required
def quiosque_matricula_excluir(request, pk: int):
    from ProjetoEstoque.models import KioskMatricula

    matricula = get_object_or_404(KioskMatricula, pk=pk)
    if request.method == "POST":
        codigo = matricula.codigo
        # Excluir o código não afeta o dispositivo já matriculado (FK SET_NULL).
        matricula.delete()
        messages.success(request, f"Matrícula {codigo} excluída.")
    return redirect("quiosque_matriculas")


@login_required
def quiosque_matricula_renomear(request, pk: int):
    """Edita a descrição (nome de identificação) de uma matrícula já gerada."""
    from ProjetoEstoque.models import KioskMatricula

    matricula = get_object_or_404(KioskMatricula, pk=pk)
    if request.method == "POST":
        matricula.descricao = (request.POST.get("descricao") or "").strip()[:120]
        matricula.save(update_fields=["descricao"])
        messages.success(request, "Descrição da matrícula atualizada.")
    return redirect("quiosque_matriculas")


@login_required
def quiosque_matricula_qrcode(request, pk: int):
    """AJAX GET — QR Code do código de matrícula, para escanear no app em vez de
    digitar os 8 caracteres na mão. Só é gerado enquanto o código ainda estiver
    disponível (não usado e não expirado): um QR de código morto induziria o
    operador a distribuir algo que o app vai recusar no enroll."""
    from ProjetoEstoque.models import KioskMatricula

    matricula = get_object_or_404(KioskMatricula, pk=pk)
    if matricula.usado or not matricula.esta_valida():
        return JsonResponse({
            "ok": False,
            "erro": "Este código já foi usado ou expirou. Gere um novo código para obter um QR válido.",
        }, status=400)

    return JsonResponse({
        "ok": True,
        "codigo": matricula.codigo,
        "descricao": matricula.descricao,
        "qr_base64": qs.gerar_qrcode_data_uri(matricula.codigo),
        "expira_em": timezone.localtime(matricula.expira_em).strftime("%d/%m/%Y %H:%M") if matricula.expira_em else None,
    })


@login_required
def quiosque_apk_upload(request):
    """POST — upload do instalador (.apk) direto pela tela, substituindo o
    arquivo anterior na pasta protegida (`KIOSK_APK_DIR`). Dispensa copiar o
    arquivo manualmente no servidor a cada nova versão do app. Se
    `version_code`/`version_name` forem informados, já registra a versão nesse
    mesmo passo, habilitando a auto-atualização (Device Owner) dos aparelhos já
    matriculados sem precisar rodar o management command à parte."""
    if request.method != "POST":
        return redirect("quiosque_matriculas")

    arquivo = request.FILES.get("apk")
    if arquivo is None:
        messages.error(request, "Selecione o arquivo .apk para enviar.")
        return redirect("quiosque_matriculas")

    version_code_raw = (request.POST.get("version_code") or "").strip()
    version_name = (request.POST.get("version_name") or "").strip()
    version_code = None
    if version_code_raw:
        try:
            version_code = int(version_code_raw)
            if version_code <= 0:
                raise ValueError
        except ValueError:
            messages.error(request, "O código da versão (version_code) precisa ser um número inteiro positivo.")
            return redirect("quiosque_matriculas")

    try:
        qs.salvar_apk_upload(arquivo, version_code=version_code, version_name=version_name)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("quiosque_matriculas")

    if version_code:
        messages.success(
            request,
            f"Instalador “{arquivo.name}” enviado (versão {version_name or version_code}) e substituiu a "
            "anterior. Auto-atualização habilitada para os aparelhos já matriculados."
        )
    else:
        messages.success(
            request,
            f"Instalador “{arquivo.name}” enviado e substituiu a versão anterior. Informe o código da "
            "versão (version_code) para habilitar a auto-atualização dos aparelhos já matriculados — ou "
            "rode depois “python manage.py assinar_apk_quiosque <version_code> <version_name>”."
        )
    return redirect("quiosque_matriculas")


@login_required
def quiosque_apk_versao_anterior_baixar(request, nome_arquivo: str):
    """GET — baixa um instalador (.apk) arquivado em `versoes_anteriores/`
    (substituído por um upload mais recente na tela de Matrículas). Acesso
    interno do TI autenticado — diferente do link público de provisionamento
    (`kiosk_instalador_download`), que usa token opaco para o celular sem sessão."""
    from django.http import FileResponse, Http404

    caminho = qs.caminho_versao_anterior(nome_arquivo)
    if caminho is None:
        raise Http404()
    return FileResponse(
        open(caminho, "rb"),
        as_attachment=True,
        filename=caminho.name,
        content_type="application/vnd.android.package-archive",
    )


@login_required
def quiosque_instalador_gerar(request):
    """POST AJAX — gera um link de instalação (token de validade curta) do .apk
    atual, com QR Code embutido na resposta. Ver kiosk_instalador_download para
    o outro lado (rota pública que o celular acessa)."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "erro": "Método não permitido."}, status=405)

    try:
        validade = int(request.POST.get("validade_minutos") or 30)
    except (TypeError, ValueError):
        validade = 30

    try:
        resultado = qs.gerar_link_instalador(validade_minutos=validade, user=request.user, request=request)
    except ValueError as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=400)

    link = resultado["link"]
    return JsonResponse({
        "ok": True,
        "pk": link.pk,
        "url": resultado["url"],
        "qr_base64": resultado["qr_base64"],
        "nome_arquivo": link.nome_arquivo,
        "expira_em": timezone.localtime(link.expira_em).strftime("%d/%m/%Y %H:%M"),
        "validade_minutos": resultado["validade_minutos"],
    })


@login_required
def quiosque_instalador_status(request, pk: int):
    """AJAX GET — confere se um link de instalador já gerado ainda está válido
    (não expirou e não foi revogado). Nunca devolve a URL/token: o servidor só
    guarda o hash do token (ver KioskInstaladorLink) — o link e o QR Code em si
    ficam apenas no sessionStorage do navegador que os gerou (botão "Visualizar
    QR Code"), este endpoint só confirma se aquele QR ainda pode ser usado."""
    from ProjetoEstoque.models import KioskInstaladorLink

    link = get_object_or_404(KioskInstaladorLink, pk=pk)
    return JsonResponse({
        "ok": True,
        "valido": link.esta_valido(),
        "revogado": link.revogado,
        "expira_em": timezone.localtime(link.expira_em).strftime("%d/%m/%Y %H:%M"),
    })


@login_required
def quiosque_instalador_revogar(request, pk: int):
    """Revoga antecipadamente um link de instalação (ex.: gerado por engano ou
    após concluir o provisionamento dos aparelhos)."""
    from ProjetoEstoque.models import KioskInstaladorLink

    link = get_object_or_404(KioskInstaladorLink, pk=pk)
    if request.method == "POST":
        link.revogado = True
        link.save(update_fields=["revogado"])
        messages.success(request, "Link de instalação revogado.")
    return redirect("quiosque_matriculas")


@login_required
def quiosque_mapa(request):
    """Mapa com a localização atual de todos os dispositivos ativos."""
    from ProjetoEstoque.models import KioskDevice

    ativos = list(KioskDevice.objects.filter(ativo=True))
    com_local = [d for d in ativos if d.tem_localizacao]
    pontos = [
        {
            "pk": d.pk,
            "nome": d.apelido or d.modelo or "Quiosque",
            "modelo": f"{d.fabricante} {d.modelo}".strip(),
            "lat": d.ultima_latitude,
            "lon": d.ultima_longitude,
            "online": d.online,
            "bateria": d.ultima_bateria,
            "rede": d.ultima_rede,
            "checkin": timezone.localtime(d.ultimo_checkin).strftime("%d/%m/%Y %H:%M") if d.ultimo_checkin else "—",
            "url": reverse("quiosque_detalhe", args=[d.pk]),
        }
        for d in com_local
    ]
    return render(request, "front/quiosque/quiosque_mapa.html", {
        "pontos": pontos,
        "total": len(pontos),
        "online": sum(1 for d in com_local if d.online),
        "sem_local": len(ativos) - len(com_local),
    })


@login_required
def quiosque_config_editar(request, pk: int):
    from ProjetoEstoque.models import KioskDevice

    device = get_object_or_404(KioskDevice, pk=pk)

    if request.method == "POST":
        device.apelido = (request.POST.get("apelido") or "").strip()[:120]
        device.wifi_only = request.POST.get("wifi_only") == "on"
        device.mensagem_quiosque = (request.POST.get("mensagem_quiosque") or "").strip()[:200]
        device.telemetria_wifi = request.POST.get("telemetria_wifi") == "on"
        # wifi_ssid/wifi_senha (rede provisionada pelo servidor) ficam de fora do
        # painel de propósito: exigiria o TI guardar a senha real da rede aqui, e a
        # infraestrutura de Wi-Fi da empresa usa autenticação/cadastro no Meraki —
        # incompatível com esse modelo. Sem SSID configurado, o app nunca provisiona
        # nem mexe na rede do aparelho (ver INFORME §4.1); só a telemetria é usada.
        try:
            # Faixa aceita pelo app: [5, 300]s. 5s = tempo real; 300s = economia de bateria/dados.
            device.intervalo_checkin_seg = min(300, max(5, int(request.POST.get("intervalo_checkin_seg") or 300)))
        except (TypeError, ValueError):
            device.intervalo_checkin_seg = 300
        # Apps permitidos = marcados no inventário (checkbox) + pacotes extras
        # digitados (avançado). A chave é o package name; deduplica preservando a
        # ordem (marcados primeiro). Guarda o `pkg`, nunca o nome amigável.
        marcados = request.POST.getlist("app_pkg")
        raw = (request.POST.get("apps_extra") or "")
        manuais = [a.strip() for a in raw.replace(",", "\n").splitlines() if a.strip()]
        apps = []
        for p in marcados + manuais:
            p = p.strip()[:255]
            if p and p not in apps:
                apps.append(p)
        device.apps_permitidos = apps
        device.config_versao = (device.config_versao or 1) + 1
        device.save()

        novo_pin = (request.POST.get("admin_pin") or "").strip()
        if novo_pin:
            qs.definir_pin(device, novo_pin)

        messages.success(request, "Configuração atualizada. Será aplicada no próximo check-in do dispositivo.")
        return redirect("quiosque_detalhe", pk=device.pk)

    # Inventário recebido do aparelho + estado de liberação de cada app (checkbox).
    permitidos = list(device.apps_permitidos or [])
    perm_set = set(permitidos)
    inventario = list(device.apps.all())
    for a in inventario:
        a.liberado = a.pkg in perm_set
    inv_pkgs = {a.pkg for a in inventario}
    # Pacotes liberados que NÃO constam do inventário (manuais/avançado) — vão no
    # textarea para não serem perdidos ao salvar.
    extras = [p for p in permitidos if p not in inv_pkgs]

    return render(request, "front/quiosque/quiosque_config.html", {
        "device": device,
        "inventario": inventario,
        "inventario_total": len(inventario),
        "apps_extra_texto": "\n".join(extras),
    })


@login_required
def quiosque_comando_novo(request, pk: int):
    from ProjetoEstoque.models import KioskDevice, KioskComando

    device = get_object_or_404(KioskDevice, pk=pk)
    if request.method == "POST":
        tipo = (request.POST.get("tipo") or "").strip()
        if tipo in KioskComando.Tipo.values:
            payload = {}
            if tipo == KioskComando.Tipo.MENSAGEM:
                payload = {"texto": (request.POST.get("mensagem") or "").strip()[:200]}
            KioskComando.objects.create(device=device, tipo=tipo, payload=payload, criado_por=request.user)
            messages.success(request, "Comando enfileirado. Será entregue no próximo check-in.")
        else:
            messages.error(request, "Tipo de comando inválido.")
    return redirect("quiosque_detalhe", pk=device.pk)


@login_required
def quiosque_revogar(request, pk: int):
    from ProjetoEstoque.models import KioskDevice

    device = get_object_or_404(KioskDevice, pk=pk)
    if request.method == "POST":
        device.ativo = False
        device.save(update_fields=["ativo", "atualizado_em"])
        messages.success(request, f"Acesso do dispositivo “{device}” revogado.")
        return redirect("quiosque_dashboard")
    return redirect("quiosque_detalhe", pk=device.pk)


@login_required
def quiosque_excluir(request, pk: int):
    """Exclui DEFINITIVAMENTE o dispositivo (e seu histórico). Exige a senha do
    usuário logado como 2ª etapa de segurança."""
    from ProjetoEstoque.models import KioskDevice

    device = get_object_or_404(KioskDevice, pk=pk)
    if request.method == "POST":
        senha = request.POST.get("senha") or ""
        if not request.user.check_password(senha):
            messages.error(request, "Senha incorreta — o dispositivo NÃO foi excluído.")
            return redirect("quiosque_detalhe", pk=device.pk)
        nome = str(device)
        device.delete()  # cascata: apaga check-ins e comandos; matrículas viram SET_NULL
        messages.success(request, f"Dispositivo “{nome}” excluído definitivamente.")
        return redirect("quiosque_dashboard")
    return redirect("quiosque_detalhe", pk=device.pk)
