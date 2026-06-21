"""
views/ninja.py — Módulo NinjaOne RMM (via importação de CSV)

A integração via API/OAuth foi removida. Os dados vêm da planilha CSV
exportada do NinjaOne e importada pelo usuário.

Views:
    ninja_dashboard     GET  /ninja/
    ninja_dispositivos  GET  /ninja/dispositivos/
    ninja_relatorio     GET  /ninja/relatorio/
    ninja_importar      POST /ninja/importar/   (upload da planilha CSV)
"""

import csv as _csv
import datetime
from collections import defaultdict

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

_LOGIN_STATUS_LABEL = {
    "confere": "Confere",
    "divergente": "Divergente",
    "sem_atribuicao": "Sem atribuição",
    "sem_login": "Sem login",
}


# ─────────────────────────────────────────────────────────────
# Exportação Excel (.xlsx) — planilha profissional de dispositivos
# ─────────────────────────────────────────────────────────────

def _dispositivos_xlsx(qs, titulo, subtitulo, filename_prefix="ninja_dispositivos"):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND_DARK, BRAND, SOFT, ZEBRA = "3A1480", "6528D8", "EEE9FB", "F6F3FC"
    INK = "1F2733"
    hair = Side(style="thin", color="DDD6F3")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    f_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color="5B6B7F")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Calibri", size=10, color=INK)
    fill_title = PatternFill("solid", fgColor=BRAND_DARK)
    fill_sub = PatternFill("solid", fgColor=SOFT)
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left", vertical="center")
    a_left_ind = Alignment(horizontal="left", vertical="center", indent=1)
    dt_fmt = "DD/MM/YYYY HH:MM"

    header = ["#", "Dispositivo", "Número de Série", "Modelo", "Local", "Organização",
              "Usuário Logado", "Status", "Vínculo no Estoque", "Centro de Custo", "Último Contato"]
    ncols = len(header)
    center_cols = {1, 8}

    wb = Workbook()
    ws = wb.active
    ws.title = "Dispositivos"
    ws.sheet_view.showGridLines = False

    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]; c.value = titulo; c.font = f_title; c.fill = fill_title; c.alignment = a_left_ind
    ws.row_dimensions[1].height = 34
    ws.merge_cells(f"A2:{last}2")
    c2 = ws["A2"]; c2.value = subtitulo; c2.font = f_sub; c2.fill = fill_sub; c2.alignment = a_left_ind
    ws.row_dimensions[2].height = 18

    HEADER_ROW = 3
    for ci, h in enumerate(header, 1):
        cc = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cc.fill = fill_header; cc.font = f_header; cc.border = border
        cc.alignment = a_center if ci in center_cols else a_left
    ws.row_dimensions[HEADER_ROW].height = 22

    row = HEADER_ROW + 1
    for i, d in enumerate(qs, start=1):
        vinc = d.item.nome if d.item_id else "—"
        cc_txt = ""
        if d.item_id and d.item.centro_custo_id:
            cc_txt = d.item.centro_custo.departamento or str(d.item.centro_custo)
        lc = d.last_contact
        if lc is not None:
            lc = timezone.localtime(lc).replace(tzinfo=None)
        valores = [i, d.display_name, d.serial_number or "", d.model_name or "", d.local or "",
                   d.organization_name or "", d.last_user or "",
                   "Online" if d.is_online else "Offline", vinc, cc_txt, lc]
        zebra = (i % 2 == 0)
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = border
            cell.font = f_cell
            cell.alignment = a_center if ci in center_cols else a_left
            if ci == 11 and val:
                cell.number_format = dt_fmt
                cell.alignment = a_center
            if ci == 8:
                cell.fill = PatternFill("solid", fgColor="E6F4EA" if d.is_online else "F0F0F2")
                cell.font = Font(name="Calibri", size=10, bold=True,
                                 color="1E8E3E" if d.is_online else "5B6B7F")
                cell.alignment = a_center
            elif zebra:
                cell.fill = fill_zebra
        row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last}{max(row - 1, HEADER_ROW)}"

    widths = {}
    for r_ in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        for idx, val in enumerate(r_, start=1):
            widths[idx] = max(widths.get(idx, 0), len(str(val)) if val is not None else 0)
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 11), 46)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    now = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{now}.xlsx"'
    return resp


def _validacao_xlsx(resultados, titulo, subtitulo, filename_prefix="ninja_validacao_login"):
    """Planilha profissional (.xlsx) da Validação de Login."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND_DARK, BRAND, SOFT, ZEBRA = "3A1480", "6528D8", "EEE9FB", "F6F3FC"
    INK = "1F2733"
    hair = Side(style="thin", color="DDD6F3")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    f_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color="5B6B7F")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Calibri", size=10, color=INK)
    fill_title = PatternFill("solid", fgColor=BRAND_DARK)
    fill_sub = PatternFill("solid", fgColor=SOFT)
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left", vertical="center")
    a_left_ind = Alignment(horizontal="left", vertical="center", indent=1)
    dt_fmt = "DD/MM/YYYY HH:MM"

    STATUS_FILL = {
        "confere": ("E6F4EA", "1E8E3E"),
        "divergente": ("FCE8E6", "D93025"),
        "sem_atribuicao": ("FEF1E0", "B35A00"),
        "sem_login": ("F0F0F2", "5B6B7F"),
    }

    header = ["#", "Status", "Dispositivo", "Número de Série", "Login no Dispositivo",
              "Atribuído no Sistema", "Detectado", "Último Contato", "Detalhe"]
    ncols = len(header)
    center_cols = {1, 2}

    wb = Workbook()
    ws = wb.active
    ws.title = "Validação de Login"
    ws.sheet_view.showGridLines = False

    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]; c.value = titulo; c.font = f_title; c.fill = fill_title; c.alignment = a_left_ind
    ws.row_dimensions[1].height = 34
    ws.merge_cells(f"A2:{last}2")
    c2 = ws["A2"]; c2.value = subtitulo; c2.font = f_sub; c2.fill = fill_sub; c2.alignment = a_left_ind
    ws.row_dimensions[2].height = 18

    HEADER_ROW = 3
    for ci, h in enumerate(header, 1):
        cc = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cc.fill = fill_header; cc.font = f_header; cc.border = border
        cc.alignment = a_center if ci in center_cols else a_left
    ws.row_dimensions[HEADER_ROW].height = 22

    row = HEADER_ROW + 1
    for i, r in enumerate(resultados, start=1):
        d = r["device"]
        lc = d.last_contact
        if lc is not None:
            lc = timezone.localtime(lc).replace(tzinfo=None)
        status_label = r.get("status_label") or _LOGIN_STATUS_LABEL.get(r["status"], r["status"])
        valores = [i, status_label, d.display_name, d.serial_number or "",
                   r.get("login") or "", r.get("usuario_nome") or "",
                   r.get("detectado_nome") or "", lc, r.get("detalhe") or ""]
        zebra = (i % 2 == 0)
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = border
            cell.font = f_cell
            cell.alignment = a_center if ci in center_cols else a_left
            if ci == 8 and val:
                cell.number_format = dt_fmt
                cell.alignment = a_center
            if ci == 2:
                bg, fg = STATUS_FILL.get(r["status"], ("FFFFFF", INK))
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=fg)
                cell.alignment = a_center
            elif zebra:
                cell.fill = fill_zebra
        row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last}{max(row - 1, HEADER_ROW)}"

    widths = {}
    for r_ in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        for idx, val in enumerate(r_, start=1):
            widths[idx] = max(widths.get(idx, 0), len(str(val)) if val is not None else 0)
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 11), 60)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    now = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{now}.xlsx"'
    return resp


# ─────────────────────────────────────────────────────────────
# Dashboard principal
# ─────────────────────────────────────────────────────────────

@login_required
def ninja_dashboard(request):
    from ProjetoEstoque.models import NinjaDevice, NinjaDeviceSnapshot, Item

    hoje = timezone.localdate()
    qs = NinjaDevice.objects.select_related("item", "item__centro_custo", "item__localidade")

    total = qs.count()
    online = qs.filter(is_online=True).count()
    matched = qs.filter(item__isnull=False).count()
    unmatched = qs.filter(item__isnull=True, serial_number__gt="").count()
    sem_serie = qs.filter(serial_number="").count()
    com_user = qs.filter(is_online=True).exclude(last_user="").count()

    # ── Cross-ref: itens do estoque SEM dispositivo no Ninja (não monitorados) ──
    itens_sem_agente_qs = (
        Item.objects.filter(ninja_device__isnull=True)
        .exclude(item_consumo="sim")
        .exclude(numero_serie="")
        .exclude(numero_serie__isnull=True)
        .select_related("localidade", "centro_custo")
        .order_by("nome")
    )
    itens_sem_agente = itens_sem_agente_qs.count()

    # ── Dispositivos por Local / Site ──
    por_local = list(
        qs.values("local")
        .annotate(
            qtd=Count("id"),
            online=Count("id", filter=Q(is_online=True)),
            vinc=Count("id", filter=Q(item__isnull=False)),
        )
        .order_by("-qtd")
    )
    for row in por_local:
        row["local"] = row["local"] or "— sem local —"
        row["offline"] = row["qtd"] - row["online"]

    online_com_user = qs.filter(is_online=True).exclude(last_user="").order_by("display_name")[:12]
    sem_match_online = qs.filter(is_online=True, item__isnull=True, serial_number__gt="").order_by("display_name")[:8]

    snapshots_hoje = NinjaDeviceSnapshot.objects.filter(timestamp__date=hoje).count()
    last_import = qs.order_by("-last_sync").values_list("last_sync", flat=True).first()

    # ── Séries para gráficos (Chart.js) ──
    chart_local = {
        "labels": [r["local"] for r in por_local[:8]],
        "online": [r["online"] for r in por_local[:8]],
        "offline": [r["offline"] for r in por_local[:8]],
    }

    context = {
        "hoje": hoje,
        "last_import": last_import,
        "snapshots_hoje": snapshots_hoje,
        "kpi": {
            "total": total,
            "online": online,
            "offline": total - online,
            "matched": matched,
            "unmatched": unmatched,
            "sem_serie": sem_serie,
            "com_user": com_user,
            "itens_sem_agente": itens_sem_agente,
            "pct_online": round(online / total * 100) if total else 0,
            "pct_matched": round(matched / total * 100) if total else 0,
        },
        "por_local": por_local,
        "online_com_user": online_com_user,
        "sem_match_online": sem_match_online,
        "itens_sem_agente_sample": itens_sem_agente_qs[:8],
        "chart_status": {"online": online, "offline": total - online},
        "chart_cobertura": {
            "vinculados": matched,
            "nao_cadastrados": unmatched,
            "sem_serie": sem_serie,
        },
        "chart_local": chart_local,
    }
    return render(request, "front/ninja/ninja_dashboard.html", context)


# ─────────────────────────────────────────────────────────────
# Lista de dispositivos
# ─────────────────────────────────────────────────────────────

@login_required
def ninja_dispositivos(request):
    from ProjetoEstoque.models import NinjaDevice

    q = (request.GET.get("q") or "").strip()
    f_status = (request.GET.get("status") or "").strip()    # online | offline
    f_match = (request.GET.get("match") or "").strip()      # matched | unmatched | sem_serie
    f_local = (request.GET.get("local") or "").strip()      # nome do site
    f_user = (request.GET.get("user") or "").strip()        # com | sem
    f_ordenar = (request.GET.get("ordenar") or "").strip()

    qs = NinjaDevice.objects.select_related("item", "item__centro_custo", "item__localidade")

    if q:
        qs = qs.filter(
            Q(display_name__icontains=q) |
            Q(hostname__icontains=q) |
            Q(serial_number__icontains=q) |
            Q(last_user__icontains=q) |
            Q(local__icontains=q) |
            Q(model_name__icontains=q) |
            Q(item__nome__icontains=q)
        )

    if f_status == "online":
        qs = qs.filter(is_online=True)
    elif f_status == "offline":
        qs = qs.filter(is_online=False)

    if f_match == "matched":
        qs = qs.filter(item__isnull=False)
    elif f_match == "unmatched":
        qs = qs.filter(item__isnull=True, serial_number__gt="")
    elif f_match == "sem_serie":
        qs = qs.filter(serial_number="")

    if f_local:
        qs = qs.filter(local=f_local)

    if f_user == "com":
        qs = qs.exclude(last_user="")
    elif f_user == "sem":
        qs = qs.filter(last_user="")

    ordem_map = {
        "nome": ("display_name",),
        "-nome": ("-display_name",),
        "contato": ("last_contact", "display_name"),
        "-contato": ("-last_contact", "display_name"),
        "local": ("local", "display_name"),
    }
    qs = qs.order_by(*ordem_map.get(f_ordenar, ("-is_online", "display_name")))

    # Exportação Excel (.xlsx) respeitando os filtros aplicados
    if request.GET.get("export") == "xlsx":
        gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
        return _dispositivos_xlsx(
            qs,
            "DISPOSITIVOS NINJAONE",
            f"Santa Colomba Agropecuária  ·  {qs.count()} dispositivo(s)  ·  gerado em {gerado}",
        )

    base = NinjaDevice.objects.all()
    _btotal = base.count()
    _bonline = base.filter(is_online=True).count()
    resumo = {
        "total": _btotal,
        "online": _bonline,
        "offline": _btotal - _bonline,
        "matched": base.filter(item__isnull=False).count(),
        "unmatched": base.filter(item__isnull=True, serial_number__gt="").count(),
    }

    total_filtrado = qs.count()
    try:
        per_page = int(request.GET.get("pp", 24))
    except ValueError:
        per_page = 24
    per_page = max(6, min(per_page, 120))

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    get_copy = request.GET.copy()
    if "page" in get_copy:
        del get_copy["page"]
    qs_keep = get_copy.urlencode()

    opt_locais = list(
        base.exclude(local="").values_list("local", flat=True).distinct().order_by("local")
    )

    context = {
        "dispositivos": page_obj.object_list,
        "page_obj": page_obj,
        "qs_keep": qs_keep,
        "total_filtrado": total_filtrado,
        "total_registros": total_filtrado,
        "per_page": per_page,
        "f_q": q,
        "f_status": f_status,
        "f_match": f_match,
        "f_local": f_local,
        "f_user": f_user,
        "f_ordenar": f_ordenar,
        "opt_locais": opt_locais,
        "resumo": resumo,
    }
    return render(request, "front/ninja/ninja_dispositivos.html", context)


# ─────────────────────────────────────────────────────────────
# Relatório de uso (séries de snapshots, alimentadas pelas importações)
# ─────────────────────────────────────────────────────────────

@login_required
def ninja_relatorio(request):
    from ProjetoEstoque.models import NinjaDeviceSnapshot

    hoje = timezone.localdate()

    try:
        data_ini = datetime.date.fromisoformat(request.GET.get("data_ini") or "")
    except ValueError:
        data_ini = hoje - datetime.timedelta(days=6)
    try:
        data_fim = datetime.date.fromisoformat(request.GET.get("data_fim") or "")
    except ValueError:
        data_fim = hoje
    if data_fim < data_ini:
        data_fim = data_ini

    snaps_qs = NinjaDeviceSnapshot.objects.filter(
        timestamp__date__gte=data_ini,
        timestamp__date__lte=data_fim,
    ).select_related("device", "device__item")

    total_snaps = snaps_qs.count()
    delta_dias = max((data_fim - data_ini).days + 1, 1)

    device_data: dict[int, dict] = defaultdict(lambda: {
        "device": None, "total_online": 0, "total_offline": 0,
        "usuarios": set(), "primeira_vez": None, "ultima_vez": None,
        "dias_ativos": set(),
    })

    for snap in snaps_qs.order_by("device_id", "timestamp"):
        d = device_data[snap.device_id]
        d["device"] = snap.device
        if snap.is_online:
            d["total_online"] += 1
            d["dias_ativos"].add(snap.timestamp.date())
            if snap.current_user:
                d["usuarios"].add(snap.current_user)
            ts_local = timezone.localtime(snap.timestamp)
            if d["primeira_vez"] is None or ts_local < d["primeira_vez"]:
                d["primeira_vez"] = ts_local
            if d["ultima_vez"] is None or ts_local > d["ultima_vez"]:
                d["ultima_vez"] = ts_local
        else:
            d["total_offline"] += 1

    relatorio = []
    for dev_id, info in device_data.items():
        device = info["device"]
        if not device:
            continue
        tot = info["total_online"] + info["total_offline"]
        pct = round(info["total_online"] / tot * 100) if tot else 0
        minutos_estimados = info["total_online"] * 15
        relatorio.append({
            "device": device,
            "total_online": info["total_online"],
            "total_offline": info["total_offline"],
            "pct_online": pct,
            "minutos_est": minutos_estimados,
            "horas_est": round(minutos_estimados / 60, 1),
            "usuarios": sorted(info["usuarios"]),
            "primeira_vez": info["primeira_vez"],
            "ultima_vez": info["ultima_vez"],
            "dias_ativos": len(info["dias_ativos"]),
        })

    relatorio.sort(key=lambda x: x["total_online"], reverse=True)

    total_devices_com_dado = len(relatorio)
    total_ativas = sum(1 for r in relatorio if r["total_online"] > 0)
    max_online = max((r["total_online"] for r in relatorio), default=0)

    context = {
        "relatorio": relatorio,
        "data_ini": data_ini,
        "data_fim": data_fim,
        "hoje": hoje,
        "total_snaps": total_snaps,
        "delta_dias": delta_dias,
        "kpi": {
            "total_devices": total_devices_com_dado,
            "total_ativas": total_ativas,
            "total_inativas": total_devices_com_dado - total_ativas,
            "max_online_snaps": max_online,
        },
    }
    return render(request, "front/ninja/ninja_relatorio.html", context)


# ─────────────────────────────────────────────────────────────
# Não cadastrados — dispositivos da planilha que NÃO existem no estoque
# ─────────────────────────────────────────────────────────────

@login_required
def ninja_nao_cadastrados(request):
    """
    Lista os dispositivos presentes na importação do NinjaOne cujo número de série
    (ou nome) NÃO corresponde a nenhum Item do estoque — ou seja, equipamentos que
    existem fisicamente mas ainda não foram cadastrados no sistema.
    """
    from ProjetoEstoque.models import NinjaDevice

    q = (request.GET.get("q") or "").strip()
    f_local = (request.GET.get("local") or "").strip()

    universo = NinjaDevice.objects.filter(item__isnull=True, serial_number__gt="")

    qs = universo
    if q:
        qs = qs.filter(
            Q(display_name__icontains=q) |
            Q(serial_number__icontains=q) |
            Q(last_user__icontains=q) |
            Q(local__icontains=q) |
            Q(model_name__icontains=q)
        )
    if f_local:
        qs = qs.filter(local=f_local)
    qs = qs.order_by("-is_online", "display_name")

    # Exportação Excel (.xlsx) — planilha profissional
    if request.GET.get("export") == "xlsx":
        gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
        return _dispositivos_xlsx(
            qs,
            "ITENS NÃO CADASTRADOS — NINJAONE",
            f"Dispositivos sem cadastro no estoque  ·  {qs.count()} item(ns)  ·  gerado em {gerado}",
            filename_prefix="ninja_nao_cadastrados",
        )

    # Exportação CSV (respeita os filtros aplicados)
    if request.GET.get("export") == "csv":
        resp = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        resp["Content-Disposition"] = 'attachment; filename="ninja_nao_cadastrados.csv"'
        writer = _csv.writer(resp)
        writer.writerow(["Dispositivo", "Numero de serie", "Modelo", "Local",
                         "Ultimo usuario", "Online", "Ultimo contato"])
        for d in qs:
            writer.writerow([
                d.display_name, d.serial_number, d.model_name, d.local, d.last_user,
                "Sim" if d.is_online else "Nao",
                d.last_contact.strftime("%d/%m/%Y %H:%M") if d.last_contact else "",
            ])
        return resp

    por_local = list(
        universo.values("local").annotate(qtd=Count("id")).order_by("-qtd")
    )
    for row in por_local:
        row["local"] = row["local"] or "— sem local —"

    total_devices = NinjaDevice.objects.count()
    total_nao_cad = universo.count()
    online_nao_cad = qs.filter(is_online=True).count()

    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    get_copy = request.GET.copy()
    if "page" in get_copy:
        del get_copy["page"]
    qs_keep = get_copy.urlencode()

    context = {
        "dispositivos": page_obj.object_list,
        "page_obj": page_obj,
        "qs_keep": qs_keep,
        "total_registros": qs.count(),
        "total_nao_cad": total_nao_cad,
        "total_devices": total_devices,
        "online_nao_cad": online_nao_cad,
        "pct_nao_cad": round(total_nao_cad / total_devices * 100) if total_devices else 0,
        "por_local": por_local,
        "opt_locais": [r["local"] for r in por_local if r["local"] != "— sem local —"],
        "f_q": q,
        "f_local": f_local,
    }
    return render(request, "front/ninja/ninja_nao_cadastrados.html", context)


# ─────────────────────────────────────────────────────────────
# Validação de login — último usuário do device × colaborador do sistema
# ─────────────────────────────────────────────────────────────

@login_required
def ninja_login_validacao(request):
    from ProjetoEstoque.models import NinjaLoginRegistro
    from services.ninja_service import avaliar_logins

    f_status = (request.GET.get("status") or "").strip()
    q = (request.GET.get("q") or "").strip()

    # ── Filtro de data sobre o último contato/login do dispositivo ──
    raw_ini = (request.GET.get("data_ini") or "").strip()
    raw_fim = (request.GET.get("data_fim") or "").strip()
    try:
        data_ini = datetime.date.fromisoformat(raw_ini) if raw_ini else None
    except ValueError:
        data_ini = None
    try:
        data_fim = datetime.date.fromisoformat(raw_fim) if raw_fim else None
    except ValueError:
        data_fim = None
    if data_ini and data_fim and data_fim < data_ini:
        data_ini, data_fim = data_fim, data_ini

    resultados = avaliar_logins()

    # Aplica o filtro de data sobre last_contact (último login/atividade capturado)
    if data_ini or data_fim:
        filtrados = []
        for r in resultados:
            lc = r["device"].last_contact
            if lc is None:
                continue
            d_local = timezone.localtime(lc).date()
            if data_ini and d_local < data_ini:
                continue
            if data_fim and d_local > data_fim:
                continue
            filtrados.append(r)
        resultados = filtrados

    # KPIs refletem a janela de data (mas não o filtro de status — são os botões de filtro)
    kpi = {"total": len(resultados), "confere": 0, "divergente": 0,
           "sem_atribuicao": 0, "sem_login": 0}
    for r in resultados:
        kpi[r["status"]] = kpi.get(r["status"], 0) + 1

    hist = {
        row["device_id"]: row["n"]
        for row in NinjaLoginRegistro.objects.values("device_id").annotate(n=Count("id"))
    }

    if f_status:
        resultados = [r for r in resultados if r["status"] == f_status]
    if q:
        ql = q.lower()
        resultados = [
            r for r in resultados
            if ql in r["device"].display_name.lower()
            or ql in (r["login"] or "").lower()
            or ql in (r["usuario_nome"] or "").lower()
            or ql in (r["detectado_nome"] or "").lower()
        ]
    for r in resultados:
        r["hist"] = hist.get(r["device"].id, 0)
        r["status_label"] = _LOGIN_STATUS_LABEL.get(r["status"], r["status"])

    # Exportação Excel (.xlsx) respeitando todos os filtros aplicados
    if request.GET.get("export") == "xlsx":
        gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
        periodo = ""
        if data_ini or data_fim:
            ini_txt = data_ini.strftime("%d/%m/%Y") if data_ini else "início"
            fim_txt = data_fim.strftime("%d/%m/%Y") if data_fim else "hoje"
            periodo = f"  ·  período {ini_txt} a {fim_txt}"
        return _validacao_xlsx(
            resultados,
            "VALIDAÇÃO DE LOGIN — NINJAONE",
            f"Último login × colaborador atribuído  ·  {len(resultados)} dispositivo(s)"
            f"{periodo}  ·  gerado em {gerado}",
        )

    paginator = Paginator(resultados, 25)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    get_copy = request.GET.copy()
    for _k in ("page", "export"):
        if _k in get_copy:
            del get_copy[_k]
    qs_keep = get_copy.urlencode()

    ultima = (
        NinjaLoginRegistro.objects.order_by("-verificado_em")
        .values_list("verificado_em", flat=True).first()
    )

    context = {
        "resultados": page_obj.object_list,
        "page_obj": page_obj,
        "qs_keep": qs_keep,
        "total_registros": len(resultados),
        "kpi": kpi,
        "f_status": f_status,
        "f_q": q,
        "f_data_ini": raw_ini,
        "f_data_fim": raw_fim,
        "ultima_validacao": ultima,
    }
    return render(request, "front/ninja/ninja_login_validacao.html", context)


@login_required
def ninja_login_revalidar(request):
    if request.method != "POST":
        return redirect("ninja_login_validacao")

    from services.ninja_service import registrar_validacao
    try:
        stats = registrar_validacao(user=request.user)
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f"Falha ao validar logins: {exc}")
        return redirect("ninja_login_validacao")

    messages.success(
        request,
        f"Validação concluída: {stats['confere']} conferem · "
        f"{stats['divergente']} divergente(s) · {stats['sem_atribuicao']} sem atribuição · "
        f"{stats['sem_login']} sem login · {stats['novos_registros']} novo(s) registro(s) no histórico."
    )
    return redirect(request.POST.get("next") or "ninja_login_validacao")


@login_required
def ninja_login_detalhe(request, pk):
    from ProjetoEstoque.models import NinjaDevice
    from services.ninja_service import avaliar_login_device

    device = get_object_or_404(
        NinjaDevice.objects.select_related("item", "item__centro_custo", "item__localidade"),
        pk=pk,
    )
    atual = avaliar_login_device(device)
    atual["status_label"] = _LOGIN_STATUS_LABEL.get(atual["status"], atual["status"])
    registros = device.login_registros.select_related("usuario_sistema").all()

    context = {
        "device": device,
        "atual": atual,
        "registros": registros,
        "total_registros": registros.count(),
    }
    return render(request, "front/ninja/ninja_login_detalhe.html", context)


# ─────────────────────────────────────────────────────────────
# Importação da planilha CSV (POST)
# ─────────────────────────────────────────────────────────────

@login_required
def ninja_importar(request):
    if request.method != "POST":
        return redirect("ninja_dashboard")

    destino = request.POST.get("next") or "ninja_dashboard"

    arquivo = request.FILES.get("arquivo_csv")
    if not arquivo:
        messages.error(request, "Selecione a planilha CSV exportada do NinjaOne.")
        return redirect(destino)

    if not arquivo.name.lower().endswith(".csv"):
        messages.error(request, "O arquivo deve estar no formato CSV (.csv).")
        return redirect(destino)

    from services.ninja_service import importar_csv

    try:
        res = importar_csv(arquivo, user=request.user)
    except Exception as exc:  # noqa: BLE001 — feedback amigável ao usuário
        messages.error(request, f"Não foi possível processar o CSV: {exc}")
        return redirect(destino)

    if not res.get("ok"):
        messages.error(request, res.get("erro") or "Falha ao importar a planilha.")
        return redirect(destino)

    messages.success(
        request,
        f"Importação concluída: {res['total']} dispositivo(s) "
        f"({res['criados']} novo(s), {res['atualizados']} atualizado(s)) · "
        f"{res['vinculados']} vinculado(s) ao estoque · "
        f"{res['sem_serie']} sem número de série."
    )
    return redirect(destino)
