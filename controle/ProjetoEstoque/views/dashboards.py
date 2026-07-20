from decimal import Decimal, ROUND_HALF_UP
from datetime import date, timedelta, datetime

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.db.models import Q, Count, Sum, F, Case, When, Value as V
from django.db.models import DecimalField
from django.db.models.functions import TruncMonth, Coalesce
from django.utils import timezone
from django.utils.dateparse import parse_date
from dateutil.relativedelta import relativedelta

from ..models import (
    Item, Licenca, LicencaLote, MovimentacaoLicenca,
    MovimentacaoItem, Preventiva, PreventivaExecucao,
    CentroCusto, Fornecedor, Localidade, Usuario, StatusItemChoices,
    PeriodicidadeChoices, TipoMovLicencaChoices, TipoMovimentacaoChoices,
    Locacao, CheckListModelo, OrdemManutencao, StatusOrdemManutencaoChoices,
)

def _month_key(dt):
    """YYYY-MM para indexação."""
    return f"{dt.year:04d}-{dt.month:02d}"

def _last_n_month_stamps(n=12):
    """Lista de (ano, mês) dos últimos n meses, do mais antigo ao mais recente."""
    now = timezone.localtime()
    y, m = now.year, now.month
    out = []
    for _ in range(n):
        out.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(out))

def _labels_pt_br(stamps):
    """Gera labels 'Mes/AnoCurto' ex.: Jan/25 a partir de (ano, mês)."""
    nomes = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    return [f"{nomes[m-1]}/{str(y)[-2:]}" for (y, m) in stamps]

def _align_series(stamps, qs_month_count, field_name="c"):
    """
    Alinha uma série mensal aos stamps fornecidos.
    TruncMonth sobre DateField retorna date; sobre DateTimeField retorna datetime.
    Ambos são tratados aqui.
    """
    m2v = {}
    for row in qs_month_count:
        mdt = row["m"]
        if isinstance(mdt, datetime):
            key = _month_key(timezone.localtime(mdt))
        elif isinstance(mdt, date):
            key = f"{mdt.year:04d}-{mdt.month:02d}"
        else:
            continue
        m2v[key] = int(row[field_name] or 0)
    out = []
    for (y, m) in stamps:
        out.append(m2v.get(f"{y:04d}-{m:02d}", 0))
    return out


# =========================
# Helpers de custo de licença
# =========================
def _custo_mensal_lic(lic: Licenca) -> Decimal:
    cm = lic.custo_mensal()  # helper implementado na sua model
    return cm if cm is not None else Decimal("0.00")

# ==============================================================================
# 1. MOTOR DE DADOS (Funções Auxiliares)
# ==============================================================================

def _generate_month_keys(months=12):
    """
    Gera as chaves (ano, mes) e os labels para os últimos N meses.
    Retorna: (lista_chaves, lista_labels)
    Ex: ([(2023, 1), ...], ['Jan/23', ...])
    """
    today = timezone.localdate()
    keys = []
    labels = []
    
    # Começa do primeiro dia do mês atual
    curr = today.replace(day=1)
    
    names = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

    for _ in range(months):
        k = (curr.year, curr.month)
        lbl = f"{names[curr.month]}/{str(curr.year)[2:]}"
        
        keys.append(k)
        labels.append(lbl)
        
        # Volta 1 mês
        curr = (curr - timedelta(days=1)).replace(day=1)
    
    # Inverte para ficar cronológico (Antigo -> Novo)
    return list(reversed(keys)), list(reversed(labels))

def _process_chart_data(keys, queryset):
    """
    Recebe um QuerySet agrupado por mês e alinha com as chaves de data.
    IMPORTANTE: O QuerySet DEVE ter o campo anotado como 'valor'.
    """
    # 1. Transforma o QuerySet em um Dicionário de busca rápida: {(ano, mes): valor}
    data_map = {}
    for item in queryset:
        dt = item.get('m') # 'm' é o TruncMonth
        val = item.get('valor') # 'valor' é o dado padronizado
        
        if dt:
            # Garante que None vire 0 (caso de Somas nulas)
            final_val = val if val is not None else 0
            data_map[(dt.year, dt.month)] = final_val

    # 2. Monta a lista final alinhada com as keys (preenche buracos com 0)
    result = []
    for k in keys:
        result.append(data_map.get(k, 0))
        
    return result

# ==============================================================================
# 2. VIEW PRINCIPAL
# ==============================================================================

# ==============================================================================
# 2.1  APRESENTAÇÃO (DASHBOARD VISÃO GERAL — estilo slides com filtros)
# ==============================================================================

# Meses por periodicidade para custo mensal de licença
_MESES_PERIODICIDADE = {
    "mensal": 1, "trimestral": 3, "semestral": 6, "anual": 12, "trienal": 36, "contrato": 12,
}

# Rótulo + cor por status de equipamento (ordem de exibição)
_STATUS_META = [
    ("ativo",      "Em uso",     "#34c759"),
    ("backup",     "Em estoque", "#5856d6"),
    ("manutencao", "Manutenção", "#ff9500"),
    ("defeito",    "Defeito",    "#ff3b30"),
    ("pausado",    "Pausado",    "#8e8e93"),
    ("descarte",   "Descarte",   "#475569"),
]


def _fmt_brl(v, dec=0):
    """Formata valor em Real (pt-BR) com separador de milhar."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        v = 0.0
    s = f"{v:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return "R$ " + s


def _dados_equipamentos(request):
    qs = Item.objects.all()
    loc = (request.GET.get("localidade") or "").strip()
    cc  = (request.GET.get("centro_custo") or "").strip()
    st  = (request.GET.get("status") or "").strip()
    if loc.isdigit():
        qs = qs.filter(localidade_id=int(loc))
    if cc.isdigit():
        qs = qs.filter(centro_custo_id=int(cc))
    if st in dict(StatusItemChoices.choices):
        qs = qs.filter(status=st)

    total = qs.count()
    by_status = dict(qs.values_list("status").annotate(c=Count("id")))
    # "Estoque" é o status equivalente a "Backup" para item de consumo (ver
    # StatusItemChoices) — soma no mesmo balde pra não sumir da contagem/gráfico.
    by_status["backup"] = by_status.get("backup", 0) + by_status.pop("estoque", 0)
    ativos  = by_status.get("ativo", 0)
    backup  = by_status.get("backup", 0)
    manut   = by_status.get("manutencao", 0)
    defeito = by_status.get("defeito", 0)
    valor_total = qs.aggregate(s=Sum("valor"))["s"] or Decimal("0")

    # Distribuição por status (na ordem de _STATUS_META)
    status_labels, status_data, status_colors = [], [], []
    for key, lbl, col in _STATUS_META:
        cnt = by_status.get(key, 0)
        if cnt:
            status_labels.append(lbl); status_data.append(cnt); status_colors.append(col)

    def top(field, limit, label_null="Não definido"):
        rows = qs.values(field).annotate(c=Count("id")).order_by("-c")[:limit]
        return ([(r[field] or label_null) for r in rows], [r["c"] for r in rows])

    sub_l, sub_d     = top("subtipo__nome", 7)
    loc_l, loc_d     = top("localidade__local", 8)
    marca_l, marca_d = top("marca", 7)

    cc_rows = (
        qs.exclude(centro_custo__isnull=True)
          .values("centro_custo__numero", "centro_custo__departamento")
          .annotate(v=Sum("valor"), c=Count("id"))
          .order_by("-v")[:7]
    )
    cc_l = [f'{r["centro_custo__numero"]} · {r["centro_custo__departamento"]}' for r in cc_rows]
    cc_v = [float(r["v"] or 0) for r in cc_rows]

    # Aquisições por mês (12) via data_compra
    keys, lbls = _generate_month_keys(12)
    aq_qs = (
        qs.filter(data_compra__isnull=False)
          .annotate(m=TruncMonth("data_compra")).values("m").annotate(valor=Count("id"))
    )
    aq_data = _process_chart_data(keys, aq_qs)

    # Movimentações filtradas pelo item
    movqs = MovimentacaoItem.objects.all()
    if loc.isdigit():
        movqs = movqs.filter(item__localidade_id=int(loc))
    if cc.isdigit():
        movqs = movqs.filter(item__centro_custo_id=int(cc))
    fy, fm = keys[0]
    start = timezone.make_aware(timezone.datetime(fy, fm, 1))
    movbase = movqs.filter(created_at__gte=start)

    def movserie(tipo):
        q = (movbase.filter(tipo_movimentacao=tipo)
             .annotate(m=TruncMonth("created_at")).values("m").annotate(valor=Count("id")))
        return _process_chart_data(keys, q)

    mov = {
        "entrada": movserie("entrada"),
        "baixa":   movserie("baixa"),
        "transf":  movserie("transferencia"),
        "manut":   movserie("envio_manutencao"),
    }

    pct_uso = f"{round(ativos / total * 100)}% do parque" if total else "—"
    kpis = [
        {"label": "Total de Ativos",  "value": str(total),   "icon": "fa-database",            "color": "blue",   "sub": "equipamentos no escopo"},
        {"label": "Em Uso",           "value": str(ativos),  "icon": "fa-circle-check",        "color": "green",  "sub": pct_uso},
        {"label": "Em Estoque",       "value": str(backup),  "icon": "fa-boxes-stacked",       "color": "purple", "sub": "disponíveis como backup"},
        {"label": "Em Manutenção",    "value": str(manut),   "icon": "fa-screwdriver-wrench",  "color": "orange", "sub": "em reparo no momento"},
        {"label": "Críticos",         "value": str(defeito), "icon": "fa-triangle-exclamation","color": "red",    "sub": "com defeito registrado"},
        {"label": "Valor do Parque",  "value": _fmt_brl(valor_total), "icon": "fa-coins",      "color": "teal",   "sub": "soma do valor de aquisição"},
    ]

    return {
        "dataset": "equipamentos",
        "titulo": "Parque de Equipamentos",
        "kpis": kpis,
        "charts": {
            "status":       {"labels": status_labels, "data": status_data, "colors": status_colors},
            "subtipo":      {"labels": sub_l, "data": sub_d},
            "marca":        {"labels": marca_l, "data": marca_d},
            "localidade":   {"labels": loc_l, "data": loc_d},
            "centro_custo": {"labels": cc_l, "data": cc_v, "moeda": True},
            "aquisicoes":   {"labels": lbls, "data": aq_data},
            "mov":          {"labels": lbls, **mov},
        },
    }


def _dados_licencas(request):
    forn = (request.GET.get("fornecedor") or "").strip()
    cc   = (request.GET.get("centro_custo") or "").strip()
    per  = (request.GET.get("periodicidade") or "").strip()
    pmb  = (request.GET.get("pmb") or "").strip().lower()

    lic_qs = Licenca.objects.all()
    if forn.isdigit():
        lic_qs = lic_qs.filter(fornecedor_id=int(forn))
    if pmb in ("sim", "nao"):
        lic_qs = lic_qs.filter(pmb=pmb)

    lotes = (
        LicencaLote.objects
        .select_related("licenca", "fornecedor", "centro_custo", "licenca__fornecedor")
        .filter(licenca__in=lic_qs)
    )
    if cc.isdigit():
        lotes = lotes.filter(centro_custo_id=int(cc))
    if per:
        lotes = lotes.filter(periodicidade=per)

    total_seats = 0
    disp = 0
    custo_mensal = Decimal("0")
    forn_cost, forn_seats, per_count, cc_cost = {}, {}, {}, {}
    lic_ids = set()

    for lote in lotes:
        qtd = int(lote.quantidade_total or 0)
        d   = int(lote.quantidade_disponivel or 0)
        meses = _MESES_PERIODICIDADE.get(lote.periodicidade, 12) or 12
        custo_ciclo = Decimal(lote.custo_ciclo or 0)
        cm = (custo_ciclo / Decimal(meses)) if meses else Decimal("0")

        total_seats += qtd
        disp += d
        custo_mensal += cm
        lic_ids.add(lote.licenca_id)

        if lote.fornecedor:
            fn = lote.fornecedor.nome
        elif lote.licenca and lote.licenca.fornecedor:
            fn = lote.licenca.fornecedor.nome
        else:
            fn = "Indefinido"
        forn_cost[fn]  = forn_cost.get(fn, Decimal("0")) + cm
        forn_seats[fn] = forn_seats.get(fn, 0) + qtd

        pl = lote.get_periodicidade_display()
        per_count[pl] = per_count.get(pl, 0) + 1

        if lote.centro_custo:
            cl = f"{lote.centro_custo.numero} · {lote.centro_custo.departamento}"
            cc_cost[cl] = cc_cost.get(cl, Decimal("0")) + cm

    em_uso = max(0, total_seats - disp)
    custo_anual = custo_mensal * Decimal("12")

    def top_dict(d, n, as_float=True):
        items = sorted(d.items(), key=lambda x: x[1], reverse=True)[:n]
        return ([k for k, _ in items], [(float(v) if as_float else v) for _, v in items])

    fc_l, fc_d = top_dict(forn_cost, 8)
    fs_l, fs_d = top_dict(forn_seats, 8, as_float=False)
    pc_l, pc_d = top_dict(per_count, 10, as_float=False)
    ccc_l, ccc_d = top_dict(cc_cost, 8)

    pct_ocup = f"{round(em_uso / total_seats * 100)}% de ocupação" if total_seats else "—"
    kpis = [
        {"label": "Licenças",        "value": str(len(lic_ids)), "icon": "fa-id-badge",     "color": "blue",   "sub": "produtos licenciados"},
        {"label": "Assentos Totais", "value": str(total_seats),  "icon": "fa-chair",        "color": "purple", "sub": "posições adquiridas"},
        {"label": "Em Uso",          "value": str(em_uso),       "icon": "fa-user-check",   "color": "green",  "sub": pct_ocup},
        {"label": "Disponíveis",     "value": str(disp),         "icon": "fa-box-open",     "color": "teal",   "sub": "assentos livres"},
        {"label": "Custo Mensal",    "value": _fmt_brl(custo_mensal), "icon": "fa-coins",   "color": "orange", "sub": "recorrência mensal"},
        {"label": "Custo Anual",     "value": _fmt_brl(custo_anual),  "icon": "fa-calendar","color": "red",    "sub": "projeção 12 meses"},
    ]

    return {
        "dataset": "licencas",
        "titulo": "Licenças de Software",
        "kpis": kpis,
        "charts": {
            "ocupacao":         {"labels": ["Em uso", "Disponível"], "data": [em_uso, disp], "colors": ["#0071e3", "#c7c7cc"]},
            "custo_fornecedor": {"labels": fc_l, "data": fc_d, "moeda": True},
            "seats_fornecedor": {"labels": fs_l, "data": fs_d},
            "periodicidade":    {"labels": pc_l, "data": pc_d},
            "custo_cc":         {"labels": ccc_l, "data": ccc_d, "moeda": True},
        },
    }


_OM_STATUS_META = [
    ("andamento",  "Em andamento", "#ff9500"),
    ("concluida",  "Concluída",    "#34c759"),
    ("cancelada",  "Cancelada",    "#8e8e93"),
    ("descartada", "Descartada",   "#ff3b30"),
]


def _dados_manutencao(request):
    forn = (request.GET.get("fornecedor") or "").strip()
    cc   = (request.GET.get("centro_custo") or "").strip()
    st   = (request.GET.get("status") or "").strip()

    qs = OrdemManutencao.objects.select_related("fornecedor", "item", "item__centro_custo")
    if forn.isdigit():
        qs = qs.filter(fornecedor_id=int(forn))
    if cc.isdigit():
        qs = qs.filter(item__centro_custo_id=int(cc))

    ordens = list(qs)
    if st == "andamento":
        ordens = [o for o in ordens if o.aberta]
    elif st == "concluida":
        ordens = [o for o in ordens if o.status == StatusOrdemManutencaoChoices.CONCLUIDO]
    elif st == "cancelada":
        ordens = [o for o in ordens if o.status in (
            StatusOrdemManutencaoChoices.CANCELADO, StatusOrdemManutencaoChoices.DESCARTADO)]

    total = len(ordens)
    by_grupo = {
        "andamento":  sum(1 for o in ordens if o.aberta),
        "concluida":  sum(1 for o in ordens if o.status == StatusOrdemManutencaoChoices.CONCLUIDO),
        "cancelada":  sum(1 for o in ordens if o.status == StatusOrdemManutencaoChoices.CANCELADO),
        "descartada": sum(1 for o in ordens if o.status == StatusOrdemManutencaoChoices.DESCARTADO),
    }
    em_garantia = sum(1 for o in ordens if o.garantia_vigente)

    concluidas = [o for o in ordens if o.status == StatusOrdemManutencaoChoices.CONCLUIDO]
    custo_total = Decimal("0.00")
    forn_cost, cc_cost = {}, {}
    for o in concluidas:
        v = o.valor_manutencao
        if not v:
            continue
        custo_total += v
        fn = o.fornecedor.nome if o.fornecedor else "Indefinido"
        forn_cost[fn] = forn_cost.get(fn, Decimal("0")) + v
        if o.item and o.item.centro_custo:
            cl = f"{o.item.centro_custo.numero} · {o.item.centro_custo.departamento}"
            cc_cost[cl] = cc_cost.get(cl, Decimal("0")) + v
    qtd_com_custo = sum(1 for o in concluidas if o.valor_manutencao)
    ticket_medio = (custo_total / qtd_com_custo) if qtd_com_custo else Decimal("0.00")

    def top_dict(d, n):
        items = sorted(d.items(), key=lambda x: x[1], reverse=True)[:n]
        return ([k for k, _ in items], [float(v) for _, v in items])

    fc_l, fc_d = top_dict(forn_cost, 8)
    ccm_l, ccm_d = top_dict(cc_cost, 8)

    status_labels, status_data, status_colors = [], [], []
    for key, lbl, col in _OM_STATUS_META:
        cnt = by_grupo.get(key, 0)
        if cnt:
            status_labels.append(lbl); status_data.append(cnt); status_colors.append(col)

    # Séries mensais (12 meses) — abertura por created_at, custo por finalizada_em
    keys, lbls = _generate_month_keys(12)

    def _mes_key(dt):
        local_dt = timezone.localtime(dt) if timezone.is_aware(dt) else dt
        return (local_dt.year, local_dt.month)

    abertura_map = {}
    for o in ordens:
        if not o.created_at:
            continue
        k = _mes_key(o.created_at)
        abertura_map[k] = abertura_map.get(k, 0) + 1
    abertura_data = [abertura_map.get(k, 0) for k in keys]

    custo_mensal_map = {}
    for o in concluidas:
        v = o.valor_manutencao
        if not v or not o.finalizada_em:
            continue
        k = _mes_key(o.finalizada_em)
        custo_mensal_map[k] = custo_mensal_map.get(k, Decimal("0")) + v
    custo_mensal_data = [float(custo_mensal_map.get(k, Decimal("0"))) for k in keys]

    pct_andamento = f"{round(by_grupo['andamento'] / total * 100)}% do total" if total else "—"
    kpis = [
        {"label": "Total de OS",     "value": str(total),                 "icon": "fa-screwdriver-wrench", "color": "blue",   "sub": "ordens no escopo"},
        {"label": "Em Andamento",    "value": str(by_grupo["andamento"]), "icon": "fa-truck-fast",         "color": "orange", "sub": pct_andamento},
        {"label": "Concluídas",      "value": str(by_grupo["concluida"]), "icon": "fa-circle-check",       "color": "green",  "sub": "finalizadas pelo fornecedor"},
        {"label": "Em Garantia",     "value": str(em_garantia),           "icon": "fa-shield-halved",      "color": "purple", "sub": "reparo/troca vigente"},
        {"label": "Custo Total",     "value": _fmt_brl(custo_total),      "icon": "fa-coins",              "color": "teal",   "sub": "concluídas com valor apurado"},
        {"label": "Ticket Médio",    "value": _fmt_brl(ticket_medio),     "icon": "fa-receipt",            "color": "red",    "sub": "custo médio por OS concluída"},
    ]

    return {
        "dataset": "manutencao",
        "titulo": "Manutenção de Equipamentos",
        "kpis": kpis,
        "charts": {
            "status":       {"labels": status_labels, "data": status_data, "colors": status_colors},
            "fornecedor":   {"labels": fc_l, "data": fc_d, "moeda": True},
            "abertura":     {"labels": lbls, "data": abertura_data},
            "custo_mensal": {"labels": lbls, "data": custo_mensal_data, "moeda": True},
            "centro_custo": {"labels": ccm_l, "data": ccm_d, "moeda": True},
        },
    }


@login_required
def dashboard_apresentacao_dados(request):
    """Retorna os dados agregados (equipamentos, licenças ou manutenção) para o deck de apresentação."""
    dataset = request.GET.get("dataset", "equipamentos")
    try:
        if dataset == "licencas":
            data = _dados_licencas(request)
        elif dataset == "manutencao":
            data = _dados_manutencao(request)
        else:
            data = _dados_equipamentos(request)
    except Exception as exc:  # noqa: BLE001 — nunca derrubar a apresentação
        return JsonResponse({"ok": False, "erro": str(exc)}, status=500)
    return JsonResponse({"ok": True, **data})


@login_required
def dashboard(request):
    """Dashboard de Visão Geral — apresentação estilo slides com filtros (dados via AJAX)."""
    context = {
        "localidades":           Localidade.objects.order_by("local"),
        "centros_custo":         CentroCusto.objects.order_by("numero"),
        "fornecedores":          Fornecedor.objects.order_by("nome"),
        "status_choices":        StatusItemChoices.choices,
        "periodicidade_choices": PeriodicidadeChoices.choices,
        "dados_url":             reverse("dashboard_apresentacao_dados"),
    }
    return render(request, "front/dashboards/dashboard.html", context)


# ==== helpers que você já usa em outros dashboards ====
def _month_key(dt):
    return f"{dt.year:04d}-{dt.month:02d}"

def _last_n_month_stamps(n=12):
    now = timezone.localtime()
    y, m = now.year, now.month
    out = []
    for _ in range(n):
        out.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(out))

def _labels_pt_br(stamps):
    nomes = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    return [f"{nomes[m-1]}/{str(y)[-2:]}" for (y, m) in stamps]

def _align_series_date(stamps, qs_month_count, field_name="c"):
    """
    Igual a _align_series mas aceita date ou datetime (TruncMonth sobre DateField → date).
    Usada nas séries do dashboard de preventivas.
    """
    m2v = {}
    for row in qs_month_count:
        mdt = row["m"]
        if isinstance(mdt, datetime):
            key = _month_key(timezone.localtime(mdt))
        elif isinstance(mdt, date):
            key = f"{mdt.year:04d}-{mdt.month:02d}"
        else:
            continue
        m2v[key] = int(row[field_name] or 0)
    out = []
    for (y, m) in stamps:
        out.append(m2v.get(f"{y:04d}-{m:02d}", 0))
    return out


@login_required
def preventiva_dashboard(request):
    """
    Dashboard de Preventivas.

    Fonte única de verdade: `_get_preventiva_dashboard_data` (a mesma usada
    no export Excel), garantindo que tela e planilha nunca divirjam.
    """
    ctx = _get_preventiva_dashboard_data(request, limit_vencidas=50)
    return render(request, "front/dashboards/preventiva_dashboard.html", ctx)


@login_required
def preventiva_dashboard_export(request):
    """Exportação Excel do Dashboard de Preventivas, com os mesmos filtros da tela."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ctx = _get_preventiva_dashboard_data(request, limit_vencidas=None)

    BRAND = "0071E3"
    DARK = "0A2540"
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Calibri", size=10, color="1D1D1F")
    fill_title = PatternFill("solid", fgColor=DARK)
    fill_sub = PatternFill("solid", fgColor="EEF2F7")
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor="F7F9FC")
    a_left = Alignment(horizontal="left", vertical="center")
    a_center = Alignment(horizontal="center", vertical="center")

    def faixa_titulo(ws, ncols, titulo, subtitulo):
        last = get_column_letter(ncols)
        ws.merge_cells(f"A1:{last}1")
        c1 = ws["A1"]
        c1.value = titulo
        c1.font = f_title
        c1.fill = fill_title
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A2:{last}2")
        c2 = ws["A2"]
        c2.value = subtitulo
        c2.font = Font(name="Calibri", size=9, color="334155")
        c2.fill = fill_sub
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18
        ws.sheet_view.showGridLines = False

    def cabecalho(ws, row, headers, center_cols=()):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = fill_header
            c.font = f_header
            c.border = border
            c.alignment = a_center if ci in center_cols else a_left
        ws.row_dimensions[row].height = 22

    def linhas(ws, hr, valores_list, center_cols=(), col_widths=None):
        row = hr + 1
        for i, valores in enumerate(valores_list):
            zebra = (i % 2 == 1)
            for ci, val in enumerate(valores, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.border = border
                c.font = f_cell
                c.alignment = a_center if ci in center_cols else a_left
                if zebra:
                    c.fill = fill_zebra
            row += 1
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{hr + 1}"

    periodo_txt = f"{ctx['dt_ini']:%d/%m/%Y} a {ctx['dt_fim']:%d/%m/%Y}"
    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")

    wb = Workbook()

    # ── Aba 1: Resumo ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    faixa_titulo(ws1, 2, "DASHBOARD DE PREVENTIVAS",
                 f"Santa Colomba Agropecuária  ·  Período {periodo_txt}  ·  Gerado em {gerado}")
    resumo_kv = [
        ("Total de Ativos", ctx["total"]),
        ("Em Dia", ctx["ok_count"]),
        ("Vencidas", ctx["vencidas_count"]),
        ("Sem Agenda", ctx["sem_agenda_count"]),
        ("Executadas (Mês)", ctx["executadas_mes"]),
        ("Com Data Agendada", ctx["agendadas_count"]),
    ]
    hr = 4
    cabecalho(ws1, hr, ["Indicador", "Valor"], center_cols=(2,))
    linhas(ws1, hr, resumo_kv, center_cols=(2,), col_widths=[28, 14])

    # ── Aba 2: Vencidas ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Vencidas")
    faixa_titulo(ws2, 5, "PREVENTIVAS VENCIDAS", f"{len(ctx['vencidas'])} preventiva(s)")
    hr = 4
    cabecalho(ws2, hr, ["Equipamento", "Localidade", "Checklist", "Próxima (Calc.)", "Dias em Atraso"], center_cols=(4, 5))
    valores = [
        [p.equipamento.nome, p.equipamento.localidade.local if p.equipamento.localidade else "—",
         p.checklist_modelo.nome if p.checklist_modelo else "—",
         p.proxima_calc.strftime("%d/%m/%Y") if p.proxima_calc else "—",
         p.dias_atraso if p.dias_atraso is not None else "—"]
        for p in ctx["vencidas"]
    ]
    linhas(ws2, hr, valores, center_cols=(4, 5), col_widths=[32, 22, 26, 16, 14])

    # ── Aba 3: Próximos 30 dias ──────────────────────────────────────────
    ws3 = wb.create_sheet("Proximos 30 Dias")
    faixa_titulo(ws3, 5, "PRÓXIMAS PREVENTIVAS (30 DIAS)", f"{len(ctx['proximas'])} preventiva(s)")
    hr = 4
    cabecalho(ws3, hr, ["Equipamento", "Localidade", "Checklist", "Próxima (Calc.)", "Dias Restantes"], center_cols=(4, 5))
    valores = [
        [p.equipamento.nome, p.equipamento.localidade.local if p.equipamento.localidade else "—",
         p.checklist_modelo.nome if p.checklist_modelo else "—",
         p.proxima_calc.strftime("%d/%m/%Y") if p.proxima_calc else "—",
         p.dias_faltam]
        for p in ctx["proximas"]
    ]
    linhas(ws3, hr, valores, center_cols=(4, 5), col_widths=[32, 22, 26, 16, 14])

    # ── Aba 4/5/6: agregados ─────────────────────────────────────────────
    def aba_agregado(nome_aba, titulo, dados, chave_nome):
        ws = wb.create_sheet(nome_aba)
        faixa_titulo(ws, 5, titulo, f"{len(dados)} grupo(s)")
        hr = 4
        cabecalho(ws, hr, ["Grupo", "Total", "Ok", "Vencidas", "Sem Agenda"], center_cols=(2, 3, 4, 5))
        valores = [
            [d.get(chave_nome) or "—", d["total"], d["ok"], d["vencidas"], d["sem_agenda"]]
            for d in dados
        ]
        linhas(ws, hr, valores, center_cols=(2, 3, 4, 5), col_widths=[30, 10, 10, 12, 14])

    aba_agregado("Por Checklist", "PREVENTIVAS POR CHECKLIST", ctx["agg_chk"], "checklist_modelo__nome")
    aba_agregado("Por Localidade", "PREVENTIVAS POR LOCALIDADE", ctx["agg_loc"], "equipamento__localidade__local")
    aba_agregado("Por Subtipo", "PREVENTIVAS POR SUBTIPO", ctx["agg_sub"], "equipamento__subtipo__nome")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arquivo = f"dashboard_preventivas_{ctx['dt_ini']:%Y%m%d}_{ctx['dt_fim']:%Y%m%d}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def _get_preventiva_dashboard_data(request, limit_vencidas=50):
    """
    Fonte única de verdade dos dados do Dashboard de Preventivas (tela e export).
    Usa _aplicar_status_preventiva para calcular proxima_calc = data_ultima + intervalo
    dinamicamente, evitando dependência de data_proxima gravada no banco.
    """
    from .preventivas import _aplicar_status_preventiva
    from collections import defaultdict

    today = timezone.localdate()
    now   = timezone.localtime()

    # -------- filtros --------
    q        = (request.GET.get("q") or "").strip()
    status   = (request.GET.get("status") or "").strip()
    chk_id   = (request.GET.get("checklist") or "").strip()
    loc      = (request.GET.get("local") or "").strip()
    subtipo  = (request.GET.get("subtipo") or "").strip()

    base_qs = Preventiva.objects.select_related(
        "equipamento", "equipamento__localidade",
        "equipamento__subtipo", "checklist_modelo",
    )
    if q:
        base_qs = base_qs.filter(Q(equipamento__nome__icontains=q) | Q(observacao__icontains=q))
    if chk_id.isdigit():
        base_qs = base_qs.filter(checklist_modelo_id=int(chk_id))
    if loc:
        base_qs = base_qs.filter(equipamento__localidade__local__icontains=loc)
    if subtipo:
        base_qs = base_qs.filter(equipamento__subtipo__nome__icontains=subtipo)

    # Processa TODAS em Python para garantir proxima_calc correto
    all_preventivas = list(base_qs)
    for p in all_preventivas:
        _aplicar_status_preventiva(p, today)

    # Filtro de status (afeta tabelas/listas, não KPIs)
    if status == "ok":
        filtered = [p for p in all_preventivas if p.status_visual in ("ok", "atencao")]
    elif status == "vencida":
        filtered = [p for p in all_preventivas if p.status_visual == "vencida"]
    elif status == "sem_agenda":
        filtered = [p for p in all_preventivas if p.status_visual == "indefinido"]
    else:
        filtered = all_preventivas

    # -------- KPIs --------
    total            = len(all_preventivas)
    vencidas_count   = sum(1 for p in all_preventivas if p.status_visual == "vencida")
    ok_count         = sum(1 for p in all_preventivas if p.status_visual in ("ok", "atencao"))
    sem_agenda_count = sum(1 for p in all_preventivas if p.status_visual == "indefinido")
    agendadas_count  = sum(1 for p in all_preventivas if p.data_agendamento)
    executadas_mes   = PreventivaExecucao.objects.filter(
        preventiva_id__in=[p.id for p in all_preventivas],
        data_execucao__year=now.year,
        data_execucao__month=now.month,
    ).count()

    # -------- Séries 12 meses --------
    stamps12   = _last_n_month_stamps(12)
    labels12   = _labels_pt_br(stamps12)
    start12_dt = date(stamps12[0][0], stamps12[0][1], 1)

    exec_qs = (PreventivaExecucao.objects
               .filter(preventiva_id__in=[p.id for p in all_preventivas],
                       data_execucao__gte=start12_dt)
               .annotate(m=TruncMonth("data_execucao"))
               .values("m").annotate(c=Count("id")).order_by("m"))
    serie_exec = _align_series_date(stamps12, exec_qs)

    prog_by_month = defaultdict(int)
    for p in all_preventivas:
        if p.proxima_calc and p.proxima_calc >= start12_dt:
            key = f"{p.proxima_calc.year:04d}-{p.proxima_calc.month:02d}"
            prog_by_month[key] += 1
    serie_prog = [prog_by_month.get(f"{y:04d}-{m:02d}", 0) for y, m in stamps12]

    # -------- Listas operacionais --------
    vencidas = sorted(
        [p for p in all_preventivas if p.status_visual == "vencida"],
        key=lambda p: p.proxima_calc or date.max,
    )[:limit_vencidas]
    for p in vencidas:
        p.dias_atraso = abs(p.dias_restantes) if p.dias_restantes is not None else None

    proximas = sorted(
        [p for p in all_preventivas
         if p.proxima_calc and 0 <= (p.proxima_calc - today).days <= 30
         and not getattr(p, "pausada", False)],
        key=lambda p: p.proxima_calc,
    )
    for p in proximas:
        p.dias_faltam = (p.proxima_calc - today).days

    historico = sorted(
        [p for p in all_preventivas if p.data_ultima],
        key=lambda p: p.data_ultima,
        reverse=True,
    )[:20]

    # -------- AGG tables (usando filtered para refletir filtro de status) --------
    chk_agg = defaultdict(lambda: {"nome": None, "total": 0, "vencidas": 0, "ok": 0, "sem_agenda": 0, "prox_30": 0})
    loc_agg = defaultdict(lambda: {"nome": None, "total": 0, "vencidas": 0, "ok": 0, "sem_agenda": 0})
    sub_agg = defaultdict(lambda: {"nome": None, "total": 0, "vencidas": 0, "ok": 0, "sem_agenda": 0})

    for p in filtered:
        ck = p.checklist_modelo_id
        chk_agg[ck]["nome"] = p.checklist_modelo.nome if p.checklist_modelo else None
        chk_agg[ck]["total"] += 1

        lk = (p.equipamento.localidade.local if p.equipamento.localidade else None) or "—"
        loc_agg[lk]["nome"] = lk
        loc_agg[lk]["total"] += 1

        sk = (p.equipamento.subtipo.nome if p.equipamento.subtipo else None) or "—"
        sub_agg[sk]["nome"] = sk
        sub_agg[sk]["total"] += 1

        sv = p.status_visual
        for d in (chk_agg[ck], loc_agg[lk], sub_agg[sk]):
            if sv == "vencida":
                d["vencidas"] += 1
            elif sv in ("ok", "atencao"):
                d["ok"] += 1
            elif sv == "indefinido":
                d["sem_agenda"] += 1

        if p.proxima_calc and 0 <= (p.proxima_calc - today).days <= 30:
            chk_agg[ck]["prox_30"] += 1

    agg_chk = sorted(
        [{"checklist_modelo__nome": d["nome"], **{k: d[k] for k in ("total","vencidas","ok","sem_agenda","prox_30")}}
         for d in chk_agg.values()],
        key=lambda x: -x["total"],
    )
    agg_loc = sorted(
        [{"equipamento__localidade__local": d["nome"], **{k: d[k] for k in ("total","vencidas","ok","sem_agenda")}}
         for d in loc_agg.values()],
        key=lambda x: -x["vencidas"],
    )
    agg_sub = sorted(
        [{"equipamento__subtipo__nome": d["nome"], **{k: d[k] for k in ("total","vencidas","ok","sem_agenda")}}
         for d in sub_agg.values()],
        key=lambda x: -x["vencidas"],
    )

    chk_labels, chk_rates = [], []
    for r in agg_chk[:8]:
        den = r["ok"] + r["vencidas"]
        taxa = (100.0 * r["ok"] / den) if den > 0 else 0.0
        chk_labels.append(r["checklist_modelo__nome"] or "Sem checklist")
        chk_rates.append(round(taxa, 2))

    checklist_opts = CheckListModelo.objects.all().order_by("nome").values("id", "nome")

    ctx = dict(
        q=q, status=status, checklist=chk_id, local=loc, subtipo=subtipo,
        checklist_opts=checklist_opts,
        dt_ini=start12_dt,
        dt_fim=today,
        total=total,
        ok_count=ok_count,
        vencidas_count=vencidas_count,
        sem_agenda_count=sem_agenda_count,
        executadas_mes=executadas_mes,
        agendadas_count=agendadas_count,
        serie_labels=labels12,
        serie_exec=serie_exec,
        serie_prog=serie_prog,
        agg_chk=agg_chk,
        agg_loc=agg_loc,
        agg_sub=agg_sub,
        vencidas=vencidas,
        proximas=proximas,
        historico=historico,
        chk_labels=chk_labels,
        chk_rates=chk_rates,
        today=today,
    )
    return ctx

# ---- helpers de data ----
def _parse_date(date_str, default):
    if not date_str: return default
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return default

def _get_meses_ciclo(periodicidade_str):
    """Auxiliar para converter periodicidade em meses"""
    if not periodicidade_str: return 1
    p = str(periodicidade_str).upper()
    if 'MEN' in p: return 1
    if 'BI' in p: return 2
    if 'TRI' in p: return 3
    if 'SEM' in p: return 6
    if 'ANU' in p: return 12
    return 1

# NOTA: a função `_get_cc_custos_data` viva é a definição mais abaixo neste
# módulo (usa `_calcular_custo_mensal_unitario_lote` + quantize). Uma versão
# duplicada/antiga existia aqui e era SOBRESCRITA por aquela — removida para
# evitar edição no lugar errado.


@login_required
def cc_custos_dashboard(request):
    """
    Dashboard de Custos por Setor (Centro de Custo).

    Fonte única de verdade: `_get_cc_custos_data` (a mesma usada no export PDF),
    garantindo que tela e relatório nunca divirjam. Suporta o filtro `pmb`
    (sim/nao) para separar centros de custo do PMB dos demais.
    """
    context = _get_cc_custos_data(request)
    return render(request, "front/dashboards/cc_custos_dashboard.html", context)

@login_required
def cc_custos_export_pdf(request):
    """Exportação PDF - Custos por Centro de Custo"""
    data = _get_cc_custos_data(request)

    data["usuario"] = request.user
    data["data_geracao"] = timezone.now()

    template_path = "front/dashboards/cc_custos_pdf.html"
    template = get_template(template_path)
    html = template.render(data)

    response = HttpResponse(content_type="application/pdf")
    filename = f'relatorio_custos_cc_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Erro ao gerar PDF")

    return response
def _calcular_custo_mensal_unitario_lote(lote):
    """
    Regra consolidada:
    - mensal: custo do lote já é mensal
    - trimestral: divide por 3
    - semestral: divide por 6
    - anual: divide por 12

    Depois:
    - mensal unitário = mensal do lote / quantidade
    - anual unitário = mensal unitário * 12
    """
    if not lote:
        return Decimal("0.00")

    qtd = Decimal(lote.quantidade_total or 0)
    if qtd <= 0:
        return Decimal("0.00")

    custo_ciclo = Decimal(lote.custo_ciclo or 0)
    periodicidade = str(lote.periodicidade or "").lower()

    if periodicidade == "mensal":
        custo_mensal_lote = custo_ciclo
    elif periodicidade == "trimestral":
        custo_mensal_lote = (custo_ciclo / Decimal("3")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    elif periodicidade == "semestral":
        custo_mensal_lote = (custo_ciclo / Decimal("6")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    elif periodicidade == "anual":
        custo_mensal_lote = (custo_ciclo / Decimal("12")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    else:
        custo_mensal_lote = custo_ciclo.quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

    custo_mensal_unitario = (custo_mensal_lote / qtd).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    return custo_mensal_unitario


def _get_cc_custos_data(request):
    hoje = timezone.localdate()
    dt_ini = _parse_date(request.GET.get("inicio"), hoje.replace(day=1))
    dt_fim = _parse_date(request.GET.get("fim"), hoje)

    # Filtro PMB (sim/nao). Vazio = todos os centros de custo.
    pmb_filtro = (request.GET.get("pmb") or "").strip().lower()
    if pmb_filtro not in ("sim", "nao"):
        pmb_filtro = ""

    # Filtro por Centro de Custo específico (id). Vazio = todos.
    cc_sel_raw = (request.GET.get("centro_custo") or "").strip()
    cc_sel = int(cc_sel_raw) if cc_sel_raw.isdigit() else None

    totals = {}

    def get_acc(cc_id):
        if not cc_id:
            return None

        if cc_id not in totals:
            totals[cc_id] = {
                "cc_obj": None,
                "qtd_usuarios": 0,
                "qtd_itens": 0,
                "qtd_licencas": 0,
                "custo_locacao": Decimal("0.00"),
                "custo_licencas": Decimal("0.00"),
                "custo_baixas": Decimal("0.00"),
            }
        return totals[cc_id]

    # ==========================================================
    # 1. CUSTO DE LOCAÇÃO (HARDWARE - RECORRENTE MENSAL)
    # ==========================================================
    locacoes = (
        Locacao.objects
        .select_related("equipamento__centro_custo")
        .filter(
            equipamento__status="ativo",
            valor_mensal__gt=0,
            equipamento__centro_custo__isnull=False
        )
    )

    for loc in locacoes:
        cc_id = loc.equipamento.centro_custo.id
        acc = get_acc(cc_id)
        if acc:
            acc["custo_locacao"] += Decimal(loc.valor_mensal or 0)

    # ==========================================================
    # 2. CUSTO DE LICENÇAS (SOFTWARE - RECORRENTE MENSAL)
    # ==========================================================
    movs_lic = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario__centro_custo", "centro_custo_destino", "lote")
        .filter(usuario__isnull=False)
        .order_by("licenca_id", "usuario_id", "created_at")
    )

    estado_atual_lic = {}
    for mov in movs_lic:
        estado_atual_lic[(mov.licenca_id, mov.usuario_id)] = mov

    for (_, _), mov in estado_atual_lic.items():
        if mov.tipo != TipoMovLicencaChoices.ATRIBUICAO:
            continue

        cc_id = None
        if mov.usuario and mov.usuario.centro_custo:
            cc_id = mov.usuario.centro_custo.id
        elif mov.centro_custo_destino:
            cc_id = mov.centro_custo_destino.id
        elif mov.licenca and mov.licenca.centro_custo:
            cc_id = mov.licenca.centro_custo.id

        acc = get_acc(cc_id)
        if not acc:
            continue

        if mov.lote:
            custo_mensal_unitario = _calcular_custo_mensal_unitario_lote(mov.lote)
        else:
            # fallback legado
            custo_mensal_unitario = Decimal(getattr(mov.licenca, "custo", 0) or 0).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

        acc["custo_licencas"] += custo_mensal_unitario
        acc["qtd_licencas"] += 1

    # ==========================================================
    # 3. CUSTO DE BAIXAS (PONTUAL NO PERÍODO)
    # ==========================================================
    baixas = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .select_related("item__centro_custo", "centro_custo_origem", "centro_custo_destino")
    )

    for baixa in baixas:
        # Baixa atribuída ao CC de DESTINO (setor consumidor); fallback origem → item.
        cc_id = None
        if baixa.centro_custo_destino:
            cc_id = baixa.centro_custo_destino.id
        elif baixa.centro_custo_origem:
            cc_id = baixa.centro_custo_origem.id
        elif baixa.item and baixa.item.centro_custo:
            cc_id = baixa.item.centro_custo.id

        acc = get_acc(cc_id)
        if not acc:
            continue

        custo_baixa = baixa.custo
        if custo_baixa is None:
            custo_baixa = Decimal(baixa.item.valor or 0) * Decimal(baixa.quantidade or 1)

        acc["custo_baixas"] += Decimal(custo_baixa or 0)

    # ==========================================================
    # 4. DADOS DE CONTEXTO
    # ==========================================================
    cc_ids = list(totals.keys())

    ccs_objs = CentroCusto.objects.filter(id__in=cc_ids)
    for cc in ccs_objs:
        if cc.id in totals:
            totals[cc.id]["cc_obj"] = cc

    users_agg = (
        Usuario.objects
        .filter(centro_custo_id__in=cc_ids, status="ativo")
        .values("centro_custo_id")
        .annotate(n=Count("id"))
    )
    for row in users_agg:
        cc_id = row["centro_custo_id"]
        if cc_id in totals:
            totals[cc_id]["qtd_usuarios"] = row["n"]

    itens_agg = (
        Item.objects
        .filter(centro_custo_id__in=cc_ids, status="ativo")
        .values("centro_custo_id")
        .annotate(n=Count("id"))
    )
    for row in itens_agg:
        cc_id = row["centro_custo_id"]
        if cc_id in totals:
            totals[cc_id]["qtd_itens"] = row["n"]

    # ==========================================================
    # 5. MONTAGEM FINAL (com split e filtro PMB)
    # ==========================================================
    linhas = []

    total_geral_itens = Decimal("0.00")
    total_geral_lics = Decimal("0.00")
    total_geral_baixas = Decimal("0.00")

    # Split PMB x Fora do PMB do recorrente mensal — SEMPRE sobre o conjunto
    # completo (ignora o filtro), para dar a visão comparativa geral.
    total_pmb = Decimal("0.00")
    total_nao_pmb = Decimal("0.00")

    for cc_id, dados in totals.items():
        cc_obj = dados["cc_obj"]
        if not cc_obj:
            continue

        custo_itens = dados["custo_locacao"].quantize(Decimal("0.01"))
        custo_licencas = dados["custo_licencas"].quantize(Decimal("0.01"))
        custo_baixas = dados["custo_baixas"].quantize(Decimal("0.01"))

        total_mensal = (custo_itens + custo_licencas).quantize(Decimal("0.01"))
        total_impacto = (total_mensal + custo_baixas).quantize(Decimal("0.01"))

        is_pmb = (cc_obj.pmb or "nao") == "sim"

        # split geral (independe do filtro de exibição)
        if is_pmb:
            total_pmb += total_mensal
        else:
            total_nao_pmb += total_mensal

        # filtro PMB: afeta tabela, KPIs e gráficos principais
        if pmb_filtro and (cc_obj.pmb or "nao") != pmb_filtro:
            continue

        # filtro por centro de custo específico
        if cc_sel and cc_obj.id != cc_sel:
            continue

        total_geral_itens += custo_itens
        total_geral_lics += custo_licencas
        total_geral_baixas += custo_baixas

        linhas.append({
            "cc": cc_obj,
            "is_pmb": is_pmb,
            "usuarios": dados["qtd_usuarios"],
            "itens": dados["qtd_itens"],
            "licencas": dados["qtd_licencas"],
            "custo_itens": custo_itens,
            "custo_licencas": custo_licencas,
            "baixas": custo_baixas,
            "total_mensal": total_mensal,
            "total_impacto": total_impacto,
        })

    linhas.sort(key=lambda x: x["total_impacto"], reverse=True)

    chart_labels = [f"{l['cc'].numero}" for l in linhas[:10]]
    chart_itens = [float(l["custo_itens"]) for l in linhas[:10]]
    chart_lics = [float(l["custo_licencas"]) for l in linhas[:10]]

    return {
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "pmb_filtro": pmb_filtro,
        "centro_custo_sel": cc_sel_raw if cc_sel else "",
        "centros_custo": CentroCusto.objects.order_by("numero"),
        "linhas": linhas,
        "kpi_cc_count": len(linhas),
        "kpi_total_mensal": (total_geral_itens + total_geral_lics).quantize(Decimal("0.01")),
        "kpi_total_baixas": total_geral_baixas.quantize(Decimal("0.01")),
        "kpi_top_cc": linhas[0]["cc"].departamento if linhas else "-",
        "total_pmb": total_pmb.quantize(Decimal("0.01")),
        "total_nao_pmb": total_nao_pmb.quantize(Decimal("0.01")),
        "js_labels": chart_labels,
        "js_itens": chart_itens,
        "js_lics": chart_lics,
        "js_mix_values": [
            float(total_geral_itens.quantize(Decimal("0.01"))),
            float(total_geral_lics.quantize(Decimal("0.01"))),
        ],
        "js_pmb_values": [
            float(total_pmb.quantize(Decimal("0.01"))),
            float(total_nao_pmb.quantize(Decimal("0.01"))),
        ],
    }


@login_required
def cc_custos_detalhe(request):
    """
    AJAX — detalhamento completo dos gastos de UM centro de custo.

    Usa exatamente a mesma cadeia de atribuição de `_get_cc_custos_data`
    (locação → licenças → baixas), garantindo que a soma exibida no drawer
    bata com a linha correspondente da tabela. Endpoint para a diretoria.
    """
    cc_id_raw = (request.GET.get("cc") or "").strip()
    if not cc_id_raw.isdigit():
        return JsonResponse({"erro": "Centro de custo inválido."}, status=400)
    cc_id = int(cc_id_raw)

    cc = CentroCusto.objects.filter(id=cc_id).first()
    if not cc:
        return JsonResponse({"erro": "Centro de custo não encontrado."}, status=404)

    hoje = timezone.localdate()
    dt_ini = _parse_date(request.GET.get("inicio"), hoje.replace(day=1))
    dt_fim = _parse_date(request.GET.get("fim"), hoje)

    # ── 1. Locação (hardware recorrente mensal) ───────────────────────────────
    locacoes = (
        Locacao.objects
        .select_related("equipamento", "equipamento__subtipo")
        .filter(
            equipamento__status="ativo",
            valor_mensal__gt=0,
            equipamento__centro_custo_id=cc_id,
        )
        .order_by("-valor_mensal")
    )
    itens_loc = []
    total_locacao = Decimal("0.00")
    qtd_locacao_total = 0
    for loc in locacoes:
        v = Decimal(loc.valor_mensal or 0)
        total_locacao += v
        qtd_locacao_total += 1
        if len(itens_loc) < 30:
            itens_loc.append({
                "nome": loc.equipamento.nome if loc.equipamento else "—",
                "tipo": (loc.equipamento.subtipo.nome if loc.equipamento and loc.equipamento.subtipo else ""),
                "valor": float(v.quantize(Decimal("0.01"))),
            })
    total_locacao = total_locacao.quantize(Decimal("0.01"))

    # ── 2. Licenças (software recorrente mensal) ──────────────────────────────
    # Reconstrói o estado atual por (licença, usuário) igual ao dashboard e
    # mantém apenas as atribuições cujo CC resolvido é este centro de custo.
    movs_lic = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "licenca__centro_custo",
                        "usuario__centro_custo", "centro_custo_destino", "lote")
        .filter(usuario__isnull=False)
        .order_by("licenca_id", "usuario_id", "created_at")
    )
    estado_lic = {}
    for m in movs_lic:
        estado_lic[(m.licenca_id, m.usuario_id)] = m

    lic_agg = {}
    total_licencas = Decimal("0.00")
    for mov in estado_lic.values():
        if mov.tipo != TipoMovLicencaChoices.ATRIBUICAO:
            continue
        resolved = None
        if mov.usuario and mov.usuario.centro_custo_id:
            resolved = mov.usuario.centro_custo_id
        elif mov.centro_custo_destino_id:
            resolved = mov.centro_custo_destino_id
        elif mov.licenca and mov.licenca.centro_custo_id:
            resolved = mov.licenca.centro_custo_id
        if resolved != cc_id:
            continue

        if mov.lote:
            custo = _calcular_custo_mensal_unitario_lote(mov.lote)
        else:
            custo = Decimal(getattr(mov.licenca, "custo", 0) or 0).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

        nome_lic = mov.licenca.nome if mov.licenca else "Desconhecida"
        if nome_lic not in lic_agg:
            lic_agg[nome_lic] = {"nome": nome_lic, "qtd_usuarios": 0, "subtotal": Decimal("0.00")}
        lic_agg[nome_lic]["qtd_usuarios"] += 1
        lic_agg[nome_lic]["subtotal"] += custo
        total_licencas += custo

    licencas = sorted(
        [{"nome": v["nome"], "qtd_usuarios": v["qtd_usuarios"],
          "subtotal": float(v["subtotal"].quantize(Decimal("0.01")))}
         for v in lic_agg.values()],
        key=lambda x: x["subtotal"], reverse=True,
    )
    qtd_lic_seats = sum(l["qtd_usuarios"] for l in licencas)
    total_licencas = total_licencas.quantize(Decimal("0.01"))

    # ── 3. Baixas / perdas no período ─────────────────────────────────────────
    baixas_qs = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .select_related("item", "item__centro_custo", "centro_custo_origem", "centro_custo_destino")
        .order_by("-created_at")
    )
    top_baixas = []
    total_baixas = Decimal("0.00")
    qtd_baixas = 0
    for b in baixas_qs:
        resolved = None
        if b.centro_custo_destino_id:
            resolved = b.centro_custo_destino_id
        elif b.centro_custo_origem_id:
            resolved = b.centro_custo_origem_id
        elif b.item and b.item.centro_custo_id:
            resolved = b.item.centro_custo_id
        if resolved != cc_id:
            continue

        custo = b.custo
        if custo is None:
            custo = (Decimal(b.item.valor or 0) * Decimal(b.quantidade or 1)) if b.item else Decimal("0")
        custo = Decimal(custo or 0)
        total_baixas += custo
        qtd_baixas += 1
        if len(top_baixas) < 15:
            top_baixas.append({
                "item": b.item.nome if b.item else "—",
                "data": b.created_at.strftime("%d/%m/%Y") if b.created_at else "",
                "custo": float(custo.quantize(Decimal("0.01"))),
            })
    total_baixas = total_baixas.quantize(Decimal("0.01"))

    # ── 4. Colaboradores e itens ativos do CC ─────────────────────────────────
    colab_rows = (
        Usuario.objects
        .filter(centro_custo_id=cc_id, status="ativo")
        .select_related("funcao")
        .order_by("nome")
    )
    colaboradores = [
        {"nome": u.nome, "funcao": (u.funcao.nome if u.funcao else "")}
        for u in colab_rows[:60]
    ]
    qtd_colab = colab_rows.count()
    qtd_itens_ativos = Item.objects.filter(centro_custo_id=cc_id, status="ativo").count()

    total_mensal = (total_locacao + total_licencas).quantize(Decimal("0.01"))
    total_impacto = (total_mensal + total_baixas).quantize(Decimal("0.01"))

    return JsonResponse({
        "cc": {
            "numero": cc.numero,
            "departamento": cc.departamento,
            "is_pmb": (cc.pmb or "nao") == "sim",
        },
        "periodo": {"inicio": dt_ini.strftime("%d/%m/%Y"), "fim": dt_fim.strftime("%d/%m/%Y")},
        "totais": {
            "custo_locacao": float(total_locacao),
            "custo_licencas": float(total_licencas),
            "custo_baixas": float(total_baixas),
            "total_mensal": float(total_mensal),
            "total_impacto": float(total_impacto),
            "qtd_colaboradores": qtd_colab,
            "qtd_itens": qtd_itens_ativos,
            "qtd_licencas": qtd_lic_seats,
            "qtd_baixas": qtd_baixas,
        },
        "locacao": itens_loc,
        "licencas": licencas,
        "baixas": top_baixas,
        "qtd_locacao_total": qtd_locacao_total,
        "colaboradores": colaboradores,
        "qtd_colab_total": qtd_colab,
    })


@login_required
def custos_diretoria_dashboard(request):
    """
    Dashboard de Gastos por Diretoria/Gestor.
    Agrega por 'diretor' quando o campo estiver populado;
    caso contrário usa 'gestor' como fallback automático.
    """
    # Detectar qual campo agrupar: diretor (quando importado) ou gestor (fallback)
    has_diretor = Usuario.objects.filter(diretor__isnull=False).exclude(diretor="").exists()
    campo = "diretor" if has_diretor else "gestor"
    label_grupo = "Diretoria" if has_diretor else "Gestor"

    totais = {}

    def get_acc(nome):
        if not nome:
            return None
        if nome not in totais:
            totais[nome] = {
                "nome": nome,
                "qtd_colaboradores": 0,
                "custo_licencas": Decimal("0.00"),
                "custo_movimentacoes": Decimal("0.00"),
                "custo_itens": Decimal("0.00"),
            }
        return totais[nome]

    # 1. Colaboradores ativos agrupados pelo campo escolhido
    colab_agg = (
        Usuario.objects
        .filter(status="ativo", **{f"{campo}__isnull": False})
        .exclude(**{campo: ""})
        .values(campo)
        .annotate(n=Count("id"))
    )
    for row in colab_agg:
        a = get_acc(row[campo])
        if a:
            a["qtd_colaboradores"] = row["n"]

    # 2. Custo mensal de licenças agrupado pelo campo
    movs_lic = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario", "lote")
        .filter(usuario__isnull=False, **{f"usuario__{campo}__isnull": False})
        .exclude(**{f"usuario__{campo}": ""})
        .order_by("licenca_id", "usuario_id", "created_at")
    )
    estado_lic = {}
    for m in movs_lic:
        estado_lic[(m.licenca_id, m.usuario_id)] = m

    for mov in estado_lic.values():
        if mov.tipo != TipoMovLicencaChoices.ATRIBUICAO:
            continue
        grupo_val = getattr(mov.usuario, campo, None) if mov.usuario else None
        a = get_acc(grupo_val)
        if not a:
            continue
        custo = (
            _calcular_custo_mensal_unitario_lote(mov.lote)
            if mov.lote
            else Decimal(getattr(mov.licenca, "custo", 0) or 0).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        )
        a["custo_licencas"] += custo

    # 3. Custo de movimentações com valor registrado, agrupado pelo campo
    movs_item = (
        MovimentacaoItem.objects
        .filter(custo__isnull=False, usuario__isnull=False,
                **{f"usuario__{campo}__isnull": False})
        .exclude(custo=0)
        .exclude(**{f"usuario__{campo}": ""})
        .select_related("usuario")
    )
    for mov in movs_item:
        grupo_val = getattr(mov.usuario, campo, None) if mov.usuario else None
        a = get_acc(grupo_val)
        if a:
            a["custo_movimentacoes"] += Decimal(mov.custo or 0)

    # 4. Custo mensal de locação de itens por diretoria via centro de custo
    cc_grupo_map = {}
    for cc_id, grupo in (
        Usuario.objects
        .filter(status="ativo", **{f"{campo}__isnull": False})
        .exclude(**{campo: ""})
        .exclude(centro_custo__isnull=True)
        .values_list("centro_custo_id", campo)
    ):
        if cc_id not in cc_grupo_map:
            cc_grupo_map[cc_id] = grupo

    locacoes_agg = (
        Locacao.objects
        .filter(
            equipamento__status="ativo",
            valor_mensal__isnull=False,
            equipamento__centro_custo__isnull=False,
        )
        .exclude(valor_mensal=0)
        .values("equipamento__centro_custo_id")
        .annotate(total=Sum("valor_mensal"))
    )
    for row in locacoes_agg:
        grupo = cc_grupo_map.get(row["equipamento__centro_custo_id"])
        if not grupo:
            continue
        a = get_acc(grupo)
        if a:
            a["custo_itens"] += Decimal(row["total"] or 0)

    # Consolidação
    linhas = []
    total_colab = 0
    total_lics = Decimal("0.00")
    total_movs = Decimal("0.00")
    total_itens = Decimal("0.00")

    for dados in totais.values():
        total_g = dados["custo_licencas"] + dados["custo_movimentacoes"] + dados["custo_itens"]
        n = dados["qtd_colaboradores"]
        linhas.append({
            "gestor": dados["nome"],
            "qtd_colaboradores": n,
            "custo_licencas": dados["custo_licencas"].quantize(Decimal("0.01")),
            "custo_movimentacoes": dados["custo_movimentacoes"].quantize(Decimal("0.01")),
            "custo_itens": dados["custo_itens"].quantize(Decimal("0.01")),
            "total_geral": total_g.quantize(Decimal("0.01")),
            "custo_por_colab": (total_g / n).quantize(Decimal("0.01")) if n > 0 else Decimal("0.00"),
        })
        total_colab += n
        total_lics += dados["custo_licencas"]
        total_movs += dados["custo_movimentacoes"]
        total_itens += dados["custo_itens"]

    linhas.sort(key=lambda x: x["total_geral"], reverse=True)

    total_geral = (total_lics + total_movs + total_itens).quantize(Decimal("0.01"))
    custo_medio = (total_geral / total_colab).quantize(Decimal("0.01")) if total_colab else Decimal("0.00")

    palette = [
        "#0071e3", "#34c759", "#ff9500", "#ff3b30", "#5856d6",
        "#af52de", "#ff2d55", "#5ac8fa", "#ffcc00", "#30b0c7",
        "#32ade6", "#fe3c30", "#64d2ff", "#bf5af2", "#ff6961",
        "#4cd964", "#007aff", "#ff375f", "#ffd60a",
    ]

    for i, l in enumerate(linhas):
        l["color"] = palette[i % len(palette)]

    js_labels = [l["gestor"] for l in linhas]
    js_lics   = [float(l["custo_licencas"]) for l in linhas]
    js_movs   = [float(l["custo_movimentacoes"]) for l in linhas]
    js_itens  = [float(l["custo_itens"]) for l in linhas]
    js_total  = [float(l["total_geral"]) for l in linhas]
    js_colors = [l["color"] for l in linhas]

    return render(request, "front/dashboards/custos_diretoria.html", {
        "linhas": linhas,
        "total_colab": total_colab,
        "total_lics": total_lics.quantize(Decimal("0.01")),
        "total_movs": total_movs.quantize(Decimal("0.01")),
        "total_itens": total_itens.quantize(Decimal("0.01")),
        "total_geral": total_geral,
        "top_diretoria": linhas[0]["gestor"] if linhas else "–",
        "custo_medio_colab": custo_medio,
        "qtd_diretorias": len(linhas),
        "label_grupo": label_grupo,
        "campo_grupo": campo,
        "js_labels": js_labels,
        "js_lics": js_lics,
        "js_movs": js_movs,
        "js_itens": js_itens,
        "js_total": js_total,
        "js_colors": js_colors,
        "sem_dados": not linhas,
    })


@login_required
def custos_diretoria_detalhe(request):
    """AJAX – detalhamento de custos de uma diretoria/gestor específico."""
    grupo = request.GET.get("grupo", "").strip()
    campo = request.GET.get("campo", "diretor")
    if not grupo or campo not in ("diretor", "gestor"):
        return JsonResponse({"erro": "Parâmetros inválidos."}, status=400)

    tipo_label = {
        "entrada": "Entrada", "baixa": "Baixa",
        "transferencia": "Transferência", "transferencia_equipamento": "Transf. Equipamento",
        "envio_manutencao": "Envio Manutenção", "retorno_manutencao": "Ret. Manutenção",
        "outros": "Outros",
    }

    # ── 1. Licenças ───────────────────────────────────────────────────────────
    movs_lic = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario", "lote")
        .filter(usuario__isnull=False, **{f"usuario__{campo}": grupo})
        .order_by("licenca_id", "usuario_id", "created_at")
    )
    estado_lic = {}
    for m in movs_lic:
        estado_lic[(m.licenca_id, m.usuario_id)] = m

    lic_agg = {}
    for mov in estado_lic.values():
        if mov.tipo != TipoMovLicencaChoices.ATRIBUICAO:
            continue
        nome_lic = mov.licenca.nome if mov.licenca else "Desconhecida"
        custo = (
            _calcular_custo_mensal_unitario_lote(mov.lote)
            if mov.lote
            else Decimal(getattr(mov.licenca, "custo", 0) or 0).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        )
        if nome_lic not in lic_agg:
            lic_agg[nome_lic] = {"nome": nome_lic, "qtd_usuarios": 0, "subtotal": Decimal("0.00")}
        lic_agg[nome_lic]["qtd_usuarios"] += 1
        lic_agg[nome_lic]["subtotal"] += custo

    licencas = sorted(
        [{"nome": k, "qtd_usuarios": v["qtd_usuarios"], "subtotal": float(v["subtotal"].quantize(Decimal("0.01")))}
         for k, v in lic_agg.items()],
        key=lambda x: x["subtotal"], reverse=True
    )[:10]
    total_lics = sum((v["subtotal"] for v in lic_agg.values()), Decimal("0.00"))

    # ── 2. Movimentações ──────────────────────────────────────────────────────
    all_movs = list(
        MovimentacaoItem.objects
        .filter(custo__isnull=False, usuario__isnull=False, **{f"usuario__{campo}": grupo})
        .exclude(custo=0)
        .select_related("item")
        .order_by("-custo")
    )

    mov_tipo_agg = {}
    for mov in all_movs:
        lbl = tipo_label.get(mov.tipo_movimentacao, mov.tipo_movimentacao)
        if lbl not in mov_tipo_agg:
            mov_tipo_agg[lbl] = {"tipo": lbl, "qtd": 0, "total": Decimal("0.00")}
        mov_tipo_agg[lbl]["qtd"] += 1
        mov_tipo_agg[lbl]["total"] += Decimal(mov.custo or 0)

    movimentacoes = sorted(
        [{"tipo": v["tipo"], "qtd": v["qtd"], "total": float(v["total"].quantize(Decimal("0.01")))}
         for v in mov_tipo_agg.values()],
        key=lambda x: x["total"], reverse=True
    )
    total_movs = sum((v["total"] for v in mov_tipo_agg.values()), Decimal("0.00"))

    top_movs = [
        {
            "item": mov.item.nome if mov.item else "—",
            "tipo": tipo_label.get(mov.tipo_movimentacao, mov.tipo_movimentacao),
            "custo": float(Decimal(mov.custo or 0).quantize(Decimal("0.01"))),
        }
        for mov in all_movs[:8]
    ]

    # ── 3. Mensalidades de locação de itens ───────────────────────────────────
    cc_ids = list(
        Usuario.objects
        .filter(status="ativo", **{f"{campo}": grupo})
        .exclude(centro_custo__isnull=True)
        .values_list("centro_custo_id", flat=True)
        .distinct()
    )

    locacoes_top = (
        Locacao.objects
        .filter(
            equipamento__status="ativo",
            valor_mensal__isnull=False,
            equipamento__centro_custo__id__in=cc_ids,
        )
        .exclude(valor_mensal=0)
        .select_related("equipamento", "equipamento__centro_custo")
        .order_by("-valor_mensal")[:10]
    )
    itens = [
        {
            "nome": loc.equipamento.nome,
            "valor": float(loc.valor_mensal),
            "centro_custo": loc.equipamento.centro_custo.numero if loc.equipamento.centro_custo else "—",
        }
        for loc in locacoes_top
    ]

    itens_totais = (
        Locacao.objects
        .filter(
            equipamento__status="ativo",
            valor_mensal__isnull=False,
            equipamento__centro_custo__id__in=cc_ids,
        )
        .exclude(valor_mensal=0)
        .aggregate(total=Sum("valor_mensal"), qtd=Count("id"))
    )
    total_itens = Decimal(itens_totais["total"] or 0).quantize(Decimal("0.01"))
    qtd_itens = itens_totais["qtd"] or 0

    total_geral = (total_lics + total_movs + total_itens).quantize(Decimal("0.01"))
    qtd_colab = Usuario.objects.filter(status="ativo", **{f"{campo}": grupo}).count()

    return JsonResponse({
        "nome": grupo,
        "totais": {
            "qtd_colaboradores": qtd_colab,
            "custo_licencas": float(total_lics.quantize(Decimal("0.01"))),
            "custo_movimentacoes": float(total_movs.quantize(Decimal("0.01"))),
            "custo_itens": float(total_itens),
            "total_geral": float(total_geral),
        },
        "licencas": licencas,
        "movimentacoes": movimentacoes,
        "top_movimentacoes": top_movs,
        "itens": itens,
        "qtd_itens_total": qtd_itens,
    })
