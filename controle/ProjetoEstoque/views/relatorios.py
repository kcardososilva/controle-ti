from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime, timedelta
from io import BytesIO

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import (
    Q, Count, Sum, F, Case, When, Value as V,
    DecimalField, IntegerField, BigIntegerField,
    CharField, DateField, DateTimeField,
    ExpressionWrapper, Window,
)
from django.db.models.functions import TruncMonth, Coalesce, Cast, Concat
from django.utils import timezone
from django.utils.dateparse import parse_date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import unicodedata

from ..models import (
    Item, Licenca, LicencaLote, MovimentacaoLicenca,
    MovimentacaoItem, TipoMovimentacaoChoices,
    CentroCusto, Fornecedor, Localidade, Subtipo,
    StatusItemChoices, SimNaoChoices, PeriodicidadeChoices,
    TipoMovLicencaChoices, ItemLote,
)

from .equipamentos import _aplicar_filtros_itens


@login_required
def toner_cc_export_excel(request):
    """
    Exporta para Excel o custo de TONER por Centro de Custo, no período filtrado.
    Regra: considerar MovimentacaoItem do tipo 'baixa' cujo item.subtipo contém 'toner',
    agrupando por centro_custo_destino.
    Custo = quantidade * item.valor (se None, trata como 0).
    """
    # --- período (fallback: mês atual até hoje) ---
    hoje = timezone.localdate()
    dt_ini = parse_date(request.GET.get("inicio") or "") or hoje.replace(day=1)
    dt_fim = parse_date(request.GET.get("fim") or "") or hoje

    # --- queryset base: BAIXAS de TONER com CC destino válido ---
    qs = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
            item__subtipo__nome__icontains="toner",
            centro_custo_destino__isnull=False,
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .select_related("item", "centro_custo_destino", "localidade_destino")
    )

    # --- anotações seguras (evita mixed types) ---
    qty_dec = Cast(F("quantidade"), DecimalField(max_digits=12, decimal_places=2))
    item_val = Coalesce(
        Cast(F("item__valor"), DecimalField(max_digits=12, decimal_places=2)),
        V(Decimal("0.00")),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    linha_total = qty_dec * item_val  # Decimal * Decimal
    # Para agrupar por Centro de Custo:
    grp = (
        qs.values(
            "centro_custo_destino",
            "centro_custo_destino__numero",
            "centro_custo_destino__departamento",
        )
        .annotate(
            total_qtd=Coalesce(Sum("quantidade"), V(0)),
            total_valor=Coalesce(
                Sum(Cast(linha_total, DecimalField(max_digits=18, decimal_places=2))),
                V(Decimal("0.00")),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ),
            itens_distintos=Count("item", distinct=True),
            movs=Count("id"),
        )
        .order_by("centro_custo_destino__numero", "centro_custo_destino__departamento")
    )

    # --- planilha ---
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo por CC"

    # estilos
    header_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    header_font = Font(bold=True, color="001e3a")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # cabeçalho
    ws1.append(["Período", f"{dt_ini.strftime('%d/%m/%Y')} a {dt_fim.strftime('%d/%m/%Y')}"])
    ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

    ws1.append(["Centro de Custo", "Departamento", "Movimentações", "Itens Distintos",
                "Quantidade Baixada", "Valor Total (R$)"])
    for c in range(1, 7):
        cell = ws1.cell(row=2, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # linhas
    total_qtd_geral = 0
    total_valor_geral = Decimal("0.00")

    r = 3
    for row in grp:
        cc_num = row["centro_custo_destino__numero"] or "-"
        cc_dep = row["centro_custo_destino__departamento"] or "-"
        movs = row["movs"] or 0
        itens_d = row["itens_distintos"] or 0
        qtd = int(row["total_qtd"] or 0)
        val = Decimal(row["total_valor"] or 0).quantize(Decimal("0.01"))

        ws1.cell(row=r, column=1, value=str(cc_num))
        ws1.cell(row=r, column=2, value=str(cc_dep))
        ws1.cell(row=r, column=3, value=movs)
        ws1.cell(row=r, column=4, value=itens_d)
        ws1.cell(row=r, column=5, value=qtd)
        c6 = ws1.cell(row=r, column=6, value=float(val))
        c6.number_format = 'R$ #,##0.00'

        for c in range(1, 7):
            ws1.cell(row=r, column=c).border = border

        total_qtd_geral += qtd
        total_valor_geral += val
        r += 1

    # totalizador
    ws1.append(["", "", "", "Totais:", total_qtd_geral, float(total_valor_geral)])
    ws1.cell(row=r, column=4).font = Font(bold=True)
    ws1.cell(row=r, column=5).font = Font(bold=True)
    ws1.cell(row=r, column=6).font = Font(bold=True)
    ws1.cell(row=r, column=6).number_format = 'R$ #,##0.00'
    for c in range(1, 7):
        ws1.cell(row=r, column=c).border = border

    # larguras
    widths = [20, 36, 18, 18, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # --- aba de detalhes (opcional, mas útil) ---
    ws2 = wb.create_sheet("Detalhes")
    ws2.append(["Data", "Centro de Custo", "Departamento", "Item", "Subtipo",
                "Quantidade", "Valor Unitário (R$)", "Valor Total (R$)"])

    for c in range(1, 9):
        cell = ws2.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    r = 2
    for m in qs.order_by("centro_custo_destino__numero", "created_at", "id"):
        valor_unit = Decimal(m.item.valor or 0).quantize(Decimal("0.01"))
        qtd = Decimal(m.quantidade or 0)
        val_total = (valor_unit * qtd).quantize(Decimal("0.01"))

        ws2.cell(row=r, column=1, value=m.created_at.strftime("%d/%m/%Y %H:%M"))
        ws2.cell(row=r, column=2, value=getattr(m.centro_custo_destino, "numero", "") or "-")
        ws2.cell(row=r, column=3, value=getattr(m.centro_custo_destino, "departamento", "") or "-")
        ws2.cell(row=r, column=4, value=m.item.nome if m.item_id else "-")
        ws2.cell(row=r, column=5, value=getattr(m.item.subtipo, "nome", "") if m.item_id and m.item.subtipo_id else "-")
        ws2.cell(row=r, column=6, value=int(m.quantidade or 0))
        c7 = ws2.cell(row=r, column=7, value=float(valor_unit))
        c7.number_format = 'R$ #,##0.00'
        c8 = ws2.cell(row=r, column=8, value=float(val_total))
        c8.number_format = 'R$ #,##0.00'

        for c in range(1, 9):
            ws2.cell(row=r, column=c).border = border
        r += 1

    for i, w in enumerate([18, 18, 28, 36, 20, 14, 22, 22], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- resposta HTTP (arquivo xlsx) ---
    filename = f"custo_toner_por_cc_{dt_ini.strftime('%Y%m%d')}_{dt_fim.strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@login_required
def equipamentos_exportar(request):
    base_qs = Item.objects.select_related(
        "centro_custo", "fornecedor", "categoria", "subtipo", "localidade", "locacao",
    )
    qs = _aplicar_filtros_itens(request, base_qs).order_by("id")

    from collections import Counter
    items = list(qs)

    # ── Helpers de formatação de texto ───────────────────────────────────────
    def _sim_nao(valor):
        return "Sim" if str(valor).lower() in ("sim", "true", "1") else "Não"

    def _fmt_cc(obj):
        if not obj:
            return "-"
        numero = getattr(obj, "numero", "") or ""
        departamento = getattr(obj, "departamento", "") or ""
        texto = f"{numero} - {departamento}".strip(" -")
        return texto or "-"

    def _texto_relacionado(obj, attr="nome", fallback="__str__"):
        if not obj:
            return "-"
        if attr and hasattr(obj, attr):
            valor = getattr(obj, attr)
            if valor:
                return str(valor)
        if fallback == "__str__":
            return str(obj)
        return "-"

    def _texto_usuario_auditoria(obj):
        if not obj:
            return "-"
        for campo in ("get_full_name", "username", "nome", "email"):
            if hasattr(obj, campo):
                valor = getattr(obj, campo)
                if callable(valor):
                    valor = valor()
                if valor:
                    return str(valor)
        return str(obj)

    # ── Agregados para o resumo executivo ────────────────────────────────────
    total_itens = len(items)
    valor_total = Decimal("0.00")
    valor_loc_mensal = Decimal("0.00")
    n_consumo = n_locados = n_preventiva = 0
    por_status = Counter()
    por_categoria = Counter()
    valor_por_status = Counter()
    for it in items:
        valor_total += (it.valor or Decimal("0.00"))
        if str(it.item_consumo).lower() == "sim":
            n_consumo += 1
        loc = getattr(it, "locacao", None)
        if str(it.locado).lower() == "sim":
            n_locados += 1
        if loc and loc.valor_mensal:
            valor_loc_mensal += loc.valor_mensal
        if str(it.precisa_preventiva).lower() == "sim":
            n_preventiva += 1
        st = it.get_status_display() if hasattr(it, "get_status_display") else (it.status or "-")
        por_status[st] += 1
        valor_por_status[st] += float(it.valor or 0)
        por_categoria[_texto_relacionado(it.categoria)] += 1

    # ── Paleta / estilos profissionais ───────────────────────────────────────
    BRAND_DARK, BRAND, SOFT, ZEBRA = "0A2540", "1D4ED8", "EEF2F7", "F4F7FB"
    INK, MUTED = "1F2733", "5B6B7F"
    hair = Side(style="thin", color="DCE3EC")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    f_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color=MUTED)
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_bold = Font(name="Calibri", size=10, bold=True, color=INK)
    f_cell = Font(name="Calibri", size=10, color=INK)
    fill_title = PatternFill("solid", fgColor=BRAND_DARK)
    fill_sub = PatternFill("solid", fgColor=SOFT)
    fill_header = PatternFill("solid", fgColor=BRAND)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left", vertical="center")
    a_right = Alignment(horizontal="right", vertical="center")
    a_left_ind = Alignment(horizontal="left", vertical="center", indent=1)
    BRL = 'R$ #,##0.00'
    data_format = 'DD/MM/YYYY'
    data_hora_format = 'DD/MM/YYYY HH:MM'

    def faixa_titulo(ws, ncols, titulo, subtitulo):
        last = get_column_letter(ncols)
        ws.merge_cells(f"A1:{last}1")
        c = ws["A1"]; c.value = titulo; c.font = f_title; c.fill = fill_title
        c.alignment = a_left_ind
        ws.row_dimensions[1].height = 34
        ws.merge_cells(f"A2:{last}2")
        c2 = ws["A2"]; c2.value = subtitulo; c2.font = f_sub; c2.fill = fill_sub
        c2.alignment = a_left_ind
        ws.row_dimensions[2].height = 18
        ws.sheet_view.showGridLines = False

    def cabecalho_tabela(ws, row, headers, center_cols=()):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = fill_header; c.font = f_header; c.border = border
            c.alignment = a_center if ci in center_cols else a_left
        ws.row_dimensions[row].height = 22

    def _status_fill(raw):
        s = (raw or "").lower()
        if s == "ativo":
            return "E6F4EA"
        if s in ("manutencao", "correcao"):
            return "FFF3E0"
        if s in ("defeito", "queimado"):
            return "FDECEA"
        if s == "descarte":
            return "E8EAEE"
        return None

    gerado = timezone.localtime().strftime("%d/%m/%Y às %H:%M")
    wb = Workbook()

    # =========================================================================
    # ABA 1 — RESUMO
    # =========================================================================
    wsr = wb.active
    wsr.title = "Resumo"
    faixa_titulo(wsr, 6, "INVENTÁRIO DE EQUIPAMENTOS",
                 f"Santa Colomba Agropecuária  ·  Relatório gerado em {gerado}  ·  {total_itens} item(ns)")

    kpis = [
        ("TOTAL DE ITENS", total_itens, "334155", None),
        ("VALOR DE AQUISIÇÃO", float(valor_total), "0A2540", BRL),
        ("ITENS DE CONSUMO", n_consumo, "EA580C", None),
        ("ITENS LOCADOS", n_locados, "7C3AED", None),
        ("LOCAÇÃO MENSAL", float(valor_loc_mensal), "1E8E3E", BRL),
        ("PRECISAM PREVENTIVA", n_preventiva, "D93025", None),
    ]
    rk = 4
    for idx, (lbl, val, color, fmt) in enumerate(kpis):
        col = idx + 1
        cl = wsr.cell(row=rk, column=col, value=lbl)
        cl.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        cl.fill = PatternFill("solid", fgColor=color); cl.alignment = a_center; cl.border = border
        cv = wsr.cell(row=rk + 1, column=col, value=val)
        cv.font = Font(name="Calibri", size=15, bold=True, color=color)
        cv.fill = PatternFill("solid", fgColor="F4F6F9"); cv.alignment = a_center; cv.border = border
        if fmt:
            cv.number_format = fmt
    wsr.row_dimensions[rk].height = 16
    wsr.row_dimensions[rk + 1].height = 28
    for c in range(1, 7):
        wsr.column_dimensions[get_column_letter(c)].width = 20

    def _hcell(rr, col, val, align):
        c = wsr.cell(row=rr, column=col, value=val)
        c.fill = fill_header; c.font = f_header; c.border = border; c.alignment = align
        return c

    # Quebra por Status (Status | Qtd. | Valor de Aquisição)
    r = rk + 3
    _hcell(r, 1, "Status", a_left)
    _hcell(r, 2, "Qtd.", a_center)
    wsr.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    _hcell(r, 3, "Valor de Aquisição", a_center)
    _hcell(r, 4, None, a_center)
    wsr.row_dimensions[r].height = 22
    r += 1
    for i, (st, qt) in enumerate(por_status.most_common()):
        zebra = (i % 2 == 1)
        c1 = wsr.cell(row=r, column=1, value=st)
        c2 = wsr.cell(row=r, column=2, value=qt)
        wsr.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c3 = wsr.cell(row=r, column=3, value=float(valor_por_status[st]))
        c1.alignment, c2.alignment, c3.alignment = a_left, a_center, a_right
        c3.number_format = BRL
        for cc in (c1, c2, c3, wsr.cell(row=r, column=4)):
            cc.border = border; cc.font = f_cell
            if zebra:
                cc.fill = fill_zebra
        r += 1

    # Quebra por Categoria (top 12) — Categoria (merge A:B) | Qtd. (C)
    r += 1
    wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    _hcell(r, 1, "Categoria (top 12)", a_left)
    _hcell(r, 2, None, a_left)
    _hcell(r, 3, "Qtd.", a_center)
    wsr.row_dimensions[r].height = 22
    r += 1
    for i, (cat, qt) in enumerate(por_categoria.most_common(12)):
        zebra = (i % 2 == 1)
        wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c1 = wsr.cell(row=r, column=1, value=cat)
        c2 = wsr.cell(row=r, column=2)
        c3 = wsr.cell(row=r, column=3, value=qt)
        c1.alignment, c3.alignment = a_left, a_center
        for cc in (c1, c2, c3):
            cc.border = border; cc.font = f_cell
            if zebra:
                cc.fill = fill_zebra
        r += 1

    # =========================================================================
    # ABA 2 — ITENS (detalhado)
    # =========================================================================
    ws = wb.create_sheet(title="Itens")
    header = [
        "#", "ID", "Nome", "Número de Série", "Marca", "Modelo", "Quantidade",
        "Item de Consumo", "PMB", "Valor de Aquisição (R$)", "Status", "Fornecedor",
        "Categoria", "Subtipo", "Localidade", "Centro de Custo", "Precisa Preventiva",
        "Data Limite Preventiva (dias)", "Data da Compra", "Número do Pedido",
        "Observações do Item", "Locado", "Tempo Locado (meses)", "Valor Locação Mensal (R$)",
        "Data de Entrada Locação", "Contrato Locação", "Observações da Locação",
        "Fornecedor da Locação", "Criado em", "Criado por", "Atualizado em", "Atualizado por",
    ]
    ncols = len(header)
    faixa_titulo(ws, ncols, "EQUIPAMENTOS — DETALHADO",
                 f"{total_itens} item(ns) exportado(s)  ·  gerado em {gerado}")
    HEADER_ROW = 3
    center_cols = {1, 2, 7, 8, 9, 11, 17, 18, 22, 23}
    cabecalho_tabela(ws, HEADER_ROW, header, center_cols=center_cols)

    row = HEADER_ROW + 1
    for i, item in enumerate(items, start=1):
        locacao = getattr(item, "locacao", None)
        status_raw = item.status or ""
        status_txt = item.get_status_display() if hasattr(item, "get_status_display") else (status_raw or "-")
        item_consumo_txt = item.get_item_consumo_display() if hasattr(item, "get_item_consumo_display") else _sim_nao(item.item_consumo)
        pmb_txt = item.get_pmb_display() if hasattr(item, "get_pmb_display") else _sim_nao(item.pmb)
        preventiva_txt = item.get_precisa_preventiva_display() if hasattr(item, "get_precisa_preventiva_display") else _sim_nao(item.precisa_preventiva)
        locado_txt = item.get_locado_display() if hasattr(item, "get_locado_display") else _sim_nao(item.locado)
        valor_item = item.valor if item.valor is not None else Decimal("0.00")
        valor_locacao = locacao.valor_mensal if (locacao and locacao.valor_mensal is not None) else Decimal("0.00")

        valores = [
            i, item.id, item.nome or "", item.numero_serie or "", item.marca or "",
            item.modelo or "", item.quantidade or 0, item_consumo_txt, pmb_txt,
            float(valor_item), status_txt, _texto_relacionado(item.fornecedor),
            _texto_relacionado(item.categoria), _texto_relacionado(item.subtipo),
            _texto_relacionado(item.localidade, attr="local"), _fmt_cc(item.centro_custo),
            preventiva_txt,
            item.data_limite_preventiva if item.data_limite_preventiva is not None else "",
            item.data_compra, item.numero_pedido or "", item.observacoes or "", locado_txt,
            locacao.tempo_locado if locacao and locacao.tempo_locado is not None else "",
            float(valor_locacao), locacao.data_entrada if locacao else "",
            locacao.contrato if locacao and locacao.contrato else "",
            locacao.observacoes if locacao and locacao.observacoes else "",
            _texto_relacionado(locacao.fornecedor) if locacao else "-",
            getattr(item, "criado_em", None), _texto_usuario_auditoria(getattr(item, "criado_por", None)),
            getattr(item, "atualizado_em", None), _texto_usuario_auditoria(getattr(item, "atualizado_por", None)),
        ]
        zebra = (i % 2 == 0)
        st_fill = _status_fill(status_raw)
        for ci, val in enumerate(valores, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = border
            c.font = f_cell
            c.alignment = a_center if ci in center_cols else a_left
            if ci == 10 or ci == 24:
                c.number_format = BRL
                c.alignment = a_right
            elif ci == 19 and val:
                c.number_format = data_format
            elif ci == 25 and val:
                c.number_format = data_format
            elif ci in (29, 31) and val:
                c.number_format = data_hora_format
            if ci == 11 and st_fill:
                c.fill = PatternFill("solid", fgColor=st_fill)
                c.font = f_bold
                c.alignment = a_center
            elif zebra:
                c.fill = fill_zebra
        row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # Largura automática a partir do cabeçalho + dados (ignora a faixa de título)
    widths = {}
    for r_ in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        for idx, val in enumerate(r_, start=1):
            texto = str(val) if val is not None else ""
            widths[idx] = max(widths.get(idx, 0), len(texto))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 11), 42)

    # ── Resposta ─────────────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    now = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    filename = f"equipamentos_{now}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def _parse_dt(value, default):
    if not value:
        return default

    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return default


def _model_has_field(model, field_name):
    try:
        model._meta.get_field(field_name)
        return True
    except FieldDoesNotExist:
        return False


def _related_model_has_field(model, fk_field, field_name):
    try:
        field = model._meta.get_field(fk_field)
        related_model = field.remote_field.model
        related_model._meta.get_field(field_name)
        return True
    except Exception:
        return False


def _money(value):
    if value is None:
        return Decimal("0.00")

    try:
        return Decimal(value)
    except Exception:
        return Decimal("0.00")


def _percent(part, total):
    part = _money(part)
    total = _money(total)

    if total <= 0:
        return Decimal("0.00")

    return (part / total) * Decimal("100.00")


def _cc_display(numero, departamento):
    numero = numero or ""
    departamento = departamento or ""

    if numero and departamento:
        return f"{numero} - {departamento}"

    if departamento:
        return departamento

    if numero:
        return str(numero)

    return "Centro de custo não informado"


def _build_toner_filter():
    """
    Filtro seguro para localizar itens de toner.
    Usa categoria somente se o model Item realmente tiver esse campo.
    """
    query = Q(item__nome__icontains="toner")

    if _related_model_has_field(Item, "subtipo", "nome"):
        query |= Q(item__subtipo__nome__icontains="toner")

    if _related_model_has_field(Item, "categoria", "nome"):
        query |= Q(item__categoria__nome__icontains="toner")

    return query


@login_required
def toner_cc_dashboard(request):
    hoje = date.today()

    dt_ini = _parse_dt(
        request.GET.get("inicio"),
        date(hoje.year, 1, 1),
    )

    dt_fim = _parse_dt(
        request.GET.get("fim"),
        hoje,
    )

    if dt_ini > dt_fim:
        dt_ini, dt_fim = dt_fim, dt_ini

    filtro_item = request.GET.get("item", "").strip()
    filtro_cc = request.GET.get("cc", "").strip()
    filtro_usuario = request.GET.get("usuario", "").strip()

    dec14_2 = DecimalField(max_digits=14, decimal_places=2)

    qtd_dec = Cast(
        Coalesce(F("quantidade"), V(0)),
        output_field=dec14_2,
    )

    preco_item = Coalesce(
        F("item__valor"),
        V(Decimal("0.00"), output_field=dec14_2),
        output_field=dec14_2,
    )

    custo_total_expr = Case(
        When(custo__gt=Decimal("0.00"), then=F("custo")),
        default=ExpressionWrapper(qtd_dec * preco_item, output_field=dec14_2),
        output_field=dec14_2,
    )

    base = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao="baixa",
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .filter(_build_toner_filter())
        .select_related(
            "item",
            "item__subtipo",
            "item__centro_custo",
            "usuario",
            "criado_por",
            "centro_custo_origem",
            "centro_custo_destino",
            "lote",
        )
        .annotate(custo_calc=custo_total_expr)
    )

    if filtro_item:
        base = base.filter(
            Q(item__nome__icontains=filtro_item)
            | Q(item__numero_serie__icontains=filtro_item)
            | Q(item__modelo__icontains=filtro_item)
            | Q(item__marca__icontains=filtro_item)
        )

    if filtro_cc:
        base = base.filter(
            Q(centro_custo_destino__numero__icontains=filtro_cc)
            | Q(centro_custo_destino__departamento__icontains=filtro_cc)
            | Q(centro_custo_origem__numero__icontains=filtro_cc)
            | Q(centro_custo_origem__departamento__icontains=filtro_cc)
            | Q(item__centro_custo__numero__icontains=filtro_cc)
            | Q(item__centro_custo__departamento__icontains=filtro_cc)
        )

    if filtro_usuario:
        base = base.filter(
            Q(usuario__nome__icontains=filtro_usuario)
            | Q(usuario__email__icontains=filtro_usuario)
            | Q(criado_por__username__icontains=filtro_usuario)
            | Q(criado_por__first_name__icontains=filtro_usuario)
            | Q(criado_por__last_name__icontains=filtro_usuario)
        )

    # Centro de custo de consumo:
    # 1º destino da baixa, 2º origem, 3º centro atual do item.
    base_cc = base.annotate(
        cc_id=Coalesce(
            F("centro_custo_destino_id"),
            F("centro_custo_origem_id"),
            F("item__centro_custo_id"),
        ),
        cc_numero=Case(
            When(centro_custo_destino__isnull=False, then=F("centro_custo_destino__numero")),
            When(centro_custo_origem__isnull=False, then=F("centro_custo_origem__numero")),
            default=F("item__centro_custo__numero"),
            output_field=CharField(),
        ),
        cc_departamento=Case(
            When(centro_custo_destino__isnull=False, then=F("centro_custo_destino__departamento")),
            When(centro_custo_origem__isnull=False, then=F("centro_custo_origem__departamento")),
            default=F("item__centro_custo__departamento"),
            output_field=CharField(),
        ),
    )

    por_cc_qs = (
        base_cc
        .values("cc_id", "cc_numero", "cc_departamento")
        .annotate(
            qtd=Coalesce(Sum("quantidade"), V(0)),
            movimentos=Count("id"),
            gasto=Coalesce(
                Sum("custo_calc", output_field=dec14_2),
                V(Decimal("0.00"), output_field=dec14_2),
            ),
        )
        .order_by("-gasto", "-qtd")
    )

    linhas = []
    cc_labels = []
    cc_gasto = []

    total_geral = Decimal("0.00")
    total_qtd = 0
    total_movimentos = 0

    top_cc_nome = "—"
    top_cc_valor = Decimal("0.00")

    for idx, row in enumerate(por_cc_qs):
        gasto = _money(row["gasto"])
        qtd = int(row["qtd"] or 0)
        movimentos = int(row["movimentos"] or 0)

        cc_nome = _cc_display(
            row.get("cc_numero"),
            row.get("cc_departamento"),
        )

        if idx == 0:
            top_cc_nome = cc_nome
            top_cc_valor = gasto

        total_geral += gasto
        total_qtd += qtd
        total_movimentos += movimentos

        linhas.append({
            "cc": cc_nome,
            "qtd": qtd,
            "movimentos": movimentos,
            "gasto": gasto,
            "ticket": (gasto / qtd) if qtd > 0 else Decimal("0.00"),
            "percentual": Decimal("0.00"),
        })

        cc_labels.append(cc_nome)
        cc_gasto.append(float(gasto))

    for linha in linhas:
        linha["percentual"] = _percent(linha["gasto"], total_geral)

    ticket_medio = (total_geral / total_qtd) if total_qtd > 0 else Decimal("0.00")

    por_user_qs = (
        base
        .annotate(
            consumidor_nome=Case(
                When(usuario__isnull=False, then=F("usuario__nome")),
                When(criado_por__isnull=False, then=F("criado_por__username")),
                default=V("Sem solicitante"),
                output_field=CharField(),
            )
        )
        .values("consumidor_nome")
        .annotate(
            qtd=Coalesce(Sum("quantidade"), V(0)),
            gasto=Coalesce(
                Sum("custo_calc", output_field=dec14_2),
                V(Decimal("0.00"), output_field=dec14_2),
            ),
        )
        .order_by("-gasto", "-qtd")[:10]
    )

    user_labels = []
    user_gasto = []
    top_user_nome = "—"

    for idx, row in enumerate(por_user_qs):
        nome = row["consumidor_nome"] or "Sem solicitante"
        gasto = _money(row["gasto"])

        if idx == 0:
            top_user_nome = nome

        user_labels.append(nome)
        user_gasto.append(float(gasto))

    por_item_qs = (
        base
        .values("item__id", "item__nome", "item__numero_serie")
        .annotate(
            qtd=Coalesce(Sum("quantidade"), V(0)),
            movimentos=Count("id"),
            gasto=Coalesce(
                Sum("custo_calc", output_field=dec14_2),
                V(Decimal("0.00"), output_field=dec14_2),
            ),
        )
        .order_by("-gasto", "-qtd")[:10]
    )

    top_itens = []
    item_labels = []
    item_gasto = []
    top_item_nome = "—"

    for idx, row in enumerate(por_item_qs):
        nome_item = row["item__nome"] or "Item sem nome"
        gasto = _money(row["gasto"])
        qtd = int(row["qtd"] or 0)

        if idx == 0:
            top_item_nome = nome_item

        top_itens.append({
            "nome": nome_item,
            "numero_serie": row["item__numero_serie"] or "—",
            "qtd": qtd,
            "movimentos": int(row["movimentos"] or 0),
            "gasto": gasto,
            "ticket": (gasto / qtd) if qtd > 0 else Decimal("0.00"),
        })

        item_labels.append(nome_item)
        item_gasto.append(float(gasto))

    mensal_qs = (
        base
        .annotate(mes=TruncMonth("created_at"))
        .values("mes")
        .annotate(
            qtd=Coalesce(Sum("quantidade"), V(0)),
            gasto=Coalesce(
                Sum("custo_calc", output_field=dec14_2),
                V(Decimal("0.00"), output_field=dec14_2),
            ),
        )
        .order_by("mes")
    )

    # Série mensal CONTÍNUA: o TruncMonth omite meses sem baixa, o que distorce
    # a "evolução do gasto". Preenchemos todos os meses do período com zero.
    mensal_map = {}
    for row in mensal_qs:
        mes = row["mes"]
        if not mes:
            continue
        mensal_map[(mes.year, mes.month)] = (
            float(_money(row["gasto"])),
            int(row["qtd"] or 0),
        )

    _MESES_PT = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                 "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

    mes_labels = []
    mes_gasto = []
    mes_qtd = []

    cursor = date(dt_ini.year, dt_ini.month, 1)
    limite = date(dt_fim.year, dt_fim.month, 1)
    guard = 0
    while cursor <= limite and guard < 120:
        g, q = mensal_map.get((cursor.year, cursor.month), (0.0, 0))
        mes_labels.append(f"{_MESES_PT[cursor.month]}/{str(cursor.year)[2:]}")
        mes_gasto.append(g)
        mes_qtd.append(q)
        cursor = (
            date(cursor.year + 1, 1, 1)
            if cursor.month == 12
            else date(cursor.year, cursor.month + 1, 1)
        )
        guard += 1

    mes_media = round(sum(mes_gasto) / len(mes_gasto), 2) if mes_gasto else 0.0

    periodo_dias = (dt_fim - dt_ini).days + 1
    gasto_dia = (total_geral / Decimal(periodo_dias)) if periodo_dias > 0 else Decimal("0.00")
    gasto_projetado_30d = gasto_dia * Decimal("30")

    # ── Tempo médio entre baixas por centro de custo (cadência de consumo) ──
    from collections import defaultdict as _dd
    _datas_cc = _dd(list)
    _nome_cc = {}
    for _r in base_cc.values("cc_id", "cc_numero", "cc_departamento", "created_at").order_by("cc_id", "created_at"):
        _cid = _r["cc_id"]
        _datas_cc[_cid].append(_r["created_at"])
        if _cid not in _nome_cc:
            _nome_cc[_cid] = _cc_display(_r.get("cc_numero"), _r.get("cc_departamento"))

    tempo_medio_cc = []
    for _cid, _datas in _datas_cc.items():
        _n = len(_datas)
        if _n >= 2:
            _span = (_datas[-1] - _datas[0]).days
            _dias = round(_span / (_n - 1), 1)
        else:
            _dias = None
        tempo_medio_cc.append({"cc": _nome_cc[_cid], "baixas": _n, "dias_medio": _dias})
    tempo_medio_cc.sort(key=lambda x: (x["dias_medio"] is None, x["dias_medio"] if x["dias_medio"] is not None else 0))

    # ── Dados para a exportação em imagem (resumo executivo) ──
    resumo_export = {
        "periodo": f"{dt_ini:%d/%m/%Y} — {dt_fim:%d/%m/%Y}",
        "periodo_dias": periodo_dias,
        "gerado_em": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
        "kpi": {
            "gasto_total": float(total_geral),
            "qtd": total_qtd,
            "movimentos": total_movimentos,
            "ticket_medio": float(ticket_medio),
        },
        "centros": [
            {"cc": l["cc"], "qtd": l["qtd"], "gasto": float(l["gasto"]), "pct": float(l["percentual"])}
            for l in linhas
        ],
        "itens": [
            {"nome": it["nome"], "qtd": it["qtd"], "gasto": float(it["gasto"])}
            for it in top_itens
        ],
        "tempo": tempo_medio_cc,
    }

    if request.GET.get("export") == "1":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        filename = f"dashboard_toner_{dt_ini:%Y%m%d}_{dt_fim:%Y%m%d}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")

        writer.writerow([
            "Centro de Custo",
            "Quantidade",
            "Movimentações",
            "Gasto R$",
            "% do Total",
            "Ticket Médio R$",
        ])

        for linha in linhas:
            writer.writerow([
                linha["cc"],
                linha["qtd"],
                linha["movimentos"],
                f"{linha['gasto']:.2f}".replace(".", ","),
                f"{linha['percentual']:.2f}".replace(".", ","),
                f"{linha['ticket']:.2f}".replace(".", ","),
            ])

        writer.writerow([])
        writer.writerow(["Total", total_qtd, total_movimentos, f"{total_geral:.2f}".replace(".", ","), "100,00", f"{ticket_medio:.2f}".replace(".", ",")])

        return response

    ctx = {
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "periodo_dias": periodo_dias,
        "f_item": filtro_item,
        "f_cc": filtro_cc,
        "f_usuario": filtro_usuario,

        "linhas": linhas,
        "top_itens": top_itens,

        "cc_labels": cc_labels,
        "cc_gasto": cc_gasto,

        "user_labels": user_labels,
        "user_gasto": user_gasto,

        "item_labels": item_labels,
        "item_gasto": item_gasto,

        "mes_labels": mes_labels,
        "mes_gasto": mes_gasto,
        "mes_qtd": mes_qtd,
        "mes_media": mes_media,

        "kpi_total_gasto": total_geral,
        "kpi_total_qtd": total_qtd,
        "kpi_total_movimentos": total_movimentos,
        "kpi_ticket_medio": ticket_medio,
        "kpi_top_cc_nome": top_cc_nome,
        "kpi_top_cc_valor": top_cc_valor,
        "kpi_top_user_nome": top_user_nome,
        "kpi_top_item_nome": top_item_nome,
        "kpi_gasto_dia": gasto_dia,
        "kpi_gasto_projetado_30d": gasto_projetado_30d,

        "resumo_export": resumo_export,
    }

    return render(request, "front/dashboards/dashboard_toner.html", ctx)

    ##### EXPORTAR EXCEL #################

@login_required
def custo_cc_export_excel(request):
    """Exporta para Excel a mesma tabela do cc_custos_dashboard, respeitando os filtros."""
    # ── filtros (iguais ao dashboard) ─────────────────────────────────────────────
    hoje = datetime.today().date()
    dt_ini = _parse_dt(request.GET.get("inicio"), hoje - timedelta(days=30))
    dt_fim = _parse_dt(request.GET.get("fim"), hoje)

    # ── imports locais (mesmos modelos usados no dashboard) ──────────────────────
    from ..models import (
        Locacao, MovimentacaoLicenca, MovimentacaoItem, CentroCusto,
        Usuario, Item, TipoMovLicencaChoices, TipoMovimentacaoChoices
    )

    # ── acumulador por CC (mesma estrutura) ──────────────────────────────────────
    totals = {}  # cc_id -> dict

    def acc(cc_id):
        if not cc_id:
            return None
        if cc_id not in totals:
            totals[cc_id] = {
                "cc": None,
                "usuarios": 0,
                "itens": 0,
                "licencas_set": set(),
                "assentos": 0,
                "custo_itens": Decimal("0.00"),
                "custo_licencas": Decimal("0.00"),
                "baixas": Decimal("0.00"),
            }
        return totals[cc_id]

    # ── custo mensal de ITENS (locações) por CC ──────────────────────────────────
    loc_qs = (
        Locacao.objects
        .select_related("equipamento", "equipamento__centro_custo")
        .exclude(valor_mensal__isnull=True)
    )
    for loc in loc_qs:
        item = loc.equipamento
        cc_id = getattr(item.centro_custo, "id", None)
        if not cc_id:
            continue
        valor = loc.valor_mensal or Decimal("0.00")
        if valor > 0:
            a = acc(cc_id)
            a["custo_itens"] += valor

    # ── assentos/licenças (último evento por par licença/usuário) ───────────────
    mov_l_qs = (
        MovimentacaoLicenca.objects
        .select_related("licenca", "usuario__centro_custo", "centro_custo_destino", "lote")
        .order_by("licenca_id", "usuario_id", "created_at", "id")
    )

    last_by_pair = {}
    for m in mov_l_qs:
        if m.usuario_id is None:
            continue
        last_by_pair[(m.licenca_id, m.usuario_id)] = m

    for (lic_id, user_id), m in last_by_pair.items():
        if m.tipo != TipoMovLicencaChoices.ATRIBUICAO:
            continue
        cc_id = (
            getattr(getattr(m.usuario, "centro_custo", None), "id", None)
            or getattr(m.centro_custo_destino, "id", None)
            or getattr(m.licenca.centro_custo, "id", None)
        )
        if not cc_id:
            continue

        cm = m.valor_unitario or Decimal("0.00")
        a = acc(cc_id)
        a["assentos"] += 1
        a["custo_licencas"] += cm
        a["licencas_set"].add(lic_id)

    # ── baixas no período ────────────────────────────────────────────────────────
    baixas_qs = (
        MovimentacaoItem.objects
        .filter(
            tipo_movimentacao=TipoMovimentacaoChoices.BAIXA,
            created_at__date__gte=dt_ini,
            created_at__date__lte=dt_fim,
        )
        .select_related("item__centro_custo", "centro_custo_origem")
    )
    for mv in baixas_qs:
        cc_id = (
            getattr(mv.centro_custo_origem, "id", None)
            or getattr(getattr(mv.item, "centro_custo", None), "id", None)
        )
        if not cc_id:
            continue
        valor_baixa = mv.custo if mv.custo is not None else (mv.item.valor or Decimal("0.00")) * (mv.quantidade or 1)
        a = acc(cc_id)
        a["baixas"] += (valor_baixa or Decimal("0.00"))

    # ── metadados: usuários e itens por CC ───────────────────────────────────────
    cc_ids = list(totals.keys())
    ccs = {cc.id: cc for cc in CentroCusto.objects.filter(id__in=cc_ids)}

    users_count = (
        Usuario.objects
        .filter(centro_custo_id__in=cc_ids, status="ativo")
        .values("centro_custo_id")
        .annotate(n=Count("id"))
    )
    itens_count = (
        Item.objects
        .filter(centro_custo_id__in=cc_ids)
        .values("centro_custo_id")
        .annotate(n=Count("id"))
    )
    map_users = {r["centro_custo_id"]: r["n"] for r in users_count}
    map_itens = {r["centro_custo_id"]: r["n"] for r in itens_count}

    # ── LINHAS (igual ao dashboard) ──────────────────────────────────────────────
    linhas = []
    for cc_id, d in totals.items():
        cc = ccs.get(cc_id)
        if not cc:
            continue
        d["cc"] = cc
        d["usuarios"] = map_users.get(cc_id, 0)
        d["itens"] = map_itens.get(cc_id, 0)
        lic_tipos = len(d["licencas_set"])
        d["licencas"] = lic_tipos
        d["total_mensal"] = (d["custo_itens"] + d["custo_licencas"])
        d["total_geral"] = (d["total_mensal"] + d["baixas"])

        linhas.append({
            "cc": cc,
            "usuarios": d["usuarios"],
            "itens": d["itens"],
            "licencas": d["licencas"],
            "assentos": d["assentos"],
            "custo_itens": d["custo_itens"],
            "custo_licencas": d["custo_licencas"],
            "baixas": d["baixas"],
            "total_mensal": d["total_mensal"],
            "total_geral": d["total_geral"],
        })

    # mesma ordenação da tela
    linhas.sort(key=lambda x: x["total_geral"], reverse=True)

    # ── EXCEL (apenas tabela detalhamento + resumo) ──────────────────────────────
    wb = Workbook()

    # Aba 1: Detalhamento (títulos iguais aos da tabela do template)
    ws = wb.active
    ws.title = "Detalhamento"

    headers = [
        "Centro de Custo",
        "Usuários",
        "Itens",
        "Licenças",
        "Assentos ativos",
        "Custo Itens (R$/mês)",
        "Custo Licenças (R$/mês)",
        "Baixas no período (R$)",
        "Total Mensal (R$)",
        "Total Geral (R$)",
    ]
    ws.append(headers)

    # estilos
    header_fill = PatternFill("solid", fgColor="FF1D4ED8")
    header_font = Font(color="FFFFFFFF", bold=True)
    thin = Side(style="thin", color="FFCBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    int_fmt = "#,##0"
    money_fmt = "[$R$-pt-BR] #,##0.00"

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    def _cc_nome(ccobj):
        try:
            return f"{ccobj.numero} - {ccobj.departamento}"
        except Exception:
            return str(ccobj) if ccobj else "—"

    for l in linhas:
        ws.append([
            _cc_nome(l["cc"]),
            int(l["usuarios"] or 0),
            int(l["itens"] or 0),
            int(l["licencas"] or 0),
            int(l["assentos"] or 0),
            Decimal(l["custo_itens"] or 0),
            Decimal(l["custo_licencas"] or 0),
            Decimal(l["baixas"] or 0),
            Decimal(l["total_mensal"] or 0),
            Decimal(l["total_geral"] or 0),
        ])

    last_row = ws.max_row
    for r in range(2, last_row + 1):
        for col in (2, 3, 4, 5):
            ws.cell(row=r, column=col).number_format = int_fmt
            ws.cell(row=r, column=col).border = border_all
        for col in (6, 7, 8, 9, 10):
            ws.cell(row=r, column=col).number_format = money_fmt
            ws.cell(row=r, column=col).border = border_all
        ws.cell(row=r, column=1).border = border_all

    # transforma em "tabela" do Excel (zebra)
    ref = f"A1:J{last_row}"
    table = Table(displayName="tb_detalhamento", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    # larguras e freeze
    widths = [28, 12, 10, 12, 16, 22, 22, 20, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "A2"

    # Aba 2: Resumo rápido
    ws2 = wb.create_sheet("Resumo")
    ws2["A1"] = "Período"
    ws2["B1"] = f"{dt_ini:%d/%m/%Y} — {dt_fim:%d/%m/%Y}"
    ws2["A1"].font = Font(bold=True)
    ws2["A3"] = "Centros de Custo (linhas)"
    ws2["B3"] = len(linhas)
    ws2["A4"] = "Total Geral (soma)"
    if last_row >= 2:
        ws2["B4"] = f"=SUM(Detalhamento!J2:J{last_row})"
        ws2["B4"].number_format = money_fmt
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 36

    # resposta HTTP
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="custo_cc_{dt_ini:%Y%m%d}_{dt_fim:%Y%m%d}.xlsx"'
    wb.save(resp)
    return resp

try:
    from ..models import LicencaLote
except Exception:
    LicencaLote = None

def _parse_date_opt(date_str):
    if not date_str: return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return None

def _get_meses_ciclo(periodicidade_str):
    """
    Converte periodicidade em meses para cálculo do custo mensal.
    """
    if not periodicidade_str: return 1
    p = str(periodicidade_str).upper()
    if 'MEN' in p: return 1
    if 'BI' in p: return 2
    if 'TRI' in p: return 3
    if 'SEM' in p: return 6
    if 'ANU' in p: return 12
    return 1 

@login_required
def licencas_dashboard(request):
    hoje = timezone.localdate()

    q = (request.GET.get("q") or "").strip()
    fornecedor_id = (request.GET.get("fornecedor") or "").strip()
    cc_id = (request.GET.get("centro_custo") or "").strip()
    periodicidade = (request.GET.get("periodicidade") or "").strip()
    pmb = (request.GET.get("pmb") or "").strip().lower()
    # Filtro PMB por COLABORADOR: "pertence ao PMB" = campo próprio do usuário
    # (Usuario.pmb) E o Centro de Custo do usuário também marcado como PMB
    # (centro_custo.pmb). Restringe a consulta ao custo das licenças atribuídas a
    # essas pessoas. É distinto do filtro `pmb` acima (que é a flag da Licenca).
    pmb_pessoa = (request.GET.get("pmb_pessoa") or "").strip().lower()

    dt_ini = _parse_date_opt(request.GET.get("inicio"))
    dt_fim = _parse_date_opt(request.GET.get("fim"))

    qs_licencas = Licenca.objects.select_related(
        "fornecedor", "centro_custo"
    ).prefetch_related(
        "lotes", "lotes__fornecedor", "lotes__centro_custo"
    )

    if q:
        qs_licencas = qs_licencas.filter(
            Q(nome__icontains=q) | Q(fornecedor__nome__icontains=q)
        )
    if fornecedor_id:
        qs_licencas = qs_licencas.filter(fornecedor_id=fornecedor_id)
    if pmb in ["sim", "nao"]:
        qs_licencas = qs_licencas.filter(pmb=pmb)

    if periodicidade:
        qs_licencas = qs_licencas.filter(lotes__periodicidade=periodicidade)
    if cc_id:
        qs_licencas = qs_licencas.filter(lotes__centro_custo_id=cc_id)
    if dt_ini:
        qs_licencas = qs_licencas.filter(lotes__data_compra__gte=dt_ini)
    if dt_fim:
        qs_licencas = qs_licencas.filter(lotes__data_compra__lte=dt_fim)

    qs_licencas = qs_licencas.distinct()
    licencas_list = list(qs_licencas)
    licenca_ids = [l.id for l in licencas_list]

    movs_ativas = (
        MovimentacaoLicenca.objects
        .filter(licenca_id__in=licenca_ids, usuario__isnull=False)
        .select_related("usuario__centro_custo", "usuario__funcao", "centro_custo_destino")
        .order_by("licenca_id", "usuario_id", "created_at")
    )

    # Filtro PMB (colaborador): pessoa PMB = Usuario.pmb='sim' E centro de custo PMB.
    if pmb_pessoa == "sim":
        movs_ativas = movs_ativas.filter(usuario__pmb="sim", usuario__centro_custo__pmb="sim")
    elif pmb_pessoa == "nao":
        movs_ativas = movs_ativas.exclude(usuario__pmb="sim", usuario__centro_custo__pmb="sim")

    estado_usuario = {}
    for m in movs_ativas:
        estado_usuario[(m.licenca_id, m.usuario_id)] = m

    uso_map_cc = {}
    for (lid, uid), mov in estado_usuario.items():
        if mov.tipo == "atribuicao":
            if lid not in uso_map_cc:
                uso_map_cc[lid] = []

            cc_nome = "Indefinido"
            if mov.centro_custo_destino:
                cc_nome = f"{mov.centro_custo_destino.numero} - {mov.centro_custo_destino.departamento}"
            elif mov.usuario and mov.usuario.centro_custo:
                cc_nome = f"{mov.usuario.centro_custo.numero} - {mov.usuario.centro_custo.departamento}"

            uso_map_cc[lid].append(cc_nome)

    kpi_total_licencas = len(licencas_list)
    kpi_assentos_em_uso = 0
    kpi_seats_total = 0
    kpi_custo_mensal = Decimal("0.00")
    # Com o filtro PMB (colaborador) ativo, a consulta passa a mostrar SÓ o custo
    # dos assentos atribuídos a pessoas PMB (sem a parcela de estoque).
    pmb_pessoa_ativo = pmb_pessoa in ("sim", "nao")

    cc_costs = {}
    forn_costs = {}
    per_counts = {}
    # Metadados por licença (só das que entram na consulta) p/ montar a visão por
    # colaborador (janela "Colaboradores"). custo_mensal_unit = custo médio mensal
    # por assento, o mesmo valor atribuído a cada pessoa que detém a licença.
    licenca_meta = {}

    linhas_tabela = []
    lotes_detalhes = []

    for lic in licencas_list:
        lotes_da_lic = list(lic.lotes.all())

        if periodicidade:
            lotes_da_lic = [l for l in lotes_da_lic if l.periodicidade == periodicidade]
        if cc_id:
            lotes_da_lic = [l for l in lotes_da_lic if l.centro_custo_id == int(cc_id)]
        if dt_ini:
            lotes_da_lic = [l for l in lotes_da_lic if l.data_compra and l.data_compra >= dt_ini]
        if dt_fim:
            lotes_da_lic = [l for l in lotes_da_lic if l.data_compra and l.data_compra <= dt_fim]

        if not lotes_da_lic:
            continue

        l_qtd_total = 0
        l_qtd_disp = 0
        l_custo_mensal_total_licenca = Decimal("0.00")
        l_custo_anual_total_licenca = Decimal("0.00")
        l_periodicidades = set()
        soma_custo_mensal_de_todos_lotes = Decimal("0.00")
        # Buffers commitados só se a licença permanecer na consulta (com o filtro
        # PMB, licenças sem nenhuma pessoa PMB são descartadas mais abaixo).
        lotes_buf = []
        per_labels_buf = []

        for lote in lotes_da_lic:
            qtd = int(lote.quantidade_total or 0)
            disp = int(lote.quantidade_disponivel or 0)
            custo_ciclo_lote = Decimal(lote.custo_ciclo or 0)

            if qtd <= 0:
                custo_unit_ciclo = Decimal("0.00")
                custo_mensal_unit = Decimal("0.00")
                custo_anual_unit = Decimal("0.00")
                custo_mensal_lote_total = Decimal("0.00")
                custo_anual_lote_total = Decimal("0.00")
            else:
                if lote.periodicidade == "mensal":
                    custo_mensal_lote_base = custo_ciclo_lote

                elif lote.periodicidade == "trimestral":
                    custo_mensal_lote_base = (custo_ciclo_lote / Decimal("3")).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )

                elif lote.periodicidade == "semestral":
                    custo_mensal_lote_base = (custo_ciclo_lote / Decimal("6")).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )

                elif lote.periodicidade == "anual":
                    custo_mensal_lote_base = (custo_ciclo_lote / Decimal("12")).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )

                else:
                    custo_mensal_lote_base = custo_ciclo_lote.quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )

                custo_mensal_unit = (custo_mensal_lote_base / Decimal(qtd)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )

                custo_anual_unit = (custo_mensal_unit * Decimal("12")).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )

                custo_mensal_lote_total = (custo_mensal_unit * Decimal(qtd)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )

                custo_anual_lote_total = (custo_mensal_lote_total * Decimal("12")).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )

                custo_unit_ciclo = (custo_ciclo_lote / Decimal(qtd)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )

            l_qtd_total += qtd
            l_qtd_disp += disp
            l_custo_mensal_total_licenca += custo_mensal_lote_total
            l_custo_anual_total_licenca += custo_anual_lote_total
            soma_custo_mensal_de_todos_lotes += custo_mensal_lote_total

            l_periodicidades.add(lote.get_periodicidade_display())

            p_label = lote.get_periodicidade_display()
            per_labels_buf.append(p_label)

            lotes_buf.append({
                "licenca": lic.nome,
                "lote": getattr(lote, "numero_lote", None) or f"Lote #{lote.id}",
                "pedido": getattr(lote, "pedido", None),
                "fornecedor": lote.fornecedor.nome if lote.fornecedor else "-",
                "qtd": qtd,
                "disp": disp,
                "periodicidade": p_label,
                "custo_total_lote": custo_ciclo_lote,
                "custo_unit_ciclo": custo_unit_ciclo,
                "custo_mensal_unit": custo_mensal_unit,
                "custo_anual_unit": custo_anual_unit,
                "custo_mensal_total": custo_mensal_lote_total,
                "custo_anual_total": custo_anual_lote_total,
            })

        if l_qtd_total > 0:
            custo_medio_unitario = (
                soma_custo_mensal_de_todos_lotes / Decimal(l_qtd_total)
            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            custo_medio_unitario = Decimal("0.00")

        ccs_ativos = uso_map_cc.get(lic.id, [])
        qtd_ativos = len(ccs_ativos)

        if pmb_pessoa_ativo:
            # Consulta focada nas PESSOAS PMB: sem parcela de estoque e sem
            # licenças que não tenham nenhuma pessoa PMB atribuída.
            if qtd_ativos == 0:
                continue
            qtd_estoque = 0
            custo_mensal_exibido = (custo_medio_unitario * Decimal(qtd_ativos)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            custo_anual_exibido = (custo_mensal_exibido * Decimal("12")).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            total_exibido = qtd_ativos
        else:
            qtd_estoque = max(0, l_qtd_total - qtd_ativos)
            custo_mensal_exibido = l_custo_mensal_total_licenca
            custo_anual_exibido = l_custo_anual_total_licenca
            total_exibido = l_qtd_total

        # A licença entra na consulta: commita os buffers acumulados no laço de lotes.
        lotes_detalhes.extend(lotes_buf)
        for p_label in per_labels_buf:
            per_counts[p_label] = per_counts.get(p_label, 0) + 1
        kpi_seats_total += total_exibido

        kpi_assentos_em_uso += qtd_ativos

        for cc_nome in ccs_ativos:
            cc_costs[cc_nome] = cc_costs.get(cc_nome, Decimal("0.00")) + custo_medio_unitario

        if qtd_estoque > 0:
            cc_estoque = "Estoque (Sem CC Definido)"
            if getattr(lic, "centro_custo", None):
                cc_estoque = f"{lic.centro_custo.numero} - {lic.centro_custo.departamento}"
            elif lotes_da_lic and getattr(lotes_da_lic[0], "centro_custo", None):
                cc_estoque = f"{lotes_da_lic[0].centro_custo.numero} - {lotes_da_lic[0].centro_custo.departamento}"

            cc_costs[cc_estoque] = cc_costs.get(cc_estoque, Decimal("0.00")) + (
                custo_medio_unitario * Decimal(qtd_estoque)
            )

        kpi_custo_mensal += custo_mensal_exibido

        f_nome = lic.fornecedor.nome if lic.fornecedor else "Indefinido"
        forn_costs[f_nome] = forn_costs.get(f_nome, Decimal("0.00")) + custo_mensal_exibido

        per_display = ", ".join(sorted(l_periodicidades)) if l_periodicidades else "-"

        licenca_meta[lic.id] = {
            "nome": lic.nome,
            "fornecedor": lic.fornecedor.nome if lic.fornecedor else "-",
            "periodicidade": per_display,
            "custo_mensal_unit": custo_medio_unitario,
        }

        linhas_tabela.append({
            "obj": lic,
            "periodicidade_display": per_display,
            "custo_mensal_total": custo_mensal_exibido,
            "custo_anual_total": custo_anual_exibido,
            "custo_mensal_unit_medio": custo_medio_unitario,
            "ativos": qtd_ativos,
            "total": total_exibido,
            "estoque": qtd_estoque,
        })

    # ── Visão por Colaborador (janela "Colaboradores") ──────────────────────────
    # Reúne, por colaborador, as licenças que ele detém (última movimentação =
    # atribuição), respeitando TODOS os filtros ativos. O custo de cada licença é o
    # custo médio mensal por assento (mesma base usada nos custos por CC).
    colaboradores_map = {}
    for (lid, uid), mov in estado_usuario.items():
        if mov.tipo != "atribuicao":
            continue
        meta = licenca_meta.get(lid)
        if not meta or mov.usuario is None:
            continue
        entry = colaboradores_map.get(uid)
        if entry is None:
            u = mov.usuario
            cc = u.centro_custo
            entry = {
                "nome": u.nome,
                "matricula": u.matricula or "",
                "cc": f"{cc.numero} - {cc.departamento}" if cc else "—",
                "cargo": u.funcao.nome if u.funcao else "",
                "pmb": (u.pmb == "sim" and cc is not None and cc.pmb == "sim"),
                "licencas": [],
                "total_mensal": Decimal("0.00"),
            }
            colaboradores_map[uid] = entry
        custo_m = meta["custo_mensal_unit"]
        entry["licencas"].append({
            "nome": meta["nome"],
            "fornecedor": meta["fornecedor"],
            "periodicidade": meta["periodicidade"],
            "custo_mensal": float(custo_m),
            "custo_anual": float(custo_m * Decimal("12")),
        })
        entry["total_mensal"] += custo_m

    colaboradores = []
    for c in sorted(colaboradores_map.values(), key=lambda x: x["total_mensal"], reverse=True):
        c["licencas"].sort(key=lambda l: l["custo_mensal"], reverse=True)
        colaboradores.append({
            "nome": c["nome"],
            "matricula": c["matricula"],
            "cc": c["cc"],
            "cargo": c["cargo"],
            "pmb": c["pmb"],
            "qtd_licencas": len(c["licencas"]),
            "total_mensal": float(c["total_mensal"]),
            "total_anual": float(c["total_mensal"] * Decimal("12")),
            "licencas": c["licencas"],
        })

    kpi_colaboradores = len(colaboradores)
    colaboradores_total_mensal = sum((Decimal(str(c["total_mensal"])) for c in colaboradores), Decimal("0.00"))

    sorted_cc = sorted(cc_costs.items(), key=lambda x: x[1], reverse=True)
    periodicidade_choices = LicencaLote._meta.get_field("periodicidade").choices

    context = {
        "f_q": q,
        "f_forn": fornecedor_id,
        "f_cc": cc_id,
        "f_per": periodicidade,
        "f_pmb": pmb,
        "f_pmb_pessoa": pmb_pessoa,
        "pmb_pessoa_ativo": pmb_pessoa_ativo,
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,

        "fornecedores": Fornecedor.objects.all().order_by("nome"),
        "centros_custo": CentroCusto.objects.all().order_by("numero"),
        "periodicidade_choices": periodicidade_choices,

        "kpi_total": kpi_total_licencas,
        "kpi_assentos": kpi_assentos_em_uso,
        "kpi_disp": kpi_seats_total - kpi_assentos_em_uso,
        "kpi_custo_mensal": kpi_custo_mensal.quantize(Decimal("0.01")),
        "kpi_custo_anual": (kpi_custo_mensal * Decimal("12")).quantize(Decimal("0.01")),
        "kpi_colaboradores": kpi_colaboradores,

        "linhas": linhas_tabela,
        "lotes_rows": lotes_detalhes,

        "colaboradores": colaboradores,
        "colaboradores_total_mensal": colaboradores_total_mensal.quantize(Decimal("0.01")),

        "cc_list": [{"label": k, "val": v} for k, v in sorted_cc],
        "chart_forn_labels": list(forn_costs.keys()),
        "chart_forn_data": [float(v) for v in forn_costs.values()],
        "chart_per_labels": list(per_counts.keys()),
        "chart_per_data": list(per_counts.values()),
    }

    return render(request, "front/dashboards/licencas_dashboard.html", context)

@login_required
def avisos_contratos_vencer(request):
    """
    Tela de avisos para contratos próximos do vencimento.
    Divide em dois rankings:
    1) Itens operacionais (ativo, backup, manutencao, defeito)
    2) Itens pausados
    """

    # =========================
    # CONFIG
    # =========================
    DIAS_ALERTA = 60  # ajuste conforme sua operação

    # Ajuste estes nomes conforme os valores reais do seu StatusItemChoices
    STATUS_OPERACIONAIS = ["ativo", "backup", "manutencao", "defeito", "queimado"]
    STATUS_PAUSADO = "pausado"

    hoje = date.today()

    # =========================
    # FILTROS
    # =========================
    f_nome = (request.GET.get("nome") or "").strip()
    f_ns = (request.GET.get("ns") or "").strip()
    f_subtipo = (request.GET.get("subtipo") or "").strip()
    f_status = (request.GET.get("status") or "").strip()
    f_fornecedor = (request.GET.get("fornecedor") or "").strip()

    qs = (
        Item.objects
        .filter(
            locado="sim",
            locacao__isnull=False,
            locacao__data_entrada__isnull=False,
            locacao__tempo_locado__isnull=False,
        )
        .select_related(
            "subtipo",
            "fornecedor",
            "centro_custo",
            "localidade",
            "locacao",
        )
        .order_by("nome")
    )

    if f_nome:
        qs = qs.filter(nome__icontains=f_nome)

    if f_ns:
        qs = qs.filter(numero_serie__icontains=f_ns)

    if f_subtipo:
        qs = qs.filter(subtipo_id=f_subtipo)

    if f_status:
        qs = qs.filter(status=f_status)

    if f_fornecedor:
        qs = qs.filter(fornecedor_id=f_fornecedor)

    itens_alerta = []

    for item in qs:
        loc = getattr(item, "locacao", None)
        if not loc or not loc.data_entrada or not loc.tempo_locado:
            continue

        try:
            data_vencimento = loc.data_entrada + relativedelta(months=int(loc.tempo_locado))
        except Exception:
            continue

        dias_restantes = (data_vencimento - hoje).days

        # traz vencidos e próximos do vencimento
        if dias_restantes <= DIAS_ALERTA:
            item.data_vencimento_contrato = data_vencimento
            item.dias_restantes_contrato = dias_restantes
            item.valor_mensal_calc = loc.valor_mensal or 0
            itens_alerta.append(item)

    # Ordenação do ranking:
    # vencidos primeiro, depois os mais próximos
    itens_alerta.sort(
        key=lambda x: (
            x.dias_restantes_contrato > 0,
            x.dias_restantes_contrato,
            x.nome.lower()
        )
    )

    ranking_operacional = [
        i for i in itens_alerta
        if (i.status or "").lower() in STATUS_OPERACIONAIS
    ]

    ranking_pausados = [
        i for i in itens_alerta
        if (i.status or "").lower() == STATUS_PAUSADO
    ]

    # KPIs
    total_alertas = len(itens_alerta)
    total_operacionais = len(ranking_operacional)
    total_pausados = len(ranking_pausados)
    vencidos = len([i for i in itens_alerta if i.dias_restantes_contrato < 0])

    subtipos = Subtipo.objects.order_by("nome")
    fornecedores = Fornecedor.objects.order_by("nome")

    # status disponíveis na própria base filtrada
    status_opcoes = (
        Item.objects.exclude(status__isnull=True)
        .exclude(status__exact="")
        .values_list("status", flat=True)
        .distinct()
        .order_by("status")
    )

    context = {
        "ranking_operacional": ranking_operacional,
        "ranking_pausados": ranking_pausados,
        "subtipos": subtipos,
        "fornecedores": fornecedores,
        "status_opcoes": status_opcoes,
        "filtros": {
            "nome": f_nome,
            "ns": f_ns,
            "subtipo": f_subtipo,
            "status": f_status,
            "fornecedor": f_fornecedor,
        },
        "kpi": {
            "total_alertas": total_alertas,
            "total_operacionais": total_operacionais,
            "total_pausados": total_pausados,
            "vencidos": vencidos,
            "dias_alerta": DIAS_ALERTA,
        }
    }

    return render(request, "front/dashboards/avisos_contrato_vencer.html", context)

@login_required
def avisos_contratos_vencer_export_excel(request):
    """
    Exporta para Excel a tela de avisos de contratos a vencer,
    respeitando os mesmos filtros da listagem.
    Inclui o usuário atual do equipamento com base na última movimentação válida.
    """

    DIAS_ALERTA = 60
    STATUS_OPERACIONAIS = ["ativo", "backup", "manutencao", "defeito", "queimado"]
    STATUS_PAUSADO = "pausado"

    hoje = date.today()

    # =========================
    # FILTROS
    # =========================
    f_nome = (request.GET.get("nome") or "").strip()
    f_ns = (request.GET.get("ns") or "").strip()
    f_subtipo = (request.GET.get("subtipo") or "").strip()
    f_status = (request.GET.get("status") or "").strip()
    f_fornecedor = (request.GET.get("fornecedor") or "").strip()

    qs = (
        Item.objects
        .filter(
            locado="sim",
            locacao__isnull=False,
            locacao__data_entrada__isnull=False,
            locacao__tempo_locado__isnull=False,
        )
        .select_related(
            "subtipo",
            "fornecedor",
            "centro_custo",
            "localidade",
            "locacao",
        )
        .order_by("nome")
    )

    if f_nome:
        qs = qs.filter(nome__icontains=f_nome)

    if f_ns:
        qs = qs.filter(numero_serie__icontains=f_ns)

    if f_subtipo:
        qs = qs.filter(subtipo_id=f_subtipo)

    if f_status:
        qs = qs.filter(status=f_status)

    if f_fornecedor:
        qs = qs.filter(fornecedor_id=f_fornecedor)

    itens_alerta = []

    for item in qs:
        loc = getattr(item, "locacao", None)
        if not loc or not loc.data_entrada or not loc.tempo_locado:
            continue

        try:
            data_vencimento = loc.data_entrada + relativedelta(months=int(loc.tempo_locado))
        except Exception:
            continue

        dias_restantes = (data_vencimento - hoje).days

        if dias_restantes <= DIAS_ALERTA:
            item.data_vencimento_contrato = data_vencimento
            item.dias_restantes_contrato = dias_restantes
            item.valor_mensal_calc = loc.valor_mensal or 0
            itens_alerta.append(item)

    itens_alerta.sort(
        key=lambda x: (
            x.dias_restantes_contrato > 0,
            x.dias_restantes_contrato,
            x.nome.lower()
        )
    )

    ranking_operacional = [
        i for i in itens_alerta
        if (i.status or "").lower() in STATUS_OPERACIONAIS
    ]

    ranking_pausados = [
        i for i in itens_alerta
        if (i.status or "").lower() == STATUS_PAUSADO
    ]

    # =========================
    # WORKBOOK
    # =========================
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=13, bold=True, color="1F1F1F")
    bold_font = Font(bold=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    left_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")

    def get_usuario_atual(item):
        """
        Busca o usuário atual do equipamento com base
        na última movimentação com campo usuario preenchido.
        Ajuste 'MovimentacaoItem' para o nome real do seu model, se necessário.
        """
        ultima_mov = (
            MovimentacaoItem.objects
            .filter(item=item, usuario__isnull=False)
            .select_related("usuario")
            .order_by("-created_at")
            .first()
        )

        if not ultima_mov or not ultima_mov.usuario:
            return {
                "nome": "-",
                "username": "-",
                "email": "-",
            }

        usuario = ultima_mov.usuario

        nome = "-"
        if getattr(usuario, "first_name", None) or getattr(usuario, "last_name", None):
            nome = f"{usuario.first_name} {usuario.last_name}".strip()
        elif getattr(usuario, "nome", None):
            nome = f"{usuario.nome} {getattr(usuario, 'last_name', '')}".strip()
        else:
            nome = getattr(usuario, "username", "-") or "-"

        return {
            "nome": nome,
            "username": getattr(usuario, "username", "-") or "-",
            "email": getattr(usuario, "email", "-") or "-",
        }

    def preencher_aba(ws, titulo, itens, grupo_nome):
        ws["A1"] = titulo
        ws["A1"].font = title_font

        ws["A2"] = "Grupo"
        ws["A2"].font = bold_font
        ws["B2"] = grupo_nome

        ws["D2"] = "Data Exportação"
        ws["D2"].font = bold_font
        ws["E2"] = hoje.strftime("%d/%m/%Y")

        headers = [
            "Ranking",
            "Grupo",
            "Item",
            "Número de Série",
            "Marca",
            "Modelo",
            "Subtipo",
            "Status",
            "Fornecedor",
            "Centro de Custo",
            "Localidade",
            "Valor Item",
            "Locado",
            "Valor Mensal",
            "Data Entrada Contrato",
            "Prazo (Meses)",
            "Data Vencimento",
            "Dias Restantes",
            "Contrato",
            "Observações Locação",
            "Usuário Atual",
            "Username Atual",
            "E-mail Usuário Atual",
        ]

        start_row = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment

        row = start_row + 1

        for idx, item in enumerate(itens, start=1):
            usuario_atual = get_usuario_atual(item)
            loc = getattr(item, "locacao", None)

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=grupo_nome)
            ws.cell(row=row, column=3, value=item.nome or "-")
            ws.cell(row=row, column=4, value=item.numero_serie or "-")
            ws.cell(row=row, column=5, value=item.marca or "-")
            ws.cell(row=row, column=6, value=item.modelo or "-")
            ws.cell(row=row, column=7, value=item.subtipo.nome if item.subtipo else "-")
            ws.cell(row=row, column=8, value=item.status or "-")
            ws.cell(row=row, column=9, value=item.fornecedor.nome if item.fornecedor else "-")
            ws.cell(row=row, column=10, value=item.centro_custo.departamento if item.centro_custo else "-")
            ws.cell(row=row, column=11, value=item.localidade.local if item.localidade else "-")
            ws.cell(row=row, column=12, value=float(item.valor or 0))
            ws.cell(row=row, column=13, value="Sim" if item.locado == "sim" else "Não")
            ws.cell(row=row, column=14, value=float(item.valor_mensal_calc or 0))
            ws.cell(row=row, column=15, value=loc.data_entrada.strftime("%d/%m/%Y") if loc and loc.data_entrada else "-")
            ws.cell(row=row, column=16, value=loc.tempo_locado if loc and loc.tempo_locado else "-")
            ws.cell(row=row, column=17, value=item.data_vencimento_contrato.strftime("%d/%m/%Y") if item.data_vencimento_contrato else "-")
            ws.cell(row=row, column=18, value=item.dias_restantes_contrato)
            ws.cell(row=row, column=19, value=loc.contrato if loc and loc.contrato else "-")
            ws.cell(row=row, column=20, value=loc.observacoes if loc and loc.observacoes else "-")
            ws.cell(row=row, column=21, value=usuario_atual["nome"])
            ws.cell(row=row, column=22, value=usuario_atual["username"])
            ws.cell(row=row, column=23, value=usuario_atual["email"])

            for col in range(1, 24):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                if col in [12, 14]:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = right_alignment
                elif col in [1, 16, 18]:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

            row += 1

        widths = {
            1: 10, 2: 18, 3: 28, 4: 22, 5: 16, 6: 18, 7: 18, 8: 16, 9: 26,
            10: 24, 11: 20, 12: 14, 13: 10, 14: 14, 15: 18, 16: 14, 17: 18,
            18: 14, 19: 24, 20: 28, 21: 24, 22: 18, 23: 28
        }

        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A5"

    # Aba 1 - Operacionais
    ws1 = wb.active
    ws1.title = "Ranking Operacional"
    preencher_aba(
        ws1,
        "Avisos de Contratos a Vencer - Ranking Operacional",
        ranking_operacional,
        "Operacional"
    )

    # Aba 2 - Pausados
    ws2 = wb.create_sheet(title="Ranking Pausados")
    preencher_aba(
        ws2,
        "Avisos de Contratos a Vencer - Ranking Pausados",
        ranking_pausados,
        "Pausado"
    )

    # Aba 3 - Resumo
    ws3 = wb.create_sheet(title="Resumo")

    ws3["A1"] = "Resumo da Exportação"
    ws3["A1"].font = title_font

    resumo = [
        ("Data da Exportação", hoje.strftime("%d/%m/%Y")),
        ("Janela de Alerta (dias)", DIAS_ALERTA),
        ("Total em Alerta", len(itens_alerta)),
        ("Ranking Operacional", len(ranking_operacional)),
        ("Ranking Pausados", len(ranking_pausados)),
        ("Vencidos", len([i for i in itens_alerta if i.dias_restantes_contrato < 0])),
        ("Filtro Nome", f_nome or "-"),
        ("Filtro NS", f_ns or "-"),
        ("Filtro Subtipo", f_subtipo or "-"),
        ("Filtro Status", f_status or "-"),
        ("Filtro Fornecedor", f_fornecedor or "-"),
    ]

    row = 3
    for label, value in resumo:
        ws3.cell(row=row, column=1, value=label).font = bold_font
        ws3.cell(row=row, column=2, value=value)
        ws3.cell(row=row, column=1).border = thin_border
        ws3.cell(row=row, column=2).border = thin_border
        row += 1

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 26

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="avisos_contratos_vencer.xlsx"'
    return response


