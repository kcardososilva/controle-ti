"""
RequisicaoService — Kanban de Solicitações de Compra/Estoque (ver
ProjetoEstoque.models.Requisicao / RequisicaoItem).

Rastreio interno do fluxo real: o usuário monta itens soltos (ideias ou
rascunhos), agrupa vários num Requisicao quando pronto pra enviar, e a
requisição segue pro Datasul (aprovação do gestor) e depois é retirada no
almoxarifado (compra recebida do Paradigma ou saída de estoque) — sem nenhuma
integração automática com esses sistemas externos; números de requisição/
compra são só anotação manual.

Cada card do Kanban é um RequisicaoItem individual. A coluna em que ele
aparece é sempre CALCULADA (nunca armazenada) a partir do status do item +
o status da requisição a que pertence (`coluna_kanban`), pra nunca haver
dessincronia entre os dois.

O quadro é um Kanban de verdade: qualquer card pode ser arrastado livremente
entre as colunas do fluxo principal (Rascunho → Solicitado → Aprovação
Pendente → Aprovados), em qualquer direção — o servidor só recusa combinações
que quebrariam uma invariante real (ex.: marcar como retirado algo que nunca
foi aprovado, ou reabrir uma requisição encerrada em definitivo).
"""
import unicodedata

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from ProjetoEstoque.models import (
    Requisicao,
    RequisicaoItem,
    StatusItemSolicitacaoChoices,
    StatusRequisicaoChoices,
    TipoRequisicaoChoices,
)

# ── Colunas do Kanban (calculadas, nunca persistidas) ──────────────────────
COLUNA_BACKLOG = "backlog"
COLUNA_RASCUNHO = "rascunho"
COLUNA_SOLICITADO = "solicitado"
COLUNA_APROVACAO = "aprovacao_pendente"
COLUNA_APROVADOS = "aprovados"
COLUNA_RECEBIDOS = "recebidos"
COLUNA_PAUSADOS = "pausados_cancelados"

COLUNA_ORDEM = (
    COLUNA_BACKLOG, COLUNA_RASCUNHO, COLUNA_SOLICITADO, COLUNA_APROVACAO,
    COLUNA_APROVADOS, COLUNA_RECEBIDOS, COLUNA_PAUSADOS,
)

COLUNA_LABELS = {
    COLUNA_BACKLOG: "Backlog",
    COLUNA_RASCUNHO: "Rascunho",
    COLUNA_SOLICITADO: "Solicitado",
    COLUNA_APROVACAO: "Aprovação Pendente",
    COLUNA_APROVADOS: "Aprovados",
    COLUNA_RECEBIDOS: "Recebidos / Retirado no Almoxarifado",
    COLUNA_PAUSADOS: "Pausados / Cancelados",
}

# Sequência linear do fluxo principal da Requisição — usada tanto para
# calcular a coluna quanto para permitir o "pulo" livre de drag-and-drop em
# qualquer direção dentro dela.
_FLUXO_STATUS = (
    StatusRequisicaoChoices.RASCUNHO,
    StatusRequisicaoChoices.SOLICITADA,
    StatusRequisicaoChoices.ENVIADA_APROVACAO,
    StatusRequisicaoChoices.APROVADA,
)
_FLUXO_COLUNA = {
    StatusRequisicaoChoices.RASCUNHO: COLUNA_RASCUNHO,
    StatusRequisicaoChoices.SOLICITADA: COLUNA_SOLICITADO,
    StatusRequisicaoChoices.ENVIADA_APROVACAO: COLUNA_APROVACAO,
    StatusRequisicaoChoices.APROVADA: COLUNA_APROVADOS,
}
_COLUNA_FLUXO_STATUS = {v: k for k, v in _FLUXO_COLUNA.items()}

# Pausada/Com erro são reversíveis (retomam pro status guardado em
# `status_anterior_pausa`); Não aprovada/Cancelada são encerramentos
# DEFINITIVOS — os itens já foram cascateados pra Reprovado e a requisição
# não volta mais pro fluxo.
_REQ_PAUSAVEL = {StatusRequisicaoChoices.PAUSADA, StatusRequisicaoChoices.COM_ERRO}
_REQ_ENCERRA = _REQ_PAUSAVEL | {StatusRequisicaoChoices.NAO_APROVADA, StatusRequisicaoChoices.CANCELADA}


def coluna_kanban(item: "RequisicaoItem") -> str:
    """Bucket do card no board — função pura, sem efeitos colaterais."""
    if item.status == StatusItemSolicitacaoChoices.RETIRADO:
        return COLUNA_RECEBIDOS
    if item.status in (StatusItemSolicitacaoChoices.PAUSADO, StatusItemSolicitacaoChoices.REPROVADO):
        return COLUNA_PAUSADOS
    if item.requisicao_id and item.requisicao.status in _REQ_ENCERRA:
        return COLUNA_PAUSADOS
    if item.status == StatusItemSolicitacaoChoices.NAO_CADASTRADO:
        return COLUNA_BACKLOG
    if not item.requisicao_id:
        return COLUNA_RASCUNHO
    return _FLUXO_COLUNA.get(item.requisicao.status, COLUNA_RASCUNHO)


def estagio_kanban(item: "RequisicaoItem"):
    """Rótulo + data do estágio mais recente do card, pra exibir uma única
    data "viva" (retirado > aprovado > solicitado) — nunca as três juntas.
    Retorna (None, None) quando o item ainda não passou por nenhum estágio
    com data (Backlog/Rascunho) ou foi pausado/reprovado antes de um deles."""
    if item.status == StatusItemSolicitacaoChoices.RETIRADO:
        return "Retirado em", item.retirado_em
    if item.requisicao_id and item.requisicao.status == StatusRequisicaoChoices.APROVADA:
        return "Aprovado em", item.requisicao.decidida_em
    if item.status == StatusItemSolicitacaoChoices.SOLICITADO and item.requisicao_id:
        data = item.requisicao.solicitada_em or item.requisicao.enviada_em
        if data:
            return "Solicitado em", data
    return None, None


# Cor do badge de status da Requisição — mesma paleta usada nos cards do
# quadro, pra "Aprovada" ser sempre verde, "Cancelada"/"Não Aprovada" sempre
# vermelho etc. em qualquer tela (lista, detalhe).
_STATUS_BADGE_CLASS = {
    StatusRequisicaoChoices.RASCUNHO: "kan-badge",
    StatusRequisicaoChoices.SOLICITADA: "kan-badge-primary",
    StatusRequisicaoChoices.ENVIADA_APROVACAO: "kan-badge-warning",
    StatusRequisicaoChoices.APROVADA: "kan-badge-success",
    StatusRequisicaoChoices.NAO_APROVADA: "kan-badge-danger",
    StatusRequisicaoChoices.PAUSADA: "kan-badge-warning",
    StatusRequisicaoChoices.COM_ERRO: "kan-badge-danger",
    StatusRequisicaoChoices.CANCELADA: "kan-badge-danger",
}


def status_badge_class(status: str) -> str:
    return _STATUS_BADGE_CLASS.get(status, "kan-badge")


_ITEM_STATUS_BADGE_CLASS = {
    StatusItemSolicitacaoChoices.NAO_CADASTRADO: "kan-badge",
    StatusItemSolicitacaoChoices.NAO_SOLICITADO: "kan-badge-primary",
    StatusItemSolicitacaoChoices.SOLICITADO: "kan-badge-warning",
    StatusItemSolicitacaoChoices.PAUSADO: "kan-badge-warning",
    StatusItemSolicitacaoChoices.REPROVADO: "kan-badge-danger",
    StatusItemSolicitacaoChoices.RETIRADO: "kan-badge-success",
}


def item_status_badge_class(status: str) -> str:
    return _ITEM_STATUS_BADGE_CLASS.get(status, "kan-badge")


class RequisicaoService:

    # ── Itens soltos (ideia / rascunho) ────────────────────────────────────

    @staticmethod
    def criar_item(*, form, user):
        """`form` é um RequisicaoItemForm válido e não salvo. Status inicial
        depende só de ter código preenchido ou não — sem código é ideia
        (`NAO_CADASTRADO`), com código já é um rascunho pronto pra agrupar
        (`NAO_SOLICITADO`)."""
        item = form.save(commit=False)
        item.status = (
            StatusItemSolicitacaoChoices.NAO_SOLICITADO if (item.codigo or "").strip()
            else StatusItemSolicitacaoChoices.NAO_CADASTRADO
        )
        item.criado_por = user
        item.atualizado_por = user
        item.save()
        return item

    @staticmethod
    def atualizar_item(*, form, user):
        """`form` é um RequisicaoItemForm válido, ligado a uma instância
        existente (não salvo). Bloqueia edição de conteúdo depois que o item
        já foi enviado (a requisição saiu de RASCUNHO) — só o status muda a
        partir daí, e só em cascata."""
        item = form.instance
        if item.requisicao_id and item.requisicao.status != StatusRequisicaoChoices.RASCUNHO:
            raise ValidationError(
                "Este item já foi enviado — o conteúdo não pode mais ser editado."
            )
        obj = form.save(commit=False)
        if obj.status == StatusItemSolicitacaoChoices.NAO_CADASTRADO and (obj.codigo or "").strip():
            obj.status = StatusItemSolicitacaoChoices.NAO_SOLICITADO
        obj.atualizado_por = user
        obj.save()
        return obj

    @staticmethod
    def mudar_status_item(*, item, novo_status, user):
        """Troca manual de status de um item AINDA solto ou em rascunho
        (ex.: pausar/despausar uma ideia). Uma vez que o item pertence a uma
        requisição já enviada, o status só muda em cascata — nunca
        manualmente item a item."""
        if item.requisicao_id and item.requisicao.status != StatusRequisicaoChoices.RASCUNHO:
            raise ValidationError(
                "Este item pertence a uma requisição já enviada — o status dele muda "
                "automaticamente junto com o status da requisição."
            )
        item.status = novo_status
        item.atualizado_por = user
        item.save(update_fields=["status", "atualizado_por", "updated_at"])
        return item

    # ── Agrupamento em Requisição ───────────────────────────────────────────
    # Regra dura, sem exceção: uma Requisição nunca mistura tipo Compra com
    # tipo Estoque — cada requisição real do Datasul é sempre de um tipo só
    # neste sistema, mesmo que categorias (Ferramentas/Infraestrutura/TI/
    # Escritório) venham misturadas dentro dela.

    @staticmethod
    @transaction.atomic
    def criar_requisicao_de_itens(*, item_ids, tipo, user):
        item_ids = list(dict.fromkeys(item_ids))  # remove duplicatas, preserva ordem
        itens = list(RequisicaoItem.objects.select_for_update().filter(pk__in=item_ids))
        if not itens:
            raise ValidationError("Selecione ao menos um item para agrupar em uma requisição.")
        if len(itens) != len(item_ids):
            raise ValidationError("Um ou mais itens selecionados não foram encontrados.")

        for item in itens:
            if item.requisicao_id:
                raise ValidationError(f'O item "{item.descricao}" já pertence a outra requisição.')
            if item.status != StatusItemSolicitacaoChoices.NAO_SOLICITADO:
                raise ValidationError(
                    f'O item "{item.descricao}" precisa estar com status "Item não solicitado" '
                    '(informe o código) antes de agrupar.'
                )
            if item.tipo != tipo:
                raise ValidationError(
                    "Não é possível agrupar itens de Compra e de Estoque na mesma requisição "
                    f'("{item.descricao}" é do tipo "{item.get_tipo_display()}").'
                )

        requisicao = Requisicao(tipo=tipo, status=StatusRequisicaoChoices.RASCUNHO)
        requisicao.criado_por = user
        requisicao.atualizado_por = user
        requisicao.full_clean()
        requisicao.save()

        for item in itens:
            item.requisicao = requisicao
            item.atualizado_por = user
            item.save(update_fields=["requisicao", "atualizado_por", "updated_at"])

        return requisicao

    @staticmethod
    def vincular_item_a_requisicao(*, requisicao, item, user):
        if requisicao.status != StatusRequisicaoChoices.RASCUNHO:
            raise ValidationError("Só é possível adicionar itens a uma requisição ainda em rascunho.")
        if item.requisicao_id:
            raise ValidationError(f'O item "{item.descricao}" já pertence a uma requisição.')
        if item.status != StatusItemSolicitacaoChoices.NAO_SOLICITADO:
            raise ValidationError(
                f'O item "{item.descricao}" precisa estar com status "Item não solicitado" antes de agrupar.'
            )
        if item.tipo != requisicao.tipo:
            raise ValidationError(
                "Não é possível agrupar itens de Compra e de Estoque na mesma requisição "
                f'("{item.descricao}" é do tipo "{item.get_tipo_display()}").'
            )

        item.requisicao = requisicao
        item.atualizado_por = user
        item.save(update_fields=["requisicao", "atualizado_por", "updated_at"])
        return item

    @staticmethod
    def desvincular_item(*, item, user):
        """Arrependimento de agrupamento — item volta a ficar solto (não
        cancela nem exclui nada)."""
        if not item.requisicao_id:
            raise ValidationError("Este item já não pertence a nenhuma requisição.")
        if item.requisicao.status != StatusRequisicaoChoices.RASCUNHO:
            raise ValidationError("Só é possível remover itens de uma requisição ainda em rascunho.")

        item.requisicao = None
        item.atualizado_por = user
        item.save(update_fields=["requisicao", "atualizado_por", "updated_at"])
        return item

    @staticmethod
    def excluir_requisicao_rascunho(*, requisicao, user):
        """Desfaz a requisição (os itens voltam a ficar soltos, como
        `NAO_SOLICITADO`) — só permitido enquanto ainda não foi enviada."""
        if requisicao.status != StatusRequisicaoChoices.RASCUNHO:
            raise ValidationError("Só é possível excluir uma requisição ainda em rascunho.")
        requisicao.delete()  # on_delete=SET_NULL solta os itens de volta pro board

    # ── Ciclo de vida da Requisição (ações nomeadas — tela de detalhe) ─────

    @classmethod
    @transaction.atomic
    def enviar_para_aprovacao(cls, *, requisicao, numero_datasul, user):
        req = Requisicao.objects.select_for_update().get(pk=requisicao.pk)
        if req.status not in (StatusRequisicaoChoices.RASCUNHO, StatusRequisicaoChoices.SOLICITADA):
            raise ValidationError("Esta requisição já foi enviada para aprovação.")
        if not req.itens.exists():
            raise ValidationError("Adicione ao menos um item antes de enviar para aprovação.")

        numero_datasul = (numero_datasul or "").strip()
        if numero_datasul:
            req.numero_datasul = numero_datasul
        req.status = StatusRequisicaoChoices.ENVIADA_APROVACAO
        req.enviada_em = timezone.now()
        req.atualizado_por = user
        req.save()

        req.itens.update(
            status=StatusItemSolicitacaoChoices.SOLICITADO,
            atualizado_por=user,
            updated_at=timezone.now(),
        )
        transaction.on_commit(lambda: _disparar_email_enviada_aprovacao(req.pk))
        return req

    # Ação -> (status de origem permitidos, status de destino, cascata no item,
    #          guarda o status atual em status_anterior_pausa antes de mudar)
    _ACOES = {
        "solicitar": {
            "de": {StatusRequisicaoChoices.RASCUNHO},
            "para": StatusRequisicaoChoices.SOLICITADA,
            "cascata_item": StatusItemSolicitacaoChoices.SOLICITADO,
        },
        "aprovar": {
            "de": {StatusRequisicaoChoices.ENVIADA_APROVACAO},
            "para": StatusRequisicaoChoices.APROVADA,
        },
        "reprovar": {
            "de": {StatusRequisicaoChoices.ENVIADA_APROVACAO},
            "para": StatusRequisicaoChoices.NAO_APROVADA,
            "cascata_item": StatusItemSolicitacaoChoices.REPROVADO,
        },
        "pausar": {
            "de": {StatusRequisicaoChoices.SOLICITADA, StatusRequisicaoChoices.ENVIADA_APROVACAO, StatusRequisicaoChoices.APROVADA},
            "para": StatusRequisicaoChoices.PAUSADA,
            "guarda_anterior": True,
        },
        "marcar_erro": {
            "de": {StatusRequisicaoChoices.SOLICITADA, StatusRequisicaoChoices.ENVIADA_APROVACAO, StatusRequisicaoChoices.APROVADA},
            "para": StatusRequisicaoChoices.COM_ERRO,
            "guarda_anterior": True,
        },
        "retomar": {
            "de": {StatusRequisicaoChoices.PAUSADA, StatusRequisicaoChoices.COM_ERRO},
            "para": None,  # restaura `status_anterior_pausa`
        },
        "cancelar": {
            "de": {
                StatusRequisicaoChoices.RASCUNHO, StatusRequisicaoChoices.SOLICITADA,
                StatusRequisicaoChoices.ENVIADA_APROVACAO, StatusRequisicaoChoices.APROVADA,
                StatusRequisicaoChoices.PAUSADA, StatusRequisicaoChoices.COM_ERRO,
            },
            "para": StatusRequisicaoChoices.CANCELADA,
            "cascata_item": StatusItemSolicitacaoChoices.REPROVADO,
        },
    }

    @classmethod
    @transaction.atomic
    def mudar_status_requisicao(cls, *, requisicao, acao, user):
        """Dispatcher central de solicitar/aprovar/reprovar/pausar/retomar/
        cancelar/marcar_erro — nunca mudar `Requisicao.status` fora daqui (ou
        de `_mover_requisicao_fluxo`, usado só pelo drag-and-drop), pra manter
        a cascata sobre os itens sempre consistente com a transição."""
        config = cls._ACOES.get(acao)
        if not config:
            raise ValidationError(f'Ação de requisição desconhecida: "{acao}".')

        req = Requisicao.objects.select_for_update().get(pk=requisicao.pk)
        if req.status not in config["de"]:
            raise ValidationError(
                f'Não é possível executar "{acao}" numa requisição com status '
                f'"{req.get_status_display()}".'
            )

        if acao == "retomar":
            novo_status = req.status_anterior_pausa or StatusRequisicaoChoices.ENVIADA_APROVACAO
            req.status_anterior_pausa = None
        else:
            novo_status = config["para"]
            if config.get("guarda_anterior"):
                req.status_anterior_pausa = req.status

        req.status = novo_status
        if acao == "solicitar":
            req.solicitada_em = timezone.now()
        if acao in ("aprovar", "reprovar"):
            req.decidida_em = timezone.now()
            req.decidida_por = user
        req.atualizado_por = user
        req.save()

        cascata = config.get("cascata_item")
        if cascata:
            req.itens.update(status=cascata, atualizado_por=user, updated_at=timezone.now())

        return req

    # ── Retirada no almoxarifado (individual ou requisição inteira) ────────

    @classmethod
    @transaction.atomic
    def marcar_item_retirado(cls, *, item, user):
        item = RequisicaoItem.objects.select_related("requisicao").select_for_update().get(pk=item.pk)
        if coluna_kanban(item) != COLUNA_APROVADOS:
            raise ValidationError('Só é possível marcar como retirado um item que esteja em "Aprovados".')
        # Compra vinculada a um item de estoque precisa passar por
        # `finalizar_compra_estoque` (gera a entrada real — NF, lote, custo).
        # Bloqueado aqui pra fechar TODOS os caminhos que levam a este método
        # (botão simples, drag-and-drop do quadro) — não só o botão da tela de
        # detalhe, que já direciona pro fluxo certo quando renderizado fresco.
        if item.tipo == TipoRequisicaoChoices.COMPRA and item.item_vinculado_id:
            raise ValidationError(
                'Este item é uma Compra vinculada a um item de estoque — use "Receber Compra e '
                'Dar Entrada no Estoque" na tela do item para registrar a entrada real, em vez de '
                "retirar diretamente."
            )
        return cls._marcar_item_retirado(item=item, user=user)

    @classmethod
    def _marcar_item_retirado(cls, *, item, user):
        """Núcleo sem a checagem de compra vinculada — usado pelo `marcar_item_retirado`
        público (após a checagem, acima) e por `finalizar_compra_estoque` (chamado
        depois que a entrada de estoque real já foi criada, quando o bloqueio já
        não se aplica)."""
        agora = timezone.now()
        item.status = StatusItemSolicitacaoChoices.RETIRADO
        item.retirado_em = agora
        item.retirado_por = user
        item.atualizado_por = user
        item.save(update_fields=["status", "retirado_em", "retirado_por", "atualizado_por", "updated_at"])
        if item.requisicao_id:
            transaction.on_commit(lambda: _disparar_email_itens_retirados(item.requisicao_id, [item.pk]))
        return item

    @classmethod
    @transaction.atomic
    def finalizar_compra_estoque(cls, *, item, fornecedor, numero_nf, custo_unitario,
                                  localidade_destino, centro_custo_destino, observacao, user):
        """Recebimento de uma compra vinculada a um item de estoque (ver
        `RequisicaoItem.item_vinculado`): dá entrada real no estoque — mesmo
        efeito colateral da Entrada em Movimentações (cria `LoteEstoque` +
        `ItemLote`, soma `Item.quantidade`) — e então marca o item da
        requisição como retirado. A quantidade da entrada é sempre a
        solicitada (`item.quantidade`), não é reeditável aqui."""
        item = (
            RequisicaoItem.objects
            .select_related("requisicao", "item_vinculado")
            .select_for_update()
            .get(pk=item.pk)
        )
        if coluna_kanban(item) != COLUNA_APROVADOS:
            raise ValidationError('Só é possível receber um item que esteja em "Aprovados".')
        if item.tipo != TipoRequisicaoChoices.COMPRA or not item.item_vinculado_id:
            raise ValidationError(
                "Esta ação só está disponível para itens de Compra vinculados a um item de estoque."
            )

        from services.movimentacao_service import MovimentacaoEstoqueService

        MovimentacaoEstoqueService.registrar_entrada(
            item=item.item_vinculado,
            fornecedor=fornecedor,
            data_entrada=timezone.now().date(),
            numero_nf=numero_nf,
            quantidade=item.quantidade,
            custo_unitario=custo_unitario,
            observacao_lote="",
            localidade_destino=localidade_destino,
            centro_custo_destino=centro_custo_destino,
            observacao=(observacao or "").strip()
                or f'Entrada gerada pelo recebimento da Requisição — item "{item.descricao}".',
            user=user,
        )

        return cls._marcar_item_retirado(item=item, user=user)

    @classmethod
    def desfazer_retirada_item(cls, *, item, user):
        if item.status != StatusItemSolicitacaoChoices.RETIRADO:
            raise ValidationError("Este item não está marcado como retirado.")
        item.status = StatusItemSolicitacaoChoices.SOLICITADO
        item.retirado_em = None
        item.retirado_por = None
        item.atualizado_por = user
        item.save(update_fields=["status", "retirado_em", "retirado_por", "atualizado_por", "updated_at"])
        return item

    @classmethod
    @transaction.atomic
    def marcar_requisicao_retirada(cls, *, requisicao, user):
        """Retirada em massa (.update() direto — sem NF/custo por item, então
        não pode cobrir Compra vinculada a estoque, que exige dados próprios
        por item). Esses itens são pulados aqui e precisam ser recebidos
        individualmente em "Receber Compra e Dar Entrada no Estoque"; a view
        avisa o usuário com a lista de pulados via o segundo item do retorno."""
        req = Requisicao.objects.select_for_update().get(pk=requisicao.pk)
        if req.status != StatusRequisicaoChoices.APROVADA:
            raise ValidationError("Só é possível marcar como retirada uma requisição aprovada.")
        pendentes = list(req.itens.filter(status=StatusItemSolicitacaoChoices.SOLICITADO))
        if not pendentes:
            raise ValidationError("Não há itens pendentes de retirada nesta requisição.")

        puladas = [
            i for i in pendentes
            if i.tipo == TipoRequisicaoChoices.COMPRA and i.item_vinculado_id
        ]
        itens = [i for i in pendentes if i not in puladas]
        if not itens:
            raise ValidationError(
                "Todos os itens pendentes são Compras vinculadas a estoque — receba cada um "
                'individualmente em "Receber Compra e Dar Entrada no Estoque", na tela de cada item.'
            )

        agora = timezone.now()
        pks = [i.pk for i in itens]
        req.itens.filter(pk__in=pks).update(
            status=StatusItemSolicitacaoChoices.RETIRADO,
            retirado_em=agora, retirado_por=user,
            atualizado_por=user, updated_at=agora,
        )
        transaction.on_commit(lambda: _disparar_email_itens_retirados(req.pk, pks))
        return req, puladas

    # ── Drag-and-drop do Kanban (movimentação livre do fluxo principal) ────

    @classmethod
    @transaction.atomic
    def mover_item_kanban(cls, *, item, coluna_destino, coluna_conhecida, user):
        """Traduz um drop do board. O fluxo principal (Rascunho → Solicitado →
        Aprovação Pendente → Aprovados) pode ser percorrido livremente em
        qualquer direção arrastando qualquer card do grupo — a requisição
        inteira acompanha. Só ficam de fora do "livre" as invariantes reais:
        um item sem requisição não pode pular direto pra Solicitado+; só um
        item em Aprovados pode virar Recebido (retirada é individual, nunca
        em massa por aqui); um item já Recebido fica travado no quadro — não
        volta mais por drag-and-drop, só pela ação explícita "Desfazer
        Retirada" na tela do item; e uma requisição encerrada em definitivo
        (Não aprovada/Cancelada) não pode ser reaberta pelo quadro."""
        if coluna_destino not in COLUNA_ORDEM:
            raise ValidationError("Coluna de destino inválida.")

        item = (
            RequisicaoItem.objects
            .select_related("requisicao")
            .select_for_update()
            .get(pk=item.pk)
        )
        coluna_atual = coluna_kanban(item)
        if coluna_conhecida and coluna_atual != coluna_conhecida:
            raise ValidationError("Este card foi alterado por outro usuário — recarregue a página.")
        if coluna_atual == coluna_destino:
            return item

        # 1) Backlog ⇄ Rascunho — toggle de item ainda solto (fora de
        #    requisição, ou requisição ainda em rascunho).
        if {coluna_atual, coluna_destino} <= {COLUNA_BACKLOG, COLUNA_RASCUNHO}:
            if item.requisicao_id and item.requisicao.status != StatusRequisicaoChoices.RASCUNHO:
                raise ValidationError("Este item já pertence a uma requisição enviada.")
            if coluna_destino == COLUNA_RASCUNHO:
                if not (item.codigo or "").strip():
                    raise ValidationError('Informe o código do item antes de movê-lo para "Rascunho".')
                novo_status = StatusItemSolicitacaoChoices.NAO_SOLICITADO
            else:
                novo_status = StatusItemSolicitacaoChoices.NAO_CADASTRADO
            return cls.mudar_status_item(item=item, novo_status=novo_status, user=user)

        # 2) Entrar em Recebidos — retirada individual (nunca em massa pelo
        #    quadro; "toda a requisição" é ação explícita na tela dela).
        if coluna_destino == COLUNA_RECEBIDOS:
            return cls.marcar_item_retirado(item=item, user=user)

        # Um item já recebido/retirado fica travado no quadro — nenhum
        # arrasto o move dali (nem de volta pra "Aprovados"), só a ação
        # explícita e deliberada "Desfazer Retirada" na tela do item. Evita
        # desfazer sem querer por um arrasto acidental — especialmente
        # crítico pra Compra vinculada a estoque, onde a entrada real já foi
        # registrada e não é desfeita junto (ver `finalizar_compra_estoque`).
        if coluna_atual == COLUNA_RECEBIDOS:
            raise ValidationError(
                'Este item já foi recebido/retirado no almoxarifado — não é mais possível movê-lo '
                'pelo quadro. Para desfazer, use "Desfazer Retirada" na tela do item.'
            )

        # 3) Sair de Pausados/Cancelados — retoma a requisição (ou o item
        #    solto pausado), exceto encerramento definitivo.
        if coluna_atual == COLUNA_PAUSADOS:
            if not item.requisicao_id:
                if item.status != StatusItemSolicitacaoChoices.PAUSADO:
                    raise ValidationError("Este item foi reprovado e não pode ser reativado.")
                cls.mudar_status_item(item=item, novo_status=StatusItemSolicitacaoChoices.NAO_SOLICITADO, user=user)
                item.refresh_from_db()
            else:
                if item.requisicao.status not in _REQ_PAUSAVEL:
                    raise ValidationError("Esta requisição foi encerrada em definitivo e não pode ser reaberta.")
                cls.mudar_status_requisicao(requisicao=item.requisicao, acao="retomar", user=user)
                item.refresh_from_db()
            coluna_atual = coluna_kanban(item)
            if coluna_atual == coluna_destino:
                return item
            # cai para o bloco 4 abaixo, já com o estado restaurado

        # 4) Entrar em Pausados/Cancelados — pausa (reversível) do item solto
        #    ou da requisição inteira. Reprovar/Cancelar continuam ações
        #    explícitas e deliberadas na tela de detalhe da requisição.
        if coluna_destino == COLUNA_PAUSADOS:
            if not item.requisicao_id:
                return cls.mudar_status_item(item=item, novo_status=StatusItemSolicitacaoChoices.PAUSADO, user=user)
            cls.mudar_status_requisicao(requisicao=item.requisicao, acao="pausar", user=user)
            item.refresh_from_db()
            return item

        # 5) Fluxo principal — Rascunho ⇄ Solicitado ⇄ Aprovação Pendente ⇄
        #    Aprovados, em qualquer direção, movendo a requisição inteira.
        if coluna_destino not in _COLUNA_FLUXO_STATUS:
            raise ValidationError("Essa movimentação não é permitida pelo quadro.")
        if not item.requisicao_id:
            raise ValidationError(
                'Agrupe este item em uma Requisição (selecione e clique em "Agrupar em Requisição") '
                "antes de movê-lo adiante no quadro."
            )
        if item.requisicao.status in _REQ_ENCERRA:
            raise ValidationError("Esta requisição está encerrada — retome-a antes de movê-la.")

        cls._mover_requisicao_fluxo(
            requisicao=item.requisicao,
            alvo=_COLUNA_FLUXO_STATUS[coluna_destino],
            user=user,
        )
        item.refresh_from_db()
        return item

    @classmethod
    @transaction.atomic
    def _mover_requisicao_fluxo(cls, *, requisicao, alvo, user):
        """Move a requisição (e todos os seus itens, em bloco) para qualquer
        ponto do fluxo principal, pra frente ou pra trás — é o que permite o
        drag-and-drop livre entre as 4 colunas centrais do quadro."""
        req = Requisicao.objects.select_for_update().get(pk=requisicao.pk)
        if req.status not in _FLUXO_STATUS:
            raise ValidationError("Esta requisição está encerrada — retome-a antes de movê-la.")
        if req.status == alvo:
            return req

        status_anterior = req.status
        indo_para_frente = _FLUXO_STATUS.index(alvo) > _FLUXO_STATUS.index(status_anterior)

        req.status = alvo
        if alvo == StatusRequisicaoChoices.SOLICITADA and indo_para_frente:
            req.solicitada_em = timezone.now()
        if alvo == StatusRequisicaoChoices.ENVIADA_APROVACAO and indo_para_frente:
            req.enviada_em = timezone.now()
        if alvo == StatusRequisicaoChoices.APROVADA:
            req.decidida_em = timezone.now()
            req.decidida_por = user
        req.atualizado_por = user
        req.save()

        novo_status_item = (
            StatusItemSolicitacaoChoices.NAO_SOLICITADO if alvo == StatusRequisicaoChoices.RASCUNHO
            else StatusItemSolicitacaoChoices.SOLICITADO
        )
        req.itens.update(status=novo_status_item, atualizado_por=user, updated_at=timezone.now())

        if alvo == StatusRequisicaoChoices.ENVIADA_APROVACAO and indo_para_frente:
            transaction.on_commit(lambda: _disparar_email_enviada_aprovacao(req.pk))

        return req

    # ── Consulta ─────────────────────────────────────────────────────────────

    @staticmethod
    def agrupar_itens_por_coluna(itens):
        buckets = {coluna: [] for coluna in COLUNA_ORDEM}
        for item in itens:
            buckets[coluna_kanban(item)].append(item)
        return buckets

    # ── Catálogo de itens padrão (código Datasul) ───────────────────────────

    @staticmethod
    @transaction.atomic
    def importar_catalogo_datasul(*, arquivo, user):
        """Importa/atualiza `ItemPadraoDatasul` a partir de uma planilha .xlsx
        com colunas Código, Descrição e Categoria (nomes flexíveis, sem
        acento/maiúscula). A categoria precisa já existir com esse nome — não
        cria categoria nova, pra não duplicar por erro de digitação na
        planilha. Retorna (criados, atualizados, erros)."""
        from openpyxl import load_workbook

        from ProjetoEstoque.models import Categoria, ItemPadraoDatasul

        def _norm(s):
            return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().strip().lower()

        try:
            wb = load_workbook(arquivo, data_only=True)
        except Exception:
            raise ValidationError("Não foi possível ler o arquivo — envie uma planilha .xlsx válida.")
        ws = wb.active

        primeira_linha = next(ws.iter_rows(min_row=1, max_row=1), None)
        if primeira_linha is None:
            raise ValidationError("A planilha está vazia.")
        header = [_norm(c.value) for c in primeira_linha]
        col_map = {}
        for idx, h in enumerate(header):
            if h in ("codigo", "cod"):
                col_map["codigo"] = idx
            elif h in ("descricao", "nome", "item"):
                col_map["descricao"] = idx
            elif h == "categoria":
                col_map["categoria"] = idx
        faltando = {"codigo", "descricao", "categoria"} - col_map.keys()
        if faltando:
            raise ValidationError(
                "Planilha sem as colunas obrigatórias: " + ", ".join(sorted(faltando)) +
                ". Use os cabeçalhos Código, Descrição e Categoria."
            )

        categorias = {_norm(c.nome): c for c in Categoria.objects.all()}
        criados = atualizados = 0
        erros = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            codigo = str(row[col_map["codigo"]].value or "").strip()
            descricao = str(row[col_map["descricao"]].value or "").strip()
            categoria_nome = row[col_map["categoria"]].value
            if not codigo and not descricao:
                continue  # linha em branco
            if not codigo or not descricao:
                erros.append(f"Linha {row_idx}: código e descrição são obrigatórios.")
                continue
            categoria = categorias.get(_norm(categoria_nome))
            if not categoria:
                erros.append(f'Linha {row_idx}: categoria "{categoria_nome}" não encontrada.')
                continue

            obj = ItemPadraoDatasul.objects.filter(codigo=codigo).first()
            if obj is None:
                obj = ItemPadraoDatasul(codigo=codigo, criado_por=user)
                criados += 1
            else:
                atualizados += 1
            obj.descricao = descricao
            obj.categoria = categoria
            obj.ativo = True
            obj.atualizado_por = user
            obj.save()

        return criados, atualizados, erros


def _disparar_email_enviada_aprovacao(requisicao_pk: int):
    from services import email_alertas
    email_alertas.alerta_requisicao_enviada_aprovacao(requisicao_pk)


def _disparar_email_itens_retirados(requisicao_pk: int, item_pks: list[int]):
    from services import email_alertas
    email_alertas.alerta_requisicao_itens_retirados(requisicao_pk, item_pks)
