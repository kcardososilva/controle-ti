"""
Tema visual compartilhado por TODOS os exportadores .xlsx do sistema
(Excel/openpyxl) — paleta oficial da empresa.

Cada exportador (relatorios.py, dashboards.py, ninja.py, equipamentos.py,
licencas.py, preventivas.py, quiosque.py, plantas.py, portal_fornecedor.py)
importa daqui em vez de definir cores/fontes próprias — assim uma planilha
de Custos e uma de Preventivas saem com a mesma identidade visual.

Paleta oficial (RGB fornecido pela empresa):
  Azul abundância  79,122,121  -> banner/título
  Azul vastidão    104,152,160 -> cabeçalho de tabela
  Verde vital      128,157,94  -> positivo / subtotal / sucesso
  Marrom café      105,88,79   -> texto padrão das células
  Laranja solar    210,107,48  -> destaque / alerta / divergência
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paleta oficial (hex, sem o prefixo de alfa) ─────────────────────────────
AZUL_ABUNDANCIA = "4F7A79"
AZUL_VASTIDAO = "6898A0"
VERDE_VITAL = "809D5E"
MARROM_CAFE = "69584F"
LARANJA_SOLAR = "D26B30"

BRANCO = "FFFFFF"
ZEBRA = "EEF3F2"          # tint muito claro do Azul abundância — linha alternada
BORDA_CLARA = "D7E2E0"    # tint claro para grade fina

FONTE = "Calibri"


def _argb(hex6):
    return f"FF{hex6}"


# ── Fills ────────────────────────────────────────────────────────────────
FILL_TITULO = PatternFill("solid", fgColor=_argb(AZUL_ABUNDANCIA))
FILL_SUBTITULO = PatternFill("solid", fgColor=_argb(ZEBRA))
FILL_HEADER = PatternFill("solid", fgColor=_argb(AZUL_VASTIDAO))
FILL_ZEBRA = PatternFill("solid", fgColor=_argb(ZEBRA))
FILL_SUBTOTAL = PatternFill("solid", fgColor=_argb(VERDE_VITAL))
FILL_DESTAQUE = PatternFill("solid", fgColor=_argb(LARANJA_SOLAR))

# ── Fontes ───────────────────────────────────────────────────────────────
FONT_TITULO = Font(name=FONTE, size=16, bold=True, color=_argb(BRANCO))
FONT_SUBTITULO = Font(name=FONTE, size=10, italic=True, color=_argb(MARROM_CAFE))
FONT_HEADER = Font(name=FONTE, size=10, bold=True, color=_argb(BRANCO))
FONT_CELL = Font(name=FONTE, size=10, color=_argb(MARROM_CAFE))
FONT_CELL_BOLD = Font(name=FONTE, size=10, bold=True, color=_argb(MARROM_CAFE))
FONT_SUBTOTAL = Font(name=FONTE, size=10, bold=True, color=_argb(BRANCO))
FONT_DESTAQUE = Font(name=FONTE, size=10, bold=True, color=_argb(BRANCO))
FONT_DESTAQUE_TEXTO = Font(name=FONTE, size=10, bold=True, color=_argb(LARANJA_SOLAR))
FONT_RODAPE = Font(name=FONTE, size=8, italic=True, color=_argb(MARROM_CAFE))

# ── Bordas / alinhamento ─────────────────────────────────────────────────
_hair = Side(style="thin", color=_argb(BORDA_CLARA))
BORDA = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_IND = Alignment(horizontal="left", vertical="center", indent=1)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Formatos numéricos padrão ────────────────────────────────────────────
MONEY_FMT = "[$R$-pt-BR] #,##0.00"
INT_FMT = "#,##0"
PCT_FMT = "0.0%"
DATE_FMT = "DD/MM/YYYY"
DATETIME_FMT = "DD/MM/YYYY HH:MM"

TABLE_STYLE_NAME = "TableStyleMedium9"


def estilizar_titulo(ws, row, num_cols, altura=28):
    """Aplica o preenchimento/fonte de título na linha `row`, colunas 1..num_cols."""
    ws.row_dimensions[row].height = altura
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_TITULO
        cell.font = FONT_TITULO
        cell.alignment = ALIGN_LEFT_IND if c == 1 else ALIGN_LEFT


def estilizar_subtitulo(ws, row, num_cols):
    """Aplica o preenchimento/fonte de subtítulo (metadados: filtros, data) na linha `row`."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_SUBTITULO
        cell.font = FONT_SUBTITULO
        cell.alignment = ALIGN_LEFT_IND if c == 1 else ALIGN_LEFT


def estilizar_cabecalho(ws, row, num_cols, wrap=True):
    """Aplica o preenchimento/fonte/borda de cabeçalho de tabela na linha `row`."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_WRAP if wrap else ALIGN_CENTER
        cell.border = BORDA


def aplicar_zebra_e_borda(ws, primeira_linha, ultima_linha, num_cols, zebra_a_cada=2):
    """Aplica borda em todas as células e zebra nas linhas alternadas do intervalo."""
    for r in range(primeira_linha, ultima_linha + 1):
        zebra = (r - primeira_linha) % zebra_a_cada == 1
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDA
            if not cell.font or cell.font.name != FONTE:
                cell.font = FONT_CELL
            if zebra and cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = FILL_ZEBRA


def criar_tabela(ws, nome, primeira_linha, ultima_linha, num_cols):
    """Registra um Excel Table nativo cobrindo o intervalo — habilita filtro/ordenação no Excel."""
    from openpyxl.utils import get_column_letter

    ref = f"A{primeira_linha}:{get_column_letter(num_cols)}{ultima_linha}"
    table = Table(displayName=nome, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE_NAME, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    return table


def ajustar_larguras(ws, larguras):
    """`larguras`: lista de ints, uma por coluna (A, B, C...)."""
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def rodape(ws, row, texto, num_cols=1):
    ws.cell(row=row, column=1, value=texto)
    ws.cell(row=row, column=1).font = FONT_RODAPE
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
