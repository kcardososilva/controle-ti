import datetime
import re
import unicodedata
from dataclasses import dataclass, field
from difflib import SequenceMatcher

from django.core.exceptions import FieldDoesNotExist
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from ProjetoEstoque.models import (
    Usuario,
    CentroCusto,
    Localidade,
    Funcao,
    StatusUsuarioChoices,
)


MESES_PT = {
    "jan": 1,
    "janeiro": 1,
    "fev": 2,
    "fevereiro": 2,
    "mar": 3,
    "marco": 3,
    "março": 3,
    "abr": 4,
    "abril": 4,
    "mai": 5,
    "maio": 5,
    "jun": 6,
    "junho": 6,
    "jul": 7,
    "julho": 7,
    "ago": 8,
    "agosto": 8,
    "set": 9,
    "setembro": 9,
    "out": 10,
    "outubro": 10,
    "nov": 11,
    "novembro": 11,
    "dez": 12,
    "dezembro": 12,
}


def model_has_field(model, field_name):
    try:
        model._meta.get_field(field_name)
        return True
    except FieldDoesNotExist:
        return False


def normalizar_texto(valor):
    if valor is None:
        return ""

    valor = str(valor).strip().lower()
    valor = unicodedata.normalize("NFKD", valor)
    valor = "".join(c for c in valor if not unicodedata.combining(c))
    valor = re.sub(r"[^a-z0-9\s]", " ", valor)
    valor = re.sub(r"\s+", " ", valor).strip()

    return valor


def normalizar_matricula(valor):
    if valor is None:
        return None

    valor = str(valor).strip()

    if not valor:
        return None

    if valor.endswith(".0"):
        valor = valor[:-2]

    valor = re.sub(r"\s+", "", valor)

    return valor or None


def parse_excel_date(valor):
    if not valor:
        return None

    if isinstance(valor, datetime.datetime):
        return valor.date()

    if isinstance(valor, datetime.date):
        return valor

    if isinstance(valor, (int, float)):
        try:
            return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(valor))
        except Exception:
            return None

    texto = str(valor).strip()

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(texto, fmt).date()
        except ValueError:
            continue

    return None


def sim_nao(valor):
    texto = normalizar_texto(valor)

    if texto in ("sim", "s", "yes", "y", "1", "verdadeiro", "true"):
        return "sim"

    return "nao"


def status_value(tipo):
    choices = dict(StatusUsuarioChoices.choices)
    values = list(choices.keys())

    if tipo == "ativo":
        for candidato in ("ativo", "ATIVO", "A"):
            if candidato in values:
                return candidato

    if tipo == "desligado":
        for candidato in ("desligado", "inativo", "INATIVO", "D", "I"):
            if candidato in values:
                return candidato

    return values[0] if values else tipo


def is_status_desligado(valor, data_desligamento=None):
    texto = normalizar_texto(valor)

    if data_desligamento:
        return True

    return texto in (
        "desligado",
        "desligada",
        "inativo",
        "inativa",
        "demitido",
        "demitida",
        "encerrado",
        "encerrada",
        "rescindido",
        "rescindida",
        "desligamento",
    )


def gerar_email_base(nome):
    partes = normalizar_texto(nome).split()

    if not partes:
        return None

    primeiro = partes[0]
    ultimo = partes[-1] if len(partes) > 1 else partes[0]

    return f"{primeiro}.{ultimo}@santacolomba.com"


def gerar_email_usuario(nome, matricula=None, usuario_id=None):
    email = gerar_email_base(nome)

    if not email:
        return None

    qs = Usuario.objects.filter(email__iexact=email)

    if usuario_id:
        qs = qs.exclude(pk=usuario_id)

    if not qs.exists():
        return email

    if matricula:
        partes = email.split("@")
        email_com_matricula = f"{partes[0]}.{matricula}@{partes[1]}"

        qs = Usuario.objects.filter(email__iexact=email_com_matricula)

        if usuario_id:
            qs = qs.exclude(pk=usuario_id)

        if not qs.exists():
            return email_com_matricula

    return email


def nome_aba_para_data(nome_aba):
    texto = normalizar_texto(nome_aba)
    partes = texto.split()

    mes = None
    ano = None

    for parte in partes:
        if parte in MESES_PT:
            mes = MESES_PT[parte]

        if parte.isdigit() and len(parte) == 4:
            ano = int(parte)

    if mes and ano:
        return datetime.date(ano, mes, 1)

    return None


def selecionar_abas(workbook, modo_importacao="ultima_aba", nome_aba=None):
    abas = workbook.sheetnames

    if modo_importacao == "todas_abas":
        return abas

    if modo_importacao == "aba_especifica":
        if nome_aba and nome_aba in abas:
            return [nome_aba]

        raise ValueError(f"A aba '{nome_aba}' não foi encontrada na planilha.")

    abas_com_data = []

    for aba in abas:
        data_ref = nome_aba_para_data(aba)
        if data_ref:
            abas_com_data.append((data_ref, aba))

    if abas_com_data:
        abas_com_data.sort(key=lambda item: item[0])
        return [abas_com_data[-1][1]]

    return [workbook.active.title]


def score_nome(a, b):
    return SequenceMatcher(None, normalizar_texto(a), normalizar_texto(b)).ratio()


def encontrar_usuario_por_nome(nome, limite=0.92):
    if not nome:
        return None

    direto = Usuario.objects.filter(nome__iexact=nome).first()

    if direto:
        return direto

    melhor = None
    melhor_score = 0

    for usuario in Usuario.objects.all().only("id", "nome", "matricula"):
        score = score_nome(nome, usuario.nome)

        if score > melhor_score:
            melhor = usuario
            melhor_score = score

    if melhor and melhor_score >= limite:
        return melhor

    return None


def buscar_ou_criar_funcao(nome):
    if not nome:
        return None

    nome = str(nome).strip()

    if not nome:
        return None

    obj = Funcao.objects.filter(nome__iexact=nome).first()

    if obj:
        return obj

    return Funcao.objects.create(nome=nome)


def buscar_ou_criar_localidade(nome):
    if not nome:
        return None

    nome = str(nome).strip()

    if not nome:
        return None

    if model_has_field(Localidade, "local"):
        obj = Localidade.objects.filter(local__iexact=nome).first()

        if obj:
            return obj

        try:
            return Localidade.objects.create(local=nome)
        except IntegrityError:
            return Localidade.objects.filter(local__icontains=nome).first()

    if model_has_field(Localidade, "nome"):
        obj = Localidade.objects.filter(nome__iexact=nome).first()

        if obj:
            return obj

        try:
            return Localidade.objects.create(nome=nome)
        except IntegrityError:
            return Localidade.objects.filter(nome__icontains=nome).first()

    return None


def buscar_ou_criar_centro_custo(numero, descricao):
    numero = normalizar_matricula(numero)
    descricao = str(descricao or "").strip()

    if not numero and not descricao:
        return None

    qs = CentroCusto.objects.all()

    if numero and model_has_field(CentroCusto, "numero"):
        obj = qs.filter(numero__iexact=numero).first()

        if obj:
            return obj

    if descricao and model_has_field(CentroCusto, "departamento"):
        obj = qs.filter(departamento__iexact=descricao).first()

        if obj:
            return obj

    data = {}

    if model_has_field(CentroCusto, "numero"):
        data["numero"] = numero or descricao[:20]

    if model_has_field(CentroCusto, "departamento"):
        data["departamento"] = descricao or numero

    if not data:
        return None

    try:
        return CentroCusto.objects.create(**data)
    except Exception:
        if numero and model_has_field(CentroCusto, "numero"):
            return CentroCusto.objects.filter(numero__icontains=numero).first()

        if descricao and model_has_field(CentroCusto, "departamento"):
            return CentroCusto.objects.filter(departamento__icontains=descricao).first()

    return None

def centro_custo_eh_tabaco(centro_custo):
    """
    Regra de negócio:
    Todo usuário vinculado a centro de custo relacionado a TABACO
    deve ser marcado automaticamente como PMB = sim.
    """

    if not centro_custo:
        return False

    campos_para_validar = [
        getattr(centro_custo, "numero", ""),
        getattr(centro_custo, "codigo", ""),
        getattr(centro_custo, "departamento", ""),
        getattr(centro_custo, "nome", ""),
        str(centro_custo),
    ]

    texto = " ".join(str(campo or "") for campo in campos_para_validar)
    texto_normalizado = normalizar_texto(texto)

    return "tabaco" in texto_normalizado


@dataclass
class ResultadoImportacao:
    abas_processadas: list = field(default_factory=list)
    criados: list = field(default_factory=list)
    atualizados: list = field(default_factory=list)
    desligados: list = field(default_factory=list)
    ignorados: list = field(default_factory=list)
    erros: list = field(default_factory=list)

    def as_dict(self):
        return {
            "abas_processadas": self.abas_processadas,
            "totais": {
                "criados": len(self.criados),
                "atualizados": len(self.atualizados),
                "desligados": len(self.desligados),
                "ignorados": len(self.ignorados),
                "erros": len(self.erros),
            },
            "criados": self.criados,
            "atualizados": self.atualizados,
            "desligados": self.desligados,
            "ignorados": self.ignorados,
            "erros": self.erros,
        }


class UsuarioImportService:
    # Diretor Geral da empresa — fixo para todos os colaboradores.
    # Quando a planilha não possui a coluna "Diretor Geral", esse valor é
    # atribuído automaticamente a todos os registros importados.
    DIRETOR_GERAL_PADRAO = "MIGUEL PRADO"

    # Estrutura hierárquica real da empresa:
    #   Diretor Geral → Diretor → Gestor → Coordenador → Supervisor → Colaborador
    #
    # A planilha atual do RH (Relação Funcionários) possui apenas as colunas:
    #   Gestor | Coordenador | Supervisor | Junção (imediato consolidado)
    # Os níveis Diretor Geral e Diretor não fazem parte desta planilha;
    # Diretor Geral é sempre MIGUEL PRADO (via DIRETOR_GERAL_PADRAO acima).
    # Diretor será preenchido automaticamente quando a planilha incluir essa coluna.
    COLUNAS_RH = {
        "matricula":              ["matrícula", "matricula"],
        "nome":                   ["nome"],
        "status":                 ["observação", "observacao", "status"],
        "data_inicio":            ["data admissão", "data admissao"],
        "funcao":                 ["cargo básico-descrição", "cargo basico-descricao",
                                   "cargo básico descrição", "cargo basico descricao"],
        "centro_custo_numero":    ["centro custo", "centro de custo"],
        "centro_custo_descricao": ["centro custo-descrição", "centro custo-descricao",
                                   "centro custo descrição", "centro custo descricao"],
        "estabelecimento":        ["estabelecimento-descrição", "estabelecimento-descricao",
                                   "estabelecimento descrição", "estabelecimento descricao"],
        "data_termino":           ["data desligamento", "data demissão", "data demissao"],
        "localidade":             ["unid lotação-descrição", "unid lotacao-descricao",
                                   "unid lotação descrição", "unid lotacao descricao",
                                   "unidade lotação descrição", "unidade lotacao descricao"],
        "email":                  ["email", "e-mail"],
        "pmb":                    ["pmb"],
        # ── Hierarquia organizacional ──
        "diretor_geral": ["diretor geral", "diretor-geral", "diretoria geral"],
        "diretor":       ["diretor", "diretoria", "outros diretores"],
        "gestor":        ["gestor", "gerente"],
        "coordenador":   ["coordenador"],
        "supervisor":    ["supervisor"],
        # "Junção - Gestor, Coordenador e Supervisor" é o imediato consolidado da planilha atual
        "responsavel": [
            "junção - gestor, coordenador e supervisor",
            "juncao gestor coordenador e supervisor",
            "junção gestor coordenador e supervisor",
            "junção",
            "juncao",
            "responsável",
            "responsavel",
        ],
    }

    def __init__(
        self,
        arquivo,
        user,
        modo_importacao="ultima_aba",
        nome_aba=None,
        desligar_ausentes=False,
    ):
        self.arquivo = arquivo
        self.user = user
        self.modo_importacao = modo_importacao or "ultima_aba"
        self.nome_aba = nome_aba
        self.desligar_ausentes = desligar_ausentes

        self.resultado = ResultadoImportacao()
        self.status_ativo = status_value("ativo")
        self.status_desligado = status_value("desligado")

        self.matriculas_processadas = set()
        self.nomes_processados = set()

    def executar(self):
        wb = load_workbook(self.arquivo, data_only=True)
        abas = selecionar_abas(wb, self.modo_importacao, self.nome_aba)

        for nome_aba in abas:
            ws = wb[nome_aba]
            self.resultado.abas_processadas.append(nome_aba)

            headers = self._mapear_headers(ws)

            if "nome" not in headers:
                self.resultado.erros.append(
                    f"Aba '{nome_aba}' ignorada: coluna Nome não encontrada."
                )
                continue

            # Passagem 1: carrega todas as linhas em memória.
            # Necessário para construir o índice de nomes antes de processar.
            linhas = []
            for row_idx in range(2, ws.max_row + 1):
                dados = self._ler_linha(ws, row_idx, headers)
                if any(dados.values()):
                    linhas.append((row_idx, dados))

            # Índice nome_completo → dados_da_linha, usado para resolver a
            # cadeia hierárquica (Gestor → Diretor → Diretor Geral). Construído
            # inteiramente a partir da planilha em memória — não depende de
            # nenhuma linha já ter sido gravada no banco, então processar cada
            # linha em sua própria transação não afeta a resolução da hierarquia.
            indice_nomes = self._construir_indice_nomes(linhas)

            # Passagem 2: processa cada linha com hierarquia completamente resolvida.
            # Uma transação por linha (savepoint curto): mantém o import resiliente
            # a erro pontual sem abortar as linhas seguintes e evita segurar o lock
            # de escrita do SQLite pela duração da planilha inteira — mesmo padrão
            # de importador_planilha.py (import de equipamentos).
            for row_idx, dados in linhas:
                try:
                    with transaction.atomic():
                        self._processar_linha(nome_aba, row_idx, dados, indice_nomes)
                except Exception as exc:
                    self.resultado.erros.append(f"Aba '{nome_aba}', linha {row_idx}: {exc}")
                    self.resultado.ignorados.append({
                        "aba": nome_aba,
                        "linha": row_idx,
                        "motivo": str(exc),
                    })

        if self.desligar_ausentes:
            self._desligar_ausentes()

        return self.resultado.as_dict()

    def _mapear_headers(self, ws):
        headers = {}

        primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        for idx, coluna in enumerate(primeira_linha, start=1):
            coluna_norm = normalizar_texto(coluna)

            if not coluna_norm:
                continue

            for campo, aliases in self.COLUNAS_RH.items():
                aliases_norm = [normalizar_texto(alias) for alias in aliases]

                if coluna_norm in aliases_norm:
                    headers[campo] = idx
                    break

        return headers

    def _ler_linha(self, ws, row_idx, headers):
        dados = {}

        for campo, col_idx in headers.items():
            dados[campo] = ws.cell(row=row_idx, column=col_idx).value

        return dados

    def _construir_indice_nomes(self, linhas):
        """
        Monta um dicionário { nome_normalizado → dados_da_linha } para todas as
        linhas da planilha. Usado para resolver nomes abreviados nas colunas de
        hierarquia (Gestor, Coordenador, Supervisor, Junção) de volta ao
        colaborador correspondente e percorrer a cadeia de liderança.
        """
        indice = {}
        for _, dados in linhas:
            nome = str(dados.get("nome") or "").strip()
            if nome:
                indice[normalizar_texto(nome)] = dados
        return indice

    def _resolver_nome_no_indice(self, nome_curto, indice):
        """
        Localiza no índice o colaborador referenciado por um nome abreviado
        (formato usado nas colunas Gestor/Coordenador/Supervisor/Junção).

        Estratégias aplicadas em ordem de confiabilidade:
          1. Match exato (nome normalizado igual).
          2. Primeiro e último palavra do abrev. == primeiro e último do nome completo.
          3. Todas as palavras do abrev. contidas (exatas) no nome completo.
          4. Todas as palavras do abrev. contidas com similaridade ≥ 0.75 por palavra
             (cobre variações ortográficas: ACASSIO↔ACACIO, TALLES↔THALES).
          5. SequenceMatcher ≥ 0.72 sobre o nome inteiro (fallback geral).

        Compostos com "/" (ex: "JUNIO TEIXEIRA /EDIVILSON") usam apenas o
        primeiro segmento para a resolução.

        Retorna None se nenhuma estratégia encontrar candidato confiável.
        """
        if not nome_curto:
            return None

        # Compostos: usar apenas o primeiro nome do conjunto
        segmento = str(nome_curto).strip().split("/")[0].strip()
        alvo = normalizar_texto(segmento)
        if not alvo:
            return None

        # 1) Match exato
        if alvo in indice:
            return indice[alvo]

        partes = alvo.split()

        # 2) Primeiro + último palavra
        if len(partes) >= 2:
            pri, ult = partes[0], partes[-1]
            cands = [
                (k, v) for k, v in indice.items()
                if k.split() and k.split()[0] == pri and k.split()[-1] == ult
            ]
            if len(cands) == 1:
                return cands[0][1]
            if len(cands) > 1:
                melhor = max(cands, key=lambda x: SequenceMatcher(None, alvo, x[0]).ratio())
                return melhor[1]

        # 3) Todas as palavras do abrev. presentes (exatas) no nome completo
        cands = [
            (k, v) for k, v in indice.items()
            if all(p in k.split() for p in partes)
        ]
        if len(cands) == 1:
            return cands[0][1]
        if len(cands) > 1:
            melhor = max(cands, key=lambda x: SequenceMatcher(None, alvo, x[0]).ratio())
            return melhor[1]

        # 4) Palavras do abrev. com correspondência fuzzy por palavra (≥ 0.75)
        #    Cobre variações ortográficas como ACASSIO↔ACACIO, TALLES↔THALES
        def _palavras_batem(k_norm):
            k_partes = k_norm.split()
            return all(
                any(SequenceMatcher(None, p, kp).ratio() >= 0.75 for kp in k_partes)
                for p in partes
            )

        cands = [(k, v) for k, v in indice.items() if _palavras_batem(k)]
        if len(cands) == 1:
            return cands[0][1]
        if len(cands) > 1:
            melhor = max(cands, key=lambda x: SequenceMatcher(None, alvo, x[0]).ratio())
            return melhor[1]

        # 5) SequenceMatcher global (fallback; threshold alto para evitar falsos positivos)
        melhor_score, melhor_dados = 0, None
        for k, v in indice.items():
            s = SequenceMatcher(None, alvo, k).ratio()
            if s > melhor_score:
                melhor_score, melhor_dados = s, v
        if melhor_dados and melhor_score >= 0.72:
            return melhor_dados

        return None

    def _processar_linha(self, nome_aba, row_idx, dados, indice_nomes=None):
        nome = str(dados.get("nome") or "").strip()
        matricula = normalizar_matricula(dados.get("matricula"))

        if not nome:
            self.resultado.ignorados.append({
                "aba": nome_aba,
                "linha": row_idx,
                "motivo": "Linha sem nome.",
            })
            return

        nome_norm = normalizar_texto(nome)

        if matricula:
            self.matriculas_processadas.add(matricula)

        if nome_norm:
            self.nomes_processados.add(nome_norm)

        usuario = None

        if matricula:
            usuario = Usuario.objects.filter(matricula__iexact=matricula).first()

        if not usuario:
            usuario = encontrar_usuario_por_nome(nome)

        data_inicio = parse_excel_date(dados.get("data_inicio"))
        data_termino = parse_excel_date(dados.get("data_termino"))
        status_planilha = dados.get("status")
        desligado = is_status_desligado(status_planilha, data_termino)

        centro_custo = buscar_ou_criar_centro_custo(
            dados.get("centro_custo_numero"),
            dados.get("centro_custo_descricao"),
        )

        localidade_nome = dados.get("localidade") or dados.get("estabelecimento")
        localidade = buscar_ou_criar_localidade(localidade_nome)
        funcao = buscar_ou_criar_funcao(dados.get("funcao"))

        # Regra automática de PMB:
        # Todo funcionário pertencente ao centro de custo TABACO recebe PMB = sim.
        if centro_custo_eh_tabaco(centro_custo):
            pmb_valor = "sim"
        else:
            pmb_valor = sim_nao(dados.get("pmb"))

        email_planilha = str(dados.get("email") or "").strip() or None

        if usuario:
            email = email_planilha or usuario.email or gerar_email_usuario(
                nome=nome,
                matricula=matricula,
                usuario_id=usuario.pk,
            )
        else:
            email = email_planilha or gerar_email_usuario(
                nome=nome,
                matricula=matricula,
            )

        # ── Hierarquia: normaliza cada campo (strip; None se vazio ou só traço) ──
        def _limpar_nome(v):
            v = str(v or "").strip()
            return v if v and v not in ("-", "–", "—") else None

        diretor_geral_val = _limpar_nome(dados.get("diretor_geral"))
        diretor_val       = _limpar_nome(dados.get("diretor"))
        gestor_val        = _limpar_nome(dados.get("gestor"))
        coordenador_val   = _limpar_nome(dados.get("coordenador"))
        supervisor_val    = _limpar_nome(dados.get("supervisor"))

        # ── Normalização de duplicatas nas colunas de hierarquia ─────────────
        # Regra 1: Coordenador == Supervisor (mesmo nome).
        #   A planilha preenche os dois campos com o mesmo nome quando a área
        #   não tem Coordenador intermediário. Nesse caso o campo Coordenador
        #   é redundante; a pessoa age somente como Supervisor.
        if coordenador_val and supervisor_val:
            if normalizar_texto(coordenador_val) == normalizar_texto(supervisor_val):
                coordenador_val = None

        # Regra 2: Coordenador sem Supervisor + Coordenador == Junção (imediato).
        #   Quando o mesmo nome aparece em Coordenador e na coluna Junção, mas
        #   o campo Supervisor está vazio, a pessoa é o Supervisor direto
        #   (não há nível de Coordenador na cadeia).
        responsavel_raw = _limpar_nome(dados.get("responsavel"))
        if coordenador_val and not supervisor_val and responsavel_raw:
            if normalizar_texto(coordenador_val) == normalizar_texto(responsavel_raw):
                supervisor_val  = coordenador_val
                coordenador_val = None

        # ── Resolução da cadeia hierárquica ──────────────────────────────────
        # Todos os colaboradores da planilha (inclusive Gestores, Diretores e o
        # Diretor Geral) possuem um imediato registrado na coluna Gestor.
        # A cadeia é:  colaborador → Gestor → Diretor → Diretor Geral (auto-ref)
        #
        # Algoritmo: percorre a cadeia gestor → gestor-do-gestor → ... até
        # encontrar auto-referência (quem é seu próprio gestor = raiz/DG) ou
        # nome não encontrado na planilha.
        #
        # Resultado:
        #   chain[-1] = Diretor Geral   (topo da cadeia)
        #   chain[-2] = Diretor         (nível imediatamente abaixo do DG)
        #
        # Quando a planilha tiver colunas explícitas de Diretor/DG, os valores
        # lidos diretamente prevalecem e a travessia é pulada.
        if indice_nomes and gestor_val and not (diretor_val and diretor_geral_val):
            chain = []
            current = gestor_val
            seen = set()

            while current:
                current_norm = normalizar_texto(current)
                if current_norm in seen:
                    break  # ciclo detectado — para aqui
                seen.add(current_norm)

                row = self._resolver_nome_no_indice(current, indice_nomes)
                chain.append(current)

                if not row:
                    break  # nome não encontrado na planilha — fim da cadeia

                next_g = _limpar_nome(row.get("gestor"))
                if next_g and normalizar_texto(next_g) == current_norm:
                    break  # auto-referência — este é o Diretor Geral (raiz)

                current = next_g

            # Atribui apenas os campos ainda não preenchidos pela planilha
            if chain and not diretor_geral_val:
                diretor_geral_val = chain[-1]
            if len(chain) >= 2 and not diretor_val:
                diretor_val = chain[-2]

        # Fallback: DG padrão quando a cadeia não consegue resolver o topo
        if not diretor_geral_val and self.DIRETOR_GERAL_PADRAO:
            diretor_geral_val = self.DIRETOR_GERAL_PADRAO

        # "Junção" = responsável imediato consolidado; se ausente, sobe a cadeia
        # hierárquica do mais específico para o mais geral.
        responsavel_val = (
            _limpar_nome(dados.get("responsavel"))
            or supervisor_val
            or coordenador_val
            or gestor_val
            or diretor_val
            or diretor_geral_val
        )

        payload = {
            "matricula": matricula,
            "nome": nome,
            "email": email,
            "status": self.status_desligado if desligado else self.status_ativo,
            "data_inicio": data_inicio,
            "data_termino": data_termino if desligado else None,
            "pmb": pmb_valor,
            "centro_custo": centro_custo,
            "localidade": localidade,
            "funcao": funcao,
            "diretor_geral": diretor_geral_val,
            "diretor": diretor_val,
            "gestor": gestor_val,
            "coordenador": coordenador_val,
            "supervisor": supervisor_val,
            "responsavel": responsavel_val,
        }

        if usuario:
            self._atualizar_usuario(usuario, payload, desligado, nome_aba, row_idx)
            return

        self._criar_usuario(payload, desligado, nome_aba, row_idx)

    def _atualizar_usuario(self, usuario, payload, desligado, nome_aba, row_idx):
        alterados = []

        for campo, valor in payload.items():
            if campo == "matricula" and not valor:
                continue

            if campo == "data_inicio" and not valor and usuario.data_inicio:
                continue

            if campo in ("centro_custo", "localidade", "funcao") and valor is None:
                continue

            if getattr(usuario, campo) != valor:
                setattr(usuario, campo, valor)
                alterados.append(campo)

        if hasattr(usuario, "atualizado_por"):
            usuario.atualizado_por = self.user
            alterados.append("atualizado_por")

        if alterados:
            usuario.save(update_fields=list(set(alterados)))

            destino = self.resultado.desligados if desligado else self.resultado.atualizados

            destino.append({
                "aba": nome_aba,
                "linha": row_idx,
                "nome": usuario.nome,
                "matricula": usuario.matricula or "—",
                "email": usuario.email or "—",
                "acao": "desligado" if desligado else "atualizado",
                "campos": ", ".join(sorted(set(alterados))),
            })
        else:
            self.resultado.ignorados.append({
                "aba": nome_aba,
                "linha": row_idx,
                "nome": usuario.nome,
                "matricula": usuario.matricula or "—",
                "motivo": "Sem alterações.",
            })

    def _criar_usuario(self, payload, desligado, nome_aba, row_idx):
        if not payload.get("data_inicio"):
            payload["data_inicio"] = timezone.localdate()

        novo = Usuario(**payload)

        if hasattr(novo, "criado_por"):
            novo.criado_por = self.user

        if hasattr(novo, "atualizado_por"):
            novo.atualizado_por = self.user

        novo.save()

        destino = self.resultado.desligados if desligado else self.resultado.criados

        destino.append({
            "aba": nome_aba,
            "linha": row_idx,
            "nome": novo.nome,
            "matricula": novo.matricula or "—",
            "email": novo.email or "—",
            "acao": "criado como desligado" if desligado else "criado",
        })

    def _desligar_ausentes(self):
        """
        Desliga colaboradores ativos ausentes na planilha.

        Regra:
        - COM matrícula → desligado automaticamente se a matrícula não aparece
          na planilha (o RH controla oficialmente pelo número de matrícula).
        - SEM matrícula → nunca desligado automaticamente; o desligamento deve
          ser feito manualmente, pois não há chave formal de verificação.
        """
        ativos = Usuario.objects.filter(status=self.status_ativo)

        for usuario in ativos:
            matricula = usuario.matricula
            nome_norm = normalizar_texto(usuario.nome)

            # Colaboradores SEM matrícula: não podem ser desligados
            # automaticamente — sem chave formal não há como confirmar ausência.
            if not matricula:
                continue

            # Colaboradores COM matrícula: verifica se a matrícula está na planilha.
            presente_por_matricula = matricula in self.matriculas_processadas

            # Fallback por nome normalizado caso a matrícula não tenha sido
            # importada nesta planilha mas o nome bata exatamente.
            presente_por_nome = nome_norm and nome_norm in self.nomes_processados

            if presente_por_matricula or presente_por_nome:
                continue

            try:
                with transaction.atomic():
                    usuario.status = self.status_desligado
                    usuario.data_termino = timezone.localdate()

                    update_fields = ["status", "data_termino"]

                    if hasattr(usuario, "atualizado_por"):
                        usuario.atualizado_por = self.user
                        update_fields.append("atualizado_por")

                    usuario.save(update_fields=update_fields)
            except Exception as exc:
                self.resultado.erros.append(
                    f"Desligamento automático de '{usuario.nome}': {exc}"
                )
                continue

            self.resultado.desligados.append({
                "aba": "ausente",
                "linha": "-",
                "nome": usuario.nome,
                "matricula": matricula,
                "email": usuario.email or "—",
                "acao": "desligado por ausência na planilha mensal",
            })